"""Build data/drivers.json from PSYWERX Driver Schema v1.1 workbooks.

The canonical spreadsheet contract is defined once in DRIVER_SCHEMA below.
Explicit IDs are preserved verbatim except for trimming surrounding whitespace.
The fallback ID helper is reserved for a future non-v1 import mode; Schema v1.1
requires an ID in every driver row.
"""

from __future__ import annotations

import json
import re
import sys
import tempfile
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook


ValueKind = Literal["scalar", "list"]


@dataclass(frozen=True)
class FieldSpec:
    json_name: str
    header: str
    kind: ValueKind
    required: bool


# Canonical PSYWERX Driver Schema v1.1. Header matching is intentionally exact
# after trimming/collapsing whitespace so taxonomy drift is reported.
DRIVER_SCHEMA: tuple[FieldSpec, ...] = (
    FieldSpec("id", "ID", "scalar", True),
    FieldSpec("name", "Name", "scalar", True),
    FieldSpec("aliases", "Other Names / Aliases", "list", False),
    FieldSpec("layer", "Layer", "scalar", True),
    FieldSpec("family", "Family", "scalar", True),
    FieldSpec("definition", "Definition", "scalar", True),
    FieldSpec("dataType", "Data Type", "scalar", True),
    FieldSpec("representationScale", "Representation / Scale", "scalar", True),
    FieldSpec(
        "polarityDirection",
        "Polarity / Direction of Interpretation",
        "scalar",
        True,
    ),
    FieldSpec("mechanism", "Mechanism", "scalar", True),
    FieldSpec(
        "likelyUpstreamInfluences",
        "Likely Upstream Influences",
        "list",
        False,
    ),
    FieldSpec(
        "likelyDownstreamInfluences",
        "Likely Downstream Influences",
        "list",
        False,
    ),
    FieldSpec(
        "moderatorsBoundaryConditions",
        "Moderators / Boundary Conditions",
        "scalar",
        True,
    ),
    FieldSpec(
        "typicalInteractionCandidates",
        "Typical Interaction Candidates",
        "list",
        False,
    ),
    FieldSpec(
        "modifiability",
        "Modifiability / Malleability",
        "scalar",
        True,
    ),
    FieldSpec("volatility", "Volatility", "scalar", True),
    FieldSpec("timeScaleOfChange", "Time Scale of Change", "list", True),
    FieldSpec("timeScaleQualifier", "Time Scale Qualifier", "scalar", False),
    FieldSpec("onsetCausalLag", "Onset / Causal Lag", "list", True),
    FieldSpec(
        "persistenceRecovery",
        "Persistence / Recovery",
        "scalar",
        True,
    ),
    FieldSpec("indicators", "Indicators", "list", False),
    FieldSpec(
        "measurementAssessmentMethods",
        "Measurement / Assessment Methods",
        "scalar",
        True,
    ),
    FieldSpec("observability", "Observability", "scalar", True),
    FieldSpec(
        "measurementCaveats",
        "Measurement Caveats",
        "scalar",
        True,
    ),
    FieldSpec("evidenceStrength", "Evidence Strength", "scalar", True),
    FieldSpec("evidenceNotes", "Evidence Notes", "scalar", True),
    FieldSpec(
        "commonMisinterpretations",
        "Common Misinterpretations",
        "scalar",
        True,
    ),
    FieldSpec("keySources", "Key Sources", "list", True),
)

CANONICAL_TEMPORAL_BANDS = (
    "Seconds–Minutes",
    "Minutes–Hours",
    "Hours–Days",
    "Days–Weeks",
    "Weeks–Months",
    "Months–Years",
    "Years–Generations",
)
MIXED_TEMPORAL_VALUE = "Mixed / Context-dependent"
STABLE_TEMPORAL_VALUE = "Stable / Not applicable"
TIME_SCALE_VALUES = CANONICAL_TEMPORAL_BANDS + (
    MIXED_TEMPORAL_VALUE,
    STABLE_TEMPORAL_VALUE,
)
ONSET_CAUSAL_LAG_VALUES = CANONICAL_TEMPORAL_BANDS + (
    MIXED_TEMPORAL_VALUE,
)
TEMPORAL_FIELDS = {"timeScaleOfChange", "onsetCausalLag"}

CANONICAL_LAYERS = (
    "Biological",
    "Psychological",
    "Social",
    "Cultural",
    "Physical / Environmental",
    "Institutional / Structural",
    "Informational",
    "Technological",
)
LAYER_ALIASES = {
    "biological": "Biological",
    "psychological": "Psychological",
    "social": "Social",
    "cultural": "Cultural",
    "physical": "Physical / Environmental",
    "environmental": "Physical / Environmental",
    "physical environmental": "Physical / Environmental",
    "institutional": "Institutional / Structural",
    "structural": "Institutional / Structural",
    "institutional structural": "Institutional / Structural",
    "informational": "Informational",
    "information": "Informational",
    "technological": "Technological",
    "technology": "Technological",
}
FILENAME_LAYER_PATTERNS = (
    (r"(?:^|_)layer_1_biological(?:_|$)", "Biological"),
    (r"(?:^|_)layer_2_psychological(?:_|$)", "Psychological"),
    (r"(?:^|_)layer_3_social(?:_|$)", "Social"),
    (r"(?:^|_)layer_4_cultural(?:_|$)", "Cultural"),
    (
        r"(?:^|_)layer_5_physical_environmental(?:_|$)",
        "Physical / Environmental",
    ),
    (
        r"(?:^|_)layer_6_institutional_structural(?:_|$)",
        "Institutional / Structural",
    ),
    (r"(?:^|_)layer_7_informational(?:_|$)", "Informational"),
    (r"(?:^|_)layer_8_technological(?:_|$)", "Technological"),
)
OUTPUT_KEYS = tuple(spec.json_name for spec in DRIVER_SCHEMA) + ("source",)
HEADER_SCAN_ROWS = 25


@dataclass
class Summary:
    workbooks: int = 0
    driver_sheets: int = 0
    rows_skipped: int = 0
    skipped_worksheets: list[str] = field(default_factory=list)
    drivers_per_layer: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def clean(value: Any) -> str:
    return "" if value is None else re.sub(r"\s+", " ", str(value)).strip()


def clean_id(value: Any) -> str:
    # Explicit IDs are immutable identifiers: do not case-fold, slug, or alter
    # whitespace. Validation still rejects an empty or whitespace-only value.
    return "" if value is None else str(value)


def normalized_key(value: Any) -> str:
    text = unicodedata.normalize("NFKC", clean(value)).casefold()
    return re.sub(r"[^\w]+", " ", text, flags=re.UNICODE).strip()


def normalize_layer(value: Any) -> str | None:
    return LAYER_ALIASES.get(normalized_key(value))


def mentioned_layers(value: Any) -> set[str]:
    normalized = f" {normalized_key(value)} "
    matches: set[str] = set()
    for compound in ("physical environmental", "institutional structural"):
        if f" {compound} " in normalized:
            matches.add(LAYER_ALIASES[compound])
            normalized = normalized.replace(f" {compound} ", " ")
    for alias, canonical in LAYER_ALIASES.items():
        if " " not in alias and f" {alias} " in normalized:
            matches.add(canonical)
    return matches


def filename_layer(path: Path) -> str | None:
    """Read a layer only from the canonical PSYWERX filename segment."""
    filename = path.stem.casefold()
    matches = {
        layer
        for pattern, layer in FILENAME_LAYER_PATTERNS
        if re.search(pattern, filename)
    }
    return next(iter(matches)) if len(matches) == 1 else None


def declared_title_layer(value: Any) -> str | None:
    """Recognize a layer only when a designated title cell declares it."""
    title = normalized_key(value)
    if not title:
        return None
    for canonical in CANONICAL_LAYERS:
        label = normalized_key(canonical)
        declarations = {
            f"canonical {label} drivers",
            f"{label} drivers",
            f"psywerx {label} driver ontology",
        }
        if title in declarations:
            return canonical
    return None


def split_list(value: Any) -> list[str]:
    if value is None or not str(value).strip():
        return []
    # Schema v1.1 prose-list cells are semicolon-separated. Newlines and pipes are
    # accepted as unambiguous alternates; commas remain part of the value.
    return [
        item
        for item in (clean(part) for part in re.split(r"[;|\r\n]+", str(value)))
        if item
    ]


def split_temporal_list(value: Any) -> list[str]:
    """Split controlled temporal values only on the canonical semicolon delimiter."""
    if value is None or not str(value).strip():
        return []
    return [item for item in (clean(part) for part in str(value).split(";")) if item]


def validate_temporal_values(
    values: list[str],
    field_name: str,
    location: str,
    summary: Summary,
) -> None:
    if field_name == "timeScaleOfChange":
        allowed = TIME_SCALE_VALUES
        exclusive = (MIXED_TEMPORAL_VALUE, STABLE_TEMPORAL_VALUE)
        label = "Time Scale of Change"
    else:
        allowed = ONSET_CAUSAL_LAG_VALUES
        exclusive = (MIXED_TEMPORAL_VALUE,)
        label = "Onset / Causal Lag"

    invalid = [value for value in values if value not in allowed]
    if invalid:
        summary.errors.append(
            f"{location}: {label} contains non-canonical value(s): "
            + ", ".join(repr(value) for value in invalid)
        )

    duplicates = sorted({value for value in values if values.count(value) > 1})
    if duplicates:
        summary.errors.append(
            f"{location}: {label} contains duplicate value(s): "
            + ", ".join(repr(value) for value in duplicates)
        )

    for value in exclusive:
        if value in values and len(values) != 1:
            summary.errors.append(
                f"{location}: {value!r} must be exclusive in {label}."
            )

    ordered_bands = [
        CANONICAL_TEMPORAL_BANDS.index(value)
        for value in values
        if value in CANONICAL_TEMPORAL_BANDS
    ]
    if ordered_bands != sorted(ordered_bands):
        summary.errors.append(
            f"{location}: {label} values must be ordered shortest to longest."
        )


def fallback_id(layer: str, name: str) -> str:
    """Return a deterministic future-format fallback; unused by strict v1.1."""
    value = f"{layer}-{name}"
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", ascii_value.casefold()).strip("-") or "driver"


def row_values(row: tuple[Any, ...]) -> list[str]:
    return [clean(value) for value in row if clean(value)]


def find_driver_sheets(workbook: Any) -> list[Any]:
    named = [
        sheet for sheet in workbook.worksheets
        if normalized_key(sheet.title) == "drivers"
    ]
    if named:
        return named
    candidates = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(
            min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True
        ):
            values = set(row_values(row))
            if {"ID", "Name"}.issubset(values):
                candidates.append(sheet)
                break
    return candidates


def find_header_row(sheet: Any) -> tuple[int, tuple[Any, ...]] | None:
    expected = {spec.header for spec in DRIVER_SCHEMA}
    best: tuple[int, tuple[Any, ...]] | None = None
    best_score = 0
    for number, row in enumerate(
        sheet.iter_rows(
            min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True
        ),
        start=1,
    ):
        score = len(expected.intersection(row_values(row)))
        if score > best_score:
            best = (number, row)
            best_score = score
    return best


def validate_headers(
    workbook_name: str,
    sheet_name: str,
    row: tuple[Any, ...],
    summary: Summary,
) -> tuple[dict[str, int], tuple[str, ...]]:
    location = f"{workbook_name} / {sheet_name}"
    actual = tuple(clean(value) for value in row if clean(value))
    counts: dict[str, int] = defaultdict(int)
    positions: dict[str, int] = {}
    for index, value in enumerate(row):
        header = clean(value)
        if header:
            counts[header] += 1
            positions.setdefault(header, index)
    for header, count in sorted(counts.items()):
        if count > 1:
            summary.errors.append(
                f"{location}: duplicate driver-table header {header!r}."
            )

    expected_headers = {spec.header for spec in DRIVER_SCHEMA}
    for spec in DRIVER_SCHEMA:
        if spec.header in positions:
            continue
        if spec.json_name == "layer":
            # Layer is a required output value but its column may be replaced
            # by unambiguous workbook/sheet/title inference.
            continue
        message = f"{location}: missing {'required' if spec.required else 'optional'} header {spec.header!r}."
        if spec.required:
            summary.errors.append(message)
        else:
            summary.warnings.append(message)
    for header in actual:
        if header not in expected_headers:
            summary.warnings.append(
                f"{location}: unexpected driver-table header {header!r}."
            )
    mapping = {
        spec.json_name: positions[spec.header]
        for spec in DRIVER_SCHEMA
        if spec.header in positions
    }
    return mapping, actual


def explicit_table_layer(
    path: Path,
    sheet: Any,
    mapping: dict[str, int],
    header_row: int,
    summary: Summary,
) -> str | None:
    """Return the one canonical layer consistently declared by data rows."""
    layer_index = mapping.get("layer")
    if layer_index is None:
        return None
    location = f"{path.name} / {sheet.title}"
    layers: set[str] = set()
    invalid: set[str] = set()
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row_values(row):
            continue
        raw = row[layer_index] if layer_index < len(row) else None
        if not clean(raw):
            continue
        canonical = normalize_layer(raw)
        if canonical is None:
            invalid.add(clean(raw))
        else:
            layers.add(canonical)
    if invalid:
        summary.errors.append(
            f"{location}: explicit Layer field contains unrecognized value(s): "
            + ", ".join(repr(value) for value in sorted(invalid))
        )
    if len(layers) > 1:
        summary.errors.append(
            f"{location}: explicit Layer field is inconsistent: "
            + ", ".join(sorted(layers))
        )
        return None
    return next(iter(layers)) if len(layers) == 1 else None


def authoritative_layer(
    path: Path,
    sheet: Any,
    workbook_title: Any,
    header_row: int,
    mapping: dict[str, int],
    summary: Summary,
) -> str | None:
    """Resolve layer identity using only ordered authoritative declarations."""
    location = f"{path.name} / {sheet.title}"
    worksheet_matches = mentioned_layers(sheet.title)
    if len(worksheet_matches) > 1:
        summary.errors.append(
            f"{location}: worksheet name declares conflicting layers: "
            + ", ".join(sorted(worksheet_matches))
        )
        return None

    title_cell_layer = declared_title_layer(sheet.cell(row=1, column=1).value)
    workbook_property_layer = declared_title_layer(workbook_title)
    if (
        title_cell_layer
        and workbook_property_layer
        and title_cell_layer != workbook_property_layer
    ):
        summary.errors.append(
            f"{location}: designated worksheet title and workbook title "
            "declare different layers."
        )
        return None
    title_layer = title_cell_layer or workbook_property_layer

    # Dict insertion order is the required precedence order.
    evidence = {
        "explicit Layer field": explicit_table_layer(
            path, sheet, mapping, header_row, summary
        ),
        "workbook filename": filename_layer(path),
        "designated title": title_layer,
        "worksheet name": (
            next(iter(worksheet_matches)) if worksheet_matches else None
        ),
    }
    declared = {name: layer for name, layer in evidence.items() if layer}
    distinct = set(declared.values())
    if len(distinct) > 1:
        details = ", ".join(
            f"{source_name}={layer}" for source_name, layer in declared.items()
        )
        summary.errors.append(
            f"{location}: conflicting authoritative layer declarations ({details})."
        )
        return None
    return next(iter(declared.values())) if declared else None


def get_cell(
    row: tuple[Any, ...], mapping: dict[str, int], json_name: str
) -> Any:
    index = mapping.get(json_name)
    return row[index] if index is not None and index < len(row) else None


def build_record(
    row: tuple[Any, ...],
    mapping: dict[str, int],
    inferred: str | None,
    path: Path,
    sheet_name: str,
    row_number: int,
    summary: Summary,
) -> dict[str, Any] | None:
    if not row_values(row):
        summary.rows_skipped += 1
        return None
    location = f"{path.name} / {sheet_name} / row {row_number}"
    record: dict[str, Any] = {}
    for spec in DRIVER_SCHEMA:
        value = get_cell(row, mapping, spec.json_name)
        if spec.json_name == "id":
            normalized: str | list[str] | None = clean_id(value) or None
        elif spec.json_name in TEMPORAL_FIELDS:
            normalized = split_temporal_list(value)
        elif spec.kind == "list":
            normalized = split_list(value)
        else:
            normalized = clean(value) or None
        record[spec.json_name] = normalized

    explicit_layer = normalize_layer(record["layer"]) if record["layer"] else None
    if record["layer"] and explicit_layer is None:
        summary.errors.append(
            f"{location}: unrecognized explicit layer {record['layer']!r}."
        )
    elif explicit_layer and inferred and explicit_layer != inferred:
        summary.errors.append(
            f"{location}: explicit layer {explicit_layer!r} conflicts with "
            f"inferred layer {inferred!r}."
        )
    record["layer"] = explicit_layer or inferred

    for spec in DRIVER_SCHEMA:
        value = record[spec.json_name]
        absent = value is None or value == []
        if spec.json_name == "id":
            absent = not clean(value)
        if spec.required and absent:
            summary.errors.append(
                f"{location}: required field {spec.header!r} is empty."
            )

    for field_name in TEMPORAL_FIELDS:
        validate_temporal_values(
            record[field_name], field_name, location, summary
        )

    record["source"] = {"workbook": path.name, "sheet": sheet_name}
    return record


def report_header_discrepancies(
    signatures: dict[str, tuple[str, ...]], summary: Summary
) -> None:
    if len(set(signatures.values())) <= 1:
        return
    baseline_name, baseline = next(iter(signatures.items()))
    baseline_set = set(baseline)
    for workbook_name, headers in list(signatures.items())[1:]:
        if headers == baseline:
            continue
        missing = [header for header in baseline if header not in headers]
        added = [header for header in headers if header not in baseline_set]
        order_changed = not missing and not added and headers != baseline
        summary.errors.append(
            f"Driver schema discrepancy: {workbook_name} differs from "
            f"{baseline_name}; missing={missing or 'none'}, "
            f"extra={added or 'none'}, order_changed={order_changed}."
        )


def read_workbooks(source_dir: Path, summary: Summary) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    signatures: dict[str, tuple[str, ...]] = {}
    paths = sorted(
        path for path in source_dir.glob("*")
        if path.is_file()
        and path.suffix.casefold() == ".xlsx"
        and not path.name.startswith("~$")
    )
    summary.workbooks = len(paths)
    if not paths:
        summary.errors.append(f"No XLSX workbooks found in {source_dir}.")
        return records

    for path in paths:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            summary.errors.append(f"{path.name}: could not open workbook: {exc}")
            continue
        try:
            driver_sheets = find_driver_sheets(workbook)
            if len(driver_sheets) != 1:
                summary.errors.append(
                    f"{path.name}: expected exactly one driver worksheet; "
                    f"found {len(driver_sheets)}."
                )
            driver_sheet_ids = {id(sheet) for sheet in driver_sheets}
            for sheet in workbook.worksheets:
                if id(sheet) not in driver_sheet_ids:
                    summary.skipped_worksheets.append(
                        f"{path.name} / {sheet.title}"
                    )
            for sheet in driver_sheets:
                header = find_header_row(sheet)
                if header is None:
                    summary.errors.append(
                        f"{path.name} / {sheet.title}: no driver header row found."
                    )
                    continue
                header_row, raw_headers = header
                mapping, signature = validate_headers(
                    path.name, sheet.title, raw_headers, summary
                )
                signatures[path.name] = signature
                layer = authoritative_layer(
                    path,
                    sheet,
                    workbook.properties.title,
                    header_row,
                    mapping,
                    summary,
                )
                summary.driver_sheets += 1
                for row_number, row in enumerate(
                    sheet.iter_rows(
                        min_row=header_row + 1, values_only=True
                    ),
                    start=header_row + 1,
                ):
                    record = build_record(
                        row,
                        mapping,
                        layer,
                        path,
                        sheet.title,
                        row_number,
                        summary,
                    )
                    if record is not None:
                        records.append(record)
        except Exception as exc:
            summary.errors.append(
                f"{path.name}: error while reading workbook: {exc}"
            )
        finally:
            workbook.close()
    report_header_discrepancies(signatures, summary)
    return records


def validate_records(
    records: list[dict[str, Any]], summary: Summary
) -> None:
    if not records:
        summary.errors.append("No valid driver records were produced.")
        return
    ids: dict[str, list[dict[str, Any]]] = defaultdict(list)
    layer_names: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    counts: dict[str, int] = defaultdict(int)
    scalar_fields = {
        spec.json_name for spec in DRIVER_SCHEMA if spec.kind == "scalar"
    }
    list_fields = {
        spec.json_name for spec in DRIVER_SCHEMA if spec.kind == "list"
    }
    for record in records:
        if tuple(record) != OUTPUT_KEYS:
            summary.errors.append(
                f"Record {record.get('id')!r} has a non-canonical key structure."
            )
        for field_name in scalar_fields:
            if record.get(field_name) is not None and not isinstance(
                record[field_name], str
            ):
                summary.errors.append(
                    f"Record {record.get('id')!r}: {field_name} must be string or null."
                )
        for field_name in list_fields:
            value = record.get(field_name)
            if not isinstance(value, list) or not all(
                isinstance(item, str) for item in value
            ):
                summary.errors.append(
                    f"Record {record.get('id')!r}: {field_name} must be a string list."
                )
        if record.get("layer") not in CANONICAL_LAYERS:
            summary.errors.append(
                f"Record {record.get('id')!r} has a non-canonical layer."
            )
            continue
        source = record.get("source")
        if (
            not isinstance(source, dict)
            or set(source) != {"workbook", "sheet"}
            or any("\\" in str(value) for value in source.values())
        ):
            summary.errors.append(
                f"Record {record.get('id')!r} has invalid public provenance."
            )
        driver_id = record.get("id")
        if driver_id:
            ids[driver_id].append(record)
        name = record.get("name")
        if name:
            layer_names[(record["layer"], normalized_key(name))].append(record)
        counts[record["layer"]] += 1

    for driver_id, matches in sorted(ids.items()):
        if len(matches) > 1:
            summary.errors.append(
                f"Duplicate Driver ID {driver_id!r} ({len(matches)} records)."
            )
    for (layer, name), matches in sorted(layer_names.items()):
        if len(matches) > 1:
            summary.warnings.append(
                f"Potential duplicate normalized name in {layer} {name!r}: "
                + ", ".join(record["id"] for record in matches)
            )
    summary.drivers_per_layer = {
        layer: counts[layer] for layer in CANONICAL_LAYERS if counts[layer]
    }


def write_atomically(records: list[dict[str, Any]], output: Path) -> None:
    content = json.dumps(records, ensure_ascii=False, indent=2) + "\n"
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            newline="\n",
            dir=output.parent,
            prefix=f".{output.name}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            handle.write(content)
            temporary = Path(handle.name)
        temporary.replace(output)
    finally:
        if temporary is not None and temporary.exists():
            temporary.unlink()


def print_summary(summary: Summary, output: Path, wrote: bool) -> None:
    print("Import statistics")
    print(f"  Workbooks: {summary.workbooks}")
    print(f"  Driver worksheets: {summary.driver_sheets}")
    print(f"  Layers: {len(summary.drivers_per_layer)}")
    for layer, count in summary.drivers_per_layer.items():
        print(f"  Drivers / {layer}: {count}")
    print(f"  Total drivers: {sum(summary.drivers_per_layer.values())}")
    print(f"  Rows skipped: {summary.rows_skipped}")
    print(f"  Warnings: {len(summary.warnings)}")
    print(f"  Errors: {len(summary.errors)}")
    print(
        "  Intentionally skipped worksheets: "
        f"{len(summary.skipped_worksheets)}"
    )
    for worksheet in summary.skipped_worksheets:
        print(f"SKIPPED: {worksheet}")
    for warning in summary.warnings:
        print(f"WARNING: {warning}")
    for error in summary.errors:
        print(f"ERROR: {error}", file=sys.stderr)
    if wrote:
        print(f"Wrote validated data to {output}.")
    else:
        print(f"Did not modify {output}.", file=sys.stderr)


def main() -> int:
    root = Path(__file__).resolve().parent.parent
    output = root / "data" / "drivers.json"
    summary = Summary()
    records = read_workbooks(root / "source-data", summary)
    validate_records(records, summary)
    records.sort(
        key=lambda record: (
            CANONICAL_LAYERS.index(record["layer"]),
            normalized_key(record["name"]),
            record["id"],
        )
    )
    wrote = False
    if not summary.errors:
        try:
            write_atomically(records, output)
            wrote = True
        except OSError as exc:
            summary.errors.append(f"Could not replace {output}: {exc}")
    print_summary(summary, output, wrote)
    return 0 if wrote else 1


if __name__ == "__main__":
    raise SystemExit(main())
