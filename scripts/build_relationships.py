"""Build public Relationship Schema v2 JSON from canonical XLSX workbooks.

The eight Relationships worksheets are authoritative. This importer validates
the complete graph in memory and atomically replaces the public artifact only
after every workbook, endpoint, evidence reference, and controlled value passes.
"""

from __future__ import annotations

import json
import re
import tempfile
from collections import Counter, defaultdict, deque
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCHEMA_VERSION = "2.0"
HEADER_ROW = 4
FIRST_DATA_ROW = 5
SHEET = "Relationships"
LAYERS = (
    "Biological", "Psychological", "Social", "Cultural",
    "Physical / Environmental", "Institutional / Structural",
    "Informational", "Technological",
)
LAYER_TOKENS = (
    ("layer_1_biological", "Biological"),
    ("layer_2_psychological", "Psychological"),
    ("layer_3_social", "Social"),
    ("layer_4_cultural", "Cultural"),
    ("layer_5_physical_environmental", "Physical / Environmental"),
    ("layer_6_institutional_structural", "Institutional / Structural"),
    ("layer_7_informational", "Informational"),
    ("layer_8_technological", "Technological"),
)
HEADERS = (
    "Relationship ID", "Source Driver ID", "Source Driver",
    "Target Driver ID", "Target Driver", "Causal Role", "Polarity",
    "Directness", "Mechanism", "Conditions / Moderators",
    "Moderator Driver IDs", "Source Level", "Target Level",
    "Level-Transition Mechanism", "Lag Profile", "Lag Lower Bound",
    "Lag Upper Bound", "Lag Unit", "Lag Narrative", "Exposure Pattern",
    "Effect Persistence", "Evidence Strength", "Confidence",
    "Generalizability / Context", "Reciprocal Process ID",
    "Governance Class", "Supporting Evidence IDs", "Notes / Caveats",
)
KEYS = (
    "id", "sourceDriverId", "sourceDriverName", "targetDriverId",
    "targetDriverName", "causalRole", "polarity", "directness",
    "mechanism", "conditionsModerators", "moderatorDriverIds",
    "sourceLevel", "targetLevel", "levelTransitionMechanism", "lagProfile",
    "lagLowerBound", "lagUpperBound", "lagUnit", "lagNarrative",
    "exposurePattern", "effectPersistence", "evidenceStrength", "confidence",
    "generalizabilityContext", "reciprocalProcessId", "governanceClass",
    "supportingEvidenceIds", "notesCaveats",
)
LIST_KEYS = {"moderatorDriverIds", "lagProfile", "supportingEvidenceIds"}
OPTIONAL = {
    "moderatorDriverIds", "levelTransitionMechanism", "lagLowerBound",
    "lagUpperBound", "lagUnit", "effectPersistence", "reciprocalProcessId",
    "notesCaveats",
}
LEVELS = {"PERSON", "DYAD_INTERPERSONAL", "SMALL_GROUP", "NETWORK", "COMMUNITY", "ORGANIZATION", "INSTITUTIONAL_FIELD", "SOCIETY", "STATE_JURISDICTION", "INFORMATION_OBJECT_CORPUS", "INFORMATION_SYSTEM", "TECHNOLOGICAL_SYSTEM", "PHYSICAL_SETTING", "ECOLOGICAL_SYSTEM"}
CONTROLLED = {
    "causalRole": {"CAUSES", "ENABLES", "CONSTRAINS", "MODERATES"},
    "polarity": {"POSITIVE", "NEGATIVE", "NON_MONOTONIC", "CONTEXT_DEPENDENT", "UNSIGNED"},
    "directness": {"DIRECT_AT_STATED_RESOLUTION", "MEDIATED_PATH", "UNKNOWN"},
    "sourceLevel": LEVELS, "targetLevel": LEVELS,
    "lagProfile": {"IMMEDIATE", "SHORT", "INTERMEDIATE", "DELAYED", "LONG", "STRUCTURAL", "INTERGENERATIONAL", "MIXED_CONTEXT_DEPENDENT"},
    "lagUnit": {"seconds", "minutes", "hours", "days", "weeks", "months", "years"},
    "exposurePattern": {"PULSE", "SUSTAINED", "CUMULATIVE", "REPEATED", "NOT_SPECIFIED"},
    "evidenceStrength": {"Strong", "Moderate", "Emerging", "Mixed", "Limited"},
    "confidence": {"HIGH", "MODERATE", "LOW"},
    "governanceClass": {"CORE", "CONTEXT_DEPENDENT", "SCENARIO_SPECIFIC", "HYPOTHESIZED"},
}
LOCAL_PATH = re.compile(r"(?:[A-Za-z]:[\\/]|\\\\|file://|source-data[\\/]|analysis[\\/])", re.I)


def txt(value: Any) -> str:
    return "" if value is None else " ".join(str(value).split())


def parts(value: Any) -> list[str]:
    return [txt(item) for item in str(value or "").split(";") if txt(item)]


def workbook_layer(name: str) -> str:
    folded = name.casefold()
    matches = [layer for token, layer in LAYER_TOKENS if token in folded]
    if len(matches) != 1:
        raise ValueError(f"{name}: filename does not identify exactly one canonical layer")
    return matches[0]


def load_drivers(path: Path) -> dict[str, dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise ValueError("data/drivers.json must be an array")
    result = {row["id"]: row for row in payload}
    if len(result) != len(payload):
        raise ValueError("data/drivers.json contains duplicate IDs")
    return result


def validate_codebook(path: Path) -> None:
    payload = json.loads(path.read_text(encoding="utf-8"))
    entries = payload.get("entries") if isinstance(payload, dict) else None
    if not isinstance(entries, list):
        raise ValueError("data/codebook.json must contain an entries array")
    by_id = {entry.get("id"): entry for entry in entries if isinstance(entry, dict)}
    governed = {
        "CB-REL-CAUSAL-ROLE": "causalRole",
        "CB-REL-POLARITY": "polarity",
        "CB-REL-DIRECTNESS": "directness",
        "CB-REL-LEVEL": "sourceLevel",
        "CB-REL-LAG-PROFILE": "lagProfile",
        "CB-REL-EXPOSURE-PATTERN": "exposurePattern",
        "CB-REL-CONFIDENCE": "confidence",
        "CB-REL-GOVERNANCE-CLASS": "governanceClass",
        "CB-REL-EVIDENCE-STRENGTH": "evidenceStrength",
    }
    errors = []
    for term_id, field in governed.items():
        entry = by_id.get(term_id)
        actual = set(entry.get("allowedValues", [])) if entry else set()
        if actual != CONTROLLED[field]:
            errors.append(
                f"{term_id}: Codebook values {sorted(actual)} do not match "
                f"Relationship Schema v2 values {sorted(CONTROLLED[field])}"
            )
    if errors:
        raise ValueError("Codebook/Relationship Schema drift:\n  " + "\n  ".join(errors))


def evidence_ids(source_dir: Path) -> set[str]:
    found: set[str] = set()
    for path in sorted(source_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        book = load_workbook(path, read_only=True, data_only=True)
        try:
            for row in book["Evidence Library"].iter_rows(min_row=5, values_only=True):
                value = txt(row[0] if row else None)
                if value:
                    if value in found:
                        raise ValueError(f"Duplicate Evidence ID {value}")
                    found.add(value)
        finally:
            book.close()
    return found


def read_records(source_dir: Path, drivers: dict[str, dict[str, Any]], evidence: set[str]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    errors: list[str] = []
    paths = [p for p in sorted(source_dir.glob("*.xlsx")) if not p.name.startswith("~$")]
    if len(paths) != 8:
        raise ValueError(f"Expected 8 workbooks; found {len(paths)}")
    for path in paths:
        layer = workbook_layer(path.name)
        book = load_workbook(path, read_only=True, data_only=True)
        try:
            if SHEET not in book.sheetnames:
                errors.append(f"{path.name}: missing {SHEET} worksheet")
                continue
            sheet = book[SHEET]
            actual = tuple(txt(cell.value) for cell in sheet[HEADER_ROW][:len(HEADERS)])
            if actual != HEADERS or sheet.max_column != len(HEADERS):
                errors.append(f"{path.name}: Relationship Schema v2 headers differ; discovered {list(actual)}")
                continue
            for row_number, row in enumerate(sheet.iter_rows(min_row=FIRST_DATA_ROW, max_col=len(HEADERS), values_only=True), FIRST_DATA_ROW):
                values = [txt(v) for v in row]
                if not any(values):
                    continue
                record: dict[str, Any] = {}
                for key, value in zip(KEYS, values):
                    if key in LIST_KEYS:
                        record[key] = parts(value)
                    elif key in {"lagLowerBound", "lagUpperBound"}:
                        if not value:
                            record[key] = None
                        else:
                            try:
                                record[key] = float(value)
                            except ValueError:
                                errors.append(f"{path.name}/{SHEET} row {row_number}: {key} is not numeric")
                                record[key] = None
                    else:
                        record[key] = value or None
                rid = record.get("id") or f"row {row_number}"
                for key in KEYS:
                    if key not in OPTIONAL and (record[key] is None or record[key] == []):
                        errors.append(f"{path.name}/{SHEET} row {row_number} ({rid}): required {key} is empty")
                for key, allowed in CONTROLLED.items():
                    vals = record[key] if key in LIST_KEYS else [record[key]]
                    for value in vals:
                        if value is not None and value not in allowed:
                            errors.append(f"{rid}: {key} value {value!r} is not controlled")
                source_id, target_id = record.get("sourceDriverId"), record.get("targetDriverId")
                for role, driver_id, name_key in (("source", source_id, "sourceDriverName"), ("target", target_id, "targetDriverName")):
                    if driver_id not in drivers:
                        errors.append(f"{rid}: {role} Driver {driver_id!r} does not exist")
                    elif record[name_key] != drivers[driver_id]["name"]:
                        errors.append(f"{rid}: {role} name does not match Driver {driver_id}")
                if source_id in drivers and drivers[source_id]["layer"] != layer:
                    errors.append(f"{rid}: source Driver belongs to {drivers[source_id]['layer']}, not source workbook {layer}")
                if source_id == target_id:
                    errors.append(f"{rid}: self-relationships are prohibited")
                if record.get("sourceLevel") != record.get("targetLevel") and not record.get("levelTransitionMechanism"):
                    errors.append(f"{rid}: cross-level edge lacks level-transition mechanism")
                bounds = (record.get("lagLowerBound"), record.get("lagUpperBound"), record.get("lagUnit"))
                if any(value is not None for value in bounds) and not all(value is not None for value in bounds):
                    errors.append(f"{rid}: numeric lag requires lower bound, upper bound, and unit")
                if bounds[0] is not None and bounds[1] is not None and bounds[0] > bounds[1]:
                    errors.append(f"{rid}: lag lower bound exceeds upper bound")
                for driver_id in record["moderatorDriverIds"]:
                    if driver_id not in drivers:
                        errors.append(f"{rid}: moderator Driver {driver_id!r} does not exist")
                for evidence_id in record["supportingEvidenceIds"]:
                    if evidence_id not in evidence:
                        errors.append(f"{rid}: Evidence ID {evidence_id!r} does not exist")
                record["source"] = {"workbook": path.name, "worksheet": SHEET, "row": row_number}
                records.append(record)
        finally:
            book.close()
    ids = Counter(row["id"] for row in records)
    pairs = Counter((row["sourceDriverId"], row["targetDriverId"]) for row in records)
    errors.extend(f"Duplicate Relationship ID {key}" for key, count in ids.items() if count > 1)
    errors.extend(f"Duplicate directed pair {key[0]} -> {key[1]}" for key, count in pairs.items() if count > 1)
    for row in records:
        if row["governanceClass"] not in {"CORE", "CONTEXT_DEPENDENT"}:
            errors.append(f"{row['id']}: public canonical graph may contain only CORE or CONTEXT_DEPENDENT edges")
        if LOCAL_PATH.search(json.dumps(row, ensure_ascii=False)):
            errors.append(f"{row['id']}: local/private path detected")
    if errors:
        raise ValueError("Relationship build failed:\n  ERROR: " + "\n  ERROR: ".join(errors))
    order = {layer: index for index, layer in enumerate(LAYERS)}
    records.sort(key=lambda row: (order[drivers[row["sourceDriverId"]]["layer"]], row["id"], row["sourceDriverId"], row["targetDriverId"]))
    return records


def graph_metrics(records: list[dict[str, Any]], drivers: dict[str, dict[str, Any]]) -> dict[str, Any]:
    adjacency: dict[str, set[str]] = defaultdict(set)
    degree = Counter(); incoming = Counter(); outgoing = Counter(); cross = Counter()
    for row in records:
        source, target = row["sourceDriverId"], row["targetDriverId"]
        adjacency[source].add(target); adjacency[target].add(source)
        degree[source] += 1; degree[target] += 1; outgoing[source] += 1; incoming[target] += 1
        sl, tl = drivers[source]["layer"], drivers[target]["layer"]
        if sl != tl:
            cross[f"{sl} -> {tl}"] += 1
    seen: set[str] = set(); components: list[int] = []
    for node in drivers:
        if node in seen:
            continue
        queue = deque([node]); seen.add(node); size = 0
        while queue:
            current = queue.popleft(); size += 1
            for other in adjacency[current]:
                if other not in seen:
                    seen.add(other); queue.append(other)
        components.append(size)
    return {
        "driverCount": len(drivers), "relationshipCount": len(records),
        "causalRoles": dict(sorted(Counter(row["causalRole"] for row in records).items())),
        "governanceClasses": dict(sorted(Counter(row["governanceClass"] for row in records).items())),
        "directness": dict(sorted(Counter(row["directness"] for row in records).items())),
        "crossLayerPairs": dict(sorted(cross.items())),
        "directedLayerPairCoverage": len(cross), "weakComponents": len(components),
        "isolatedDrivers": sum(degree[node] == 0 for node in drivers),
        "degreeOneDrivers": sum(degree[node] == 1 for node in drivers),
        "driversWithIncomingAndOutgoing": sum(incoming[node] > 0 and outgoing[node] > 0 for node in drivers),
        "largestWeakComponent": max(components, default=0),
    }


def write_atomic(payload: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", newline="\n", delete=False, dir=path.parent, suffix=".tmp") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
        temporary = Path(handle.name)
    temporary.replace(path)


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    try:
        drivers = load_drivers(root / "data" / "drivers.json")
        validate_codebook(root / "data" / "codebook.json")
        records = read_records(root / "source-data", drivers, evidence_ids(root / "source-data"))
        graph = graph_metrics(records, drivers)
        write_atomic({"schemaVersion": SCHEMA_VERSION, "relationships": records}, root / "data" / "relationships.json")
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 1
    print("Relationship import summary")
    print(f"  Schema: {SCHEMA_VERSION}")
    print(f"  Relationships: {len(records)}")
    print(f"  Cross-layer relationships: {sum(graph['crossLayerPairs'].values())}")
    print(f"  Weak components: {graph['weakComponents']}")
    print("  Warnings: 0")
    print("  Errors: 0")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
