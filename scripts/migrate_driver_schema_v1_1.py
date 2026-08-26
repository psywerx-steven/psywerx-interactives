"""Migrate PSYWERX Driver Ontology workbooks from Schema v1.0 to v1.1.

The migration stages all eight XLSX outputs, validates the staged collection,
and replaces the source workbooks only after every validation succeeds. It is
safe to rerun against the resulting Schema v1.1 workbooks.
"""

from __future__ import annotations

import hashlib
import os
import re
import sys
import tempfile
from collections import Counter
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange


V1_HEADERS = (
    "ID",
    "Name",
    "Other Names / Aliases",
    "Layer",
    "Family",
    "Definition",
    "Data Type",
    "Representation / Scale",
    "Polarity / Direction of Interpretation",
    "Mechanism",
    "Likely Upstream Influences",
    "Likely Downstream Influences",
    "Moderators / Boundary Conditions",
    "Typical Interaction Candidates",
    "Modifiability / Malleability",
    "Volatility",
    "Time Scale of Change",
    "Onset / Causal Lag",
    "Persistence / Recovery",
    "Indicators",
    "Measurement / Assessment Methods",
    "Observability",
    "Measurement Caveats",
    "Evidence Strength",
    "Evidence Notes",
    "Common Misinterpretations",
    "Key Sources",
)
QUALIFIER_HEADER = "Time Scale Qualifier"
QUALIFIER_COLUMN = V1_HEADERS.index("Time Scale of Change") + 2
V1_1_HEADERS = (
    V1_HEADERS[: QUALIFIER_COLUMN - 1]
    + (QUALIFIER_HEADER,)
    + V1_HEADERS[QUALIFIER_COLUMN - 1 :]
)

CANONICAL_BANDS = (
    "Seconds–Minutes",
    "Minutes–Hours",
    "Hours–Days",
    "Days–Weeks",
    "Weeks–Months",
    "Months–Years",
    "Years–Generations",
)
MIXED = "Mixed / Context-dependent"
STABLE = "Stable / Not applicable"
TIME_VALUES = CANONICAL_BANDS + (MIXED, STABLE)
ONSET_VALUES = CANONICAL_BANDS + (MIXED,)

TIME_VALUE_MAP = {
    "Days–weeks": ("Days–Weeks",),
    "Hours–days": ("Hours–Days",),
    "Minutes–hours": ("Minutes–Hours",),
    "Months–years": ("Months–Years",),
    "Weeks–months": ("Weeks–Months",),
    "Minutes–Days": ("Minutes–Hours", "Hours–Days"),
    "Minutes–days": ("Minutes–Hours", "Hours–Days"),
    "Hours–weeks": ("Hours–Days", "Days–Weeks"),
    "Days–months": ("Days–Weeks", "Weeks–Months"),
    "Days–years": ("Days–Weeks", "Weeks–Months", "Months–Years"),
    "Weeks–years": ("Weeks–Months", "Months–Years"),
}
ONSET_VALUE_MAP = {
    "Days–Months": ("Days–Weeks", "Weeks–Months"),
    "Minutes–Days": ("Minutes–Hours", "Hours–Days"),
    "Seconds–Days": ("Seconds–Minutes", "Minutes–Hours", "Hours–Days"),
    "Seconds–Hours": ("Seconds–Minutes", "Minutes–Hours"),
}

GOVERNED_TIME_SCALES = {
    "BIO-002": ("Hours–Days",),
    "BIO-006": ("Days–Weeks",),
    "BIO-012": ("Minutes–Hours",),
    "BIO-014": ("Minutes–Hours",),
    "BIO-018": ("Minutes–Hours",),
    "BIO-026": ("Minutes–Hours",),
    "BIO-032": (MIXED,),
    "BIO-033": (MIXED,),
    "BIO-034": (MIXED,),
    "BIO-035": (MIXED,),
    "BIO-036": (MIXED,),
    "BIO-039": ("Weeks–Months",),
    "BIO-044": ("Months–Years",),
    "BIO-045": ("Weeks–Months",),
    "BIO-046": ("Months–Years",),
    "BIO-049": ("Months–Years",),
    "BIO-050": ("Years–Generations",),
    "BIO-052": (STABLE,),
    "BIO-057": ("Months–Years",),
    "BIO-059": ("Minutes–Hours",),
    "BIO-061": ("Minutes–Hours",),
    "BIO-062": ("Minutes–Hours",),
    "BIO-063": ("Minutes–Hours",),
    "BIO-064": ("Minutes–Hours",),
    "BIO-070": (MIXED,),
}

TIME_SCALE_QUALIFIERS = {
    "BIO-002": "May be recurrent or chronic; the course depends on the cause of sleep fragmentation.",
    "BIO-006": "Repeated mismatch or persistent social jetlag may continue chronically; adaptation rate varies with schedule and context.",
    "BIO-012": "Typically develops over hours and may resolve rapidly after food intake.",
    "BIO-014": "Typically accumulates over hours and may resolve rapidly after drinking.",
    "BIO-018": "Acute episodes may occur against a chronic baseline.",
    "BIO-026": "Meaningful fatigue may emerge after tens of minutes of sustained cognitive demand.",
    "BIO-032": "Often chronic; acute injury and relatively rapid assistive-technology changes are possible.",
    "BIO-033": "May reflect a chronic baseline or change after acute injury, treatment, or assistive correction.",
    "BIO-034": "May involve acute loss, a recovery trajectory, or chronic or progressive impairment.",
    "BIO-035": "May fluctuate episodically over minutes or reflect a chronic vestibular deficit.",
    "BIO-036": "May involve chronic impairment, cause-specific onset, and training or intervention response.",
    "BIO-039": "May be chronic; acute blood loss can produce faster change.",
    "BIO-044": "Developmental, stage-dependent progression typically unfolds over multiple years.",
    "BIO-045": "Pregnancy-bounded course varies by gestational phase.",
    "BIO-046": "A multi-year developmental transition with symptom-mediated fluctuation.",
    "BIO-049": "Developmental stage, sensitive periods, and multi-year maturation shape the course.",
    "BIO-050": "Typically changes over years to decades within a lifespan; cross-generational change is not implied.",
    "BIO-052": "Stable or invariant across the modeled lifespan; change speed is ordinarily not applicable.",
    "BIO-057": "A progressive, usually irreversible, disease-specific course may extend across many years.",
    "BIO-059": "Effects commonly last hours; route, dose, tolerance, and residual effects alter duration.",
    "BIO-061": "Effects usually last hours; tolerance, withdrawal reversal, and later sleep effects may alter the broader course.",
    "BIO-062": "Chronic exposure, repeated dosing, tolerance, and long-term effects are distinct from acute drug-effect change speed.",
    "BIO-063": "Duration varies by agent and route; chronic therapy, tolerance, and pain-relief context also matter.",
    "BIO-064": "Duration varies by agent, route, dose, tolerance, and withdrawal; acute effects are typically hour-scale.",
    "BIO-070": "May remain chronic while a regimen persists; change depends on drug-specific half-lives and dose or regimen changes.",
}

TIME_CODEBOOK_ROW = (
    "Drivers",
    "Time Scale of Change",
    "Typical elapsed time over which the driver state meaningfully changes.",
    "; ".join(TIME_VALUES) + "; semicolon-separated canonical combinations",
    "Yes",
    "Select one or more canonical bands. Broad contiguous ranges must list every spanned band, separated by semicolons, shortest to longest. Use canonical capitalization and en dashes. Mixed / Context-dependent and Stable / Not applicable are each exclusive. Chronicity and persistence do not themselves define change speed. Use Time Scale Qualifier to preserve narrative nuance.",
)
QUALIFIER_CODEBOOK_ROW = (
    "Drivers",
    QUALIFIER_HEADER,
    "Optional narrative information qualifying Time Scale of Change when meaningful timing information cannot be represented fully by canonical bands. It must not replace or redefine Persistence / Recovery.",
    "Free text",
    "No",
    "Preserve meaningful temporal nuance such as chronic baseline, developmental course, stable status, cause-dependent timing, or precise bounds. Do not use as a faceted/filtering value.",
)
ONSET_CODEBOOK_ROW = (
    "Drivers",
    "Onset / Causal Lag",
    "Delay from driver-state change to a downstream consequence.",
    "; ".join(ONSET_VALUES) + "; semicolon-separated canonical combinations",
    "Yes",
    "Use the same canonical delimiter, ordering, duplicate, and exclusivity rules as Time Scale of Change, except Stable / Not applicable is not permitted. Broad contiguous ranges must list every band they span. Keep causal lag distinct from driver change speed.",
)


@dataclass
class Baseline:
    ids: tuple[str, ...]
    driver_values: dict[str, dict[str, Any]]
    preserved_sheet_hashes: dict[str, str]
    layer_summary_hash: str
    layer_standard_row: int


@dataclass
class MigrationSummary:
    workbooks: int = 0
    drivers: int = 0
    time_rows_changed: int = 0
    onset_rows_changed: int = 0
    qualifiers_written: int = 0
    time_changes: Counter[tuple[str, str]] = field(default_factory=Counter)
    onset_changes: Counter[tuple[str, str]] = field(default_factory=Counter)


def clean(value: Any) -> str:
    return "" if value is None else re.sub(r"\s+", " ", str(value)).strip()


def split_temporal(value: Any) -> list[str]:
    return [part for part in (clean(item) for item in str(value or "").split(";")) if part]


def ordered_unique(values: list[str]) -> list[str]:
    result: list[str] = []
    for value in values:
        if value not in result:
            result.append(value)
    if MIXED in result or STABLE in result:
        if len(result) != 1:
            raise ValueError(f"Exclusive temporal value combined with other values: {result}")
        return result
    return sorted(result, key=CANONICAL_BANDS.index)


def normalize_temporal(
    raw: Any,
    replacements: dict[str, tuple[str, ...]],
    allowed: tuple[str, ...],
) -> tuple[str, ...]:
    normalized: list[str] = []
    for value in split_temporal(raw):
        normalized.extend(replacements.get(value, (value,)))
    unknown = [value for value in normalized if value not in allowed]
    if unknown:
        raise ValueError(f"Unapproved temporal value(s): {unknown}")
    return tuple(ordered_unique(normalized))


def header_values(sheet: Any) -> tuple[str, ...]:
    return tuple(clean(cell.value) for cell in sheet[4] if clean(cell.value))


def header_positions(sheet: Any) -> dict[str, int]:
    return {
        clean(cell.value): cell.column
        for cell in sheet[4]
        if clean(cell.value)
    }


def hash_sheet_values(sheet: Any, excluded: set[str] | None = None) -> str:
    excluded = excluded or set()
    digest = hashlib.sha256()
    digest.update(sheet.title.encode("utf-8"))
    for row in sheet.iter_rows():
        for cell in row:
            if cell.coordinate in excluded or cell.value is None:
                continue
            payload = (cell.coordinate, cell.data_type, repr(cell.value))
            digest.update(repr(payload).encode("utf-8"))
    digest.update(repr(sorted(str(item) for item in sheet.merged_cells.ranges)).encode("utf-8"))
    return digest.hexdigest()


def collect_driver_values(sheet: Any) -> tuple[tuple[str, ...], dict[str, dict[str, Any]]]:
    headers = header_values(sheet)
    positions = header_positions(sheet)
    values: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=5):
        identifier = clean(row[positions["ID"] - 1].value)
        if not identifier:
            continue
        if identifier in values:
            raise ValueError(f"Duplicate Driver ID {identifier!r} in {sheet.title}")
        values[identifier] = {
            header: row[column - 1].value for header, column in positions.items()
        }
    return headers, values


def collect_baseline(workbook: Any, path: Path) -> Baseline:
    required = {"Drivers", "Codebook", "Layer Summary"}
    missing = sorted(required - set(workbook.sheetnames))
    if missing:
        raise ValueError(f"{path.name}: missing worksheet(s): {missing}")
    headers, driver_values = collect_driver_values(workbook["Drivers"])
    if headers not in {V1_HEADERS, V1_1_HEADERS}:
        raise ValueError(
            f"{path.name}: expected Schema v1.0 or v1.1 headers; found {headers}"
        )
    summary_sheet = workbook["Layer Summary"]
    standard_rows = [
        cell.row for cell in summary_sheet["A"]
        if clean(cell.value) == "Ontology Standard Version"
    ]
    if len(standard_rows) != 1:
        raise ValueError(f"{path.name}: expected one Ontology Standard Version row")
    standard_row = standard_rows[0]
    preserved_hashes = {
        name: hash_sheet_values(workbook[name])
        for name in workbook.sheetnames
        if name not in {"Drivers", "Codebook", "Layer Summary"}
    }
    return Baseline(
        ids=tuple(driver_values),
        driver_values=driver_values,
        preserved_sheet_hashes=preserved_hashes,
        layer_summary_hash=hash_sheet_values(
            summary_sheet, {"A2", f"B{standard_row}"}
        ),
        layer_standard_row=standard_row,
    )


def copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def shift_data_validations(sheet: Any, insertion_column: int) -> None:
    if not sheet.data_validations:
        return
    for validation in sheet.data_validations.dataValidation:
        shifted: list[CellRange] = []
        for current in validation.ranges.ranges:
            cell_range = CellRange(str(current))
            if cell_range.min_col >= insertion_column:
                cell_range.shift(col_shift=1)
            elif cell_range.max_col >= insertion_column:
                cell_range.max_col += 1
            shifted.append(cell_range)
        validation.sqref = MultiCellRange(shifted)


def add_qualifier_column(sheet: Any) -> None:
    headers = header_values(sheet)
    if headers == V1_1_HEADERS:
        return
    if headers != V1_HEADERS:
        raise ValueError(f"{sheet.title}: cannot add qualifier to unexpected headers")

    old_max_column = sheet.max_column
    old_widths = {
        column: copy(sheet.column_dimensions[get_column_letter(column)])
        for column in range(1, old_max_column + 1)
    }
    merged_top_rows = [
        str(item) for item in sheet.merged_cells.ranges
        if item.min_row in {1, 2} and item.max_row == item.min_row
    ]
    for merged in merged_top_rows:
        sheet.unmerge_cells(merged)

    sheet.insert_cols(QUALIFIER_COLUMN, 1)
    shift_data_validations(sheet, QUALIFIER_COLUMN)

    for old_column in range(QUALIFIER_COLUMN, old_max_column + 1):
        new_letter = get_column_letter(old_column + 1)
        dimension = copy(old_widths[old_column])
        dimension.index = new_letter
        sheet.column_dimensions[new_letter] = dimension
    qualifier_letter = get_column_letter(QUALIFIER_COLUMN)
    qualifier_dimension = copy(old_widths[QUALIFIER_COLUMN + 1])
    qualifier_dimension.index = qualifier_letter
    qualifier_dimension.width = max(38.0, qualifier_dimension.width or 0)
    sheet.column_dimensions[qualifier_letter] = qualifier_dimension

    sheet.cell(row=4, column=QUALIFIER_COLUMN).value = QUALIFIER_HEADER
    copy_cell_style(
        sheet.cell(row=4, column=QUALIFIER_COLUMN - 1),
        sheet.cell(row=4, column=QUALIFIER_COLUMN),
    )
    for row_number in range(5, sheet.max_row + 1):
        copy_cell_style(
            sheet.cell(row=row_number, column=QUALIFIER_COLUMN + 2),
            sheet.cell(row=row_number, column=QUALIFIER_COLUMN),
        )

    new_max_column = old_max_column + 1
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=new_max_column)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=new_max_column)


def migrate_drivers(sheet: Any, summary: MigrationSummary) -> None:
    add_qualifier_column(sheet)
    if header_values(sheet) != V1_1_HEADERS:
        raise ValueError(f"{sheet.title}: Schema v1.1 header construction failed")
    positions = header_positions(sheet)
    seen: set[str] = set()
    for row in sheet.iter_rows(min_row=5):
        identifier = clean(row[positions["ID"] - 1].value)
        if not identifier:
            continue
        if identifier in seen:
            raise ValueError(f"Duplicate Driver ID {identifier!r}")
        seen.add(identifier)
        summary.drivers += 1

        time_cell = row[positions["Time Scale of Change"] - 1]
        onset_cell = row[positions["Onset / Causal Lag"] - 1]
        qualifier_cell = row[positions[QUALIFIER_HEADER] - 1]
        old_time = clean(time_cell.value)
        old_onset = clean(onset_cell.value)

        if identifier in GOVERNED_TIME_SCALES:
            time_values = GOVERNED_TIME_SCALES[identifier]
        else:
            time_values = normalize_temporal(
                time_cell.value, TIME_VALUE_MAP, TIME_VALUES
            )
        onset_values = normalize_temporal(
            onset_cell.value, ONSET_VALUE_MAP, ONSET_VALUES
        )
        new_time = "; ".join(time_values)
        new_onset = "; ".join(onset_values)
        time_cell.value = new_time
        onset_cell.value = new_onset
        if old_time != new_time:
            summary.time_rows_changed += 1
            summary.time_changes[(old_time, new_time)] += 1
        if old_onset != new_onset:
            summary.onset_rows_changed += 1
            summary.onset_changes[(old_onset, new_onset)] += 1

        approved_qualifier = TIME_SCALE_QUALIFIERS.get(identifier)
        existing_qualifier = clean(qualifier_cell.value)
        if approved_qualifier:
            if existing_qualifier and existing_qualifier != approved_qualifier:
                raise ValueError(
                    f"{identifier}: existing Time Scale Qualifier differs from approved text"
                )
            qualifier_cell.value = approved_qualifier
            summary.qualifiers_written += 1
        elif existing_qualifier:
            raise ValueError(
                f"{identifier}: unapproved Time Scale Qualifier is populated"
            )

    sheet["A2"] = (
        "All 28 fields governed by PSYWERX Driver Ontology Standard v1.1. "
        "Semicolon-separated prose fields summarize likely influences; "
        "evidence-bearing edges are stored separately."
    )


def write_row(sheet: Any, row_number: int, values: tuple[str, ...]) -> None:
    for column, value in enumerate(values, start=1):
        sheet.cell(row=row_number, column=column).value = value


def migrate_codebook(sheet: Any) -> None:
    fields = {
        clean(cell.value): cell.row for cell in sheet["B"] if clean(cell.value)
    }
    if QUALIFIER_HEADER not in fields:
        onset_row = fields.get("Onset / Causal Lag")
        if onset_row is None:
            raise ValueError("Codebook is missing Onset / Causal Lag")
        sheet.insert_rows(onset_row, 1)
        qualifier_row = onset_row
        sheet.row_dimensions[onset_row].height = sheet.row_dimensions[onset_row - 1].height
        for column in range(1, 7):
            copy_cell_style(
                sheet.cell(row=onset_row - 1, column=column),
                sheet.cell(row=onset_row, column=column),
            )
    else:
        qualifier_row = fields[QUALIFIER_HEADER]

    fields = {
        clean(cell.value): cell.row for cell in sheet["B"] if clean(cell.value)
    }
    write_row(sheet, fields["Time Scale of Change"], TIME_CODEBOOK_ROW)
    write_row(sheet, qualifier_row, QUALIFIER_CODEBOOK_ROW)
    write_row(sheet, fields["Onset / Causal Lag"], ONSET_CODEBOOK_ROW)
    sheet["A1"] = "PSYWERX Driver Ontology Codebook v1.1"
    standard_row = fields.get("Ontology Standard Version")
    if standard_row is None:
        raise ValueError("Codebook is missing Ontology Standard Version")
    sheet.cell(row=standard_row, column=4).value = (
        "PSYWERX Driver Ontology Standard v1.1"
    )


def migrate_layer_summary(sheet: Any, expected_row: int) -> None:
    if "v1.0" in clean(sheet["A2"].value):
        sheet["A2"] = str(sheet["A2"].value).replace("v1.0", "v1.1")
    elif "v1.1" not in clean(sheet["A2"].value):
        raise ValueError("Layer Summary governance text has an unexpected version")
    if clean(sheet.cell(row=expected_row, column=1).value) != "Ontology Standard Version":
        raise ValueError("Layer Summary Ontology Standard Version row moved unexpectedly")
    sheet.cell(row=expected_row, column=2).value = (
        "PSYWERX Driver Ontology Standard v1.1"
    )


def codebook_signature(sheet: Any) -> tuple[tuple[str, ...], ...]:
    return tuple(
        tuple(clean(cell.value) for cell in row)
        for row in sheet.iter_rows()
        if any(clean(cell.value) for cell in row)
    )


def validate_staged_workbook(path: Path, baseline: Baseline) -> tuple[tuple[str, ...], ...]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if set(workbook.sheetnames) != set(baseline.preserved_sheet_hashes) | {
            "Drivers", "Codebook", "Layer Summary"
        }:
            raise ValueError(f"{path.name}: worksheet set changed")
        drivers = workbook["Drivers"]
        headers, values = collect_driver_values(drivers)
        if headers != V1_1_HEADERS:
            raise ValueError(f"{path.name}: staged Drivers headers are not Schema v1.1")
        if tuple(values) != baseline.ids:
            raise ValueError(f"{path.name}: Driver IDs or row order changed")
        for identifier, old_record in baseline.driver_values.items():
            new_record = values[identifier]
            for header in V1_HEADERS:
                if header in {"Time Scale of Change", "Onset / Causal Lag"}:
                    continue
                if new_record[header] != old_record[header]:
                    raise ValueError(
                        f"{path.name}: unrelated field {header!r} changed for {identifier}"
                    )
            time_values = split_temporal(new_record["Time Scale of Change"])
            onset_values = split_temporal(new_record["Onset / Causal Lag"])
            if tuple(ordered_unique(time_values)) != tuple(time_values) or any(
                value not in TIME_VALUES for value in time_values
            ):
                raise ValueError(f"{identifier}: invalid staged Time Scale of Change")
            if tuple(ordered_unique(onset_values)) != tuple(onset_values) or any(
                value not in ONSET_VALUES for value in onset_values
            ):
                raise ValueError(f"{identifier}: invalid staged Onset / Causal Lag")
            expected_qualifier = TIME_SCALE_QUALIFIERS.get(identifier)
            actual_qualifier = clean(new_record[QUALIFIER_HEADER]) or None
            if actual_qualifier != expected_qualifier:
                raise ValueError(f"{identifier}: staged qualifier does not match approval")
        if sorted(str(item) for item in drivers.merged_cells.ranges) != ["A1:AB1", "A2:AB2"]:
            raise ValueError(f"{path.name}: Drivers merged title ranges are invalid")

        for name, expected_hash in baseline.preserved_sheet_hashes.items():
            if hash_sheet_values(workbook[name]) != expected_hash:
                raise ValueError(f"{path.name}: unrelated worksheet {name!r} changed")
        layer_summary = workbook["Layer Summary"]
        if hash_sheet_values(
            layer_summary, {"A2", f"B{baseline.layer_standard_row}"}
        ) != baseline.layer_summary_hash:
            raise ValueError(f"{path.name}: unrelated Layer Summary content changed")
        if clean(layer_summary.cell(row=baseline.layer_standard_row, column=2).value) != (
            "PSYWERX Driver Ontology Standard v1.1"
        ):
            raise ValueError(f"{path.name}: Layer Summary version was not updated")
        return codebook_signature(workbook["Codebook"])
    finally:
        workbook.close()


def cleanup_staging(directory: Path) -> None:
    resolved = directory.resolve()
    source_root = directory.parent.resolve()
    if resolved.parent != source_root or not resolved.name.startswith(".schema-v1-1-"):
        raise RuntimeError(f"Refusing to clean unexpected staging path: {resolved}")
    for child in resolved.iterdir():
        if child.is_file():
            child.unlink()
        else:
            raise RuntimeError(f"Unexpected directory in staging area: {child}")
    resolved.rmdir()


def main() -> int:
    root = Path(__file__).resolve().parent.parent
    source_dir = root / "source-data"
    paths = sorted(
        path for path in source_dir.glob("*.xlsx")
        if not path.name.startswith("~$")
    )
    if len(paths) != 8:
        print(f"ERROR: expected 8 XLSX workbooks; found {len(paths)}", file=sys.stderr)
        return 1

    summary = MigrationSummary(workbooks=len(paths))
    baselines: dict[str, Baseline] = {}
    staging = Path(tempfile.mkdtemp(prefix=".schema-v1-1-", dir=source_dir))
    try:
        global_ids: set[str] = set()
        for path in paths:
            workbook = load_workbook(path, read_only=False, data_only=False)
            try:
                baseline = collect_baseline(workbook, path)
                overlap = global_ids.intersection(baseline.ids)
                if overlap:
                    raise ValueError(
                        f"{path.name}: duplicate global Driver ID(s): {sorted(overlap)}"
                    )
                global_ids.update(baseline.ids)
                baselines[path.name] = baseline
                migrate_drivers(workbook["Drivers"], summary)
                migrate_codebook(workbook["Codebook"])
                migrate_layer_summary(
                    workbook["Layer Summary"], baseline.layer_standard_row
                )
                workbook.save(staging / path.name)
            finally:
                workbook.close()

        if len(global_ids) != 762 or summary.drivers != 762:
            raise ValueError(
                f"Expected 762 unique drivers; found {len(global_ids)} IDs and "
                f"{summary.drivers} rows"
            )
        signatures = {
            validate_staged_workbook(staging / path.name, baselines[path.name])
            for path in paths
        }
        if len(signatures) != 1:
            raise ValueError("Staged Codebook worksheets are not identical")

        for path in paths:
            os.replace(staging / path.name, path)

        print("Schema v1.1 workbook migration complete")
        print(f"  Workbooks: {summary.workbooks}")
        print(f"  Drivers: {summary.drivers}")
        print(f"  Time Scale rows changed: {summary.time_rows_changed}")
        print(f"  Onset / Causal Lag rows changed: {summary.onset_rows_changed}")
        print(f"  Time Scale Qualifiers populated: {summary.qualifiers_written}")
        print("Time Scale changes")
        for (old, new), count in sorted(summary.time_changes.items()):
            print(f"  {count}: {old} -> {new}")
        print("Onset / Causal Lag changes")
        for (old, new), count in sorted(summary.onset_changes.items()):
            print(f"  {count}: {old} -> {new}")
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        print("No source workbook was replaced.", file=sys.stderr)
        return 1
    finally:
        if staging.exists():
            cleanup_staging(staging)


if __name__ == "__main__":
    raise SystemExit(main())
