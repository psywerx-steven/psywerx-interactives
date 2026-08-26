"""Build data/families.json from canonical PSYWERX Families worksheets."""

from __future__ import annotations

import json
import re
import sys
import tempfile
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCHEMA_VERSION = "1.0"
HEADER_SCAN_ROWS = 25
CANONICAL_LAYERS = (
    "Biological",
    "Psychological",
    "Social",
    "Cultural",
    "Physical / Environmental",
    "Institutional / Structural",
    "Informational",
    "Technological",
)
FILENAME_LAYER_PATTERNS = (
    (r"(?:^|_)layer_1_biological(?:_|$)", "Biological"),
    (r"(?:^|_)layer_2_psychological(?:_|$)", "Psychological"),
    (r"(?:^|_)layer_3_social(?:_|$)", "Social"),
    (r"(?:^|_)layer_4_cultural(?:_|$)", "Cultural"),
    (
        r"(?:^|_)layer_5_physical_environmental(?:_|$)",
        "Physical / Environmental",
    ),
    (
        r"(?:^|_)layer_6_institutional_structural(?:_|$)",
        "Institutional / Structural",
    ),
    (r"(?:^|_)layer_7_informational(?:_|$)", "Informational"),
    (r"(?:^|_)layer_8_technological(?:_|$)", "Technological"),
)

FAMILY_HEADERS = (
    "Family ID",
    "Family Name",
    "Layer",
    "Definition",
    "Inclusion Rule",
    "Exclusion / Boundary Rule",
    "Representative Drivers",
    "Driver Count",
)
OUTPUT_KEYS = (
    "id",
    "name",
    "layer",
    "definition",
    "includes",
    "exclusions",
    "representativeDrivers",
    "representativeDriverIds",
    "driverCount",
    "source",
)

CODEBOOK_FAMILY_ID_ROW = (
    "Families",
    "Family ID",
    "Permanent globally unique family identifier.",
    "Explicit canonical Family ID (e.g., BIO-F01)",
    "Yes",
    "Preserve exactly; never regenerate, rename, reuse, or reassign. Globally unique across the ontology and suitable as a stable foreign key.",
)
CODEBOOK_REPRESENTATIVE_ROW = (
    "Families",
    "Representative Drivers",
    "Illustrative canonical Driver names.",
    "Semicolon-separated canonical Driver names",
    "No",
    "Each name must resolve to exactly one canonical Driver in the same Family. Public data derives representative Driver IDs without replacing the source names.",
)
CODEBOOK_DRIVER_COUNT_ROW = (
    "Families",
    "Driver Count",
    "Number of canonical Driver records assigned to the Family.",
    "Non-negative integer",
    "Yes",
    "Required validation/derived field. The workbook value must equal the authoritative count derived from Drivers; mismatches are errors and are never silently corrected.",
)


@dataclass
class Summary:
    workbooks: int = 0
    family_sheets: int = 0
    drivers_total: int = 0
    drivers_matched: int = 0
    representative_references: int = 0
    representatives_resolved: int = 0
    count_mismatches: int = 0
    duplicate_family_ids: int = 0
    duplicate_family_names: int = 0
    duplicate_layer_names: int = 0
    orphan_families: int = 0
    unmatched_drivers: int = 0
    ambiguous_driver_matches: int = 0
    unresolved_representatives: int = 0
    ambiguous_representatives: int = 0
    wrong_family_representatives: int = 0
    families_per_layer: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def clean(value: Any) -> str:
    return "" if value is None else re.sub(r"\s+", " ", str(value)).strip()


def clean_id(value: Any) -> str:
    """Preserve an explicit Family ID exactly; validate emptiness separately."""
    return "" if value is None else str(value)


def normalized_key(value: Any) -> str:
    text = unicodedata.normalize("NFKC", clean(value)).casefold()
    return re.sub(r"[^\w]+", " ", text, flags=re.UNICODE).strip()


def split_representative_drivers(value: Any) -> list[str]:
    if value is None or not str(value).strip():
        return []
    return [item for item in (clean(part) for part in str(value).split(";")) if item]


def parse_driver_count(value: Any) -> int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, int):
        return value if value >= 0 else None
    if isinstance(value, float) and value.is_integer() and value >= 0:
        return int(value)
    text = clean(value)
    if re.fullmatch(r"\d+", text):
        return int(text)
    return None


def filename_layer(path: Path) -> str | None:
    filename = path.stem.casefold()
    matches = {
        layer
        for pattern, layer in FILENAME_LAYER_PATTERNS
        if re.search(pattern, filename)
    }
    return next(iter(matches)) if len(matches) == 1 else None


def row_values(row: tuple[Any, ...]) -> tuple[str, ...]:
    return tuple(clean(value) for value in row if clean(value))


def find_family_sheets(workbook: Any) -> list[Any]:
    named = [
        sheet
        for sheet in workbook.worksheets
        if normalized_key(sheet.title) == "families"
    ]
    if named:
        return named
    candidates = []
    expected = set(FAMILY_HEADERS)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(
            min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True
        ):
            if expected.issubset(set(row_values(row))):
                candidates.append(sheet)
                break
    return candidates


def find_header_row(sheet: Any) -> tuple[int, tuple[Any, ...]] | None:
    expected = set(FAMILY_HEADERS)
    best: tuple[int, tuple[Any, ...]] | None = None
    best_score = 0
    for number, row in enumerate(
        sheet.iter_rows(
            min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True
        ),
        start=1,
    ):
        score = len(expected.intersection(row_values(row)))
        if score > best_score:
            best = (number, row)
            best_score = score
    return best


def codebook_signature(workbook: Any, path: Path, summary: Summary) -> tuple[tuple[str, ...], ...]:
    if "Codebook" not in workbook.sheetnames:
        summary.errors.append(f"{path.name}: missing Codebook worksheet.")
        return ()
    signature = tuple(
        tuple(clean(value) for value in row)
        for row in workbook["Codebook"].iter_rows(values_only=True)
        if any(clean(value) for value in row)
    )
    governed = {
        row[1]: row for row in signature if row[0] == "Families" and len(row) >= 6
    }
    expected = {
        "Family ID": CODEBOOK_FAMILY_ID_ROW,
        "Representative Drivers": CODEBOOK_REPRESENTATIVE_ROW,
        "Driver Count": CODEBOOK_DRIVER_COUNT_ROW,
    }
    for field_name, expected_row in expected.items():
        actual = governed.get(field_name)
        if actual != expected_row:
            summary.errors.append(
                f"{path.name} / Codebook: Family governance row "
                f"{field_name!r} is missing or does not match Family Schema v1.0."
            )
    return signature


def build_family_record(
    row: tuple[Any, ...],
    positions: dict[str, int],
    path: Path,
    sheet_name: str,
    row_number: int,
    inferred_layer: str | None,
    summary: Summary,
) -> dict[str, Any] | None:
    if not row_values(row):
        return None
    location = f"{path.name} / {sheet_name} / row {row_number}"

    def cell(header: str) -> Any:
        index = positions[header]
        return row[index] if index < len(row) else None

    family_id = clean_id(cell("Family ID"))
    name = clean(cell("Family Name"))
    layer = clean(cell("Layer"))
    definition = clean(cell("Definition"))
    includes = clean(cell("Inclusion Rule"))
    exclusions = clean(cell("Exclusion / Boundary Rule"))
    representatives = split_representative_drivers(cell("Representative Drivers"))
    driver_count = parse_driver_count(cell("Driver Count"))

    required = {
        "Family ID": family_id.strip(),
        "Family Name": name,
        "Layer": layer,
        "Definition": definition,
        "Inclusion Rule": includes,
        "Exclusion / Boundary Rule": exclusions,
    }
    for header, value in required.items():
        if not value:
            summary.errors.append(f"{location}: required field {header!r} is empty.")
    if driver_count is None:
        summary.errors.append(
            f"{location}: Driver Count must be a non-negative integer."
        )
    if layer and layer not in CANONICAL_LAYERS:
        summary.errors.append(
            f"{location}: Layer {layer!r} is not an exact canonical layer value."
        )
    if layer and inferred_layer and layer != inferred_layer:
        summary.errors.append(
            f"{location}: explicit Layer {layer!r} conflicts with workbook "
            f"filename layer {inferred_layer!r}."
        )

    return {
        "id": family_id or None,
        "name": name or None,
        "layer": layer or inferred_layer,
        "definition": definition or None,
        "includes": includes or None,
        "exclusions": exclusions or None,
        "representativeDrivers": representatives,
        "representativeDriverIds": [],
        "driverCount": driver_count,
        "source": {
            "workbook": path.name,
            "sheet": sheet_name,
            "row": row_number,
        },
    }


def read_workbooks(source_dir: Path, summary: Summary) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    header_signatures: dict[str, tuple[str, ...]] = {}
    codebook_signatures: dict[str, tuple[tuple[str, ...], ...]] = {}
    paths = sorted(
        path
        for path in source_dir.rglob("*")
        if path.is_file()
        and path.suffix.casefold() == ".xlsx"
        and not path.name.startswith("~$")
    )
    summary.workbooks = len(paths)
    if not paths:
        summary.errors.append(f"No XLSX workbooks found in {source_dir}.")
        return records

    for path in paths:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            summary.errors.append(f"{path.name}: could not open workbook: {exc}")
            continue
        try:
            codebook_signatures[path.name] = codebook_signature(
                workbook, path, summary
            )
            sheets = find_family_sheets(workbook)
            if len(sheets) != 1:
                summary.errors.append(
                    f"{path.name}: expected exactly one Families worksheet; "
                    f"found {len(sheets)}."
                )
            for sheet in sheets:
                header = find_header_row(sheet)
                if header is None:
                    summary.errors.append(
                        f"{path.name} / {sheet.title}: no Family header row found."
                    )
                    continue
                header_row, raw_headers = header
                actual_headers = row_values(raw_headers)
                header_signatures[path.name] = actual_headers
                if actual_headers != FAMILY_HEADERS:
                    summary.errors.append(
                        f"{path.name} / {sheet.title}: expected exact headers "
                        f"{FAMILY_HEADERS}; found {actual_headers}."
                    )
                    continue
                positions = {
                    clean(value): index
                    for index, value in enumerate(raw_headers)
                    if clean(value)
                }
                inferred_layer = filename_layer(path)
                if inferred_layer is None:
                    summary.errors.append(
                        f"{path.name}: filename does not identify one canonical layer."
                    )
                summary.family_sheets += 1
                for row_number, row in enumerate(
                    sheet.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    record = build_family_record(
                        row,
                        positions,
                        path,
                        sheet.title,
                        row_number,
                        inferred_layer,
                        summary,
                    )
                    if record is not None:
                        records.append(record)
        except Exception as exc:
            summary.errors.append(
                f"{path.name}: error while reading workbook: {exc}"
            )
        finally:
            workbook.close()

    if len(set(header_signatures.values())) > 1:
        summary.errors.append("Families worksheet header signatures are not identical.")
    if len(set(codebook_signatures.values())) > 1:
        summary.errors.append("Codebook worksheet content is not identical across workbooks.")
    return records


def load_drivers(path: Path, summary: Summary) -> list[dict[str, Any]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        summary.errors.append(f"Could not load {path}: {exc}")
        return []
    if not isinstance(payload, list):
        summary.errors.append(f"{path}: expected a top-level JSON array.")
        return []
    drivers: list[dict[str, Any]] = []
    for index, value in enumerate(payload):
        if not isinstance(value, dict):
            summary.errors.append(f"{path}: record {index} is not an object.")
            continue
        missing = [
            field_name
            for field_name in ("id", "name", "layer", "family")
            if not isinstance(value.get(field_name), str) or not value[field_name]
        ]
        if missing:
            summary.errors.append(
                f"{path}: record {index} lacks valid field(s): {missing}."
            )
            continue
        drivers.append(value)
    summary.drivers_total = len(drivers)
    return drivers


def validate_records(
    families: list[dict[str, Any]],
    drivers: list[dict[str, Any]],
    summary: Summary,
) -> None:
    if not families:
        summary.errors.append("No valid Family records were produced.")
        return
    if not drivers:
        summary.errors.append("No valid Driver records were available for validation.")
        return
    family_ids: dict[str, list[dict[str, Any]]] = defaultdict(list)
    family_names: dict[str, list[dict[str, Any]]] = defaultdict(list)
    family_keys: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    counts: dict[str, int] = defaultdict(int)
    for family in families:
        if tuple(family) != OUTPUT_KEYS:
            summary.errors.append(
                f"Family {family.get('id')!r} has a non-canonical key structure."
            )
        if family.get("id"):
            family_ids[family["id"]].append(family)
        if family.get("name"):
            family_names[family["name"]].append(family)
        if family.get("layer") and family.get("name"):
            family_keys[(family["layer"], family["name"])].append(family)
        if family.get("layer") in CANONICAL_LAYERS:
            counts[family["layer"]] += 1

    for family_id, matches in sorted(family_ids.items()):
        if len(matches) > 1:
            summary.duplicate_family_ids += 1
            summary.errors.append(
                f"Duplicate Family ID {family_id!r} ({len(matches)} records)."
            )
    for name, matches in sorted(family_names.items()):
        if len(matches) > 1:
            summary.duplicate_family_names += 1
            summary.errors.append(
                f"Duplicate Family name {name!r} ({len(matches)} records)."
            )
    for key, matches in sorted(family_keys.items()):
        if len(matches) > 1:
            summary.duplicate_layer_names += 1
            summary.errors.append(
                f"Duplicate Family layer/name pair {key!r} ({len(matches)} records)."
            )

    driver_ids: dict[str, list[dict[str, Any]]] = defaultdict(list)
    driver_names: dict[str, list[dict[str, Any]]] = defaultdict(list)
    actual_counts: dict[str, int] = defaultdict(int)
    for driver in drivers:
        driver_ids[driver["id"]].append(driver)
        driver_names[driver["name"]].append(driver)
        matches = family_keys.get((driver["layer"], driver["family"]), [])
        if not matches:
            summary.unmatched_drivers += 1
            summary.errors.append(
                f"Driver {driver['id']!r} does not match a Family for "
                f"{driver['layer']!r} / {driver['family']!r}."
            )
        elif len(matches) > 1:
            summary.ambiguous_driver_matches += 1
            summary.errors.append(
                f"Driver {driver['id']!r} ambiguously matches {len(matches)} Families."
            )
        else:
            summary.drivers_matched += 1
            actual_counts[matches[0]["id"]] += 1

    for driver_id, matches in sorted(driver_ids.items()):
        if len(matches) > 1:
            summary.errors.append(
                f"drivers.json contains duplicate Driver ID {driver_id!r}."
            )

    for family in families:
        actual = actual_counts.get(family["id"], 0)
        if actual == 0:
            summary.orphan_families += 1
            summary.errors.append(
                f"Family {family['id']!r} has no canonical Drivers."
            )
        if family["driverCount"] != actual:
            summary.count_mismatches += 1
            summary.errors.append(
                f"Family {family['id']!r}: workbook Driver Count "
                f"{family['driverCount']!r} does not equal derived count {actual}."
            )

        representatives = family["representativeDrivers"]
        duplicates = sorted(
            {name for name in representatives if representatives.count(name) > 1}
        )
        if duplicates:
            summary.errors.append(
                f"Family {family['id']!r} repeats Representative Driver(s): "
                + ", ".join(repr(name) for name in duplicates)
            )
        resolved_ids: list[str] = []
        for name in representatives:
            summary.representative_references += 1
            matches = driver_names.get(name, [])
            if not matches:
                summary.unresolved_representatives += 1
                summary.errors.append(
                    f"Family {family['id']!r}: Representative Driver "
                    f"{name!r} does not resolve."
                )
                continue
            if len(matches) > 1:
                summary.ambiguous_representatives += 1
                summary.errors.append(
                    f"Family {family['id']!r}: Representative Driver "
                    f"{name!r} resolves ambiguously to {len(matches)} Drivers."
                )
                continue
            driver = matches[0]
            if (
                driver["layer"] != family["layer"]
                or driver["family"] != family["name"]
            ):
                summary.wrong_family_representatives += 1
                summary.errors.append(
                    f"Family {family['id']!r}: Representative Driver {name!r} "
                    f"belongs to {driver['layer']!r} / {driver['family']!r}."
                )
                continue
            resolved_ids.append(driver["id"])
            summary.representatives_resolved += 1
        family["representativeDriverIds"] = resolved_ids

    summary.families_per_layer = {
        layer: counts[layer] for layer in CANONICAL_LAYERS if counts[layer]
    }


def family_sort_key(family: dict[str, Any]) -> tuple[int, str, str]:
    layer = family.get("layer")
    layer_index = (
        CANONICAL_LAYERS.index(layer)
        if layer in CANONICAL_LAYERS
        else len(CANONICAL_LAYERS)
    )
    return (
        layer_index,
        normalized_key(family.get("name")),
        str(family.get("id") or ""),
    )


def write_atomically(payload: dict[str, Any], output: Path) -> None:
    content = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            newline="\n",
            dir=output.parent,
            prefix=f".{output.name}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            handle.write(content)
            temporary = Path(handle.name)
        temporary.replace(output)
    finally:
        if temporary is not None and temporary.exists():
            temporary.unlink()


def print_summary(summary: Summary, output: Path, families: int, wrote: bool) -> None:
    print("Family import statistics")
    print(f"  Workbooks: {summary.workbooks}")
    print(f"  Families worksheets: {summary.family_sheets}")
    print(f"  Layers: {len(summary.families_per_layer)}")
    for layer, count in summary.families_per_layer.items():
        print(f"  Families / {layer}: {count}")
    print(f"  Total Families: {families}")
    print(
        f"  Drivers matched to Families: {summary.drivers_matched}/"
        f"{summary.drivers_total}"
    )
    print(
        f"  Representative Drivers resolved: "
        f"{summary.representatives_resolved}/{summary.representative_references}"
    )
    print(f"  Driver Count mismatches: {summary.count_mismatches}")
    print(f"  Duplicate Family IDs: {summary.duplicate_family_ids}")
    print(f"  Duplicate Family names: {summary.duplicate_family_names}")
    print(f"  Duplicate layer/name pairs: {summary.duplicate_layer_names}")
    print(f"  Orphan Families: {summary.orphan_families}")
    print(f"  Unmatched Drivers: {summary.unmatched_drivers}")
    print(f"  Ambiguous Driver/Family matches: {summary.ambiguous_driver_matches}")
    print(f"  Unresolved Representative Drivers: {summary.unresolved_representatives}")
    print(f"  Ambiguous Representative Drivers: {summary.ambiguous_representatives}")
    print(
        f"  Wrong-Family Representative Drivers: "
        f"{summary.wrong_family_representatives}"
    )
    print(f"  Warnings: {len(summary.warnings)}")
    print(f"  Errors: {len(summary.errors)}")
    for warning in summary.warnings:
        print(f"WARNING: {warning}")
    for error in summary.errors:
        print(f"ERROR: {error}", file=sys.stderr)
    if wrote:
        print(f"Wrote validated data to {output}.")
    else:
        print(f"Did not modify {output}.", file=sys.stderr)


def main() -> int:
    root = Path(__file__).resolve().parent.parent
    output = root / "data" / "families.json"
    summary = Summary()
    families = read_workbooks(root / "source-data", summary)
    drivers = load_drivers(root / "data" / "drivers.json", summary)
    validate_records(families, drivers, summary)
    families.sort(key=family_sort_key)
    payload = {"schemaVersion": SCHEMA_VERSION, "families": families}
    wrote = False
    if not summary.errors:
        try:
            write_atomically(payload, output)
            wrote = True
        except OSError as exc:
            summary.errors.append(f"Could not replace {output}: {exc}")
    print_summary(summary, output, len(families), wrote)
    return 0 if wrote else 1


if __name__ == "__main__":
    raise SystemExit(main())
