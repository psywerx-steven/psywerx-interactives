"""Apply the governed PSYWERX ontology-remediation release v1.

This one-time migration stages all eight XLSX workbooks, writes a complete
machine-readable diff and rollback manifest, validates the staged ontology,
and only then replaces the canonical private workbooks. Public JSON is rebuilt
separately by the normal repository builders.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import sys
from collections import Counter, defaultdict, deque
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "source-data"
RC = ROOT / "analysis" / "ontology_remediation_release_candidate_v1"
AUDIT = ROOT / "analysis" / "ontology_completeness_connectivity_audit_v1"
RELEASE = ROOT / "analysis" / "ontology_remediation_release_v1"
STAGING = RELEASE / "staging"
ROLLBACK = RELEASE / "rollback"

LAYERS = (
    "Biological", "Psychological", "Social", "Cultural",
    "Physical / Environmental", "Institutional / Structural",
    "Informational", "Technological",
)
PREFIX = {
    "Biological": "BIO", "Psychological": "PSY", "Social": "SOC",
    "Cultural": "CUL", "Physical / Environmental": "ENV",
    "Institutional / Structural": "INS", "Informational": "INF",
    "Technological": "TEC",
}
LEVEL = {
    "Biological": "PERSON", "Psychological": "PERSON", "Social": "SMALL_GROUP",
    "Cultural": "COMMUNITY", "Physical / Environmental": "PHYSICAL_SETTING",
    "Institutional / Structural": "INSTITUTIONAL_FIELD",
    "Informational": "INFORMATION_OBJECT_CORPUS",
    "Technological": "TECHNOLOGICAL_SYSTEM",
}
DRIVER_HEADERS = (
    "ID", "Name", "Other Names / Aliases", "Layer", "Family", "Definition",
    "Data Type", "Representation / Scale", "Polarity / Direction of Interpretation",
    "Mechanism", "Likely Upstream Influences", "Likely Downstream Influences",
    "Moderators / Boundary Conditions", "Typical Interaction Candidates",
    "Modifiability / Malleability", "Volatility", "Time Scale of Change",
    "Time Scale Qualifier", "Onset / Causal Lag", "Persistence / Recovery",
    "Indicators", "Measurement / Assessment Methods", "Observability",
    "Measurement Caveats", "Evidence Strength", "Evidence Notes",
    "Common Misinterpretations", "Key Sources",
)
REL_V1_HEADERS = (
    "Relationship ID", "Source Driver ID", "Source Driver", "Target Driver ID",
    "Target Driver", "Relationship Type", "Expected Direction", "Functional Form",
    "Moderators / Conditions", "Time Lag", "Evidence Strength", "Evidence Notes",
    "Evidence IDs",
)
REL_V2_HEADERS = (
    "Relationship ID", "Source Driver ID", "Source Driver", "Target Driver ID",
    "Target Driver", "Causal Role", "Polarity", "Directness", "Mechanism",
    "Conditions / Moderators", "Moderator Driver IDs", "Source Level", "Target Level",
    "Level-Transition Mechanism", "Lag Profile", "Lag Lower Bound", "Lag Upper Bound",
    "Lag Unit", "Lag Narrative", "Exposure Pattern", "Effect Persistence",
    "Evidence Strength", "Confidence", "Generalizability / Context",
    "Reciprocal Process ID", "Governance Class", "Supporting Evidence IDs",
    "Notes / Caveats",
)
FINAL_NAMES = {
    "PDC-001": "Cardiorespiratory Functional Reserve",
    "PDC-011": "Collective-Action Participation Threshold Distribution",
    "PDC-012": "Unverified-Claim Retransmission Rate",
    "PDC-033": "Content-Moderation Technical Enforcement Capacity",
    "PDC-036": "Coordinated Inauthentic Account Operation Capacity",
}
ID_MAP = {
    "PDC-001": "BIO-071", "PDC-003": "BIO-072",
    "PDC-005": "PSY-132", "PDC-006": "PSY-133", "PDC-007": "PSY-134", "PDC-008": "PSY-135",
    "PDC-009": "SOC-098", "PDC-011": "SOC-099", "PDC-012": "SOC-100",
    "PDC-013": "CUL-089", "PDC-014": "CUL-090", "PDC-015": "CUL-091",
    "PDC-017": "ENV-106", "PDC-018": "ENV-107", "PDC-019": "ENV-108", "PDC-020": "ENV-109",
    "PDC-021": "INS-110", "PDC-022": "INS-111", "PDC-023": "INS-112", "PDC-024": "INS-113", "PDC-025": "INS-114",
    "PDC-026": "INF-074", "PDC-027": "INF-075", "PDC-028": "INF-076",
    "PDC-030": "TEC-090", "PDC-031": "TEC-091", "PDC-032": "TEC-092", "PDC-033": "TEC-093",
    "PDC-034": "TEC-094", "PDC-035": "TEC-095", "PDC-036": "TEC-096",
}
MEDIATORS = {
    "PRC-004": "PSY-024", "PRC-005": "PSY-054", "PRC-007": "PSY-057",
    "PRC-031": "SOC-034", "PRC-033": "SOC-001", "PRC-034": "PSY-110",
    "PRC-038": "PSY-097", "PRC-039": "SOC-048", "PRC-041": "SOC-006",
    "PRC-042": "SOC-002", "PRC-043": "INF-019", "PRC-048": "PSY-134",
    "PRC-049": "PSY-134", "PRC-054": "SOC-076", "PRC-055": "INF-075",
    "PRC-056": "SOC-001", "PRC-067": "TEC-048", "PRC-083": "INS-063",
    "PRC-084": "SOC-070", "PRC-085": "PSY-057", "PRC-087": "ENV-109",
}
SME = {"PRC-044", "PRC-057", "PRC-088"}
RECIPROCAL = {"PRC-022", "PRC-032"}
CODEBOOK_ROWS = (
    ("CB-REL-CAUSAL-ROLE", "Causal Role", "Minimal causal semantic class.", "CAUSES; ENABLES; CONSTRAINS; MODERATES"),
    ("CB-REL-POLARITY", "Polarity", "Expected local sign or form under stated conditions.", "POSITIVE; NEGATIVE; NON_MONOTONIC; CONTEXT_DEPENDENT; UNSIGNED"),
    ("CB-REL-DIRECTNESS", "Directness", "Whether the edge is direct at its governed resolution or a mediated-path segment.", "DIRECT_AT_STATED_RESOLUTION; MEDIATED_PATH; UNKNOWN"),
    ("CB-REL-LEVEL", "Source Level / Target Level", "Entity or system level instantiated by an edge endpoint.", "PERSON; DYAD_INTERPERSONAL; SMALL_GROUP; NETWORK; COMMUNITY; ORGANIZATION; INSTITUTIONAL_FIELD; SOCIETY; STATE_JURISDICTION; INFORMATION_OBJECT_CORPUS; INFORMATION_SYSTEM; TECHNOLOGICAL_SYSTEM; PHYSICAL_SETTING; ECOLOGICAL_SYSTEM"),
    ("CB-REL-LAG-PROFILE", "Lag Profile", "Categorical source-to-target onset lag, distinct from Driver change speed.", "IMMEDIATE; SHORT; INTERMEDIATE; DELAYED; LONG; STRUCTURAL; INTERGENERATIONAL; MIXED_CONTEXT_DEPENDENT"),
    ("CB-REL-EXPOSURE-PATTERN", "Exposure Pattern", "Pattern by which the source acts over the governed edge.", "PULSE; SUSTAINED; CUMULATIVE; REPEATED; NOT_SPECIFIED"),
    ("CB-REL-CONFIDENCE", "Confidence", "Governance confidence in the edge specification.", "HIGH; MODERATE; LOW"),
    ("CB-REL-GOVERNANCE-CLASS", "Governance Class", "Controls whether and how an edge belongs in the canonical graph.", "CORE; CONTEXT_DEPENDENT; SCENARIO_SPECIFIC; HYPOTHESIZED"),
)


def text(value: Any) -> str:
    return "" if value is None else " ".join(str(value).split())


def split(value: Any) -> list[str]:
    return [text(part) for part in str(value or "").split(";") if text(part)]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], headers: list[str] | tuple[str, ...] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(headers or (rows[0].keys() if rows else []))
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader(); writer.writerows(rows)


def copy_row_style(sheet: Any, source_row: int, target_row: int, width: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, width + 1):
        source = sheet.cell(source_row, column); target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def layer_from_filename(name: str) -> str:
    folded = name.casefold()
    for layer, prefix in PREFIX.items():
        tokens = layer.replace(" / ", "_").replace(" ", "_").casefold()
        if f"layer_{LAYERS.index(layer)+1}_{tokens}" in folded:
            return layer
    raise ValueError(f"Cannot infer layer from {name}")


def lag_profile(value: str) -> str:
    normalized = value.casefold().replace("—", "–").replace("-", "–")
    units = ["seconds", "minutes", "hours", "days", "weeks", "months", "years", "generations"]
    bands = ["IMMEDIATE", "SHORT", "INTERMEDIATE", "DELAYED", "LONG", "STRUCTURAL", "INTERGENERATIONAL"]
    positions = [index for index, unit in enumerate(units) if unit in normalized]
    if not positions:
        return "MIXED_CONTEXT_DEPENDENT"
    low, high = min(positions), max(positions)
    selected = bands[low:min(high, len(bands) - 1) + 1]
    return "; ".join(selected or [bands[min(low, len(bands) - 1)]])


def relationship_role(value: str) -> str:
    folded = value.casefold()
    if "constrain" in folded or "buffer" in folded or "protect" in folded:
        return "CONSTRAINS"
    if "enabl" in folded or "opportunity" in folded:
        return "ENABLES"
    # A legacy label saying "moderating influence" is not sufficient to use
    # Schema v2 MODERATES, which requires an identified governed relationship
    # or relationship pattern. Preserve these as causal contributions unless a
    # future adjudication supplies that missing identity.
    if "moderat" in folded:
        return "CAUSES"
    return "CAUSES"


def relationship_polarity(value: str) -> str:
    return {"positive": "POSITIVE", "negative": "NEGATIVE", "inverted-u": "NON_MONOTONIC", "contingent": "CONTEXT_DEPENDENT", "bidirectional": "CONTEXT_DEPENDENT"}.get(value.casefold(), "UNSIGNED")


def source_registers() -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    pattern = re.compile(r"^- \*\*(SRC-R\d+|RC-R\d+)\s+—\s+(.+?)\*\*(.*)$")
    url = re.compile(r"https?://[^)\s]+")
    for path in (AUDIT / "research_sources.md", RC / "research_sources.md"):
        for line in path.read_text(encoding="utf-8").splitlines():
            match = pattern.match(line)
            if not match:
                continue
            identifier, title, rest = match.groups()
            found = url.findall(rest)
            result[identifier] = {"title": title.rstrip("."), "url": found[0] if found else "", "finding": text(re.sub(r"\[[^]]+\]\([^)]+\)", "", rest)).lstrip(". ")}
    return result


def prepare_release() -> tuple[list[Path], list[dict[str, Any]]]:
    paths = [p for p in sorted(SOURCE.glob("*.xlsx")) if not p.name.startswith("~$")]
    if len(paths) != 8:
        raise ValueError(f"Expected eight canonical XLSX workbooks; found {len(paths)}")
    if RELEASE.exists():
        raise ValueError(f"Release directory already exists: {RELEASE.relative_to(ROOT)}")
    STAGING.mkdir(parents=True); ROLLBACK.mkdir(parents=True)
    rollback_rows: list[dict[str, Any]] = []
    protected = paths + sorted((ROOT / "data").glob("*.json"))
    for path in protected:
        relative = path.relative_to(ROOT)
        destination = ROLLBACK / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, destination)
        rollback_rows.append({
            "canonicalPath": relative.as_posix(),
            "rollbackPath": destination.relative_to(ROOT).as_posix(),
            "preMigrationSha256": sha256(path),
        })
    for path in paths:
        shutil.copy2(path, STAGING / path.name)
    write_csv(RELEASE / "rollback_manifest.csv", rollback_rows)
    return paths, rollback_rows


def existing_driver_index(paths: list[Path]) -> tuple[dict[str, dict[str, Any]], dict[str, Path]]:
    drivers: dict[str, dict[str, Any]] = {}
    layer_paths: dict[str, Path] = {}
    for path in paths:
        layer = layer_from_filename(path.name); layer_paths[layer] = path
        book = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = book["Drivers"]
            headers = tuple(text(cell.value) for cell in sheet[4][:28])
            if headers != DRIVER_HEADERS:
                raise ValueError(f"{path.name}: Driver Schema v1.1 header mismatch")
            for row_number, row in enumerate(sheet.iter_rows(min_row=5, max_col=28, values_only=True), 5):
                values = [text(v) for v in row]
                if not any(values):
                    continue
                identifier = values[0]
                if identifier in drivers:
                    raise ValueError(f"Duplicate pre-existing Driver ID {identifier}")
                drivers[identifier] = {header: value for header, value in zip(DRIVER_HEADERS, values)} | {"row": row_number, "workbook": path.name}
        finally:
            book.close()
    if len(drivers) != 762:
        raise ValueError(f"Expected 762 pre-existing Drivers; found {len(drivers)}")
    return drivers, layer_paths


def build_driver_rows(evidence_map: dict[str, str]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    audit = {row["candidateKey"]: row for row in csv_rows(AUDIT / "proposed_driver_candidates.csv")}
    adjudicated = {row["candidateKey"]: row for row in csv_rows(RC / "driver_candidate_adjudication.csv")}
    if set(ID_MAP) != {key for key, row in adjudicated.items() if row["disposition"] in {"ADD", "ADD WITH REVISION"}}:
        raise ValueError("Permanent Driver ID manifest does not exactly cover approved additions")
    records: list[dict[str, Any]] = []
    manifest: list[dict[str, Any]] = []
    for key, identifier in ID_MAP.items():
        source = audit[key]; decision = adjudicated[key]
        name = FINAL_NAMES.get(key, source["provisionalName"])
        definition = source["shortDefinition"]
        if key == "PDC-001":
            definition = "Current integrated cardiac and pulmonary reserve available to sustain oxygen delivery and physical work above resting demand for a specified person and context."
        elif key == "PDC-011":
            definition = "Distribution of participation thresholds at which members of a specified population become willing to join a collective action given observed participation by others."
        elif key == "PDC-012":
            definition = "Rate at which exposed recipients retransmit an unverified claim through a specified social network, population, channel, and interval."
        elif key == "PDC-033":
            definition = "Technical capacity of a platform or system to detect, prioritize, restrict, label, or remove governed content at the scale and speed required by its policy."
        elif key == "PDC-036":
            definition = "Operational capacity to deploy and coordinate automated or semi-automated accounts that misrepresent independent human participation across a specified platform and period."
        data_type = "Magnitude / level"
        if key == "PDC-011": data_type = "Multidimensional"
        if key == "PDC-012": data_type = "Rate"
        upstream = source["expectedUpstreamCauses"]
        downstream = source["expectedDownstreamConsequences"]
        sources = [evidence_map[item] for item in split(decision["supportingSourceRefs"])]
        representation = "Continuous or ordinal construct-specific measure with population, setting, reference period, and unit explicitly specified."
        indicators = f"Construct-valid indicators of {name.casefold()} measured for the specified unit, setting, and reference period"
        mechanism = source["causalRationale"].rstrip(".") + "."
        record = {
            "ID": identifier, "Name": name, "Other Names / Aliases": source["provisionalName"] if name != source["provisionalName"] else "",
            "Layer": source["proposedLayer"], "Family": source["proposedFamily"], "Definition": definition,
            "Data Type": data_type, "Representation / Scale": representation,
            "Polarity / Direction of Interpretation": f"Higher values = greater {name.casefold()}.",
            "Mechanism": mechanism, "Likely Upstream Influences": upstream,
            "Likely Downstream Influences": downstream,
            "Moderators / Boundary Conditions": "Interpret only for the specified population, unit, setting, exposure, reference period, and measurement method; effects may vary across contexts.",
            "Typical Interaction Candidates": "; ".join(split(upstream)[:2] + split(downstream)[:2]),
            "Modifiability / Malleability": "Moderate", "Volatility": "Moderate",
            "Time Scale of Change": "Hours–Days", "Time Scale Qualifier": "Typical state-change speed is context-dependent; do not infer persistence from this band.",
            "Onset / Causal Lag": "Mixed / Context-dependent", "Persistence / Recovery": "Persistence and recovery depend on exposure, baseline state, intervention, and system feedback.",
            "Indicators": indicators, "Measurement / Assessment Methods": "Use a validated construct-specific instrument, administrative or sensor measure, or transparent composite with explicit unit and reference period.",
            "Observability": "Moderate", "Measurement Caveats": decision["measurementFeasibility"].replace("_", " ").title() + "; proxy measures require construct-validity review.",
            "Evidence Strength": "Strong" if decision["decisionConfidence"] == "HIGH" else "Moderate",
            "Evidence Notes": decision["adjudicationRationale"],
            "Common Misinterpretations": f"Do not treat {name} as interchangeable with {decision['nearestExistingDriverIds']} or as a universal, context-free cause.",
            "Key Sources": "; ".join(sources),
        }
        records.append(record)
        manifest.append({
            "candidateKey": key, "permanentDriverId": identifier, "canonicalName": name,
            "layer": record["Layer"], "family": record["Family"],
            "idConvention": f"next unused {PREFIX[record['Layer']]} sequence",
            "validationStatus": "VALIDATED",
        })
    return records, manifest


def apply_duplicate_repairs(book_by_layer: dict[str, Any], evidence_map: dict[str, str]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for group in csv_rows(RC / "duplicate_resolution_plan.csv"):
        ids = split(group["driverIds"])
        for index, identifier in enumerate(ids):
            layer = next(layer for layer, prefix in PREFIX.items() if identifier.startswith(prefix + "-"))
            sheet = book_by_layer[layer]["Drivers"]
            row = next((r for r in range(5, sheet.max_row + 1) if text(sheet.cell(r, 1).value) == identifier), None)
            if row is None:
                raise ValueError(f"Duplicate repair target {identifier} not found")
            definition = group["proposedDefinitionA" if index == 0 else "proposedDefinitionB"]
            mechanism = group["proposedMechanismA" if index == 0 else "proposedMechanismB"]
            sheet.cell(row, 6).value = definition
            sheet.cell(row, 10).value = mechanism
            sheet.cell(row, 13).value = group["boundaryRule"]
            sheet.cell(row, 24).value = group["causalUse"]
            sheet.cell(row, 27).value = group["boundaryRule"]
            existing_sources = split(sheet.cell(row, 28).value)
            for ref in split(group["evidenceRefs"]):
                if evidence_map[ref] not in existing_sources:
                    existing_sources.append(evidence_map[ref])
            sheet.cell(row, 28).value = "; ".join(existing_sources)
        output.append({
            "groupId": group["groupId"], "driverIds": group["driverIds"],
            "finalDisposition": group["disposition"], "implementationResult": "APPLIED",
            "boundaryRule": group["boundaryRule"], "causalUse": group["causalUse"],
            "evidenceIds": "; ".join(evidence_map[ref] for ref in split(group["evidenceRefs"])),
        })
    return output


def append_evidence(book_by_layer: dict[str, Any], registers: dict[str, dict[str, str]], evidence_map: dict[str, str], driver_records: list[dict[str, Any]]) -> None:
    usage: dict[str, list[str]] = defaultdict(list)
    home: dict[str, str] = {}
    adjudicated = {row["candidateKey"]: row for row in csv_rows(RC / "driver_candidate_adjudication.csv")}
    for record, key in zip(driver_records, ID_MAP):
        for ref in split(adjudicated[key]["supportingSourceRefs"]):
            usage[ref].append(record["ID"]); home.setdefault(ref, record["Layer"])
    relationships = csv_rows(RC / "relationship_candidate_adjudication.csv")
    existing_json = json.loads((ROOT / "data" / "drivers.json").read_text(encoding="utf-8"))
    old_layers = {record["id"]: record["layer"] for record in existing_json}
    for candidate in relationships:
        for ref in split(candidate["supportingSourceRefs"]):
            usage[ref].extend([candidate["sourceDriverId"], candidate["targetDriverId"]])
            home.setdefault(ref, old_layers[candidate["sourceDriverId"]])
    for group in csv_rows(RC / "duplicate_resolution_plan.csv"):
        first_id = split(group["driverIds"])[0]
        for ref in split(group["evidenceRefs"]):
            usage[ref].extend(split(group["driverIds"])); home.setdefault(ref, old_layers[first_id])
    for ref in sorted(evidence_map, key=lambda item: int(item.split("R")[-1]) + (100 if item.startswith("RC-") else 0)):
        layer = home.get(ref, "Informational")
        sheet = book_by_layer[layer]["Evidence Library"]
        row = sheet.max_row + 1; copy_row_style(sheet, row - 1, row, 13)
        source = registers[ref]
        values = (
            evidence_map[ref], "; ".join(dict.fromkeys(usage.get(ref, []))), "Ontology remediation",
            source["title"], "", "Governed scientific or technical source", "Varies by source",
            source["finding"] or "Supports the governed construct or causal specification described in the remediation release.",
            "Context-dependent", "See source and remediation adjudication for scope and limitations.",
            "Moderate", source["url"], "Verified in ontology-remediation source register; retrieval date 2026-08-27.",
        )
        for column, value in enumerate(values, 1): sheet.cell(row, column).value = value


def append_drivers(book_by_layer: dict[str, Any], records: list[dict[str, Any]]) -> None:
    for record in records:
        sheet = book_by_layer[record["Layer"]]["Drivers"]
        row = sheet.max_row + 1; copy_row_style(sheet, row - 1, row, len(DRIVER_HEADERS))
        for column, header in enumerate(DRIVER_HEADERS, 1): sheet.cell(row, column).value = record[header]
    for layer, book in book_by_layer.items():
        sheet = book["Drivers"]
        counts = Counter(text(sheet.cell(row, 5).value) for row in range(5, sheet.max_row + 1) if text(sheet.cell(row, 1).value))
        families = book["Families"]
        for row in range(5, families.max_row + 1):
            family = text(families.cell(row, 2).value)
            if family: families.cell(row, 8).value = counts[family]
        summary = book["Layer Summary"]
        for row in range(5, summary.max_row + 1):
            metric = text(summary.cell(row, 4).value).casefold()
            if metric == "canonical drivers": summary.cell(row, 5).value = sum(counts.values())
            elif metric == "families": summary.cell(row, 5).value = len(counts)
            elif metric == "evidence records": summary.cell(row, 5).value = sum(bool(text(book["Evidence Library"].cell(r, 1).value)) for r in range(5, book["Evidence Library"].max_row + 1))


def append_codebook(book_by_layer: dict[str, Any]) -> None:
    for book in book_by_layer.values():
        sheet = book["Codebook"]
        existing = {text(sheet.cell(row, 1).value) for row in range(5, sheet.max_row + 1)}
        for identifier, field, definition, values in CODEBOOK_ROWS:
            if identifier in existing:
                continue
            row = sheet.max_row + 1; copy_row_style(sheet, row - 1, row, 7)
            cells = (identifier, "Relationships", field, definition, values, "Yes", "Use exact controlled values; preserve narrative qualifications in the dedicated context or notes field.")
            for column, value in enumerate(cells, 1): sheet.cell(row, column).value = value


def migrate_existing_relationships(book_by_layer: dict[str, Any], drivers: dict[str, dict[str, str]]) -> dict[tuple[str, str], dict[str, Any]]:
    pairs: dict[tuple[str, str], dict[str, Any]] = {}
    for layer, book in book_by_layer.items():
        sheet = book["Relationships"]
        actual = tuple(text(sheet.cell(4, column).value) for column in range(1, 14))
        if actual != REL_V1_HEADERS:
            raise ValueError(f"{layer}: expected Relationship Schema v1 before migration")
        old_rows = []
        for row in sheet.iter_rows(min_row=5, max_col=13, values_only=True):
            values = [text(value) for value in row]
            if any(values): old_rows.append(values)
        for column, header in enumerate(REL_V2_HEADERS, 1): sheet.cell(4, column).value = header
        for row in range(5, max(sheet.max_row, 5) + 1):
            for column in range(1, len(REL_V2_HEADERS) + 1): sheet.cell(row, column).value = None
        for index, old in enumerate(old_rows, 5):
            source, target = old[1], old[3]
            source_level, target_level = LEVEL[drivers[source]["layer"]], LEVEL[drivers[target]["layer"]]
            transition = "The source state reaches the target unit through exposure, communication, allocation, aggregation, or embodied response specified by the edge mechanism." if source_level != target_level else ""
            evidence = old[11] or "Canonical source worksheet evidence notes."
            values = [
                old[0], source, drivers[source]["name"], target, drivers[target]["name"],
                relationship_role(old[5]), relationship_polarity(old[6]), "DIRECT_AT_STATED_RESOLUTION",
                evidence, old[8] or "Applies only under the population, setting, baseline, and exposure conditions represented by the supporting evidence.",
                "", source_level, target_level, transition, lag_profile(old[9]), "", "", "", old[9],
                "NOT_SPECIFIED", "", old[10], "HIGH" if old[10] == "Strong" else "MODERATE" if old[10] in {"Moderate", "Mixed"} else "LOW",
                "Generalizability is limited to contexts compatible with the stated mechanism, conditions, and evidence.",
                "", "CONTEXT_DEPENDENT", old[12], f"Migrated from v1 relationship type: {old[5]}. Legacy functional form: {old[7]}.",
            ]
            for column, value in enumerate(values, 1): sheet.cell(index, column).value = value
            pairs[(source, target)] = {"layer": layer, "row": index, "id": old[0], "values": values}
    return pairs


def append_relationships(book_by_layer: dict[str, Any], drivers: dict[str, dict[str, Any]], evidence_map: dict[str, str], pairs: dict[tuple[str, str], dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    candidates = csv_rows(RC / "relationship_candidate_adjudication.csv")
    maximum: dict[str, int] = defaultdict(int)
    for item in pairs.values():
        match = re.search(r"(\d+)$", item["id"])
        if match: maximum[item["layer"]] = max(maximum[item["layer"]], int(match.group(1)))
    changes: list[dict[str, Any]] = []
    mediated: list[dict[str, Any]] = []
    blocked: list[dict[str, Any]] = []

    def add_or_enrich(source: str, target: str, candidate: dict[str, str], origin: str, segment: int | None = None, polarity: str | None = None) -> str:
        pair = (source, target)
        evidence_ids = [evidence_map[ref] for ref in split(candidate["supportingSourceRefs"])]
        governance = "CORE" if candidate["contextClass"] == "CORE_STRUCTURAL" else "CONTEXT_DEPENDENT"
        causal_role = candidate["proposedCausalRole"].split("_")[0]
        if causal_role not in {"CAUSES", "ENABLES", "CONSTRAINS", "MODERATES"}: causal_role = "CAUSES"
        edge_polarity = polarity or ({"CONTINGENT": "CONTEXT_DEPENDENT"}.get(candidate["polarity"], candidate["polarity"]))
        source_level, target_level = LEVEL[drivers[source]["layer"]], LEVEL[drivers[target]["layer"]]
        transition = "The source state reaches the target unit through the exposure, communication, allocation, aggregation, or embodied-response process stated in the mechanism." if source_level != target_level else ""
        mechanism = candidate["mechanism"]
        if segment == 1: mechanism = f"First governed segment of {origin}: {drivers[source]['name']} changes the mediator state {drivers[target]['name']} through the scoped process in the approved candidate."
        elif segment == 2: mechanism = f"Second governed segment of {origin}: change in mediator {drivers[source]['name']} contributes to {drivers[target]['name']} under the approved conditions."
        notes = f"Implemented from remediation candidate {origin}."
        if segment: notes += f" Mediated-path segment {segment}; the original shortcut edge is not canonical."
        if pair in pairs:
            item = pairs[pair]; values = item["values"]
            values[5] = causal_role; values[6] = edge_polarity
            if segment: values[7] = "MEDIATED_PATH"
            values[8] = mechanism; values[9] = candidate["conditionsModerators"]
            values[11] = source_level; values[12] = target_level; values[13] = transition
            values[14] = lag_profile(candidate["timeLag"]); values[18] = candidate["timeLag"]
            values[21] = "Strong" if candidate["evidenceStrength"] == "HIGH" else "Moderate"
            values[22] = candidate["decisionConfidence"]; values[23] = candidate["adjudicationRationale"]
            values[24] = "RP-NORM-PREVALENCE-001" if origin in RECIPROCAL else values[24]
            values[25] = governance
            current_evidence = split(values[26]); values[26] = "; ".join(current_evidence + [eid for eid in evidence_ids if eid not in current_evidence])
            values[27] = text(str(values[27] or "") + " " + notes)
            sheet = book_by_layer[item["layer"]]["Relationships"]
            for column, value in enumerate(values, 1): sheet.cell(item["row"], column).value = value
            action = "ENRICH_EXISTING"
            relationship_id = item["id"]
        else:
            layer = drivers[source]["layer"]; maximum[layer] += 1
            relationship_id = f"REL-{PREFIX[layer]}-{maximum[layer]:03d}"
            values = [
                relationship_id, source, drivers[source]["name"], target, drivers[target]["name"], causal_role,
                edge_polarity, "MEDIATED_PATH" if segment else "DIRECT_AT_STATED_RESOLUTION", mechanism,
                candidate["conditionsModerators"], "", source_level, target_level, transition,
                lag_profile(candidate["timeLag"]), "", "", "", candidate["timeLag"], "NOT_SPECIFIED", "",
                "Strong" if candidate["evidenceStrength"] == "HIGH" else "Moderate", candidate["decisionConfidence"],
                candidate["adjudicationRationale"], "RP-NORM-PREVALENCE-001" if origin in RECIPROCAL else "",
                governance, "; ".join(evidence_ids), notes,
            ]
            sheet = book_by_layer[layer]["Relationships"]
            row = sheet.max_row + 1; copy_row_style(sheet, row - 1, row, len(REL_V2_HEADERS))
            for column, value in enumerate(values, 1): sheet.cell(row, column).value = value
            pairs[pair] = {"layer": layer, "row": row, "id": relationship_id, "values": values}
            action = "ADD"
        changes.append({
            "candidateRelationshipKey": origin, "relationshipId": relationship_id,
            "sourceDriverId": source, "targetDriverId": target, "action": action,
            "pathSegment": segment or "", "causalRole": causal_role, "polarity": edge_polarity,
            "directness": "MEDIATED_PATH" if segment else "DIRECT_AT_STATED_RESOLUTION",
            "governanceClass": governance, "implementationStatus": "IMPLEMENTED",
        })
        return relationship_id

    polarity_overrides = {
        "PRC-005": ("NEGATIVE", "NEGATIVE"), "PRC-041": ("NEGATIVE", "NEGATIVE"),
        "PRC-042": ("POSITIVE", "NEGATIVE"), "PRC-054": ("NEGATIVE", "POSITIVE"),
        "PRC-085": ("POSITIVE", "NEGATIVE"),
    }
    for candidate in candidates:
        key = candidate["candidateRelationshipKey"]
        if key in SME:
            blocked.append({
                "itemType": "RELATIONSHIP", "itemKey": key, "sourceDriverId": candidate["sourceDriverId"],
                "targetDriverId": candidate["targetDriverId"], "status": "SME_REVIEW_REQUIRED",
                "issue": candidate["adjudicationRationale"], "evidence": candidate["supportingSourceRefs"],
                "specificQuestion": "Does evidence support this direction, mechanism, polarity, and scope strongly enough for a canonical cross-layer edge?",
            })
            continue
        if candidate["disposition"] == "REPLACE WITH MEDIATED PATH":
            mediator = MEDIATORS[key]
            pol1, pol2 = polarity_overrides.get(key, (candidate["polarity"].replace("CONTINGENT", "CONTEXT_DEPENDENT"), "POSITIVE"))
            first = add_or_enrich(candidate["sourceDriverId"], mediator, candidate, key, 1, pol1)
            second = add_or_enrich(mediator, candidate["targetDriverId"], candidate, key, 2, pol2)
            mediated.append({
                "candidateRelationshipKey": key, "sourceDriverId": candidate["sourceDriverId"],
                "mediatorDriverIds": mediator, "targetDriverId": candidate["targetDriverId"],
                "segment1RelationshipId": first, "segment2RelationshipId": second,
                "shortcutImplemented": "NO", "status": "IMPLEMENTED",
            })
        else:
            add_or_enrich(candidate["sourceDriverId"], candidate["targetDriverId"], candidate, key)

    # Three scientifically causal duplicate-boundary links. Taxonomy hierarchy
    # alone (for ENV-060/061) is deliberately not encoded as a causal edge.
    repair_edges = (
        ("ENV-064", "ENV-065", "CAUSES", "POSITIVE", "RC-R08", "Persistent indoor dampness enables mould growth and allergen accumulation."),
        ("ENV-016", "ENV-017", "CAUSES", "POSITIVE", "RC-R10", "Road congestion produces travel-time delay relative to the specified reference condition."),
        ("INS-080", "INS-079", "ENABLES", "POSITIVE", "RC-R05; RC-R13", "Bribery opportunities enable one specific class of institutional corruption without being synonymous with completed corruption."),
    )
    for index, (source, target, role, polarity, refs, mechanism) in enumerate(repair_edges, 1):
        pseudo = {
            "supportingSourceRefs": refs, "contextClass": "CONTEXT_DEPENDENT", "proposedCausalRole": role,
            "polarity": polarity, "mechanism": mechanism,
            "conditionsModerators": "Use only when the source and target are separately observed for the same setting and reference period.",
            "timeLag": "Mixed / Context-dependent", "evidenceStrength": "MODERATE", "decisionConfidence": "HIGH",
            "adjudicationRationale": "Approved duplicate-boundary causal link; constructs remain distinct.",
        }
        add_or_enrich(source, target, pseudo, f"DDG-EDGE-{index:03d}")
    return changes, mediated, blocked


def validate_staged(paths: list[Path], pre_ids: set[str]) -> dict[str, Any]:
    drivers: dict[str, dict[str, str]] = {}; families: dict[tuple[str, str], int] = {}
    evidence: set[str] = set(); relationship_ids: set[str] = set(); pairs: set[tuple[str, str]] = set()
    codebook_signatures = []
    layer_driver_counts = Counter(); relationship_count = 0
    for path in paths:
        layer = layer_from_filename(path.name)
        book = load_workbook(path, read_only=True, data_only=True)
        try:
            driver_sheet = book["Drivers"]
            if tuple(text(cell.value) for cell in driver_sheet[4][:28]) != DRIVER_HEADERS:
                raise ValueError(f"{path.name}: staged Driver headers invalid")
            for row in driver_sheet.iter_rows(min_row=5, max_col=28, values_only=True):
                values = [text(value) for value in row]
                if not any(values): continue
                identifier = values[0]
                if identifier in drivers: raise ValueError(f"Duplicate staged Driver ID {identifier}")
                if values[3] != layer: raise ValueError(f"{identifier}: staged Layer mismatch")
                if any(not values[index] for index in (0, 1, 3, 4, 5, 6, 7, 8, 9, 12, 14, 15, 16, 18, 19, 21, 22, 23, 24, 25, 26, 27)):
                    raise ValueError(f"{identifier}: required Driver field missing")
                drivers[identifier] = {"name": values[1], "layer": values[3], "family": values[4]}; layer_driver_counts[layer] += 1
            family_sheet = book["Families"]
            for row in family_sheet.iter_rows(min_row=5, max_col=8, values_only=True):
                if text(row[0]): families[(layer, text(row[1]))] = int(row[7])
            evidence_sheet = book["Evidence Library"]
            for row in evidence_sheet.iter_rows(min_row=5, max_col=13, values_only=True):
                identifier = text(row[0])
                if identifier:
                    if identifier in evidence: raise ValueError(f"Duplicate Evidence ID {identifier}")
                    evidence.add(identifier)
            codebook = book["Codebook"]
            signature = tuple(tuple(text(value) for value in row) for row in codebook.iter_rows(values_only=True) if any(text(value) for value in row))
            codebook_signatures.append(signature)
        finally:
            book.close()
    if len(drivers) != 793 or not pre_ids.issubset(drivers) or set(drivers) - pre_ids != set(ID_MAP.values()):
        raise ValueError(f"Staged Driver identity validation failed: {len(drivers)} records")
    normalized_names = Counter((record["layer"], re.sub(r"[^a-z0-9]+", " ", record["name"].casefold()).strip()) for record in drivers.values())
    duplicates = [key for key, count in normalized_names.items() if count > 1]
    if duplicates: raise ValueError(f"Duplicate normalized Driver names within Layer: {duplicates}")
    actual_family_counts = Counter((record["layer"], record["family"]) for record in drivers.values())
    if actual_family_counts != Counter(families): raise ValueError("Staged Family Driver Counts do not match Drivers")
    if len(codebook_signatures) != 8 or len(set(codebook_signatures)) != 1:
        raise ValueError("Staged Codebooks are not identical")
    semantic_terms = [row for row in codebook_signatures[0] if row and row[0].startswith("CB-")]
    if len(semantic_terms) != 56 or len({row[0] for row in semantic_terms}) != 56:
        raise ValueError(f"Expected 56 unique Codebook terms; found {len(semantic_terms)}")

    adjacency: dict[str, set[str]] = defaultdict(set); degree = Counter(); incoming = Counter(); outgoing = Counter(); cross = Counter()
    roles = Counter(); governance = Counter(); directness = Counter()
    for path in paths:
        layer = layer_from_filename(path.name)
        book = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = book["Relationships"]
            if tuple(text(sheet.cell(4, column).value) for column in range(1, 29)) != REL_V2_HEADERS:
                raise ValueError(f"{path.name}: staged Relationship Schema v2 headers invalid")
            for row in sheet.iter_rows(min_row=5, max_col=28, values_only=True):
                values = [text(value) for value in row]
                if not any(values): continue
                identifier, source, target = values[0], values[1], values[3]
                if identifier in relationship_ids: raise ValueError(f"Duplicate Relationship ID {identifier}")
                if (source, target) in pairs: raise ValueError(f"Duplicate directed pair {source} -> {target}")
                if source not in drivers or target not in drivers: raise ValueError(f"{identifier}: dangling endpoint")
                if drivers[source]["layer"] != layer: raise ValueError(f"{identifier}: source-workbook Layer mismatch")
                if values[2] != drivers[source]["name"] or values[4] != drivers[target]["name"]: raise ValueError(f"{identifier}: endpoint name mismatch")
                if values[11] != values[12] and not values[13]: raise ValueError(f"{identifier}: missing cross-level mechanism")
                for evidence_id in split(values[26]):
                    if evidence_id not in evidence: raise ValueError(f"{identifier}: missing Evidence {evidence_id}")
                relationship_ids.add(identifier); pairs.add((source, target)); relationship_count += 1
                roles[values[5]] += 1; directness[values[7]] += 1; governance[values[25]] += 1
                adjacency[source].add(target); adjacency[target].add(source); degree[source] += 1; degree[target] += 1; outgoing[source] += 1; incoming[target] += 1
                if drivers[source]["layer"] != drivers[target]["layer"]: cross[f"{drivers[source]['layer']} -> {drivers[target]['layer']}"] += 1
        finally: book.close()
    seen: set[str] = set(); components: list[int] = []
    for identifier in drivers:
        if identifier in seen: continue
        queue = deque([identifier]); seen.add(identifier); size = 0
        while queue:
            current = queue.popleft(); size += 1
            for other in adjacency[current]:
                if other not in seen: seen.add(other); queue.append(other)
        components.append(size)
    return {
        "driverCount": len(drivers), "driversPerLayer": dict(layer_driver_counts), "familyCount": len(families),
        "evidenceCount": len(evidence), "codebookTermCount": len(semantic_terms), "relationshipCount": relationship_count,
        "causalRoles": dict(sorted(roles.items())), "governanceClasses": dict(sorted(governance.items())),
        "directness": dict(sorted(directness.items())), "crossLayerPairs": dict(sorted(cross.items())),
        "directedLayerPairCoverage": len(cross), "weakComponents": len(components),
        "isolatedDrivers": sum(degree[node] == 0 for node in drivers), "degreeOneDrivers": sum(degree[node] == 1 for node in drivers),
        "driversWithIncomingAndOutgoing": sum(incoming[node] and outgoing[node] for node in drivers), "largestWeakComponent": max(components),
    }


def pre_metrics() -> dict[str, Any]:
    drivers = json.loads((ROOT / "data" / "drivers.json").read_text(encoding="utf-8"))
    payload = json.loads((ROOT / "data" / "relationships.json").read_text(encoding="utf-8"))
    relationships = payload["relationships"]
    ids = {row["id"] for row in drivers}; layers = {row["id"]: row["layer"] for row in drivers}
    adjacency: dict[str, set[str]] = defaultdict(set); degree = Counter(); incoming = Counter(); outgoing = Counter(); cross = Counter()
    for row in relationships:
        source, target = row["sourceDriverId"], row["targetDriverId"]
        adjacency[source].add(target); adjacency[target].add(source); degree[source] += 1; degree[target] += 1; outgoing[source] += 1; incoming[target] += 1
        if layers[source] != layers[target]: cross[f"{layers[source]} -> {layers[target]}"] += 1
    seen = set(); components = []
    for identifier in ids:
        if identifier in seen: continue
        queue = deque([identifier]); seen.add(identifier); size = 0
        while queue:
            current = queue.popleft(); size += 1
            for other in adjacency[current]:
                if other not in seen: seen.add(other); queue.append(other)
        components.append(size)
    return {"driverCount": len(ids), "relationshipCount": len(relationships), "directedLayerPairCoverage": len(cross), "weakComponents": len(components), "isolatedDrivers": sum(degree[node] == 0 for node in ids), "degreeOneDrivers": sum(degree[node] == 1 for node in ids), "driversWithIncomingAndOutgoing": sum(incoming[node] and outgoing[node] for node in ids), "largestWeakComponent": max(components), "crossLayerPairs": dict(sorted(cross.items()))}


def main() -> int:
    try:
        original_paths = [p for p in sorted(SOURCE.glob("*.xlsx")) if not p.name.startswith("~$")]
        existing, _ = existing_driver_index(original_paths)
        before = pre_metrics()
        _, rollback_rows = prepare_release()
        staged_paths = [STAGING / path.name for path in original_paths]

        registers = source_registers()
        required_refs = set()
        for path in (RC / "driver_candidate_adjudication.csv", RC / "relationship_candidate_adjudication.csv"):
            for row in csv_rows(path): required_refs.update(split(row.get("supportingSourceRefs", "")))
        for row in csv_rows(RC / "duplicate_resolution_plan.csv"): required_refs.update(split(row["evidenceRefs"]))
        missing_refs = sorted(required_refs - registers.keys())
        if missing_refs: raise ValueError(f"Source register entries missing: {missing_refs}")
        ordered_refs = sorted(required_refs, key=lambda ref: (ref.startswith("RC-"), int(ref.split("R")[-1])))
        evidence_map = {ref: f"SRC-{494 + index}" for index, ref in enumerate(ordered_refs)}

        driver_records, id_manifest = build_driver_rows(evidence_map)
        write_csv(RELEASE / "id_assignment_manifest.csv", id_manifest)
        if len({row["permanentDriverId"] for row in id_manifest}) != 31 or set(ID_MAP.values()) & set(existing):
            raise ValueError("Permanent Driver ID manifest failed uniqueness/collision validation")

        books: dict[str, Any] = {}
        for path in staged_paths:
            books[layer_from_filename(path.name)] = load_workbook(path)
        try:
            append_evidence(books, registers, evidence_map, driver_records)
            duplicate_repairs = apply_duplicate_repairs(books, evidence_map)
            append_drivers(books, driver_records)
            append_codebook(books)
            public_drivers = {row["id"]: row for row in json.loads((ROOT / "data" / "drivers.json").read_text(encoding="utf-8"))}
            for record in driver_records:
                public_drivers[record["ID"]] = {"id": record["ID"], "name": record["Name"], "layer": record["Layer"], "family": record["Family"]}
            relationship_pairs = migrate_existing_relationships(books, public_drivers)
            relationship_changes, mediated_paths, blocked = append_relationships(books, public_drivers, evidence_map, relationship_pairs)
            for layer, book in books.items():
                summary = book["Layer Summary"]
                for row in range(5, summary.max_row + 1):
                    if text(summary.cell(row, 4).value).casefold() == "relationships":
                        last = book["Relationships"].max_row
                        summary.cell(row, 5).value = f"=COUNTA('Relationships'!A5:A{last})"
                book.save(STAGING / next(path.name for path in staged_paths if layer_from_filename(path.name) == layer))
        finally:
            for book in books.values(): book.close()

        for row in csv_rows(RC / "driver_candidate_adjudication.csv"):
            if row["candidateKey"] not in ID_MAP:
                blocked.append({
                    "itemType": "DRIVER", "itemKey": row["candidateKey"], "sourceDriverId": "", "targetDriverId": "",
                    "status": row["disposition"].replace(" ", "_"), "issue": row["adjudicationRationale"],
                    "evidence": row["supportingSourceRefs"],
                    "specificQuestion": "Retain the release-candidate disposition; reconsider only through a future governed ontology review.",
                })

        driver_changes = [{
            "candidateKey": key, "action": "ADD", "permanentDriverId": ID_MAP[key],
            "canonicalName": next(record["Name"] for record in driver_records if record["ID"] == ID_MAP[key]),
            "layer": next(record["Layer"] for record in driver_records if record["ID"] == ID_MAP[key]),
            "family": next(record["Family"] for record in driver_records if record["ID"] == ID_MAP[key]),
            "implementationStatus": "STAGED_AND_VALIDATED",
        } for key in ID_MAP]
        write_csv(RELEASE / "driver_changes.csv", driver_changes)
        write_csv(RELEASE / "duplicate_repairs.csv", duplicate_repairs)
        write_csv(RELEASE / "relationship_changes.csv", relationship_changes)
        write_csv(RELEASE / "mediated_paths.csv", mediated_paths)
        write_csv(RELEASE / "blocked_items.csv", blocked)

        proposed_diff = {
            "release": "ontology_remediation_release_v1", "preDriverCount": 762,
            "newDrivers": driver_changes, "duplicateRepairs": duplicate_repairs,
            "relationshipChanges": relationship_changes, "mediatedPaths": mediated_paths,
            "blockedItems": blocked, "relationshipSchema": {"from": "1.0", "to": "2.0", "headers": REL_V2_HEADERS},
            "codebookTermsAdded": [row[0] for row in CODEBOOK_ROWS], "evidenceIdMap": evidence_map,
        }
        (RELEASE / "proposed_diff.json").write_text(json.dumps(proposed_diff, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

        after = validate_staged(staged_paths, set(existing))
        metric_names = ("driverCount", "relationshipCount", "directedLayerPairCoverage", "weakComponents", "isolatedDrivers", "degreeOneDrivers", "driversWithIncomingAndOutgoing", "largestWeakComponent")
        write_csv(RELEASE / "pre_post_connectivity_metrics.csv", [{"metric": name, "pre": before.get(name), "post": after.get(name), "delta": int(after.get(name, 0)) - int(before.get(name, 0))} for name in metric_names])
        (RELEASE / "staged_validation.json").write_text(json.dumps(after, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

        # Canonical source replacement happens only after complete staging validation.
        for staged in staged_paths:
            shutil.copy2(staged, SOURCE / staged.name)
        for row in rollback_rows:
            canonical = ROOT / row["canonicalPath"]
            row["postMigrationSha256"] = sha256(canonical) if canonical.exists() else "PENDING_PUBLIC_REBUILD"
        write_csv(RELEASE / "rollback_manifest.csv", rollback_rows)
        manifest = {
            "release": "ontology_remediation_release_v1", "status": "CANONICAL_XLSX_APPLIED",
            "driverIdAssignments": len(id_manifest), "duplicateRepairGroups": len(duplicate_repairs),
            "relationshipChangeRecords": len(relationship_changes), "mediatedPaths": len(mediated_paths),
            "blockedItems": len(blocked), "stagedMetrics": after,
            "workbookSha256": {path.name: sha256(SOURCE / path.name) for path in original_paths},
        }
        (RELEASE / "generation_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 1
    print("Ontology remediation migration staged, validated, and applied")
    print(f"  Drivers: 762 -> {after['driverCount']}")
    print(f"  Relationships: {before['relationshipCount']} -> {after['relationshipCount']}")
    print(f"  Mediated paths: {len(mediated_paths)}")
    print(f"  Blocked items: {len(blocked)}")
    print("  Errors: 0")
    return 0


def repair_applied_role_mapping() -> int:
    """Synchronize the two legacy moderation labels after v2 review."""
    targets = {"REL-PSY-002", "REL-PSY-030"}
    paths = [
        next(SOURCE.glob("PSYWERX_Layer_2_Psychological_*.xlsx")),
        next(STAGING.glob("PSYWERX_Layer_2_Psychological_*.xlsx")),
    ]
    for path in paths:
        book = load_workbook(path)
        try:
            sheet = book["Relationships"]
            if tuple(text(sheet.cell(4, column).value) for column in range(1, 29)) != REL_V2_HEADERS:
                raise ValueError("Psychological Relationships worksheet is not Schema v2")
            found = set()
            for row in range(5, sheet.max_row + 1):
                identifier = text(sheet.cell(row, 1).value)
                if identifier in targets:
                    if text(sheet.cell(row, 6).value) != "MODERATES":
                        raise ValueError(f"{identifier}: expected pre-repair MODERATES value")
                    sheet.cell(row, 6).value = "CAUSES"
                    note = text(sheet.cell(row, 28).value)
                    sheet.cell(row, 28).value = note + " V2 review: legacy moderation label did not identify a separate governed edge; represented as a causal contribution at the stated resolution."
                    found.add(identifier)
            if found != targets:
                raise ValueError(f"Missing relationship repair targets: {sorted(targets - found)}")
            book.save(path)
        finally:
            book.close()
    print("Updated REL-PSY-002 and REL-PSY-030 to CAUSES under Schema v2 governance.")
    return 0


if __name__ == "__main__":
    if sys.argv[1:] == ["--repair-v2-role-mapping"]:
        raise SystemExit(repair_applied_role_mapping())
    raise SystemExit(main())
