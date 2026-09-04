"""Deterministic, private-first Cognitive Security canonical re-synthesis.

This module builds a governed analytical overlay from the immutable historical
normalization and reconciliation products.  It deliberately does not mutate
those products or any live public Explorer payload.  Item text and historical
migration lineage are written only to the ignored analysis directory.
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
import subprocess
from collections import Counter, defaultdict
from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path
from typing import Any

from .reconcile import GOVERNED_DISTINCT_PUBLICATION_REUSES

SCHEMA_VERSION = "0.1-draft"
METHOD_VERSION = "deduplicated-canonical-resynthesis-v0.1"
COUNTERPART_ADJUDICATION_FILE = "tension_alias_item_counterparts_adjudication.json"
SPLIT_ALLOCATION_ADJUDICATION_FILE = "tension_split_item_allocation_adjudication.json"
COLLISION_ADJUDICATION_FILE = "tension_cross_tension_collision_adjudication.json"
SAME_TENSION_ADJUDICATION_FILE = "tension_same_tension_duplicate_adjudication.json"
TENSION_ADJUDICATION_FILE = "canonical_tension_adjudication.json"
FAMILY_ADJUDICATION_FILE = "family_mapping_adjudication.json"
THEME_NARRATIVE_ADJUDICATION_FILE = "theme_narrative_lineage_adjudication.json"
SCENARIO_ADJUDICATION_FILE = "scenario_adjudication.json"
SPLIT_HISTORICAL_TENSION_IDS = frozenset({"TD-001", "TD-002", "TD-007", "TD-009"})
SUPPORT_INTERPRETATION = (
    "Corpus support reflects recurrence and breadth within this practitioner "
    "discourse corpus. It does not indicate scientific validity, consensus, "
    "importance, prevalence, or real-world effect size."
)

EXPECTED_COUNTS = {
    "historicalSourceIdentities": 269,
    "publicReleases": 242,
    "canonicalContentUnits": 241,
    "historicalItems": 14397,
    "canonicalItems": 12933,
    "historicalClusters": 127,
}

RELATIONSHIP_SEMANTICS = (
    ("direct-coded-support", "A retained item was directly coded to the cluster."),
    (
        "direct-content-representation",
        "A public release directly represents the selected analytical content unit.",
    ),
    (
        "primary-family-membership",
        "The cluster's principal within-category canonical family.",
    ),
    (
        "secondary-family-relationship",
        "A reviewed, meaningful relationship to a nonprimary family.",
    ),
    (
        "primary-theme-support",
        "The family or cluster is central to defining the theme.",
    ),
    (
        "secondary-theme-support",
        "The family or cluster supports but does not define the theme.",
    ),
    (
        "conceptual-framing",
        "A Key Concepts family explains an entity without being treated as independent empirical proof.",
    ),
    (
        "future-extension",
        "A Future Trends family extends a present pattern into a future-oriented proposition.",
    ),
    (
        "tension-evidence-pole-a",
        "A retained item supports or illustrates Pole A without implying endorsement.",
    ),
    (
        "tension-evidence-pole-b",
        "A retained item supports or illustrates Pole B without implying endorsement.",
    ),
    (
        "integrates",
        "A narrative integrates the referenced themes, tensions, and families.",
    ),
    (
        "activated-tension",
        "A scenario makes a tension consequential or changes the conditions around its poles.",
    ),
    (
        "scenario-amplifies",
        "One scenario may conditionally intensify another; this is not a causal claim.",
    ),
    (
        "scenario-mitigates",
        "One scenario may conditionally reduce another's severity; this is not a causal claim.",
    ),
    (
        "contextual-connection",
        "A relevant association that is neither direct evidence nor a causal assertion.",
    ),
    (
        "shared-content-inheritance",
        "A re-release displays relationships inherited from a content-equivalent original without adding analytical weight.",
    ),
)

STOPWORDS = frozenset(
    "a an and are as at be been being by can could did do does for from had has "
    "have how if in into is it its may more most not of on or our should than that "
    "the their them these they this those through to under use using was were what "
    "when where which while who will with would".split()
)


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _json_bytes(value: Any) -> bytes:
    return (
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    ).encode("utf-8")


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(_json_bytes(value))


def _write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value.rstrip() + "\n", encoding="utf-8", newline="\n")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _natural(value: Any) -> tuple[Any, ...]:
    return tuple(
        int(part) if part.isdigit() else part.lower()
        for part in re.split(r"(\d+)", str(value))
    )


def _tokens(*values: Any) -> list[str]:
    words: list[str] = []
    for value in values:
        for token in re.findall(r"[a-z][a-z0-9-]{2,}", str(value or "").lower()):
            if token not in STOPWORDS:
                words.append(token)
    return words


def _split_ids(value: Any) -> list[str]:
    return [part.strip() for part in str(value or "").split(";") if part.strip()]


def _percent_change(original: int, corrected: int) -> float | None:
    if not original:
        return None
    return round((corrected - original) * 100.0 / original, 6)


def _round(value: float) -> float:
    return round(value, 6)


def load_private_dataset(normalized_dir: Path) -> dict[str, Any]:
    """Load only governed collections required by the re-synthesis."""

    names = (
        "categories",
        "category_findings",
        "category_summaries",
        "clusters",
        "cluster_summaries",
        "cluster_meta_mappings",
        "episode_source_identities",
        "episode_source_mappings",
        "episodes",
        "items",
        "item_cluster_assignments",
        "meta_clusters",
        "meta_narratives",
        "scenario_actions",
        "scenario_indicators",
        "scenario_pathways",
        "scenarios",
        "theme_cluster_evidence",
        "themes",
        "tension_mappings",
        "tensions",
    )
    return {name: _read_json(normalized_dir / f"{name}.json") for name in names}


def load_design(design_dir: Path) -> dict[str, Any]:
    architecture_path = design_dir / (
        "PSYWERX_Cognitive_Security_Canonical_Architecture_v0.1.json"
    )
    mapping_path = design_dir / (
        "PSYWERX_Cognitive_Security_Cluster_to_Family_Mapping_v0.1.csv"
    )
    working_package_path = design_dir / (
        "PSYWERX_Cognitive_Security_Canonical_Resynthesis_Working_Package_v0.1.md"
    )
    # The working package is a governed design input even though its prose is
    # not copied into the generated analytical records. Reading it here makes
    # absence or unreadability a fail-closed precondition rather than merely a
    # difference in the input-manifest hash.
    working_package_path.read_text(encoding="utf-8-sig")
    architecture = _read_json(architecture_path)
    with mapping_path.open(encoding="utf-8-sig", newline="") as handle:
        mappings = list(csv.DictReader(handle))
    architecture["clusterFamilyMappingsCsv"] = mappings
    return architecture


def _input_manifest(
    normalized_dir: Path,
    reconciliation_dir: Path,
    design_dir: Path,
    tension_workbook: Path,
    transcript_summary_dir: Path,
) -> dict[str, str]:
    logical_paths = {
        **{
            f"normalized/{path.name}": path
            for path in sorted(normalized_dir.glob("*.json"), key=lambda row: row.name)
        },
        **{
            f"reconciliation/{path.name}": path
            for path in sorted(
                reconciliation_dir.glob("*.json"), key=lambda row: row.name
            )
        },
        **{
            f"design/{path.name}": path
            for path in sorted(design_dir.iterdir(), key=lambda row: row.name)
            if path.is_file()
        },
        "source/tensions_debates_rebuilt.xlsx": tension_workbook,
        "transcripts/manifest_report.json": transcript_summary_dir
        / "manifest_report.json",
    }
    return {name: _sha256(path) for name, path in sorted(logical_paths.items())}


def build_corpus_selection(
    dataset: Mapping[str, Sequence[Mapping[str, Any]]],
    transcript_manifest_report: Mapping[str, Any],
    reconciliation_report: Mapping[str, Any] | None = None,
) -> tuple[dict[str, Any], dict[str, Any], set[str], dict[str, set[str]]]:
    identities = list(dataset["episode_source_identities"])
    mappings = list(dataset["episode_source_mappings"])
    episodes = list(dataset["episodes"])
    items = list(dataset["items"])
    mapping_by_source = {str(row["sourceIdentityId"]): row for row in mappings}
    item_counts = Counter(str(row["sourceIdentityId"]) for row in items)

    canonical_sources = {str(row["canonicalSourceIdentityId"]) for row in episodes}
    alias_sources = {
        str(row["sourceIdentityId"])
        for row in mappings
        if row.get("mappingRole") == "alias"
        and row.get("mappingStatus") == "confirmed-alias"
        and bool(row.get("collapseEligible"))
    }
    if canonical_sources & alias_sources:
        raise ValueError(
            "A confirmed alias identity is incorrectly designated as a public "
            "release's canonical source."
        )
    reuse_rule = GOVERNED_DISTINCT_PUBLICATION_REUSES[0]
    original_source = str(reuse_rule["originalSourceIdentityId"])
    reuse_source = str(reuse_rule["reuseSourceIdentityId"])
    if (
        original_source not in canonical_sources
        or reuse_source not in canonical_sources
    ):
        raise ValueError("Governed episode-83 shared-content identities are absent.")
    selected_sources = canonical_sources - {reuse_source}
    selected_item_ids = {
        str(row["itemId"])
        for row in items
        if str(row["sourceIdentityId"]) in selected_sources
    }

    public_release_to_content: dict[str, str] = {}
    release_by_content: dict[str, set[str]] = defaultdict(set)
    for episode in episodes:
        release_id = str(episode["episodeId"])
        source_id = str(episode["canonicalSourceIdentityId"])
        content_id = original_source if source_id == reuse_source else source_id
        public_release_to_content[release_id] = content_id
        release_by_content[content_id].add(release_id)

    identity_records: list[dict[str, Any]] = []
    for identity in sorted(
        identities, key=lambda row: _natural(row["sourceIdentityId"])
    ):
        source_id = str(identity["sourceIdentityId"])
        mapping = mapping_by_source[source_id]
        if source_id in alias_sources:
            status = "confirmed-alias-excluded"
            content_id = str(mapping["canonicalEpisodeId"])
            contributes = False
        elif source_id == reuse_source:
            status = "shared-content-reuse-excluded"
            content_id = original_source
            contributes = False
        else:
            status = "canonical-content-representative"
            content_id = source_id
            contributes = True
        identity_records.append(
            {
                "sourceIdentityId": source_id,
                "identityKind": identity.get("identityKind"),
                "publicReleaseId": mapping.get("canonicalEpisodeId"),
                "canonicalContentUnitId": content_id,
                "analyticalStatus": status,
                "contributesAnalyticalWeight": contributes,
                "historicalItemCount": item_counts[source_id],
            }
        )

    item_records: list[dict[str, Any]] = []
    for item in sorted(items, key=lambda row: _natural(row["itemId"])):
        item_id = str(item["itemId"])
        source_id = str(item["sourceIdentityId"])
        if source_id in alias_sources:
            status = "excluded-confirmed-alias"
            content_id = str(mapping_by_source[source_id]["canonicalEpisodeId"])
            contributes = False
            reason = "confirmed-alias-source-identity"
        elif source_id == reuse_source:
            status = "excluded-shared-content-reuse"
            content_id = original_source
            contributes = False
            reason = "shared-recording-already-weighted-through-original-release"
        elif source_id in selected_sources:
            status = "included-canonical-content"
            content_id = source_id
            contributes = True
            reason = "governed-canonical-content-representation"
        else:
            raise ValueError(
                f"Item source lacks a governed selection state: {item_id}."
            )
        item_records.append(
            {
                "itemId": item_id,
                "sourceIdentityId": source_id,
                "canonicalContentUnitId": content_id,
                "analyticalStatus": status,
                "contributesAnalyticalWeight": contributes,
                "selectionReason": reason,
            }
        )

    public_releases = []
    for episode in sorted(episodes, key=lambda row: _natural(row["episodeId"])):
        release_id = str(episode["episodeId"])
        content_id = public_release_to_content[release_id]
        direct = str(episode["canonicalSourceIdentityId"]) != reuse_source
        public_releases.append(
            {
                "publicReleaseId": release_id,
                "canonicalContentUnitId": content_id,
                "relationshipRole": (
                    "direct-content-representation"
                    if direct
                    else "shared-content-inheritance"
                ),
                "contributesAnalyticalWeight": direct,
            }
        )
    content_unit_records = [
        {
            "canonicalContentUnitId": content_id,
            "publicReleaseIds": sorted(
                release_by_content.get(content_id, set()), key=_natural
            ),
            "selectedRepresentationId": content_id,
            "analyticalStatus": (
                "selected-with-shared-release-inheritance"
                if len(release_by_content.get(content_id, set())) > 1
                else "selected-canonical-content"
            ),
            "contributesAnalyticalWeight": True,
            "selectionReason": "governed-canonical-content-representation",
        }
        for content_id in sorted(selected_sources, key=_natural)
    ]

    selected_items = [row for row in items if str(row["itemId"]) in selected_item_ids]
    selected_scope = Counter(str(row.get("scope")) for row in selected_items)
    alias_item_ids = {
        str(row["itemId"])
        for row in items
        if str(row["sourceIdentityId"]) in alias_sources
    }
    reuse_item_ids = {
        str(row["itemId"])
        for row in items
        if str(row["sourceIdentityId"]) == reuse_source
    }
    counts = {
        "historicalSourceIdentityCount": len(identities),
        "publicReleaseCount": len(episodes),
        "canonicalAnalyticalContentUnitCount": len(selected_sources),
        "historicalItemCount": len(items),
        "canonicalItemCount": len(selected_item_ids),
        "canonicalFocalItemCount": selected_scope["focal"],
        "canonicalContextualItemCount": selected_scope["contextual"],
        "excludedIdentityCount": len(alias_sources) + 1,
        "excludedAliasIdentityCount": len(alias_sources),
        "excludedSharedContentIdentityCount": 1,
        "excludedItemCount": len(items) - len(selected_item_ids),
        "excludedAliasItemCount": len(alias_item_ids),
        "excludedSharedContentItemCount": len(reuse_item_ids),
        "historicalClusterCount": len(dataset.get("clusters", ())),
        "historicalMetaClusterCount": len(dataset.get("meta_clusters", ())),
        "historicalThemeCount": len(dataset.get("themes", ())),
        "historicalTensionCount": len(dataset.get("tensions", ())),
        "historicalNarrativeCount": len(dataset.get("meta_narratives", ())),
        "historicalCategoryFindingCount": len(dataset.get("category_findings", ())),
        "historicalScenarioCount": len(dataset.get("scenarios", ())),
    }
    transcript_expected = {
        "status": "pass",
        "coverageComplete": True,
        "sourceIdentityCount": counts["historicalSourceIdentityCount"],
        "canonicalReleaseCount": counts["publicReleaseCount"],
        "selectedTranscriptCount": counts["publicReleaseCount"],
        "strictUniqueContentUnits": counts["canonicalAnalyticalContentUnitCount"],
        "excludedAliasIdentityCount": counts["excludedAliasIdentityCount"],
        "excludedOutsideGovernedCorpusCount": 2,
        "filenameDriftResolutions": 1,
        "contentReuseGroupCount": 1,
    }
    if "sourceIdentityCount" in transcript_manifest_report:
        transcript_mismatches = {
            key: {
                "expected": expected,
                "observed": transcript_manifest_report.get(key),
            }
            for key, expected in transcript_expected.items()
            if transcript_manifest_report.get(key) != expected
        }
        if transcript_mismatches:
            raise ValueError(
                "Transcript manifest does not match the governed corpus selection: "
                + ", ".join(sorted(transcript_mismatches))
            )

    reconciliation_audit: dict[str, Any] = {}
    if reconciliation_report is not None:
        reconciliation_counts = reconciliation_report.get("reconciliationCounts", {})
        original_counts = reconciliation_report.get("originalCounts", {})
        reconciled_counts = reconciliation_report.get("reconciledSensitivityCounts", {})
        reconciliation_expected = {
            "sourceIdentities": counts["historicalSourceIdentityCount"],
            "canonicalEpisodes": counts["publicReleaseCount"],
            "confirmedAliasGroups": counts["excludedAliasIdentityCount"],
        }
        reconciliation_observed = {
            key: reconciliation_counts.get(key) for key in reconciliation_expected
        }
        if reconciliation_observed != reconciliation_expected:
            raise ValueError(
                "Reconciliation report corpus counts do not match normalized lineage."
            )
        expected_original = {
            "items": counts["historicalItemCount"],
            "focalItems": sum(str(row.get("scope")) == "focal" for row in items),
            "contextualItems": sum(
                str(row.get("scope")) == "contextual" for row in items
            ),
        }
        expected_reconciled = {
            "items": counts["historicalItemCount"] - counts["excludedAliasItemCount"],
            "focalItems": sum(
                str(row.get("scope")) == "focal"
                and str(row["sourceIdentityId"]) not in alias_sources
                for row in items
            ),
            "contextualItems": sum(
                str(row.get("scope")) == "contextual"
                and str(row["sourceIdentityId"]) not in alias_sources
                for row in items
            ),
        }
        if (
            original_counts != expected_original
            or reconciled_counts != expected_reconciled
        ):
            raise ValueError(
                "Reconciliation report item counts do not reproduce from normalized lineage."
            )
        if not reconciliation_report.get("releaseEligible"):
            raise ValueError(
                "The governed reconciliation report is not release eligible."
            )
        reconciliation_audit = {
            "schemaVersion": reconciliation_report.get("schemaVersion"),
            "methodVersion": reconciliation_report.get("methodVersion"),
            "releaseEligible": True,
            "reconciliationCounts": reconciliation_observed,
            "originalCounts": expected_original,
            "reconciledSensitivityCounts": expected_reconciled,
            "crossValidationStatus": "pass",
        }
    selection = {
        "schemaVersion": SCHEMA_VERSION,
        "methodVersion": METHOD_VERSION,
        "selectionRule": (
            "Select the governed canonical source identity for each of 242 public "
            "releases, then map the episode-83 re-release to the original recording "
            "and give only the original content representation analytical weight."
        ),
        "counts": counts,
        "selectedCanonicalContentUnitIds": sorted(selected_sources, key=_natural),
        "selectedItemIds": sorted(selected_item_ids, key=_natural),
        "canonicalContentUnitSelection": content_unit_records,
        "historicalIdentitySelection": identity_records,
        "historicalItemSelection": item_records,
        "publicReleaseContentMap": public_releases,
        "sharedContentRelationships": [
            {
                "sourcePublicReleaseId": reuse_source,
                "targetPublicReleaseId": original_source,
                "canonicalContentUnitId": original_source,
                "semanticRole": "shared-content-inheritance",
                "contributesEvidence": False,
            }
        ],
        "reconciliationAudit": reconciliation_audit,
        "transcriptManifestAudit": {
            key: transcript_manifest_report.get(key) for key in transcript_expected
        },
    }
    report = {
        "schemaVersion": SCHEMA_VERSION,
        "methodVersion": METHOD_VERSION,
        "status": "pass",
        "counts": counts,
        "identityReconciliationIssues": [],
        "reconciliationAudit": reconciliation_audit,
        "episode83Treatment": {
            "originalSourceIdentityId": original_source,
            "reuseSourceIdentityId": reuse_source,
            "publicReleasesRetained": 2,
            "analyticalContentUnits": 1,
            "reuseItemsExcludedFromWeight": len(reuse_item_ids),
            "inheritanceSemanticRole": "shared-content-inheritance",
        },
        "transcriptCoverage": {
            key: transcript_manifest_report.get(key)
            for key in (
                "status",
                "coverageComplete",
                "sourceIdentityCount",
                "canonicalReleaseCount",
                "selectedTranscriptCount",
                "strictUniqueContentUnits",
                "excludedAliasIdentityCount",
                "excludedOutsideGovernedCorpusCount",
                "filenameDriftResolutions",
                "contentReuseGroupCount",
            )
        },
        "interpretiveBoundary": SUPPORT_INTERPRETATION,
        "provenanceModel": [
            "historical source identity",
            "source candidate where applicable",
            "retained item and canonical content unit",
            "fixed canonical cluster",
            "canonical family",
            "canonical higher-order entity",
        ],
    }
    return selection, report, selected_item_ids, dict(release_by_content)


def _support_profile(
    item_ids: Iterable[str],
    *,
    historical_item_ids: Iterable[str] | None,
    direct_item_ids: Iterable[str],
    category_item_ids: Iterable[str] | None = None,
    item_by_id: Mapping[str, Mapping[str, Any]],
    release_by_content: Mapping[str, set[str]],
    cluster_ids: Iterable[str],
    family_ids: Iterable[str],
    adjudication_status: str,
    limitations: Sequence[str] = (),
) -> dict[str, Any]:
    retained = {str(value) for value in item_ids if str(value) in item_by_id}
    historical_available = historical_item_ids is not None
    historical = (
        {str(value) for value in historical_item_ids}
        if historical_item_ids is not None
        else set()
    )
    direct = retained & {str(value) for value in direct_item_ids}
    counts = Counter(
        str(item_by_id[item_id]["sourceIdentityId"]) for item_id in retained
    )
    total = sum(counts.values())
    shares = (
        sorted((count / total for count in counts.values()), reverse=True)
        if total
        else []
    )
    herfindahl = sum(share * share for share in shares)
    releases = set()
    for content_id in counts:
        releases.update(release_by_content.get(content_id, set()))
    category_evidence = (
        retained
        if category_item_ids is None
        else retained & {str(value) for value in category_item_ids}
    )
    categories = {
        str(item_by_id[item_id]["categoryId"])
        for item_id in category_evidence
        if item_by_id[item_id].get("scope") == "focal"
    }
    inherited_release_count = max(0, len(releases) - len(counts))
    sensitivity_percent = (
        _percent_change(len(historical), len(retained))
        if historical_available
        else None
    )
    if not historical_available:
        sensitivity_status = "unassessable"
    elif sensitivity_percent is None or abs(sensitivity_percent) <= 5:
        sensitivity_status = "stable"
    elif abs(sensitivity_percent) <= 10:
        sensitivity_status = "mild"
    elif abs(sensitivity_percent) <= 25:
        sensitivity_status = "moderate"
    else:
        sensitivity_status = "high"
    historical_source_count = (
        len(
            {
                str(item_by_id[item_id]["sourceIdentityId"])
                for item_id in historical
                if item_id in item_by_id
            }
        )
        if historical_available
        else None
    )
    return {
        "uniqueContentUnitSupportCount": len(counts),
        "publicReleaseCoverageCount": len(releases),
        "inheritedPublicReleaseCoverageCount": inherited_release_count,
        "itemSupportCount": len(retained),
        "clusterSupportCount": len(set(cluster_ids)),
        "familySupportCount": len(set(family_ids)),
        "categoryBreadth": len(categories),
        "topOneContentUnitShare": _round(shares[0]) if shares else 0.0,
        "topTwoContentUnitShare": _round(sum(shares[:2])) if shares else 0.0,
        "topFiveContentUnitShare": _round(sum(shares[:5])) if shares else 0.0,
        "effectiveContentUnitCount": _round(1.0 / herfindahl) if herfindahl else 0.0,
        "directSupportItemCount": len(direct),
        "derivedSupportItemCount": len(retained - direct),
        "directSupportShare": _round(len(direct) / len(retained)) if retained else 0.0,
        "historicalToCorrectedSensitivity": {
            "status": sensitivity_status,
            "historicalItemSupportCount": (
                len(historical) if historical_available else None
            ),
            "correctedItemSupportCount": len(retained),
            "absoluteChange": (
                len(retained) - len(historical) if historical_available else None
            ),
            "percentChange": sensitivity_percent,
            "historicalSourceIdentityCoverageCount": historical_source_count,
            "correctedContentUnitCoverageCount": len(counts),
            "coverageAbsoluteChange": (
                len(counts) - historical_source_count
                if historical_source_count is not None
                else None
            ),
            "coveragePercentChange": (
                _percent_change(historical_source_count, len(counts))
                if historical_source_count is not None
                else None
            ),
        },
        "adjudicationStatus": adjudication_status,
        "limitations": list(limitations),
        "interpretation": SUPPORT_INTERPRETATION,
    }


def build_cluster_support(
    dataset: Mapping[str, Sequence[Mapping[str, Any]]],
    selected_item_ids: set[str],
    release_by_content: Mapping[str, set[str]],
) -> tuple[list[dict[str, Any]], dict[str, set[str]], dict[str, set[str]]]:
    items = list(dataset["items"])
    item_by_id = {str(row["itemId"]): row for row in items}
    assignments = list(dataset["item_cluster_assignments"])
    categories = {str(row["categoryId"]): row for row in dataset["categories"]}
    historical_by_cluster: dict[str, set[str]] = defaultdict(set)
    retained_by_cluster: dict[str, set[str]] = defaultdict(set)
    historical_primary: Counter[str] = Counter()
    historical_secondary: Counter[str] = Counter()
    retained_primary: Counter[str] = Counter()
    retained_secondary: Counter[str] = Counter()
    retained_item_role: dict[str, dict[str, str]] = defaultdict(dict)

    for assignment in assignments:
        item_id = str(assignment["itemId"])
        primary = str(assignment.get("primaryClusterId") or "")
        secondary = str(assignment.get("secondaryClusterId") or "")
        if primary:
            historical_primary[primary] += 1
            historical_by_cluster[primary].add(item_id)
            if item_id in selected_item_ids:
                retained_primary[primary] += 1
                retained_by_cluster[primary].add(item_id)
                retained_item_role[primary][item_id] = "primary"
        if secondary:
            historical_secondary[secondary] += 1
            historical_by_cluster[secondary].add(item_id)
            if item_id in selected_item_ids:
                retained_secondary[secondary] += 1
                retained_by_cluster[secondary].add(item_id)
                retained_item_role[secondary].setdefault(item_id, "secondary")

    records: list[dict[str, Any]] = []
    for cluster in sorted(
        dataset["clusters"], key=lambda row: _natural(row["clusterId"])
    ):
        cluster_id = str(cluster["clusterId"])
        historical_items = historical_by_cluster[cluster_id]
        retained_items = retained_by_cluster[cluster_id]
        content_counts = Counter(
            str(item_by_id[item_id]["sourceIdentityId"]) for item_id in retained_items
        )
        total = sum(content_counts.values())
        shares = (
            sorted((count / total for count in content_counts.values()), reverse=True)
            if total
            else []
        )
        original_weighted = (
            2 * historical_primary[cluster_id] + historical_secondary[cluster_id]
        )
        corrected_weighted = (
            2 * retained_primary[cluster_id] + retained_secondary[cluster_id]
        )
        historical_identity_coverage = len(
            {
                str(item_by_id[item_id]["sourceIdentityId"])
                for item_id in historical_items
            }
        )
        weighted_loss = (
            (original_weighted - corrected_weighted) / original_weighted
            if original_weighted
            else 0.0
        )
        breadth_loss = (
            (historical_identity_coverage - len(content_counts))
            / historical_identity_coverage
            if historical_identity_coverage
            else 0.0
        )
        dominated = bool(shares and sum(shares[:2]) >= 0.5)
        if not retained_items:
            status = "lost-all-support"
        elif weighted_loss >= 0.25 or breadth_loss >= 0.25 or dominated:
            status = "review-required"
        elif weighted_loss >= 0.10:
            status = "moderate-sensitivity"
        else:
            status = "stable"
        category_id = str(cluster["categoryId"])
        releases = set()
        for content_id in content_counts:
            releases.update(release_by_content.get(content_id, set()))
        profile = _support_profile(
            retained_items,
            historical_item_ids=historical_items,
            direct_item_ids=retained_items,
            item_by_id=item_by_id,
            release_by_content=release_by_content,
            cluster_ids=[cluster_id],
            family_ids=[],
            adjudication_status="recomputed",
            limitations=(
                ["Support is concentrated in one or two content units."]
                if dominated
                else []
            ),
        )
        records.append(
            {
                "clusterId": cluster_id,
                "clusterName": cluster["name"],
                "focalCategoryMembership": {
                    "categoryId": category_id,
                    "categoryName": categories[category_id]["name"],
                },
                "canonicalPrimaryItemCount": retained_primary[cluster_id],
                "canonicalSecondaryItemCount": retained_secondary[cluster_id],
                "governedWeightedCount": corrected_weighted,
                "uniqueContentUnitCoverage": len(content_counts),
                "publicReleaseCoverage": len(releases),
                "supportConcentration": {
                    "topOneContentUnitShare": profile["topOneContentUnitShare"],
                    "topTwoContentUnitShare": profile["topTwoContentUnitShare"],
                    "topFiveContentUnitShare": profile["topFiveContentUnitShare"],
                    "effectiveContentUnitCount": profile["effectiveContentUnitCount"],
                    "dominatedByOneOrTwoContentUnits": dominated,
                },
                "historicalToCorrectedChange": {
                    "historicalPrimaryItemCount": historical_primary[cluster_id],
                    "historicalSecondaryItemCount": historical_secondary[cluster_id],
                    "historicalWeightedCount": original_weighted,
                    "weightedAbsoluteChange": corrected_weighted - original_weighted,
                    "weightedPercentChange": _percent_change(
                        original_weighted, corrected_weighted
                    ),
                    "historicalDistinctSourceIdentityCoverage": historical_identity_coverage,
                    "correctedContentUnitCoverage": len(content_counts),
                    "coverageAbsoluteChange": len(content_counts)
                    - historical_identity_coverage,
                    "coveragePercentChange": _percent_change(
                        historical_identity_coverage, len(content_counts)
                    ),
                },
                "sensitivityStatus": status,
                "investigationTriggers": [
                    trigger
                    for trigger, present in (
                        ("lost-all-support", not retained_items),
                        (
                            "weighted-support-loss-at-least-25-percent",
                            weighted_loss >= 0.25,
                        ),
                        (
                            "content-unit-breadth-loss-at-least-25-percent",
                            breadth_loss >= 0.25,
                        ),
                        ("dominated-by-one-or-two-content-units", dominated),
                    )
                    if present
                ],
                "corpusSupportProfile": profile,
            }
        )
    return records, dict(historical_by_cluster), dict(retained_by_cluster)


# Family relationships are deliberately sparse.  A medium-confidence primary
# assignment is not converted into a secondary relationship merely because a
# neighboring family shares vocabulary.
SPECIAL_MAPPING_RATIONALES = {
    "CRB-10": (
        "The small surviving evidence set concerns forecasting, uncertainty, and "
        "limits of inference. It fits Evidence, Observability, Authenticity, and "
        "Uncertainty rather than measurement failure, narrative failure, or model "
        "reliability. The historical orphan is resolved without reviving CRB-M05."
    ),
    "FTP-13": (
        "Surviving items consistently treat identity, cohesion, and social change as "
        "system-level transformations. They remain distinguishable as a cluster but "
        "share a family with nonlinear systemic uncertainty; a new singleton family "
        "would not add a stable functional boundary."
    ),
    "KCFT-20": (
        "A stratified surviving-item review shows strategic culture and ideology "
        "operating primarily through identity, narrative, and macro-cultural audience "
        "models. Persistent campaigning and normative order remain neighboring but "
        "secondary interpretations, not the primary mechanism."
    ),
}


def _round_robin_evidence(
    item_ids: Iterable[str],
    item_by_id: Mapping[str, Mapping[str, Any]],
    maximum: int = 8,
) -> list[str]:
    by_source: dict[str, list[str]] = defaultdict(list)
    for item_id in sorted(set(item_ids), key=_natural):
        if item_id in item_by_id:
            by_source[str(item_by_id[item_id]["sourceIdentityId"])].append(item_id)
    selected: list[str] = []
    source_ids = sorted(by_source, key=_natural)
    index = 0
    while len(selected) < maximum and source_ids:
        remaining = []
        for source_id in source_ids:
            values = by_source[source_id]
            if index < len(values) and len(selected) < maximum:
                selected.append(values[index])
            if index + 1 < len(values):
                remaining.append(source_id)
        source_ids = remaining
        index += 1
    return selected


def _top_terms(
    item_ids: Iterable[str],
    item_by_id: Mapping[str, Mapping[str, Any]],
    maximum: int = 12,
) -> list[str]:
    counts: Counter[str] = Counter()
    for item_id in sorted(set(item_ids), key=_natural):
        item = item_by_id.get(item_id, {})
        counts.update(
            _tokens(
                item.get("item"),
                item.get("summary"),
                item.get("strategicSignificance"),
                item.get("operationalImplications"),
            )
        )
    return [term for term, _ in counts.most_common(maximum)]


def _family_neighbor_comparisons(
    cluster: Mapping[str, Any],
    proposed_family_id: str,
    families: Sequence[Mapping[str, Any]],
) -> list[dict[str, Any]]:
    cluster_terms = set(
        _tokens(
            cluster.get("name"),
            cluster.get("definition"),
            cluster.get("inclusionCriteria"),
            cluster.get("exclusionCriteria"),
            cluster.get("nearNeighborDistinctions"),
        )
    )
    candidates = []
    for family in families:
        family_id = str(family["familyId"])
        if family_id == proposed_family_id or family.get("category") != cluster.get(
            "categoryName"
        ):
            continue
        family_terms = set(
            _tokens(
                family.get("name"),
                family.get("definition"),
                family.get("inclusionRule"),
                family.get("exclusionRule"),
            )
        )
        union = cluster_terms | family_terms
        overlap = len(cluster_terms & family_terms) / len(union) if union else 0.0
        candidates.append((overlap, family_id, family.get("name")))
    return [
        {
            "familyId": family_id,
            "familyName": name,
            "definitionTermOverlap": _round(overlap),
            "decision": "not-primary",
        }
        for overlap, family_id, name in sorted(
            candidates, key=lambda row: (-row[0], _natural(row[1]))
        )[:3]
    ]


def _validated_family_adjudication_index(
    dataset: Mapping[str, Sequence[Mapping[str, Any]]],
    design: Mapping[str, Any],
    retained_by_cluster: Mapping[str, set[str]],
    release_by_content: Mapping[str, set[str]],
    adjudication: Mapping[str, Any],
) -> dict[str, dict[str, Any]]:
    """Validate the complete private cluster-to-family evidence review."""

    clusters = {str(row["clusterId"]): row for row in dataset["clusters"]}
    categories = {
        str(row["categoryId"]): str(row["name"]) for row in dataset["categories"]
    }
    families = {str(row["familyId"]): row for row in design["families"]}
    proposed = {
        str(row["clusterId"]): row for row in design["clusterFamilyMappingsCsv"]
    }
    items = {str(row["itemId"]): row for row in dataset["items"]}
    selected_item_ids = set().union(*retained_by_cluster.values())
    primary_counts: Counter[str] = Counter()
    secondary_counts: Counter[str] = Counter()
    for assignment in dataset["item_cluster_assignments"]:
        item_id = str(assignment["itemId"])
        if item_id not in selected_item_ids:
            continue
        primary = str(assignment.get("primaryClusterId") or "")
        secondary = str(assignment.get("secondaryClusterId") or "")
        if primary:
            primary_counts[primary] += 1
        if secondary:
            secondary_counts[secondary] += 1

    records = adjudication.get("records")
    if not isinstance(records, list):
        raise ValueError("Family mapping adjudication must contain records.")
    index: dict[str, dict[str, Any]] = {}
    for raw_record in records:
        record = dict(raw_record)
        cluster_id = str(record.get("clusterId") or "")
        if not cluster_id or cluster_id in index or cluster_id not in clusters:
            raise ValueError(
                f"Family adjudication has a missing, duplicate, or stale cluster: {cluster_id!r}."
            )
        proposed_family = str(record.get("proposedFamilyId") or "")
        final_family = str(record.get("finalFamilyId") or "")
        governed_cluster_name = str(record.get("clusterName") or "")
        source_cluster_name = str(clusters[cluster_id].get("name") or "")
        if (
            proposed_family != str(proposed[cluster_id]["proposedFamilyId"])
            or final_family not in families
            or not governed_cluster_name.strip()
            or governed_cluster_name != source_cluster_name
        ):
            raise ValueError(f"Family adjudication target is stale for {cluster_id}.")
        category_name = categories[str(clusters[cluster_id]["categoryId"])]
        if str(families[final_family]["category"]) != category_name:
            raise ValueError(
                f"Family adjudication crosses categories for {cluster_id}."
            )
        if record.get("decision") not in {"confirm", "reassign"}:
            raise ValueError(f"Family decision is invalid for {cluster_id}.")
        if (record["decision"] == "confirm") != (proposed_family == final_family):
            raise ValueError(f"Family decision semantics are invalid for {cluster_id}.")
        if record.get("confidence") not in {"high", "moderate", "review"}:
            raise ValueError(f"Family confidence is invalid for {cluster_id}.")
        if record.get("proposedMappingConfidence") not in {"high", "medium"}:
            raise ValueError(f"Proposed confidence is invalid for {cluster_id}.")
        evidence = record.get("evidenceReview", {})
        support = evidence.get("support", {})
        expected_primary = primary_counts[cluster_id]
        expected_secondary = secondary_counts[cluster_id]
        if support != {
            "canonicalPrimaryItemCount": expected_primary,
            "canonicalSecondaryItemCount": expected_secondary,
            "governedWeightedCount": 2 * expected_primary + expected_secondary,
        }:
            raise ValueError(f"Family evidence support is stale for {cluster_id}.")
        retained_items = retained_by_cluster.get(cluster_id, set())
        content_ids = {
            str(items[item_id]["sourceIdentityId"]) for item_id in retained_items
        }
        releases = {
            release_id
            for content_id in content_ids
            for release_id in release_by_content.get(content_id, set())
        }
        breadth = evidence.get("breadth", {})
        if breadth.get("canonicalContentUnitCount") != len(content_ids) or breadth.get(
            "publicReleaseCoverageCount"
        ) != len(releases):
            raise ValueError(f"Family evidence breadth is stale for {cluster_id}.")
        representative_ids = {
            str(value) for value in evidence.get("privateRepresentativeItemIds", [])
        }
        if not representative_ids or not representative_ids <= retained_items:
            raise ValueError(
                f"Family representative evidence is invalid for {cluster_id}."
            )
        review_depth = str(evidence.get("reviewDepth") or "")
        if (
            record["proposedMappingConfidence"] == "medium"
            and review_depth != "deep-item-level"
        ) or review_depth not in {"deep-item-level", "boundary-check"}:
            raise ValueError(f"Family review depth is invalid for {cluster_id}.")
        expected_finding_results = {
            "definitionFinding": "fit",
            "inclusionFinding": "fit",
            "exclusionFinding": "pass",
            "boundaryFinding": "resolved",
        }
        if not all(
            isinstance(record.get(field), Mapping)
            and record[field].get("result") == expected_result
            and str(record[field].get("finding") or "").strip()
            for field, expected_result in expected_finding_results.items()
        ):
            raise ValueError(f"Family boundary review is incomplete for {cluster_id}.")
        alternatives = record.get("alternativesConsidered")
        if (
            not isinstance(alternatives, list)
            or not alternatives
            or any(
                str(row.get("familyId") or "") not in families
                or str(row.get("familyId")) == final_family
                or not str(row.get("finding") or "").strip()
                for row in alternatives
            )
        ):
            raise ValueError(f"Family alternatives are incomplete for {cluster_id}.")
        if (
            not str(record.get("rationale") or "").strip()
            or record.get("reviewerStatus")
            != "approved-after-independent-evidence-review"
        ):
            raise ValueError(f"Family review approval is incomplete for {cluster_id}.")
        index[cluster_id] = record

    if set(index) != set(clusters):
        raise ValueError("Family adjudication does not cover all 127 clusters.")
    final_counts = Counter(str(row["finalFamilyId"]) for row in index.values())
    if set(final_counts) != set(families) or any(
        not count for count in final_counts.values()
    ):
        raise ValueError("Family adjudication creates an empty family.")
    proposed_medium = [
        row for row in index.values() if row["proposedMappingConfidence"] == "medium"
    ]
    if len(proposed_medium) != 33 or not all(
        row["evidenceReview"]["reviewDepth"] == "deep-item-level"
        for row in proposed_medium
    ):
        raise ValueError("All 33 proposed-medium mappings require deep review.")
    meta_decisions = adjudication.get("historicalMetaClusterDecisions", [])
    if not any(
        row.get("metaClusterId") == "CRB-M05"
        and row.get("decision") == "do-not-revive"
        and row.get("finalTreatment") == "private-provenance-only-no-canonical-family"
        for row in meta_decisions
    ):
        raise ValueError("CRB-M05 provenance-only treatment is missing.")
    if adjudication.get("validation", {}).get("passed") is not True:
        raise ValueError("Family adjudication validation status is not pass.")
    return index


def build_families(
    dataset: Mapping[str, Sequence[Mapping[str, Any]]],
    design: Mapping[str, Any],
    historical_by_cluster: Mapping[str, set[str]],
    retained_by_cluster: Mapping[str, set[str]],
    release_by_content: Mapping[str, set[str]],
    governed_adjudication: Mapping[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, Any], dict[str, str], dict[str, set[str]]]:
    items = list(dataset["items"])
    item_by_id = {str(row["itemId"]): row for row in items}
    clusters = {str(row["clusterId"]): row for row in dataset["clusters"]}
    cluster_summaries = {
        str(row["clusterId"]): row for row in dataset["cluster_summaries"]
    }
    families_seed = list(design["families"])
    csv_mappings = list(design["clusterFamilyMappingsCsv"])
    json_mappings = list(design["clusterFamilyMappings"])
    json_pairs = {
        (str(row["clusterId"]), str(row["proposedFamilyId"])) for row in json_mappings
    }
    csv_pairs = {
        (str(row["clusterId"]), str(row["proposedFamilyId"])) for row in csv_mappings
    }
    if json_pairs != csv_pairs:
        raise ValueError("Design JSON and CSV cluster-family mappings disagree.")
    proposed_cluster_to_family = {
        str(row["clusterId"]): str(row["proposedFamilyId"]) for row in csv_mappings
    }
    if set(proposed_cluster_to_family) != set(clusters):
        raise ValueError("The proposed family map does not account for all clusters.")
    governed_index = _validated_family_adjudication_index(
        dataset,
        design,
        retained_by_cluster,
        release_by_content,
        governed_adjudication,
    )
    cluster_to_family = {
        cluster_id: str(record["finalFamilyId"])
        for cluster_id, record in governed_index.items()
    }

    decisions = []
    for mapping in sorted(csv_mappings, key=lambda row: _natural(row["clusterId"])):
        cluster_id = str(mapping["clusterId"])
        governed = governed_index[cluster_id]
        family_id = str(governed["finalFamilyId"])
        cluster = clusters[cluster_id]
        retained_items = retained_by_cluster.get(cluster_id, set())
        sample_ids = sorted(
            {
                str(value)
                for value in governed["evidenceReview"]["privateRepresentativeItemIds"]
            },
            key=_natural,
        )
        evidence = [
            {
                "itemId": item_id,
                "canonicalContentUnitId": item_by_id[item_id]["sourceIdentityId"],
                "summary": item_by_id[item_id].get("summary"),
                "evidenceExcerpt": item_by_id[item_id].get("evidenceExcerpt"),
            }
            for item_id in sample_ids
        ]
        proposed_confidence = str(governed["proposedMappingConfidence"])
        adjudicated_confidence = str(governed["confidence"])
        rationale = str(governed["rationale"])
        decisions.append(
            {
                "clusterId": cluster_id,
                "clusterName": cluster["name"],
                "proposedPrimaryFamilyId": governed["proposedFamilyId"],
                "finalPrimaryFamilyId": family_id,
                "secondaryFamilyIds": [],
                "mappingConfidence": proposed_confidence,
                "adjudicatedConfidence": adjudicated_confidence,
                "reviewDepth": governed["evidenceReview"]["reviewDepth"],
                "definitionFit": governed["definitionFinding"]["result"],
                "inclusionRuleFit": governed["inclusionFinding"]["result"],
                "exclusionRuleConflicts": (
                    []
                    if governed["exclusionFinding"]["result"] == "pass"
                    else [governed["exclusionFinding"]["finding"]]
                ),
                "neighborFamilyComparisons": governed["alternativesConsidered"],
                "observedEvidenceTerms": _top_terms(retained_items, item_by_id),
                "reviewedEvidence": evidence,
                "reviewedCanonicalItemCount": len(retained_items),
                "reviewedCanonicalContentUnitCount": len(
                    {
                        item_by_id[item_id]["sourceIdentityId"]
                        for item_id in retained_items
                    }
                ),
                "clusterDefinition": cluster.get("definition"),
                "clusterInclusionCriteria": cluster.get("inclusionCriteria"),
                "clusterExclusionCriteria": cluster.get("exclusionCriteria"),
                "historicalClusterSummary": cluster_summaries.get(cluster_id, {}).get(
                    "summary"
                ),
                "decision": (
                    "confirm-primary-family"
                    if governed["decision"] == "confirm"
                    else "reassign-primary-family"
                ),
                "definitionFinding": governed["definitionFinding"],
                "inclusionFinding": governed["inclusionFinding"],
                "exclusionFinding": governed["exclusionFinding"],
                "boundaryFinding": governed["boundaryFinding"],
                "evidenceFinding": governed["evidenceReview"]["evidenceFinding"],
                "rationale": rationale,
                "adjudicationStatus": "analyst-reviewed-draft",
                "limitations": (
                    (
                        [
                            "Small surviving support base; interpret the boundary cautiously."
                        ]
                        if len(retained_items) < 10
                        else []
                    )
                    + (
                        ["Final mapping confidence remains moderate after deep review."]
                        if adjudicated_confidence == "moderate"
                        else []
                    )
                ),
            }
        )

    decisions_by_cluster = {row["clusterId"]: row for row in decisions}
    family_item_ids: dict[str, set[str]] = defaultdict(set)
    historical_family_item_ids: dict[str, set[str]] = defaultdict(set)
    for cluster_id, family_id in cluster_to_family.items():
        family_item_ids[family_id].update(retained_by_cluster.get(cluster_id, set()))
        historical_family_item_ids[family_id].update(
            historical_by_cluster.get(cluster_id, set())
        )

    families: list[dict[str, Any]] = []
    for seed in sorted(families_seed, key=lambda row: _natural(row["familyId"])):
        family_id = str(seed["familyId"])
        member_ids = sorted(
            [cid for cid, fid in cluster_to_family.items() if fid == family_id],
            key=_natural,
        )
        retained_items = family_item_ids[family_id]
        confidence_counts = Counter(
            decisions_by_cluster[cluster_id]["adjudicatedConfidence"]
            for cluster_id in member_ids
        )
        proposed_medium_count = sum(
            decisions_by_cluster[cluster_id]["mappingConfidence"] == "medium"
            for cluster_id in member_ids
        )
        limitations = []
        if len(member_ids) == 1:
            limitations.append(
                "Singleton family retained because the member cluster has a distinct "
                "functional boundary and nonzero corrected support."
            )
        if len(retained_items) < 10:
            limitations.append("Small corrected evidence base.")
        profile = _support_profile(
            retained_items,
            historical_item_ids=historical_family_item_ids[family_id],
            direct_item_ids=retained_items,
            item_by_id=item_by_id,
            release_by_content=release_by_content,
            cluster_ids=member_ids,
            family_ids=[family_id],
            adjudication_status="analyst-reviewed-draft",
            limitations=limitations,
        )
        families.append(
            {
                "familyId": family_id,
                "name": seed["name"],
                "category": seed["category"],
                "definition": seed["definition"],
                "inclusionRules": [seed["inclusionRule"]],
                "exclusionRules": [seed["exclusionRule"]],
                "distinguishingBoundaries": (
                    "Apply the primary functional mechanism in the fixed cluster "
                    "definition; lexical overlap with neighboring families is not sufficient."
                ),
                "memberClusterIds": member_ids,
                "secondaryRelatedClusterIds": [],
                "corpusSupportProfile": profile,
                "mappingConfidence": {
                    "highClusterAssignments": confidence_counts["high"],
                    "moderateClusterAssignments": confidence_counts["moderate"],
                    "reviewClusterAssignments": confidence_counts["review"],
                    "proposedMediumAssignmentsEvidenceReviewed": proposed_medium_count,
                    "minimum": (
                        "review"
                        if confidence_counts["review"]
                        else ("moderate" if confidence_counts["moderate"] else "high")
                    ),
                },
                "adjudicationStatus": "analyst-reviewed-draft",
                "limitations": limitations,
            }
        )

    adjudication = {
        "schemaVersion": SCHEMA_VERSION,
        "methodVersion": METHOD_VERSION,
        "decisionSummary": {
            "proposedFamilyCount": len(families_seed),
            "finalDraftFamilyCount": len(families),
            "confirmedWithoutFamilyChange": sum(
                row["decision"] == "confirm-primary-family" for row in decisions
            ),
            "reassignedAfterReview": sum(
                row["decision"] == "reassign-primary-family" for row in decisions
            ),
            "mediumConfidenceEvidenceReviews": sum(
                row["mappingConfidence"] == "medium" for row in decisions
            ),
            "deepItemLevelEvidenceReviews": sum(
                row["reviewDepth"] == "deep-item-level" for row in decisions
            ),
            "finalHighConfidenceMappings": sum(
                row["adjudicatedConfidence"] == "high" for row in decisions
            ),
            "finalModerateConfidenceMappings": sum(
                row["adjudicatedConfidence"] == "moderate" for row in decisions
            ),
            "secondaryFamilyRelationshipsAccepted": 0,
            "historicalEmptyMetaClusterTreatment": (
                "CRB-M05 retained only as private provenance; no empty canonical family created."
            ),
        },
        "historicalMetaClusterDecisions": governed_adjudication[
            "historicalMetaClusterDecisions"
        ],
        "governedReviewSummary": governed_adjudication["reviewSummary"],
        "governedValidation": governed_adjudication["validation"],
        "mappingDecisions": decisions,
        "allClustersAssignedExactlyOnce": len(cluster_to_family) == len(clusters),
        "allFamiliesNonempty": all(row["memberClusterIds"] for row in families),
    }
    return families, adjudication, cluster_to_family, dict(family_item_ids)


# This constant preserves the public-safe working-package proposal for review
# and tests.  The build itself is driven by the private, analyst-reviewed
# adjudication table validated below.
THEME_FAMILY_ROLES: dict[str, dict[str, tuple[str, ...]]] = {
    "TH-01": {
        "primary": (
            "ORG-F01",
            "ORG-F04",
            "ORG-F07",
            "CRB-F01",
            "CRB-F02",
            "CRB-F03",
            "KCF-F06",
            "KCF-F07",
            "KEH-F03",
            "FTP-F04",
            "OPP-F01",
            "OPP-F02",
            "OPP-F03",
        ),
        "secondary": ("TTP-F06", "ORG-F03", "ORG-F08", "KCF-F08", "OPP-F07"),
    },
    "TH-02": {
        "primary": (
            "TTP-F01",
            "CRB-F04",
            "KCF-F01",
            "KCF-F05",
            "KCF-F06",
            "FTP-F01",
            "FTP-F02",
            "OPP-F06",
        ),
        "secondary": (
            "TTP-F03",
            "ORG-F04",
            "ORG-F08",
            "KEH-F02",
            "KEH-F06",
            "FTP-F03",
            "OPP-F07",
        ),
    },
    "TH-03": {
        "primary": ("TTP-F02", "ORG-F06", "CRB-F05", "KCF-F02", "FTP-F01", "FTP-F06"),
        "secondary": ("TTP-F05", "CRB-F06", "KCF-F03", "KEH-F04", "OPP-F04"),
    },
    "TH-04": {
        "primary": ("TTP-F03", "CRB-F04", "CRB-F05", "KCF-F04", "FTP-F03", "FTP-F06"),
        "secondary": (
            "TTP-F01",
            "TTP-F04",
            "TTP-F06",
            "ORG-F06",
            "KCF-F05",
            "OPP-F02",
            "OPP-F06",
            "OPP-F07",
        ),
    },
    "TH-05": {
        "primary": (
            "CRB-F06",
            "KCF-F01",
            "KCF-F02",
            "KCF-F05",
            "KEH-F05",
            "FTP-F07",
            "OPP-F05",
        ),
        "secondary": ("TTP-F02", "TTP-F04", "ORG-F05", "KEH-F04", "OPP-F04", "OPP-F07"),
    },
    "TH-06": {
        "primary": ("TTP-F04", "CRB-F04", "CRB-F06", "KCF-F02", "KCF-F04", "FTP-F07"),
        "secondary": (
            "TTP-F03",
            "ORG-F06",
            "ORG-F08",
            "CRB-F05",
            "FTP-F02",
            "FTP-F06",
            "OPP-F07",
        ),
    },
    "TH-07": {
        "primary": (
            "ORG-F03",
            "ORG-F05",
            "ORG-F06",
            "ORG-F07",
            "CRB-F01",
            "KCF-F08",
            "FTP-F04",
            "FTP-F05",
            "FTP-F06",
            "OPP-F03",
        ),
        "secondary": (
            "ORG-F01",
            "ORG-F02",
            "ORG-F04",
            "ORG-F08",
            "CRB-F02",
            "KCF-F07",
            "OPP-F01",
            "OPP-F07",
        ),
    },
    "TH-08": {
        "primary": (
            "TTP-F05",
            "TTP-F06",
            "CRB-F05",
            "KCF-F08",
            "KEH-F01",
            "KEH-F02",
            "OPP-F08",
        ),
        "secondary": (
            "TTP-F02",
            "TTP-F03",
            "ORG-F01",
            "ORG-F06",
            "FTP-F03",
            "FTP-F05",
            "OPP-F02",
        ),
    },
    "TH-09": {
        "primary": (
            "ORG-F04",
            "CRB-F03",
            "KCF-F06",
            "KEH-F03",
            "KEH-F06",
            "FTP-F04",
            "OPP-F01",
            "OPP-F02",
            "OPP-F06",
        ),
        "secondary": ("TTP-F06", "ORG-F01", "ORG-F07", "CRB-F04", "KCF-F05", "OPP-F03"),
    },
    "TH-10": {
        "primary": (
            "ORG-F02",
            "ORG-F03",
            "CRB-F02",
            "CRB-F07",
            "KCF-F03",
            "KCF-F08",
            "KEH-F01",
            "FTP-F05",
            "FTP-F06",
            "OPP-F03",
            "OPP-F04",
        ),
        "secondary": (
            "TTP-F03",
            "TTP-F05",
            "ORG-F06",
            "KEH-F04",
            "FTP-F01",
            "FTP-F03",
            "OPP-F07",
            "OPP-F08",
        ),
    },
    "TH-11": {
        "primary": (
            "ORG-F01",
            "ORG-F05",
            "ORG-F08",
            "CRB-F02",
            "CRB-F04",
            "CRB-F06",
            "KCF-F07",
            "KEH-F05",
            "FTP-F02",
            "OPP-F07",
        ),
        "secondary": (
            "TTP-F02",
            "TTP-F03",
            "TTP-F04",
            "ORG-F06",
            "ORG-F07",
            "KCF-F01",
            "KCF-F02",
            "FTP-F07",
            "OPP-F05",
        ),
    },
}


def _validate_theme_narrative_adjudication(
    dataset: Mapping[str, Sequence[Mapping[str, Any]]],
    design: Mapping[str, Any],
    families: Sequence[Mapping[str, Any]],
    adjudication: Mapping[str, Any],
) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    """Fail closed unless all governed cross-level decisions and lineage resolve."""

    theme_ids = {str(row["id"]) for row in design["themes"]}
    narrative_ids = {str(row["id"]) for row in design["narratives"]}
    family_ids = {str(row["familyId"]) for row in families}
    tension_ids = {str(row["id"]) for row in design["tensions"]}
    scenario_ids = {str(row["id"]) for row in design["scenarios"]}
    historical_theme_ids = {str(row["themeId"]) for row in dataset["themes"]}
    historical_narrative_ids = {
        str(row["narrativeId"]) for row in dataset["meta_narratives"]
    }
    historical_meta_cluster_ids = {
        str(row["metaClusterId"]) for row in dataset["meta_clusters"]
    }
    historical_scenario_ids = {str(row["scenarioId"]) for row in dataset["scenarios"]}

    def exact_index(
        field: str, id_field: str, expected: set[str]
    ) -> dict[str, dict[str, Any]]:
        rows = adjudication.get(field)
        if not isinstance(rows, list):
            raise ValueError(f"Cross-level adjudication is missing {field}.")
        index: dict[str, dict[str, Any]] = {}
        for raw_row in rows:
            row = dict(raw_row)
            identifier = str(row.get(id_field) or "")
            if not identifier or identifier in index:
                raise ValueError(f"Duplicate or missing {id_field} in {field}.")
            index[identifier] = row
        if set(index) != expected:
            raise ValueError(
                f"Cross-level adjudication coverage mismatch for {field}: "
                f"missing={sorted(expected - set(index), key=_natural)}, "
                f"extra={sorted(set(index) - expected, key=_natural)}."
            )
        return index

    theme_index = exact_index("themeDecisions", "themeId", theme_ids)
    narrative_index = exact_index("narrativeDecisions", "narrativeId", narrative_ids)
    historical_theme_index = exact_index(
        "historicalThemeLineage", "historicalThemeId", historical_theme_ids
    )
    historical_narrative_index = exact_index(
        "historicalNarrativeLineage",
        "historicalNarrativeId",
        historical_narrative_ids,
    )
    historical_meta_index = exact_index(
        "historicalMetaClusterLineage",
        "historicalMetaClusterId",
        historical_meta_cluster_ids,
    )
    historical_scenario_index = exact_index(
        "historicalScenarioLineage",
        "historicalScenarioId",
        historical_scenario_ids,
    )

    for theme_id, row in theme_index.items():
        primary = [str(value) for value in row.get("primaryFamilyIds", [])]
        secondary = [str(value) for value in row.get("secondaryFamilyIds", [])]
        if (
            not primary
            or len(primary) != len(set(primary))
            or len(secondary) != len(set(secondary))
            or set(primary) & set(secondary)
            or not (set(primary) | set(secondary)) <= family_ids
            or str(row.get("historicalThemeId") or "") not in historical_theme_ids
        ):
            raise ValueError(f"Invalid governed family roles for theme {theme_id}.")
        if row.get("confidence") not in {"high", "medium", "review"}:
            raise ValueError(f"Invalid governed confidence for theme {theme_id}.")
        if row.get("adjudicationStatus") != "analyst-reviewed-draft":
            raise ValueError(f"Theme {theme_id} is not analyst reviewed.")
        if not all(
            str(row.get(field) or "").strip()
            for field in ("decision", "rationale", "boundaryConditions")
        ) or not isinstance(row.get("limitations"), list):
            raise ValueError(f"Theme decision detail is incomplete for {theme_id}.")

    for narrative_id, row in narrative_index.items():
        themes = [str(value) for value in row.get("integratesThemeIds", [])]
        tensions = [str(value) for value in row.get("integratesTensionIds", [])]
        predecessors = [str(value) for value in row.get("historicalNarrativeIds", [])]
        if (
            len(set(themes)) < 2
            or not set(themes) <= theme_ids
            or not tensions
            or not set(tensions) <= tension_ids
            or not predecessors
            or not set(predecessors) <= historical_narrative_ids
        ):
            raise ValueError(f"Invalid governed integration for {narrative_id}.")
        if row.get("confidence") not in {"high", "medium", "review"}:
            raise ValueError(f"Invalid governed confidence for {narrative_id}.")
        if row.get("adjudicationStatus") != "analyst-reviewed-draft":
            raise ValueError(f"Narrative {narrative_id} is not analyst reviewed.")
        if not all(
            str(row.get(field) or "").strip()
            for field in ("decision", "rationale", "boundaryConditions")
        ) or not isinstance(row.get("limitations"), list):
            raise ValueError(
                f"Narrative decision detail is incomplete for {narrative_id}."
            )

    for historical_id, row in historical_theme_index.items():
        targets = {str(value) for value in row.get("canonicalThemeIds", [])}
        if not targets or not targets <= theme_ids:
            raise ValueError(
                f"Historical theme lineage is invalid for {historical_id}."
            )
    for historical_id, row in historical_narrative_index.items():
        targets = {str(value) for value in row.get("canonicalNarrativeIds", [])}
        if not targets or not targets <= narrative_ids:
            raise ValueError(
                f"Historical narrative lineage is invalid for {historical_id}."
            )
    for historical_id, row in historical_meta_index.items():
        target_families = {str(value) for value in row.get("canonicalFamilyIds", [])}
        target_themes = {
            str(value)
            for field in (
                "primaryCanonicalThemeIds",
                "secondaryCanonicalThemeIds",
            )
            for value in row.get(field, [])
        }
        if not target_families <= family_ids or not target_themes <= theme_ids:
            raise ValueError(
                f"Historical meta-cluster lineage is invalid for {historical_id}."
            )
        if historical_id == "CRB-M05" and (
            str(row.get("lineageSemantics"))
            != "provenance-only-no-direct-cluster-support"
            or not str(row.get("rationale") or "").strip()
        ):
            raise ValueError("CRB-M05 must remain provenance-only and weightless.")
    for historical_id, row in historical_scenario_index.items():
        targets = {str(value) for value in row.get("canonicalScenarioIds", [])}
        if not targets or not targets <= scenario_ids:
            raise ValueError(
                f"Historical scenario lineage is invalid for {historical_id}."
            )

    review_summary = adjudication.get("reviewSummary", {})
    expected_summary = {
        "canonicalThemeDecisionCount": len(theme_index),
        "canonicalNarrativeDecisionCount": len(narrative_index),
        "historicalThemeLineageCount": len(historical_theme_index),
        "historicalNarrativeLineageCount": len(historical_narrative_index),
        "historicalMetaClusterLineageCount": len(historical_meta_index),
        "historicalScenarioLineageCount": len(historical_scenario_index),
    }
    if any(review_summary.get(key) != value for key, value in expected_summary.items()):
        raise ValueError("Cross-level adjudication review summary is stale.")
    return theme_index, narrative_index


def build_themes(
    dataset: Mapping[str, Sequence[Mapping[str, Any]]],
    design: Mapping[str, Any],
    families: Sequence[Mapping[str, Any]],
    cluster_to_family: Mapping[str, str],
    historical_by_cluster: Mapping[str, set[str]],
    retained_by_cluster: Mapping[str, set[str]],
    release_by_content: Mapping[str, set[str]],
    cross_level_adjudication: Mapping[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, Any], dict[str, set[str]]]:
    item_by_id = {str(row["itemId"]): row for row in dataset["items"]}
    historical_themes = {str(row["themeId"]): row for row in dataset["themes"]}
    theme_decisions, _ = _validate_theme_narrative_adjudication(
        dataset, design, families, cross_level_adjudication
    )
    theme_item_ids: dict[str, set[str]] = {}
    records = []
    decisions = []

    for seed in sorted(design["themes"], key=lambda row: _natural(row["id"])):
        theme_id = str(seed["id"])
        governed = theme_decisions[theme_id]
        primary_families = [str(value) for value in governed["primaryFamilyIds"]]
        secondary_families = [str(value) for value in governed["secondaryFamilyIds"]]
        primary_clusters = sorted(
            [cid for cid, fid in cluster_to_family.items() if fid in primary_families],
            key=_natural,
        )
        secondary_clusters = sorted(
            [
                cid
                for cid, fid in cluster_to_family.items()
                if fid in secondary_families
            ],
            key=_natural,
        )
        primary_items = set().union(
            *(retained_by_cluster.get(cid, set()) for cid in primary_clusters)
        )
        secondary_items = (
            set().union(
                *(retained_by_cluster.get(cid, set()) for cid in secondary_clusters)
            )
            - primary_items
        )
        all_items = primary_items | secondary_items
        direct_clusters = [
            cluster_id
            for cluster_id in primary_clusters + secondary_clusters
            if not cluster_to_family[cluster_id].startswith(("KCF-", "FTP-"))
        ]
        direct_items = set().union(
            *(retained_by_cluster.get(cid, set()) for cid in direct_clusters)
        )
        historical_items = set().union(
            *(
                historical_by_cluster.get(cid, set())
                for cid in primary_clusters + secondary_clusters
            )
        )
        theme_item_ids[theme_id] = all_items
        historical = historical_themes[str(governed["historicalThemeId"])]
        limitations = [
            "Theme support is an adjudicated synthesis of cluster and family lineage, not an independent item coding pass.",
            *[str(value) for value in governed["limitations"]],
        ]
        profile = _support_profile(
            all_items,
            historical_item_ids=historical_items,
            direct_item_ids=direct_items,
            category_item_ids=direct_items,
            item_by_id=item_by_id,
            release_by_content=release_by_content,
            cluster_ids=primary_clusters + secondary_clusters,
            family_ids=primary_families + secondary_families,
            adjudication_status=str(governed["adjudicationStatus"]),
            limitations=limitations,
        )
        relationships = []
        for family_id in primary_families + secondary_families:
            if family_id.startswith("KCF-"):
                semantic = "conceptual-framing"
            elif family_id.startswith("FTP-"):
                semantic = "future-extension"
            elif family_id in primary_families:
                semantic = "primary-theme-support"
            else:
                semantic = "secondary-theme-support"
            relationships.append(
                {
                    "familyId": family_id,
                    "semanticRole": semantic,
                    "analyticalWeight": "primary"
                    if family_id in primary_families
                    else "secondary",
                }
            )
        records.append(
            {
                "themeId": theme_id,
                "name": seed["name"],
                "definition": seed["definition"],
                "boundaryConditions": governed["boundaryConditions"],
                "internalAnalyticalRole": seed.get("role"),
                "publicLevel": "theme",
                "primaryFamilyIds": primary_families,
                "secondaryFamilyIds": secondary_families,
                "primaryClusterIds": primary_clusters,
                "secondaryClusterIds": secondary_clusters,
                "familyRelationships": relationships,
                "categoryBreadth": profile["categoryBreadth"],
                "corpusSupportProfile": profile,
                "strategicSignificance": historical.get("strategicSignificance"),
                "operationalImplications": historical.get("operationalImplications"),
                "limitations": limitations,
                "adjudicationStatus": governed["adjudicationStatus"],
                "adjudicationConfidence": governed["confidence"],
            }
        )
        decisions.append(
            {
                "themeId": theme_id,
                "decision": governed["decision"],
                "historicalThemeId": governed["historicalThemeId"],
                "historicalClusterCount": len(historical.get("linkedClusterIds", ())),
                "canonicalPrimaryFamilyCount": len(primary_families),
                "canonicalSecondaryFamilyCount": len(secondary_families),
                "keyConceptFamiliesAssessed": [
                    fid
                    for fid in primary_families + secondary_families
                    if fid.startswith("KCF-")
                ],
                "futureTrendFamiliesAssessed": [
                    fid
                    for fid in primary_families + secondary_families
                    if fid.startswith("FTP-")
                ],
                "historicalEvidenceStrength": governed["historicalEvidenceStrength"],
                "retainedHistoricalClusterCount": governed[
                    "retainedHistoricalClusterCount"
                ],
                "excludedHistoricalContextClusterCount": governed[
                    "excludedHistoricalContextClusterCount"
                ],
                "canonicalLineageClusterCount": governed[
                    "canonicalLineageClusterCount"
                ],
                "confidence": governed["confidence"],
                "boundaryConditions": governed["boundaryConditions"],
                "rationale": governed["rationale"],
                "limitations": governed["limitations"],
                "adjudicationStatus": governed["adjudicationStatus"],
            }
        )

    adjudication = {
        "schemaVersion": SCHEMA_VERSION,
        "methodVersion": METHOD_VERSION,
        "publicThemeLevelCount": 1,
        "finalDraftThemeCount": len(records),
        "decisions": decisions,
        "historicalThemeLineage": cross_level_adjudication["historicalThemeLineage"],
        "historicalMetaClusterLineage": cross_level_adjudication[
            "historicalMetaClusterLineage"
        ],
        "historicalNarrativeLineage": cross_level_adjudication[
            "historicalNarrativeLineage"
        ],
        "historicalScenarioLineage": cross_level_adjudication[
            "historicalScenarioLineage"
        ],
        "reviewSummary": cross_level_adjudication["reviewSummary"],
        "keyConceptsIncludedInEveryCrossLevelAudit": all(
            row["keyConceptFamiliesAssessed"] for row in decisions
        ),
        "futureTrendsIncludedInEveryCrossLevelAudit": all(
            row["futureTrendFamiliesAssessed"] for row in decisions
        ),
    }
    return records, adjudication, theme_item_ids


def _load_tension_candidates(workbook_path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - governed build dependency
        raise RuntimeError(
            "openpyxl is required for tension-candidate lineage."
        ) from exc

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook["Tension Evidence"]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        records = []
        for values in rows:
            row = dict(zip(headers, values))
            if row.get("tension_id") and row.get("source_candidate_id"):
                records.append(row)
        return records
    finally:
        workbook.close()


def _historical_to_canonical_targets(
    tensions: Sequence[Mapping[str, Any]],
) -> dict[str, list[str]]:
    targets: dict[str, list[str]] = defaultdict(list)
    for tension in tensions:
        for historical_id in re.findall(r"TD-\d{3}", str(tension.get("legacy", ""))):
            targets[historical_id].append(str(tension["id"]))
    return {key: sorted(set(value), key=_natural) for key, value in targets.items()}


def _validated_counterpart_index(
    dataset: Mapping[str, Sequence[Mapping[str, Any]]],
    selected_item_ids: set[str],
    adjudication: Mapping[str, Any],
) -> dict[str, dict[str, Any]]:
    """Validate the private, human-adjudicated alias-item counterpart table."""

    item_by_id = {str(row["itemId"]): row for row in dataset["items"]}
    mapping_by_source = {
        str(row["sourceIdentityId"]): row for row in dataset["episode_source_mappings"]
    }
    assignment_by_item = {
        str(row["itemId"]): row for row in dataset["item_cluster_assignments"]
    }
    records = adjudication.get("records")
    if not isinstance(records, list):
        raise ValueError("Counterpart adjudication must contain a records list.")
    index: dict[str, dict[str, Any]] = {}
    for raw_record in records:
        record = dict(raw_record)
        historical_item_id = str(record.get("historicalItemId") or "")
        canonical_item_id = str(record.get("canonicalItemId") or "")
        historical_tension_id = str(record.get("historicalTensionId") or "")
        if not historical_item_id or not canonical_item_id or not historical_tension_id:
            raise ValueError(
                "Counterpart adjudication record is missing a governed key."
            )
        if historical_item_id in index:
            raise ValueError(
                f"Duplicate counterpart adjudication for historical item {historical_item_id}."
            )
        historical_item = item_by_id.get(historical_item_id)
        canonical_item = item_by_id.get(canonical_item_id)
        if historical_item is None or canonical_item is None:
            raise ValueError(
                f"Counterpart adjudication references an unknown item: {historical_item_id}."
            )
        if (
            historical_item_id in selected_item_ids
            or canonical_item_id not in selected_item_ids
        ):
            raise ValueError(
                f"Counterpart adjudication does not map excluded-to-retained: {historical_item_id}."
            )
        historical_source_id = str(historical_item["sourceIdentityId"])
        source_mapping = mapping_by_source.get(historical_source_id, {})
        canonical_source_id = str(canonical_item["sourceIdentityId"])
        if not (
            source_mapping.get("mappingRole") == "alias"
            and source_mapping.get("mappingStatus") == "confirmed-alias"
            and source_mapping.get("collapseEligible")
            and str(source_mapping.get("canonicalEpisodeId")) == canonical_source_id
        ):
            raise ValueError(
                "Counterpart adjudication crosses the governed alias identity boundary: "
                f"{historical_item_id}."
            )
        category_drift = str(historical_item.get("categoryId")) != str(
            canonical_item.get("categoryId")
        )
        if bool(record.get("categoryDrift", False)) != category_drift:
            raise ValueError(
                f"Counterpart category-drift flag is incorrect for {historical_item_id}."
            )
        if record.get("confidence") not in {"high", "medium", "review"}:
            raise ValueError(
                f"Counterpart confidence is not governed for {historical_item_id}."
            )
        for overlap_item_id in record.get("overlapsCanonicalItemIds", []):
            overlap_item = item_by_id.get(str(overlap_item_id))
            if (
                overlap_item is None
                or str(overlap_item_id) == canonical_item_id
                or str(overlap_item_id) not in selected_item_ids
                or str(overlap_item["sourceIdentityId"]) != canonical_source_id
            ):
                raise ValueError(
                    f"Invalid governed overlap counterpart for {historical_item_id}."
                )
        canonical_assignment = assignment_by_item.get(canonical_item_id, {})
        if not canonical_assignment.get("primaryClusterId"):
            historical_assignment = assignment_by_item.get(historical_item_id, {})
            approved_cluster_id = str(record.get("approvedLineageClusterId") or "")
            if (
                not approved_cluster_id
                or approved_cluster_id
                != str(historical_assignment.get("primaryClusterId") or "")
                or not str(record.get("approvedLineageRationale") or "").strip()
            ):
                raise ValueError(
                    "Contextual counterpart lacks an approved proposition-lineage "
                    f"cluster bridge: {historical_item_id}."
                )
        record["historicalItemId"] = historical_item_id
        record["canonicalItemId"] = canonical_item_id
        record["historicalTensionId"] = historical_tension_id
        index[historical_item_id] = record
    return index


def _validated_split_allocation_index(
    design: Mapping[str, Any],
    adjudication: Mapping[str, Any],
) -> dict[tuple[str, str, str, str], dict[str, Any]]:
    """Validate explicit item-and-pole routing for the four split tensions."""

    historical_targets = _historical_to_canonical_targets(design["tensions"])
    canonical_tension_ids = {str(row["id"]) for row in design["tensions"]}
    records = adjudication.get("records")
    if not isinstance(records, list):
        raise ValueError("Split-tension adjudication must contain a records list.")
    index: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for raw_record in records:
        record = dict(raw_record)
        historical_tension_id = str(record.get("historicalTensionId") or "")
        source_candidate_id = str(record.get("sourceCandidateId") or "")
        historical_item_id = str(record.get("historicalItemId") or "")
        historical_pole = str(record.get("historicalPole") or "")
        canonical_item_id = str(record.get("canonicalItemId") or "")
        canonical_tension_id = str(record.get("canonicalTensionId") or "")
        canonical_pole = str(record.get("canonicalPole") or "")
        key = (
            historical_tension_id,
            source_candidate_id,
            historical_item_id,
            historical_pole,
        )
        if historical_tension_id not in SPLIT_HISTORICAL_TENSION_IDS:
            raise ValueError(
                f"Unexpected historical tension in split adjudication: {historical_tension_id}."
            )
        if not source_candidate_id or not historical_item_id or not canonical_item_id:
            raise ValueError(
                "Split-tension adjudication record is missing a lineage key."
            )
        if historical_pole not in {"A", "B"} or canonical_pole not in {"A", "B"}:
            raise ValueError(f"Invalid split-tension pole for {key}.")
        if canonical_tension_id not in canonical_tension_ids:
            raise ValueError(f"Split-tension target does not resolve for {key}.")
        if record.get("confidence") not in {"high", "medium", "review"}:
            raise ValueError(f"Split-tension confidence is not governed for {key}.")
        if not str(record.get("allocationRationale") or "").strip():
            raise ValueError(f"Split-tension rationale is missing for {key}.")
        included = record.get("included")
        if not isinstance(included, bool):
            raise ValueError(f"Split-tension included flag must be Boolean for {key}.")
        if not included and not str(record.get("exclusionReason") or "").strip():
            raise ValueError(
                f"Excluded split-tension occurrence lacks a reason: {key}."
            )
        if key in index:
            raise ValueError(f"Duplicate split-tension adjudication for {key}.")
        record.update(
            {
                "historicalTensionId": historical_tension_id,
                "sourceCandidateId": source_candidate_id,
                "historicalItemId": historical_item_id,
                "historicalPole": historical_pole,
                "canonicalItemId": canonical_item_id,
                "canonicalTensionId": canonical_tension_id,
                "canonicalPole": canonical_pole,
                "included": included,
                "departsFromProposedLegacyTargets": canonical_tension_id
                not in historical_targets.get(historical_tension_id, []),
            }
        )
        index[key] = record
    return index


def _tension_lineage_key(record: Mapping[str, Any]) -> str:
    """Return the stable occurrence key used by private collision decisions."""

    return "|".join(
        (
            str(record.get("historicalTensionId") or ""),
            str(record.get("sourceCandidateId") or ""),
            str(record.get("historicalItemId") or ""),
            str(record.get("sourcePoleOccurrence") or ""),
            str(record.get("canonicalTensionId") or ""),
            str(record.get("normalizedPole") or ""),
        )
    )


def _validated_collision_adjudication_index(
    occurrences: Sequence[Mapping[str, Any]],
    adjudication: Mapping[str, Any],
) -> dict[str, dict[str, Any]]:
    """Validate every semantic collision before global item deduplication.

    A collision is a retained canonical item routed to more than one canonical
    tension.  The private decision table must cover the complete detected set
    and every eligible occurrence, so identifier ordering can never decide a
    cross-tension allocation.
    """

    eligible_by_item: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for occurrence in occurrences:
        if occurrence.get("included"):
            eligible_by_item[str(occurrence["itemId"])].append(occurrence)
    detected = {
        item_id: rows
        for item_id, rows in eligible_by_item.items()
        if len({str(row["canonicalTensionId"]) for row in rows}) > 1
    }

    records = adjudication.get("records")
    if not isinstance(records, list):
        raise ValueError("Cross-tension collision adjudication must contain records.")
    index: dict[str, dict[str, Any]] = {}
    for raw_record in records:
        record = dict(raw_record)
        item_id = str(record.get("canonicalItemId") or "")
        if not item_id or item_id in index:
            raise ValueError(
                "Cross-tension collision adjudication has a missing or duplicate "
                f"canonical item key: {item_id!r}."
            )
        actual_rows = detected.get(item_id)
        if actual_rows is None:
            raise ValueError(
                "Cross-tension collision adjudication does not match a detected "
                f"collision: {item_id}."
            )
        actual_keys = {_tension_lineage_key(row) for row in actual_rows}
        governed_occurrences = record.get("occurrences")
        if not isinstance(governed_occurrences, list):
            raise ValueError(f"Collision occurrence lineage is missing for {item_id}.")
        governed_keys = {
            str(row.get("lineageKey") or "") for row in governed_occurrences
        }
        if actual_keys != governed_keys:
            raise ValueError(
                f"Collision occurrence lineage is incomplete or stale for {item_id}."
            )
        for governed in governed_occurrences:
            expected_key = "|".join(
                (
                    str(governed.get("historicalTensionId") or ""),
                    str(governed.get("sourceCandidateId") or ""),
                    str(governed.get("historicalItemId") or ""),
                    str(governed.get("historicalPole") or ""),
                    str(governed.get("canonicalTensionId") or ""),
                    str(governed.get("canonicalPole") or ""),
                )
            )
            if str(governed.get("lineageKey") or "") != expected_key:
                raise ValueError(f"Malformed governed collision lineage for {item_id}.")

        decision = record.get("decision")
        if not isinstance(decision, Mapping) or decision.get("mode") != "single-use":
            raise ValueError(
                f"Collision {item_id} requires an explicit single-use decision."
            )
        chosen_key = str(decision.get("chosenLineageKey") or "")
        if chosen_key not in actual_keys:
            raise ValueError(f"Collision winner is not eligible for {item_id}.")
        chosen = next(
            row for row in actual_rows if _tension_lineage_key(row) == chosen_key
        )
        if str(decision.get("chosenCanonicalTensionId") or "") != str(
            chosen["canonicalTensionId"]
        ) or str(decision.get("chosenCanonicalPole") or "") != str(
            chosen["normalizedPole"]
        ):
            raise ValueError(f"Collision winner target or pole is stale for {item_id}.")
        if (
            float(decision.get("itemWeightBudget", -1)) != 1.0
            or float(decision.get("chosenSupportWeight", -1)) != 1.0
            or float(decision.get("rejectedOccurrenceSupportWeight", -1)) != 0.0
            or bool(decision.get("distinctPropositionMultiUseJustified"))
        ):
            raise ValueError(f"Collision weight budget is invalid for {item_id}.")
        if record.get("confidence") not in {"high", "medium", "review"}:
            raise ValueError(f"Collision confidence is not governed for {item_id}.")
        if not str(record.get("rationale") or "").strip():
            raise ValueError(f"Collision rationale is missing for {item_id}.")
        if (
            bool(record.get("reviewRequired"))
            and not str(record.get("reviewReason") or "").strip()
        ):
            raise ValueError(f"Collision review reason is missing for {item_id}.")
        record["canonicalItemId"] = item_id
        index[item_id] = record

    if set(index) != set(detected):
        missing = sorted(set(detected) - set(index), key=_natural)
        extra = sorted(set(index) - set(detected), key=_natural)
        raise ValueError(
            "Cross-tension collision adjudication coverage mismatch: "
            f"missing={missing}, extra={extra}."
        )
    return index


def _validated_same_tension_adjudication_index(
    occurrences: Sequence[Mapping[str, Any]],
    adjudication: Mapping[str, Any],
) -> dict[str, dict[str, Any]]:
    """Validate repeated lineage inside one canonical tension and its weight budget."""

    eligible_by_item: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for occurrence in occurrences:
        if occurrence.get("included"):
            eligible_by_item[str(occurrence["itemId"])].append(occurrence)
    detected = {
        item_id: rows
        for item_id, rows in eligible_by_item.items()
        if len(rows) > 1 and len({str(row["canonicalTensionId"]) for row in rows}) == 1
    }
    records = adjudication.get("records")
    if not isinstance(records, list):
        raise ValueError("Same-tension duplicate adjudication must contain records.")
    index: dict[str, dict[str, Any]] = {}
    for raw_record in records:
        record = dict(raw_record)
        item_id = str(record.get("canonicalItemId") or "")
        if not item_id or item_id in index or item_id not in detected:
            raise ValueError(
                "Same-tension adjudication has a missing, duplicate, or stale item: "
                f"{item_id!r}."
            )
        actual_rows = detected[item_id]
        actual_keys = {_tension_lineage_key(row) for row in actual_rows}
        governed_occurrences = record.get("occurrences")
        if not isinstance(governed_occurrences, list):
            raise ValueError(
                f"Same-tension occurrence lineage is missing for {item_id}."
            )
        governed_keys = {
            str(row.get("lineageKey") or "") for row in governed_occurrences
        }
        if actual_keys != governed_keys:
            raise ValueError(
                f"Same-tension occurrence lineage is incomplete or stale for {item_id}."
            )
        if str(record.get("canonicalTensionId") or "") != str(
            actual_rows[0]["canonicalTensionId"]
        ):
            raise ValueError(f"Same-tension target is stale for {item_id}.")
        weights: dict[str, float] = {}
        for governed in governed_occurrences:
            expected_key = "|".join(
                (
                    str(governed.get("historicalTensionId") or ""),
                    str(governed.get("sourceCandidateId") or ""),
                    str(governed.get("historicalItemId") or ""),
                    str(governed.get("sourcePoleOccurrence") or ""),
                    str(record["canonicalTensionId"]),
                    str(governed.get("canonicalPole") or ""),
                )
            )
            lineage_key = str(governed.get("lineageKey") or "")
            weight = float(governed.get("occurrenceSupportWeight", -1))
            if lineage_key != expected_key or not 0.0 <= weight <= 1.0:
                raise ValueError(f"Malformed same-tension lineage for {item_id}.")
            weights[lineage_key] = weight
        decision = record.get("decision", {})
        mode = str(decision.get("mode") or "")
        if (
            mode
            not in {
                "single-pole-single-lineage",
                "same-pole-shared-provenance",
                "genuine-dual-pole-bridge",
            }
            or float(decision.get("canonicalItemWeightBudget", -1)) != 1.0
            or int(decision.get("quantitativeItemCount", -1)) != 1
            or not math.isclose(sum(weights.values()), 1.0, abs_tol=1e-9)
        ):
            raise ValueError(f"Same-tension weight budget is invalid for {item_id}.")
        poles = {
            str(row.get("canonicalPole") or "")
            for row in governed_occurrences
            if float(row.get("occurrenceSupportWeight", 0)) > 0
        }
        if mode == "genuine-dual-pole-bridge" and poles != {"A", "B"}:
            raise ValueError(f"Dual-pole bridge is malformed for {item_id}.")
        if mode != "genuine-dual-pole-bridge" and len(poles) != 1:
            raise ValueError(f"Single-pole duplicate is malformed for {item_id}.")
        if record.get("confidence") not in {"high", "medium", "review"}:
            raise ValueError(f"Same-tension confidence is invalid for {item_id}.")
        if not str(record.get("rationale") or "").strip():
            raise ValueError(f"Same-tension rationale is missing for {item_id}.")
        if (
            bool(record.get("reviewRequired"))
            and not str(record.get("reviewReason") or "").strip()
        ):
            raise ValueError(f"Same-tension review reason is missing for {item_id}.")
        record["canonicalItemId"] = item_id
        record["occurrenceWeights"] = weights
        index[item_id] = record
    if set(index) != set(detected):
        raise ValueError(
            "Same-tension adjudication coverage mismatch: "
            f"missing={sorted(set(detected) - set(index), key=_natural)}, "
            f"extra={sorted(set(index) - set(detected), key=_natural)}."
        )
    return index


def _validated_tension_adjudication_index(
    tension_seeds: Mapping[str, Mapping[str, Any]],
    included_occurrences: Sequence[Mapping[str, Any]],
    item_by_id: Mapping[str, Mapping[str, Any]],
    adjudication: Mapping[str, Any],
) -> dict[str, dict[str, Any]]:
    """Validate independent construct-level decisions against final evidence."""

    expected_policy = {
        "proposedCountIsBinding": False,
        "allowedDecisions": ["retain", "merge", "split", "reject"],
        "retentionRequiresBothPolesAndDistinctBoundaries": True,
        "reviewFlagsAreNonBlockingButRequireHumanConfirmation": True,
    }
    expected_accounting = {
        "occurrenceCountsIncludePositiveWeightLineageRows": True,
        "analyticalWeightsRespectOneUnitPerCanonicalItem": True,
        "uniqueItemsAreGloballySingleUseAcrossTensions": True,
        "sourceConcentrationIsALimitationNotAValidityScore": True,
    }

    def exact_governance_mapping(value: Any, expected: Mapping[str, Any]) -> bool:
        return (
            isinstance(value, Mapping)
            and set(value) == set(expected)
            and all(
                type(value[key]) is type(expected_value)
                and value[key] == expected_value
                for key, expected_value in expected.items()
            )
        )

    scope = adjudication.get("scope")
    if (
        adjudication.get("schemaVersion") != "1.0.0"
        or adjudication.get("methodVersion")
        != "canonical-tension-construct-adjudication-v1"
        or adjudication.get("adjudicationStatus")
        != "analyst-reviewed-private-governed-input"
        or not isinstance(scope, str)
        or not scope.strip()
        or not exact_governance_mapping(
            adjudication.get("decisionPolicy"), expected_policy
        )
        or not exact_governance_mapping(
            adjudication.get("evidenceAccounting"), expected_accounting
        )
    ):
        raise ValueError("Canonical tension adjudication governance header is invalid.")
    raw_records = adjudication.get("records")
    if not isinstance(raw_records, list):
        raise ValueError("Canonical tension adjudication must contain records.")
    expected_ids = set(tension_seeds)
    index: dict[str, dict[str, Any]] = {}
    for raw_record in raw_records:
        record = dict(raw_record)
        tension_id = str(record.get("canonicalTensionId") or "")
        if not tension_id or tension_id in index or tension_id not in expected_ids:
            raise ValueError(
                "Canonical tension adjudication has a missing, duplicate, or "
                f"stale target: {tension_id!r}."
            )
        seed = tension_seeds[tension_id]
        governed_definition = {}
        for field in (
            "name",
            "tensionType",
            "definition",
            "poleALabel",
            "poleBLabel",
        ):
            value = record.get(field)
            if not isinstance(value, str) or not value.strip():
                raise ValueError(
                    f"Canonical tension definition is stale for {tension_id}."
                )
            governed_definition[field] = value
        expected_definition = {
            "name": seed["name"],
            "tensionType": seed["type"],
            "definition": seed["definition"],
            "poleALabel": seed["poleA"],
            "poleBLabel": seed["poleB"],
        }
        if governed_definition != expected_definition:
            raise ValueError(f"Canonical tension definition is stale for {tension_id}.")
        decision = str(record.get("decision") or "")
        if decision not in {"retain", "merge", "split", "reject"}:
            raise ValueError(f"Canonical tension decision is invalid for {tension_id}.")
        if decision != "retain":
            raise ValueError(
                f"Canonical tension {tension_id} was not retained; the output "
                "architecture must be revised before this builder can continue."
            )
        confidence = str(record.get("confidence") or "")
        if confidence not in {"high", "medium", "review"}:
            raise ValueError(
                f"Canonical tension confidence is invalid for {tension_id}."
            )
        review_required = record.get("reviewRequired")
        review_flags = record.get("reviewFlags")
        if (
            not isinstance(review_required, bool)
            or not isinstance(review_flags, list)
            or any(
                not isinstance(value, str) or not value.strip()
                for value in review_flags
            )
            or (review_required and not review_flags)
            or (not review_required and review_flags)
            or (confidence in {"medium", "review"} and not review_required)
        ):
            raise ValueError(
                f"Canonical tension review state is invalid for {tension_id}."
            )
        rationale = record.get("rationale")
        if not isinstance(rationale, str) or not rationale.strip():
            raise ValueError(
                f"Canonical tension rationale is missing for {tension_id}."
            )

        proposed_historical_ids = sorted(
            set(re.findall(r"TD-\d{3}", str(seed.get("legacy", "")))),
            key=_natural,
        )
        evidence = [
            row
            for row in included_occurrences
            if str(row["canonicalTensionId"]) == tension_id
        ]
        actual_historical_ids = sorted(
            {
                str(row.get("historicalTensionId") or "")
                for row in evidence
                if float(row.get("analyticalSupportWeight", 0.0)) > 0.0
            },
            key=_natural,
        )
        if not actual_historical_ids or "" in actual_historical_ids:
            raise ValueError(
                f"Historical tension evidence lineage is missing for {tension_id}."
            )
        for field, expected in (
            ("proposedHistoricalTensionIds", proposed_historical_ids),
            ("historicalTensionIds", actual_historical_ids),
        ):
            values = record.get(field)
            if (
                not isinstance(values, list)
                or any(
                    not isinstance(value, str) or not re.fullmatch(r"TD-\d{3}", value)
                    for value in values
                )
                or values != expected
            ):
                raise ValueError(
                    f"Historical tension lineage is stale for {tension_id}/{field}."
                )

        item_ids = {str(row["itemId"]) for row in evidence}
        content_units = {
            str(item_by_id[item_id]["sourceIdentityId"]) for item_id in item_ids
        }
        focal_categories = {
            str(item_by_id[item_id]["categoryId"])
            for item_id in item_ids
            if item_by_id[item_id].get("scope") == "focal"
        }
        observed_evidence = {
            "poleAOccurrences": sum(row["normalizedPole"] == "A" for row in evidence),
            "poleBOccurrences": sum(row["normalizedPole"] == "B" for row in evidence),
            "poleAWeight": _round(
                sum(
                    float(row["analyticalSupportWeight"])
                    for row in evidence
                    if row["normalizedPole"] == "A"
                )
            ),
            "poleBWeight": _round(
                sum(
                    float(row["analyticalSupportWeight"])
                    for row in evidence
                    if row["normalizedPole"] == "B"
                )
            ),
            "uniqueItems": len(item_ids),
            "contentUnits": len(content_units),
            "categoryBreadth": len(focal_categories),
            "clusterCount": len(
                {
                    str(row["primaryClusterId"])
                    for row in evidence
                    if row.get("primaryClusterId")
                }
            ),
        }
        governed_evidence = record.get("evidenceSummary")
        assessment = (
            governed_evidence.get("assessment")
            if isinstance(governed_evidence, Mapping)
            else None
        )
        if (
            not isinstance(governed_evidence, Mapping)
            or not isinstance(assessment, str)
            or not assessment.strip()
        ):
            raise ValueError(
                f"Canonical tension evidence review is missing for {tension_id}."
            )
        for field, expected in observed_evidence.items():
            actual = governed_evidence.get(field)
            if isinstance(expected, float):
                matches = isinstance(actual, (int, float)) and math.isclose(
                    float(actual), expected, abs_tol=1e-9
                )
            else:
                matches = actual == expected
            if not matches:
                raise ValueError(
                    f"Canonical tension evidence is stale for {tension_id}/{field}: "
                    f"expected {expected!r}, found {actual!r}."
                )

        neighbors = record.get("neighborDistinctions")
        if (
            not isinstance(neighbors, Mapping)
            or len(neighbors) < 2
            or any(
                str(neighbor_id) not in expected_ids
                or str(neighbor_id) == tension_id
                or not isinstance(neighbor_id, str)
                or not isinstance(distinction, str)
                or not distinction.strip()
                for neighbor_id, distinction in neighbors.items()
            )
        ):
            raise ValueError(
                f"Canonical tension boundaries are incomplete for {tension_id}."
            )
        if any(
            not isinstance(record.get(field), str) or not record[field].strip()
            for field in (
                "poleAAssumption",
                "poleBAssumption",
                "falseDichotomyCaveat",
            )
        ):
            raise ValueError(
                f"Canonical tension pole logic is incomplete for {tension_id}."
            )
        for field in ("conditionsFavoringA", "conditionsFavoringB", "limitations"):
            values = record.get(field)
            minimum = 2 if field.startswith("conditions") else 1
            if (
                not isinstance(values, list)
                or len(values) < minimum
                or any(
                    not isinstance(value, str) or not value.strip() for value in values
                )
            ):
                raise ValueError(
                    f"Canonical tension {field} is incomplete for {tension_id}."
                )
        index[tension_id] = record

    if set(index) != expected_ids:
        raise ValueError("Canonical tension adjudication coverage is incomplete.")
    decisions = Counter(str(row["decision"]) for row in index.values())
    confidences = Counter(str(row["confidence"]) for row in index.values())
    summary = adjudication.get("decisionSummary", {})
    expected_summary = {
        "candidateCount": len(expected_ids),
        "retainCount": decisions["retain"],
        "mergeCount": decisions["merge"],
        "splitCount": decisions["split"],
        "rejectCount": decisions["reject"],
        "highConfidenceCount": confidences["high"],
        "mediumConfidenceCount": confidences["medium"],
        "reviewConfidenceCount": confidences["review"],
        "reviewRequiredCount": sum(
            bool(row["reviewRequired"]) for row in index.values()
        ),
        "unresolvedStructuralDecisionCount": 0,
    }
    if not exact_governance_mapping(summary, expected_summary):
        raise ValueError("Canonical tension adjudication summary is stale.")
    expected_completeness = {
        "allProposedTensionsReviewed": True,
        "allRetainedTensionsHaveBothPoles": True,
        "allRetainedTensionsHaveDistinctNeighborBoundaries": True,
        "allMediumConfidenceDecisionsHaveReviewFlags": True,
        "noStructuralDecisionUnresolved": True,
    }
    if not exact_governance_mapping(
        adjudication.get("completenessRequirements"), expected_completeness
    ):
        raise ValueError(
            "Canonical tension adjudication completeness checks did not pass."
        )
    return index


def build_tensions(
    dataset: Mapping[str, Sequence[Mapping[str, Any]]],
    design: Mapping[str, Any],
    tension_workbook: Path,
    selected_item_ids: set[str],
    cluster_to_family: Mapping[str, str],
    release_by_content: Mapping[str, set[str]],
    counterpart_adjudication: Mapping[str, Any],
    split_allocation_adjudication: Mapping[str, Any],
    collision_adjudication: Mapping[str, Any],
    same_tension_adjudication: Mapping[str, Any],
    tension_adjudication: Mapping[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, Any], dict[str, set[str]]]:
    item_by_id = {str(row["itemId"]): row for row in dataset["items"]}
    assignments = {
        str(row["itemId"]): row for row in dataset["item_cluster_assignments"]
    }
    tension_seeds = {str(row["id"]): row for row in design["tensions"]}
    historical_targets = _historical_to_canonical_targets(design["tensions"])
    counterpart_index = _validated_counterpart_index(
        dataset, selected_item_ids, counterpart_adjudication
    )
    split_index = _validated_split_allocation_index(
        design, split_allocation_adjudication
    )
    candidate_rows = _load_tension_candidates(tension_workbook)
    occurrences: list[dict[str, Any]] = []
    used_counterpart_ids: set[str] = set()
    used_split_keys: set[tuple[str, str, str, str]] = set()

    for candidate in candidate_rows:
        historical_id = str(candidate["tension_id"])
        candidate_id = str(candidate["source_candidate_id"])
        targets = tuple(historical_targets.get(historical_id, ()))
        if not targets:
            raise ValueError(
                f"No canonical target for tension candidate {candidate_id}."
            )
        if len(targets) > 1 and historical_id not in SPLIT_HISTORICAL_TENSION_IDS:
            raise ValueError(
                f"Multiple canonical targets lack split adjudication: {historical_id}."
            )
        pole_items = {
            "A": _split_ids(candidate.get("supporting_item_ids_pole_a")),
            "B": _split_ids(candidate.get("supporting_item_ids_pole_b")),
        }
        both = set(pole_items["A"]) & set(pole_items["B"])
        for historical_pole, item_ids in pole_items.items():
            for item_id in item_ids:
                split_key = (historical_id, candidate_id, item_id, historical_pole)
                split_record = None
                if historical_id in SPLIT_HISTORICAL_TENSION_IDS:
                    split_record = split_index.get(split_key)
                    if split_record is None:
                        raise ValueError(
                            f"Missing governed split-tension allocation for {split_key}."
                        )
                    used_split_keys.add(split_key)
                    target = str(split_record["canonicalTensionId"])
                    canonical_pole = str(split_record["canonicalPole"])
                    orientation = str(
                        split_record.get("orientationTreatment")
                        or "governed-item-pole-adjudication"
                    )
                    allocation_authority = "governed-split-item-adjudication"
                    allocation_rationale = str(split_record["allocationRationale"])
                else:
                    if len(targets) != 1:
                        raise ValueError(
                            f"Unadjudicated split remains for {historical_id}."
                        )
                    target = targets[0]
                    canonical_pole = historical_pole
                    orientation = "preserved"
                    if historical_id == "TD-024":
                        canonical_pole = "B" if historical_pole == "A" else "A"
                        orientation = "reversed-to-canonical-orientation"
                    allocation_authority = "canonical-architecture-candidate-lineage"
                    allocation_rationale = (
                        "The historical tension has one governed canonical target; "
                        "the source candidate and item retain their original lineage."
                    )
                item = item_by_id.get(item_id)
                if item is None:
                    raise ValueError(
                        "Tension evidence references an unknown historical item: "
                        f"{historical_id}/{candidate_id}/{historical_pole}/{item_id}."
                    )
                evidence_item = item
                counterpart_record = None
                lineage_treatment = "direct-retained-item"
                if item_id not in selected_item_ids:
                    counterpart_record = counterpart_index.get(item_id)
                    if counterpart_record is None:
                        raise ValueError(
                            "Excluded tension evidence lacks a governed canonical "
                            f"counterpart adjudication: {item_id}."
                        )
                    if counterpart_record["historicalTensionId"] != historical_id:
                        raise ValueError(
                            "Counterpart adjudication tension provenance mismatch for "
                            f"{item_id}."
                        )
                    evidence_item = item_by_id[counterpart_record["canonicalItemId"]]
                    used_counterpart_ids.add(item_id)
                    lineage_treatment = "governed-canonical-counterpart-substitution"
                canonical_item_id = str(evidence_item["itemId"])
                if (
                    split_record
                    and split_record["canonicalItemId"] != canonical_item_id
                ):
                    raise ValueError(
                        "Split-tension and counterpart adjudications disagree for "
                        f"{split_key}."
                    )
                assignment = assignments.get(canonical_item_id, {})
                cluster_id = str(assignment.get("primaryClusterId") or "")
                cluster_lineage_treatment = "canonical-item-primary-assignment"
                if not cluster_id and counterpart_record:
                    cluster_id = str(
                        counterpart_record.get("approvedLineageClusterId") or ""
                    )
                    if cluster_id:
                        cluster_lineage_treatment = (
                            "governed-alias-proposition-historical-cluster-lineage"
                        )
                family_id = cluster_to_family.get(cluster_id)
                counterpart_confidence = (
                    counterpart_record.get("confidence") if counterpart_record else None
                )
                split_confidence = (
                    split_record.get("confidence") if split_record else None
                )
                occurrences.append(
                    {
                        "historicalTensionId": historical_id,
                        "sourceCandidateId": candidate_id,
                        "historicalItemId": item_id,
                        "itemId": canonical_item_id,
                        "canonicalContentUnitId": evidence_item["sourceIdentityId"],
                        "sourcePoleOccurrence": historical_pole,
                        "historicalPole": "both"
                        if item_id in both
                        else historical_pole,
                        "normalizedPole": canonical_pole,
                        "canonicalTensionId": target,
                        "primaryClusterId": cluster_id or None,
                        "primaryFamilyId": family_id,
                        "clusterLineageTreatment": cluster_lineage_treatment,
                        "clusterLineageRationale": (
                            counterpart_record.get("approvedLineageRationale")
                            if counterpart_record
                            and counterpart_record.get("approvedLineageClusterId")
                            else None
                        ),
                        "semanticRole": f"tension-evidence-pole-{canonical_pole.lower()}",
                        "orientationTreatment": orientation,
                        "lineageTreatment": lineage_treatment,
                        "counterpartMappingConfidence": counterpart_confidence,
                        "counterpartCategoryDrift": bool(
                            counterpart_record
                            and counterpart_record.get("categoryDrift")
                        ),
                        "counterpartBothPoleBridge": bool(
                            counterpart_record
                            and counterpart_record.get("bothPoleBridge")
                        ),
                        "counterpartOverlapCanonicalItemIds": sorted(
                            (
                                str(value)
                                for value in counterpart_record.get(
                                    "overlapsCanonicalItemIds", []
                                )
                            )
                            if counterpart_record
                            else [],
                            key=_natural,
                        ),
                        "counterpartAdjudicationNote": (
                            counterpart_record.get("adjudicationNote")
                            if counterpart_record
                            else None
                        ),
                        "splitAllocationConfidence": split_confidence,
                        "splitAdjudicationIncluded": (
                            split_record.get("included", True) if split_record else None
                        ),
                        "splitAdjudicationExclusionReason": (
                            split_record.get("exclusionReason")
                            if split_record
                            else None
                        ),
                        "splitBothPoleBridge": bool(
                            split_record and split_record.get("bothPoleBridge")
                        ),
                        "splitCategoryDrift": bool(
                            split_record and split_record.get("categoryDrift")
                        ),
                        "splitDepartsFromProposedLegacyTargets": bool(
                            split_record
                            and split_record.get("departsFromProposedLegacyTargets")
                        ),
                        "splitReviewNote": (
                            split_record.get("reviewNote") if split_record else None
                        ),
                        "allocationAuthority": allocation_authority,
                        "included": (
                            canonical_item_id in selected_item_ids
                            and not (
                                split_record and split_record.get("included") is False
                            )
                        ),
                        "exclusionReason": (
                            str(split_record.get("exclusionReason"))
                            if split_record and split_record.get("included") is False
                            else (
                                None
                                if canonical_item_id in selected_item_ids
                                else "governed-counterpart-is-not-selected"
                            )
                        ),
                        "allocationRationale": allocation_rationale,
                        "analyticalSupportWeight": 0.0,
                        "adjudicationStatus": (
                            "revised"
                            if split_record and split_record.get("included") is False
                            else "analyst-reviewed-draft"
                        ),
                    }
                )

    unused_counterparts = sorted(
        set(counterpart_index) - used_counterpart_ids, key=_natural
    )
    if unused_counterparts:
        raise ValueError(
            "Counterpart adjudications are not represented in tension evidence: "
            + ", ".join(unused_counterparts)
        )
    unused_split_allocations = sorted(
        set(split_index) - used_split_keys,
        key=lambda key: tuple(_natural(value) for value in key),
    )
    if unused_split_allocations:
        raise ValueError(
            "Split-tension adjudications are not represented in tension evidence: "
            f"{len(unused_split_allocations)} unused records."
        )
    occurrence_keys = [
        (
            str(row.get("historicalTensionId")),
            str(row.get("sourceCandidateId")),
            str(row.get("historicalItemId")),
            str(row.get("sourcePoleOccurrence")),
        )
        for row in occurrences
    ]
    if len(occurrence_keys) != len(set(occurrence_keys)):
        raise ValueError("Tension source-pole occurrence lineage is not unique.")

    # A historical item can recur in several candidates or on both source poles.
    # Cross-tension collisions require an explicit semantic adjudication; they
    # must never be resolved by canonical identifier or source ordering.
    by_item: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in occurrences:
        if record.get("included"):
            by_item[str(record["itemId"])].append(record)
    collision_index = _validated_collision_adjudication_index(
        occurrences, collision_adjudication
    )
    same_tension_index = _validated_same_tension_adjudication_index(
        occurrences, same_tension_adjudication
    )
    for item_id, records in by_item.items():
        if len(records) <= 1:
            records[0]["analyticalSupportWeight"] = 1.0
            continue
        cross_tension = (
            len({str(record["canonicalTensionId"]) for record in records}) > 1
        )
        if cross_tension:
            adjudication = collision_index[item_id]
            chosen_key = str(adjudication["decision"]["chosenLineageKey"])
            winner = next(
                record
                for record in records
                if _tension_lineage_key(record) == chosen_key
            )
            for record in records:
                record["crossTensionCollision"] = True
                record["collisionDecisionMode"] = "single-use"
                record["collisionAdjudicationConfidence"] = adjudication["confidence"]
                record["collisionReviewRequired"] = bool(
                    adjudication.get("reviewRequired")
                )
                record["collisionReviewReason"] = adjudication.get("reviewReason")
                record["collisionRationale"] = adjudication["rationale"]
                if record is winner:
                    record["analyticalSupportWeight"] = 1.0
                    continue
                record["included"] = False
                record["analyticalSupportWeight"] = 0.0
                record["exclusionReason"] = (
                    "cross-tension-collision-governed-to-"
                    f"{winner['canonicalTensionId']}-pole-"
                    f"{str(winner['normalizedPole']).lower()}"
                )
                record["adjudicationStatus"] = "revised"
            continue

        governed = same_tension_index[item_id]
        weights = governed["occurrenceWeights"]
        for record in records:
            weight = float(weights[_tension_lineage_key(record)])
            record["sameTensionDuplicate"] = True
            record["sameTensionDecisionMode"] = governed["decision"]["mode"]
            record["sameTensionAdjudicationConfidence"] = governed["confidence"]
            record["sameTensionReviewRequired"] = bool(governed.get("reviewRequired"))
            record["sameTensionReviewReason"] = governed.get("reviewReason")
            record["sameTensionRationale"] = governed["rationale"]
            record["analyticalSupportWeight"] = weight
            record["included"] = weight > 0.0
            if not record["included"]:
                record["exclusionReason"] = "same-tension-governed-zero-weight-lineage"
                record["adjudicationStatus"] = "revised"

    included = [row for row in occurrences if row.get("included")]
    tension_item_ids: dict[str, set[str]] = defaultdict(set)
    for row in included:
        tension_item_ids[str(row["canonicalTensionId"])].add(str(row["itemId"]))
    tension_decision_index = _validated_tension_adjudication_index(
        tension_seeds, included, item_by_id, tension_adjudication
    )

    records = []
    for tension_id in sorted(tension_seeds, key=_natural):
        governed = tension_decision_index[tension_id]
        evidence = [row for row in included if row["canonicalTensionId"] == tension_id]
        pole_a = {
            str(row["itemId"]) for row in evidence if row["normalizedPole"] == "A"
        }
        pole_b = {
            str(row["itemId"]) for row in evidence if row["normalizedPole"] == "B"
        }
        pole_a_weight = sum(
            float(row.get("analyticalSupportWeight", 0.0))
            for row in evidence
            if row["normalizedPole"] == "A"
        )
        pole_b_weight = sum(
            float(row.get("analyticalSupportWeight", 0.0))
            for row in evidence
            if row["normalizedPole"] == "B"
        )
        item_ids = pole_a | pole_b
        cluster_ids = sorted(
            {
                str(row["primaryClusterId"])
                for row in evidence
                if row.get("primaryClusterId")
            },
            key=_natural,
        )
        family_ids = sorted(
            {
                str(row["primaryFamilyId"])
                for row in evidence
                if row.get("primaryFamilyId")
            },
            key=_natural,
        )
        content_units = sorted(
            {str(item_by_id[item_id]["sourceIdentityId"]) for item_id in item_ids},
            key=_natural,
        )
        all_historical_ids = {
            str(row.get("historicalItemId") or row["itemId"])
            for row in occurrences
            if row.get("canonicalTensionId") == tension_id
        }
        limitations = [
            "Tension evidence is a governed allocation of historical candidate evidence, not a new independent coding pass.",
            *[str(value) for value in governed["limitations"]],
        ]
        if not pole_a or not pole_b:
            limitations.append(
                "One pole lacks retained direct evidence and requires human review."
            )
        profile = _support_profile(
            item_ids,
            historical_item_ids=all_historical_ids,
            direct_item_ids=item_ids,
            item_by_id=item_by_id,
            release_by_content=release_by_content,
            cluster_ids=cluster_ids,
            family_ids=family_ids,
            adjudication_status="analyst-reviewed-draft",
            limitations=limitations,
        )
        total_weight = pole_a_weight + pole_b_weight
        profile["poleBalance"] = {
            "poleAItemCount": len(pole_a),
            "poleBItemCount": len(pole_b),
            "sharedAcrossPolesItemCount": len(pole_a & pole_b),
            "poleAAnalyticalWeight": _round(pole_a_weight),
            "poleBAnalyticalWeight": _round(pole_b_weight),
            "totalAnalyticalWeight": _round(total_weight),
            "poleAShare": (
                _round(pole_a_weight / total_weight) if total_weight else 0.0
            ),
            "poleBShare": (
                _round(pole_b_weight / total_weight) if total_weight else 0.0
            ),
            "bothPolesDirectlySupported": bool(pole_a and pole_b),
        }
        records.append(
            {
                "tensionId": tension_id,
                "name": governed["name"],
                "tensionType": governed["tensionType"],
                "definition": governed["definition"],
                "poleALabel": governed["poleALabel"],
                "poleAAssumption": governed["poleAAssumption"],
                "poleBLabel": governed["poleBLabel"],
                "poleBAssumption": governed["poleBAssumption"],
                "conditionsFavoringA": governed["conditionsFavoringA"],
                "conditionsFavoringB": governed["conditionsFavoringB"],
                "falseDichotomyCaveat": governed["falseDichotomyCaveat"],
                "sourceCandidateIds": sorted(
                    {str(row["sourceCandidateId"]) for row in evidence}, key=_natural
                ),
                "historicalTensionIds": governed["historicalTensionIds"],
                "proposedHistoricalTensionIds": governed[
                    "proposedHistoricalTensionIds"
                ],
                "supportingFamilyIds": family_ids,
                "supportingClusterIds": cluster_ids,
                "supportingCanonicalContentUnits": content_units,
                "corpusSupportProfile": profile,
                "evidenceBalanceAcrossPoles": profile["poleBalance"],
                "neighborDistinctions": governed["neighborDistinctions"],
                "limitations": limitations,
                "adjudicationStatus": "analyst-reviewed-draft",
                "adjudicationConfidence": governed["confidence"],
                "adjudicationDecision": governed["decision"],
                "adjudicationRationale": governed["rationale"],
                "evidenceAssessment": governed["evidenceSummary"]["assessment"],
                "reviewRequired": governed["reviewRequired"],
                "reviewFlags": governed["reviewFlags"],
            }
        )

    split_decision_summary = []
    for historical_id in sorted(SPLIT_HISTORICAL_TENSION_IDS, key=_natural):
        source_rows = [
            row
            for row in occurrences
            if row.get("historicalTensionId") == historical_id
        ]
        assigned_counts = Counter(
            (str(row["canonicalTensionId"]), str(row["normalizedPole"]))
            for row in source_rows
        )
        included_counts = Counter(
            (str(row["canonicalTensionId"]), str(row["normalizedPole"]))
            for row in source_rows
            if row.get("included")
        )
        split_decision_summary.append(
            {
                "historicalTensionId": historical_id,
                "sourcePoleOccurrenceCount": len(source_rows),
                "assignedByCanonicalTensionAndPole": [
                    {
                        "canonicalTensionId": tension_id,
                        "canonicalPole": pole,
                        "occurrenceCount": count,
                    }
                    for (tension_id, pole), count in sorted(
                        assigned_counts.items(),
                        key=lambda row: (_natural(row[0][0]), row[0][1]),
                    )
                ],
                "includedAfterGlobalDeduplication": [
                    {
                        "canonicalTensionId": tension_id,
                        "canonicalPole": pole,
                        "uniqueItemCount": count,
                    }
                    for (tension_id, pole), count in sorted(
                        included_counts.items(),
                        key=lambda row: (_natural(row[0][0]), row[0][1]),
                    )
                ],
            }
        )

    counterpart_occurrences = [
        row
        for row in occurrences
        if row.get("lineageTreatment") == "governed-canonical-counterpart-substitution"
    ]
    allocation = {
        "schemaVersion": SCHEMA_VERSION,
        "methodVersion": METHOD_VERSION,
        "allocationRules": {
            "sourceCandidateFirst": True,
            "itemLevelForOverBroadCandidates": True,
            "globalItemDoubleCountingPermitted": False,
            "crossTensionCollisionsRequireSemanticAdjudication": True,
            "canonicalIdentifierOrderingUsedForCrossTensionCollisions": False,
            "rejectedCollisionOccurrencesRetainZeroWeightLineage": True,
            "sameTensionRepeatedLineageRequiresAdjudication": True,
            "dualPoleItemsUseSharedUnitWeightBudget": True,
            "td024OrientationNormalized": True,
            "lexicalSimilarityUsedForCounterpartIdentity": False,
            "lexicalOrKeywordRoutingUsedForSplitTensions": False,
            "governedPrivateAdjudicationRequired": True,
        },
        "canonicalTensionConstructAdjudication": {
            "methodVersion": tension_adjudication["methodVersion"],
            "decisionSummary": tension_adjudication["decisionSummary"],
        },
        "counts": {
            "sourceCandidateRows": len(candidate_rows),
            "sourceCandidateIdsRepresented": len(
                {row["sourceCandidateId"] for row in occurrences}
            ),
            "allocationOccurrences": len(occurrences),
            "includedUniqueItems": len({row["itemId"] for row in included}),
            "counterpartAdjudicationRecords": len(counterpart_index),
            "uniqueGovernedCounterpartMappingsUsed": len(used_counterpart_ids),
            "distinctGovernedCounterpartTargets": len(
                {row["canonicalItemId"] for row in counterpart_index.values()}
            ),
            "governedCanonicalCounterpartSubstitutionOccurrences": len(
                counterpart_occurrences
            ),
            "counterpartHighConfidenceRecords": sum(
                row.get("confidence") == "high" for row in counterpart_index.values()
            ),
            "counterpartMediumConfidenceRecords": sum(
                row.get("confidence") == "medium" for row in counterpart_index.values()
            ),
            "counterpartReviewRecords": sum(
                row.get("confidence") == "review" for row in counterpart_index.values()
            ),
            "counterpartCategoryDriftRecords": sum(
                bool(row.get("categoryDrift")) for row in counterpart_index.values()
            ),
            "counterpartCompoundOverlapRecords": sum(
                bool(row.get("overlapsCanonicalItemIds"))
                for row in counterpart_index.values()
            ),
            "counterpartApprovedClusterLineageBridgeRecords": sum(
                bool(row.get("approvedLineageClusterId"))
                for row in counterpart_index.values()
            ),
            "counterpartBothPoleBridgeRecords": sum(
                bool(row.get("bothPoleBridge")) for row in counterpart_index.values()
            ),
            "counterpartBothPoleBridgeOccurrences": sum(
                bool(row.get("counterpartBothPoleBridge"))
                for row in counterpart_occurrences
            ),
            "unresolvedExcludedSourceIdentityOccurrences": sum(
                row.get("exclusionReason")
                in {
                    "excluded-source-identity-without-governed-counterpart",
                    "governed-counterpart-is-not-selected",
                }
                for row in occurrences
            ),
            "splitAdjudicationRecords": len(split_index),
            "splitAdjudicationRecordsUsed": len(used_split_keys),
            "splitSourcePoleOccurrences": sum(
                row.get("historicalTensionId") in SPLIT_HISTORICAL_TENSION_IDS
                for row in occurrences
            ),
            "splitOccurrencesIncludedByAdjudication": sum(
                record.get("included", True) is not False
                for record in split_index.values()
            ),
            "splitDuplicateBridgeOccurrencesExcludedByAdjudication": sum(
                record.get("included") is False and bool(record.get("bothPoleBridge"))
                for record in split_index.values()
            ),
            "splitReviewRecords": sum(
                record.get("confidence") == "review" for record in split_index.values()
            ),
            "splitHighConfidenceRecords": sum(
                record.get("confidence") == "high" for record in split_index.values()
            ),
            "splitMediumConfidenceRecords": sum(
                record.get("confidence") == "medium" for record in split_index.values()
            ),
            "splitOccurrencesDepartingFromProposedLegacyTargets": sum(
                bool(record.get("departsFromProposedLegacyTargets"))
                for record in split_index.values()
            ),
            "crossTensionCollisionAdjudicationRecords": len(collision_index),
            "crossTensionCollisionEligibleOccurrences": sum(
                len(record["occurrences"]) for record in collision_index.values()
            ),
            "crossTensionCollisionExcludedOccurrences": sum(
                str(row.get("exclusionReason", "")).startswith(
                    "cross-tension-collision-governed-to-"
                )
                for row in occurrences
            ),
            "crossTensionCollisionHighConfidenceRecords": sum(
                record.get("confidence") == "high"
                for record in collision_index.values()
            ),
            "crossTensionCollisionMediumConfidenceRecords": sum(
                record.get("confidence") == "medium"
                for record in collision_index.values()
            ),
            "crossTensionCollisionReviewConfidenceRecords": sum(
                record.get("confidence") == "review"
                for record in collision_index.values()
            ),
            "crossTensionCollisionReviewRequiredRecords": sum(
                bool(record.get("reviewRequired"))
                for record in collision_index.values()
            ),
            "crossTensionCollisionDecisionsChangingMechanicalWinner": sum(
                (
                    str(
                        record.get("currentMechanicalWinner", {}).get(
                            "canonicalTensionId", ""
                        )
                    )
                    != str(record["decision"]["chosenCanonicalTensionId"])
                    or str(
                        record.get("currentMechanicalWinner", {}).get(
                            "canonicalPole", ""
                        )
                    )
                    != str(record["decision"]["chosenCanonicalPole"])
                )
                for record in collision_index.values()
            ),
            "sameTensionDuplicateAdjudicationRecords": len(same_tension_index),
            "sameTensionDuplicateEligibleOccurrences": sum(
                len(record["occurrences"]) for record in same_tension_index.values()
            ),
            "sameTensionDuplicateHighConfidenceRecords": sum(
                record.get("confidence") == "high"
                for record in same_tension_index.values()
            ),
            "sameTensionDuplicateReviewRequiredRecords": sum(
                bool(record.get("reviewRequired"))
                for record in same_tension_index.values()
            ),
            "sameTensionSharedProvenanceRecords": sum(
                record["decision"]["mode"] == "same-pole-shared-provenance"
                for record in same_tension_index.values()
            ),
            "sameTensionDualPoleBridgeRecords": sum(
                record["decision"]["mode"] == "genuine-dual-pole-bridge"
                for record in same_tension_index.values()
            ),
            "sameTensionPoleConflictItemCount": sum(
                len({str(row["normalizedPole"]) for row in rows}) > 1
                for rows in by_item.values()
                if len({str(row["canonicalTensionId"]) for row in rows}) == 1
            ),
            "deduplicatedOccurrences": sum(
                str(row.get("exclusionReason", "")).startswith(
                    (
                        "cross-tension-collision-governed-to-",
                        "same-tension-governed-zero-weight-lineage",
                    )
                )
                for row in occurrences
            ),
        },
        "crossTensionCollisionDecisionSummary": [
            {
                "canonicalItemId": item_id,
                "chosenLineageKey": record["decision"]["chosenLineageKey"],
                "chosenCanonicalTensionId": record["decision"][
                    "chosenCanonicalTensionId"
                ],
                "chosenCanonicalPole": record["decision"]["chosenCanonicalPole"],
                "confidence": record["confidence"],
                "reviewRequired": bool(record.get("reviewRequired")),
                "rationale": record["rationale"],
                "reviewReason": record.get("reviewReason"),
            }
            for item_id, record in sorted(
                collision_index.items(), key=lambda row: _natural(row[0])
            )
        ],
        "sameTensionDuplicateDecisionSummary": [
            {
                "canonicalItemId": item_id,
                "canonicalTensionId": record["canonicalTensionId"],
                "decisionMode": record["decision"]["mode"],
                "confidence": record["confidence"],
                "reviewRequired": bool(record.get("reviewRequired")),
                "rationale": record["rationale"],
                "reviewReason": record.get("reviewReason"),
                "occurrenceWeights": record["occurrenceWeights"],
            }
            for item_id, record in sorted(
                same_tension_index.items(), key=lambda row: _natural(row[0])
            )
        ],
        "splitDecisionSummary": split_decision_summary,
        "records": sorted(
            occurrences,
            key=lambda row: (
                _natural(row.get("historicalTensionId", "")),
                _natural(row.get("sourceCandidateId", "")),
                _natural(row.get("historicalItemId", row.get("itemId", ""))),
                row.get("sourcePoleOccurrence", row.get("historicalPole", "")),
            ),
        ),
    }
    return records, allocation, dict(tension_item_ids)


NARRATIVE_UNRESOLVED_ISSUES = {
    "CN-01": "How should durable ownership and distributed execution share authority, resources, and accountability without recreating fragmentation?",
    "CN-02": "Which combinations of technical verification, contextual interpretation, and platform governance remain credible under adversarial uncertainty?",
    "CN-03": "How can identity-aware protection and personalization avoid turning people and communities into surveilled or manipulated objects?",
    "CN-04": "Where should human judgment remain decisive when machine tempo creates both advantage and correlated failure risk?",
    "CN-05": "How can democratic actors adapt to asymmetric competition without spending the legitimacy, credibility, and pluralism that underpin durable influence?",
}


def build_narratives(
    dataset: Mapping[str, Sequence[Mapping[str, Any]]],
    design: Mapping[str, Any],
    themes: Sequence[Mapping[str, Any]],
    theme_item_ids: Mapping[str, set[str]],
    tensions: Sequence[Mapping[str, Any]],
    tension_item_ids: Mapping[str, set[str]],
    historical_by_cluster: Mapping[str, set[str]],
    release_by_content: Mapping[str, set[str]],
    cross_level_adjudication: Mapping[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, set[str]]]:
    item_by_id = {str(row["itemId"]): row for row in dataset["items"]}
    theme_by_id = {str(row["themeId"]): row for row in themes}
    tension_by_id = {str(row["tensionId"]): row for row in tensions}
    governed_rows = cross_level_adjudication.get("narrativeDecisions")
    if not isinstance(governed_rows, list):
        raise ValueError("Governed narrative decisions are missing.")
    governed_by_id = {str(row["narrativeId"]): row for row in governed_rows}
    expected_narrative_ids = {str(row["id"]) for row in design["narratives"]}
    if (
        len(governed_by_id) != len(governed_rows)
        or set(governed_by_id) != expected_narrative_ids
    ):
        raise ValueError("Governed narrative decision coverage is incomplete.")
    narrative_items: dict[str, set[str]] = {}
    records = []
    for seed in sorted(design["narratives"], key=lambda row: _natural(row["id"])):
        narrative_id = str(seed["id"])
        governed = governed_by_id[narrative_id]
        theme_ids = [str(value) for value in governed["integratesThemeIds"]]
        tension_ids = [str(value) for value in governed["integratesTensionIds"]]
        if not set(theme_ids) <= set(theme_by_id) or not set(tension_ids) <= set(
            tension_by_id
        ):
            raise ValueError(
                f"Governed narrative endpoints do not resolve for {narrative_id}."
            )
        family_ids = sorted(
            {
                family_id
                for theme_id in theme_ids
                for family_id in (
                    theme_by_id[theme_id]["primaryFamilyIds"]
                    + theme_by_id[theme_id]["secondaryFamilyIds"]
                )
            },
            key=_natural,
        )
        cluster_ids = sorted(
            {
                cluster_id
                for theme_id in theme_ids
                for cluster_id in (
                    theme_by_id[theme_id]["primaryClusterIds"]
                    + theme_by_id[theme_id]["secondaryClusterIds"]
                )
            },
            key=_natural,
        )
        primary_items = set().union(
            *(theme_item_ids[theme_id] for theme_id in theme_ids)
        )
        tension_items = set().union(
            *(tension_item_ids.get(tension_id, set()) for tension_id in tension_ids)
        )
        all_items = primary_items | tension_items
        narrative_items[narrative_id] = all_items
        historical_items = set().union(
            *(
                historical_by_cluster.get(cluster_id, set())
                for cluster_id in cluster_ids
            )
        )
        limitations = [
            "Narrative support is integrative and derived from canonical themes, families, clusters, and allocated tension evidence.",
            *[str(value) for value in governed["limitations"]],
        ]
        profile = _support_profile(
            all_items,
            historical_item_ids=historical_items,
            direct_item_ids=tension_items,
            item_by_id=item_by_id,
            release_by_content=release_by_content,
            cluster_ids=cluster_ids,
            family_ids=family_ids,
            adjudication_status=str(governed["adjudicationStatus"]),
            limitations=limitations,
        )
        records.append(
            {
                "narrativeId": narrative_id,
                "name": seed["name"],
                "shortVersion": seed["short"],
                "coreClaim": seed["short"],
                "integratesThemeIds": theme_ids,
                "integratesTensionIds": tension_ids,
                "supportingFamilyIds": family_ids,
                "supportingClusterIds": cluster_ids,
                "categoryBreadth": profile["categoryBreadth"],
                "unresolvedIssue": NARRATIVE_UNRESOLVED_ISSUES[narrative_id],
                "boundaryConditions": governed["boundaryConditions"],
                "corpusSupportProfile": profile,
                "limitations": limitations,
                "adjudicationStatus": governed["adjudicationStatus"],
                "adjudicationConfidence": governed["confidence"],
                "adjudicationDecision": governed["decision"],
                "adjudicationRationale": governed["rationale"],
                "historicalNarrativeIds": governed["historicalNarrativeIds"],
            }
        )
    return records, narrative_items


CATEGORY_INTEGRATIVE_FINDINGS = {
    "Technologies / Tools / Platforms": "The corrected corpus treats the technical environment as an interdependent stack: sensing, analysis, channels, automation, identity systems, contested access, and operational preparation shape one another rather than operating as isolated tools.",
    "Organizations / Actors / Communities": "Information capability is distributed across public institutions, adversaries, alliances, knowledge networks, communities, platforms, and rule-setting bodies; effectiveness depends on governed interfaces among actors with different authority and legitimacy.",
    "Challenges / Risks / Barriers": "The principal barriers combine institutional fragmentation, authority and resource constraints, weak evidence, technical exposure, human vulnerability, and asymmetric competition; no single technical or organizational remedy resolves the system.",
    "Key Concepts / Frameworks / Theories": "The field's conceptual base links human-centered and epistemic security, narrative and identity models, persistent competition, behavioral mechanisms, systems thinking, learning, governance, and cross-domain orchestration.",
    "Key Events / Historical Examples": "Historical cases contribute through strategic shocks, technical incidents, institutional adaptation, symbolic contests, societal resilience, and comparative analogy; they provide context and mechanism clues rather than universal templates.",
    "Future Trends / Predictions": "Future-oriented evidence clusters around contested reality, trust infrastructure, human-machine operations, institutional adaptation, gray-zone competition, nonstate technological power, and systemic social uncertainty.",
    "Opportunities / Recommended Actions": "Recommended action forms a portfolio spanning institutional design, capability transition, coordination, campaign design, societal resilience, assessment, governance, and mission continuity; sequencing and legitimacy constraints determine whether the portfolio coheres.",
}

CATEGORY_OPEN_QUESTIONS = {
    "Technologies / Tools / Platforms": "Which technical architectures preserve human judgment, provenance, and continuity as automation and adversarial pressure increase?",
    "Organizations / Actors / Communities": "What governance arrangements align distributed actors without erasing their distinct mandates, access, and legitimacy?",
    "Challenges / Risks / Barriers": "Which barriers are causal bottlenecks, which are correlated symptoms, and how does that vary across operational contexts?",
    "Key Concepts / Frameworks / Theories": "Which concepts produce reliable practitioner decisions across cultures and missions, and which remain useful only as contextual lenses?",
    "Key Events / Historical Examples": "Under what conditions do historical analogies transfer, and where do technological and institutional differences break the comparison?",
    "Future Trends / Predictions": "Which proposed trajectories have observable leading indicators, and how should uncertainty be represented without false precision?",
    "Opportunities / Recommended Actions": "Which combinations of actions are feasible, legitimate, and mutually reinforcing under real authority, workforce, and resource constraints?",
}


def build_category_findings(
    dataset: Mapping[str, Sequence[Mapping[str, Any]]],
    families: Sequence[Mapping[str, Any]],
    family_item_ids: Mapping[str, set[str]],
    historical_by_cluster: Mapping[str, set[str]],
    release_by_content: Mapping[str, set[str]],
) -> tuple[list[dict[str, Any]], dict[str, set[str]], list[dict[str, Any]]]:
    item_by_id = {str(row["itemId"]): row for row in dataset["items"]}
    category_by_name = {str(row["name"]): row for row in dataset["categories"]}
    historical_findings = list(dataset["category_findings"])
    findings: list[dict[str, Any]] = []
    finding_items: dict[str, set[str]] = {}

    for family in sorted(families, key=lambda row: _natural(row["familyId"])):
        family_id = str(family["familyId"])
        category_name = str(family["category"])
        category_id = str(category_by_name[category_name]["categoryId"])
        item_ids = set(family_item_ids[family_id])
        historical_items = set().union(
            *(
                historical_by_cluster.get(cluster_id, set())
                for cluster_id in family["memberClusterIds"]
            )
        )
        finding_id = f"CF-{family_id}"
        profile = _support_profile(
            item_ids,
            historical_item_ids=historical_items,
            direct_item_ids=item_ids,
            item_by_id=item_by_id,
            release_by_content=release_by_content,
            cluster_ids=family["memberClusterIds"],
            family_ids=[family_id],
            adjudication_status="analyst-reviewed-draft",
            limitations=family["limitations"],
        )
        finding_items[finding_id] = item_ids
        historical_ids = sorted(
            {
                str(row["findingId"])
                for row in historical_findings
                if str(row.get("categoryId")) == category_id
                and set(map(str, row.get("supportingClusterIds", ())))
                & set(family["memberClusterIds"])
            },
            key=_natural,
        )
        findings.append(
            {
                "findingId": finding_id,
                "categoryId": category_id,
                "categoryName": category_name,
                "findingType": "family-finding",
                "title": family["name"],
                "finding": (
                    f"Within {category_name}, corrected evidence recurs around "
                    f"{str(family['definition']).rstrip('.').lower()}."
                ),
                "supportingFamilyIds": [family_id],
                "supportingClusterIds": family["memberClusterIds"],
                "supportingContentUnitCount": profile["uniqueContentUnitSupportCount"],
                "corpusSupportProfile": profile,
                "openQuestions": [],
                "historicalFindingIds": historical_ids,
                "limitations": family["limitations"],
                "adjudicationStatus": "analyst-reviewed-draft",
            }
        )

    families_by_category: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for family in families:
        families_by_category[str(family["category"])].append(family)
    for index, category_name in enumerate(CATEGORY_INTEGRATIVE_FINDINGS, start=1):
        category = category_by_name[category_name]
        category_id = str(category["categoryId"])
        category_families = sorted(
            families_by_category[category_name],
            key=lambda row: _natural(row["familyId"]),
        )
        family_ids = [str(row["familyId"]) for row in category_families]
        cluster_ids = sorted(
            {cid for row in category_families for cid in row["memberClusterIds"]},
            key=_natural,
        )
        item_ids = set().union(*(family_item_ids[fid] for fid in family_ids))
        historical_items = set().union(
            *(
                historical_by_cluster.get(cluster_id, set())
                for cluster_id in cluster_ids
            )
        )
        for finding_type, suffix, title, finding_text in (
            (
                "integrative-category-finding",
                "I",
                f"Integrated {category_name}",
                CATEGORY_INTEGRATIVE_FINDINGS[category_name],
            ),
            (
                "open-question",
                "Q",
                f"Open question: {category_name}",
                CATEGORY_OPEN_QUESTIONS[category_name],
            ),
        ):
            finding_id = f"CF-CAT-{index:02d}-{suffix}"
            direct_ids = (
                item_ids if finding_type == "integrative-category-finding" else set()
            )
            profile = _support_profile(
                item_ids,
                historical_item_ids=historical_items,
                direct_item_ids=direct_ids,
                item_by_id=item_by_id,
                release_by_content=release_by_content,
                cluster_ids=cluster_ids,
                family_ids=family_ids,
                adjudication_status="analyst-reviewed-draft",
                limitations=[
                    "Category-level synthesis spans heterogeneous mechanisms and should not be read as a single causal claim."
                ],
            )
            finding_items[finding_id] = item_ids
            findings.append(
                {
                    "findingId": finding_id,
                    "categoryId": category_id,
                    "categoryName": category_name,
                    "findingType": finding_type,
                    "title": title,
                    "finding": finding_text,
                    "supportingFamilyIds": family_ids,
                    "supportingClusterIds": cluster_ids,
                    "supportingContentUnitCount": profile[
                        "uniqueContentUnitSupportCount"
                    ],
                    "corpusSupportProfile": profile,
                    "openQuestions": (
                        [finding_text]
                        if finding_type == "open-question"
                        else [CATEGORY_OPEN_QUESTIONS[category_name]]
                    ),
                    "historicalFindingIds": sorted(
                        {
                            str(row["findingId"])
                            for row in historical_findings
                            if str(row.get("categoryId")) == category_id
                        },
                        key=_natural,
                    ),
                    "limitations": profile["limitations"],
                    "adjudicationStatus": "analyst-reviewed-draft",
                }
            )
    sorted_findings = sorted(findings, key=lambda row: _natural(row["findingId"]))
    finding_ids = {str(row["findingId"]) for row in sorted_findings}
    family_by_cluster = {
        str(cluster_id): str(family["familyId"])
        for family in families
        for cluster_id in family["memberClusterIds"]
    }
    category_synthesis_ids = {
        str(row["categoryId"]): str(row["findingId"])
        for row in sorted_findings
        if row["findingType"] == "integrative-category-finding"
    }
    category_question_ids = {
        str(row["categoryId"]): str(row["findingId"])
        for row in sorted_findings
        if row["findingType"] == "open-question"
    }
    historical_lineage = []
    for historical in sorted(
        historical_findings, key=lambda row: _natural(row["findingId"])
    ):
        historical_id = str(historical["findingId"])
        category_id = str(historical["categoryId"])
        target_families = sorted(
            {
                family_by_cluster[str(cluster_id)]
                for cluster_id in historical.get("supportingClusterIds", [])
                if str(cluster_id) in family_by_cluster
            },
            key=_natural,
        )
        canonical_finding_ids = sorted(
            {
                *(f"CF-{family_id}" for family_id in target_families),
                category_synthesis_ids[category_id],
                category_question_ids[category_id],
            },
            key=_natural,
        )
        if not target_families or not set(canonical_finding_ids) <= finding_ids:
            raise ValueError(
                f"Historical category-finding lineage is incomplete for {historical_id}."
            )
        historical_lineage.append(
            {
                "historicalFindingId": historical_id,
                "categoryId": category_id,
                "supportingHistoricalClusterIds": sorted(
                    {
                        str(value)
                        for value in historical.get("supportingClusterIds", [])
                    },
                    key=_natural,
                ),
                "canonicalFamilyIds": target_families,
                "canonicalFindingIds": canonical_finding_ids,
                "lineageSemantics": (
                    "cluster-descendant family finding plus category synthesis and "
                    "open-question provenance"
                ),
                "contributesAdditionalAnalyticalWeight": False,
            }
        )
    if len(historical_lineage) != 42:
        raise ValueError("All 42 historical category findings require lineage.")
    return sorted_findings, finding_items, historical_lineage


def _validated_scenario_decision_index(
    dataset: Mapping[str, Sequence[Mapping[str, Any]]],
    design: Mapping[str, Any],
    themes: Sequence[Mapping[str, Any]],
    adjudication: Mapping[str, Any],
) -> dict[str, dict[str, Any]]:
    """Validate complete, evidence-dispositioned scenario decisions."""

    expected_scenarios = {str(row["id"]) for row in design["scenarios"]}
    historical_scenarios = {str(row["scenarioId"]): row for row in dataset["scenarios"]}
    theme_ids = {str(row["themeId"]) for row in themes}
    tension_ids = {str(row["id"]) for row in design["tensions"]}
    family_ids = {
        str(family_id)
        for theme in themes
        for family_id in theme["primaryFamilyIds"] + theme["secondaryFamilyIds"]
    }
    records = adjudication.get("scenarioDecisions")
    if not isinstance(records, list):
        raise ValueError("Scenario adjudication must contain scenarioDecisions.")
    index: dict[str, dict[str, Any]] = {}
    seen_historical: set[str] = set()

    pathway_ordinals: dict[str, set[int]] = defaultdict(set)
    indicator_ordinals: dict[str, set[int]] = defaultdict(set)
    action_ordinals: dict[str, set[int]] = defaultdict(set)
    for row in dataset["scenario_pathways"]:
        pathway_ordinals[str(row["scenarioId"])].add(int(row["stepNumber"]))
    for row in dataset["scenario_indicators"]:
        indicator_ordinals[str(row["scenarioId"])].add(int(row["ordinal"]))
    for row in dataset["scenario_actions"]:
        action_ordinals[str(row["scenarioId"])].add(int(row["ordinal"]))

    for raw_record in records:
        record = dict(raw_record)
        scenario_id = str(record.get("scenarioId") or "")
        if not scenario_id or scenario_id in index:
            raise ValueError(
                f"Duplicate or missing scenario decision: {scenario_id!r}."
            )
        mapping = record.get("historicalMapping", {})
        historical_id = str(mapping.get("historicalScenarioId") or "")
        if (
            historical_id not in historical_scenarios
            or historical_id in seen_historical
        ):
            raise ValueError(
                f"Historical scenario mapping is invalid for {scenario_id}."
            )
        seen_historical.add(historical_id)
        source = historical_scenarios[historical_id]
        if str(mapping.get("historicalScenarioName") or "") != str(source["name"]):
            raise ValueError(f"Historical scenario name is stale for {scenario_id}.")

        inclusions = record.get("finalInclusions", {})
        governed_themes = {str(value) for value in inclusions.get("themeIds", [])}
        governed_tensions = {str(value) for value in inclusions.get("tensionIds", [])}
        governed_families = {str(value) for value in inclusions.get("familyIds", [])}
        if (
            not governed_themes
            or not governed_themes <= theme_ids
            or not governed_tensions
            or not governed_tensions <= tension_ids
            or not governed_families
            or not governed_families <= family_ids
            or not str(inclusions.get("rationale") or "").strip()
        ):
            raise ValueError(f"Scenario inclusions are invalid for {scenario_id}.")

        selections = record.get("historicalEvidenceSelections", {})
        for collection_name, actual_ordinals in (
            ("pathways", pathway_ordinals[historical_id]),
            ("indicators", indicator_ordinals[historical_id]),
            ("actions", action_ordinals[historical_id]),
        ):
            decision = selections.get(collection_name, {})
            included = {int(value) for value in decision.get("includedOrdinals", [])}
            excluded_rows = decision.get("excluded", [])
            excluded = {int(row["ordinal"]) for row in excluded_rows}
            if (
                included & excluded
                or included | excluded != actual_ordinals
                or len(excluded) != len(excluded_rows)
                or any(
                    not str(row.get("reason") or "").strip() for row in excluded_rows
                )
                or not str(decision.get("selectionRationale") or "").strip()
            ):
                raise ValueError(
                    f"Scenario {collection_name} disposition is incomplete for {scenario_id}."
                )

        def valid_source_ordinals(
            field: str, maximum: int, current: Mapping[str, Any] = record
        ) -> bool:
            values = [int(value) for value in current.get(field, [])]
            return (
                bool(values)
                and len(values) == len(set(values))
                and all(1 <= value <= maximum for value in values)
            )

        if (
            not valid_source_ordinals(
                "triggerConditionOrdinals", len(source.get("drivingForces", []))
            )
            or not valid_source_ordinals(
                "branchPointOrdinals", len(source.get("alternativeOutcomes", []))
            )
            or not valid_source_ordinals(
                "counterSignpostAlternativeOutcomeOrdinals",
                len(source.get("alternativeOutcomes", [])),
            )
        ):
            raise ValueError(f"Scenario source ordinals are invalid for {scenario_id}.")
        mitigating = {
            int(value) for value in record.get("mitigatingActionOrdinals", [])
        }
        included_actions = {
            int(value) for value in selections["actions"].get("includedOrdinals", [])
        }
        if not mitigating or not mitigating <= included_actions:
            raise ValueError(
                f"Scenario mitigating actions are invalid for {scenario_id}."
            )

        dynamics = record.get("tensionPoleDynamics", [])
        if (
            {str(row.get("tensionId")) for row in dynamics} != governed_tensions
            or len(dynamics) != len(governed_tensions)
            or any(
                not all(
                    str(row.get(field) or "").strip()
                    for field in ("direction", "dynamic", "rationale")
                )
                for row in dynamics
            )
        ):
            raise ValueError(
                f"Scenario tension dynamics are incomplete for {scenario_id}."
            )
        relations = record.get("relationships", [])
        for relation in relations:
            if (
                str(relation.get("targetScenarioId")) not in expected_scenarios
                or str(relation.get("targetScenarioId")) == scenario_id
                or relation.get("semanticRole")
                not in {
                    "scenario-amplifies",
                    "scenario-mitigates",
                    "contextual-connection",
                }
                or relation.get("causalClaim") is not False
                or not str(relation.get("rationale") or "").strip()
            ):
                raise ValueError(f"Scenario relationship is invalid for {scenario_id}.")
        if record.get("confidence") not in {"high", "medium", "review"}:
            raise ValueError(f"Scenario confidence is invalid for {scenario_id}.")
        if bool(record.get("reviewRequired")) and not record.get("reviewFlags"):
            raise ValueError(f"Scenario review flags are missing for {scenario_id}.")
        final_scenario = record.get("finalScenario", {})
        if not all(
            str(final_scenario.get(field) or "").strip()
            for field in ("title", "scenarioType", "core")
        ) or not isinstance(record.get("limitations"), list):
            raise ValueError(f"Scenario definition is incomplete for {scenario_id}.")
        index[scenario_id] = record

    if set(index) != expected_scenarios or seen_historical != set(historical_scenarios):
        raise ValueError("Scenario adjudication coverage is incomplete.")
    portfolio = adjudication.get("portfolioReview", {})
    if (
        portfolio.get("scenarioCount") != len(index)
        or portfolio.get("historicalScenarioCount") != len(historical_scenarios)
        or set(portfolio.get("reviewRequiredScenarioIds", []))
        != {
            scenario_id
            for scenario_id, row in index.items()
            if row.get("reviewRequired")
        }
    ):
        raise ValueError("Scenario portfolio review summary is stale.")
    return index


def build_scenarios(
    dataset: Mapping[str, Sequence[Mapping[str, Any]]],
    design: Mapping[str, Any],
    themes: Sequence[Mapping[str, Any]],
    theme_item_ids: Mapping[str, set[str]],
    tension_item_ids: Mapping[str, set[str]],
    cluster_to_family: Mapping[str, str],
    retained_by_cluster: Mapping[str, set[str]],
    release_by_content: Mapping[str, set[str]],
    scenario_adjudication: Mapping[str, Any],
) -> tuple[list[dict[str, Any]], dict[str, set[str]]]:
    item_by_id = {str(row["itemId"]): row for row in dataset["items"]}
    source_scenarios = {str(row["scenarioId"]): row for row in dataset["scenarios"]}
    decision_by_id = _validated_scenario_decision_index(
        dataset, design, themes, scenario_adjudication
    )
    pathways_by_source: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    indicators_by_source: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    actions_by_source: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for row in dataset["scenario_pathways"]:
        pathways_by_source[str(row["scenarioId"])].append(row)
    for row in dataset["scenario_indicators"]:
        indicators_by_source[str(row["scenarioId"])].append(row)
    for row in dataset["scenario_actions"]:
        actions_by_source[str(row["scenarioId"])].append(row)

    records = []
    scenario_items: dict[str, set[str]] = {}
    for seed in sorted(design["scenarios"], key=lambda row: _natural(row["id"])):
        scenario_id = str(seed["id"])
        governed = decision_by_id[scenario_id]
        source_id = str(governed["historicalMapping"]["historicalScenarioId"])
        source = source_scenarios[source_id]
        inclusions = governed["finalInclusions"]
        theme_ids = [str(value) for value in inclusions["themeIds"]]
        tension_ids = [str(value) for value in inclusions["tensionIds"]]
        family_ids = sorted(
            {str(value) for value in inclusions["familyIds"]}, key=_natural
        )
        cluster_ids = sorted(
            {
                cid
                for cid, family_id in cluster_to_family.items()
                if family_id in family_ids
            },
            key=_natural,
        )
        primary_items = set().union(
            *(retained_by_cluster.get(cluster_id, set()) for cluster_id in cluster_ids)
        )
        direct_items = set().union(
            *(tension_item_ids.get(tension_id, set()) for tension_id in tension_ids)
        )
        item_ids = primary_items | direct_items
        scenario_items[scenario_id] = item_ids
        profile = _support_profile(
            item_ids,
            historical_item_ids=None,
            direct_item_ids=direct_items,
            item_by_id=item_by_id,
            release_by_content=release_by_content,
            cluster_ids=cluster_ids,
            family_ids=family_ids,
            adjudication_status="analyst-reviewed-draft",
            limitations=[
                "Scenario support establishes plausibility within the corpus, not probability or predictive validity.",
                *[str(value) for value in governed["limitations"]],
            ],
        )
        selections = governed["historicalEvidenceSelections"]
        included_pathways = {
            int(value) for value in selections["pathways"]["includedOrdinals"]
        }
        included_indicators = {
            int(value) for value in selections["indicators"]["includedOrdinals"]
        }
        included_actions = {
            int(value) for value in selections["actions"]["includedOrdinals"]
        }
        pathways = [
            row
            for row in sorted(
                pathways_by_source[source_id],
                key=lambda row: int(row.get("stepNumber") or 0),
            )
            if int(row["stepNumber"]) in included_pathways
        ]
        indicators = [
            row
            for row in sorted(
                indicators_by_source[source_id],
                key=lambda row: int(row.get("ordinal") or 0),
            )
            if int(row["ordinal"]) in included_indicators
        ]
        actions = [
            row
            for row in sorted(
                actions_by_source[source_id],
                key=lambda row: int(row.get("ordinal") or 0),
            )
            if int(row["ordinal"]) in included_actions
        ]

        def select_one_based(
            values: Sequence[Any], ordinals: Sequence[Any]
        ) -> list[Any]:
            return [values[int(ordinal) - 1] for ordinal in ordinals]

        trigger_conditions = select_one_based(
            list(source.get("drivingForces", [])),
            governed["triggerConditionOrdinals"],
        )
        branch_points = select_one_based(
            list(source.get("alternativeOutcomes", [])),
            governed["branchPointOrdinals"],
        )
        counter_signposts = [
            f"Evidence develops toward an alternative outcome: {value}"
            for value in select_one_based(
                list(source.get("alternativeOutcomes", [])),
                governed["counterSignpostAlternativeOutcomeOrdinals"],
            )
        ]
        mitigating_ordinals = {
            int(value) for value in governed["mitigatingActionOrdinals"]
        }
        mitigating_conditions = [
            row["policyOrPracticeAction"]
            for row in actions
            if int(row["ordinal"]) in mitigating_ordinals
        ]
        final_scenario = governed["finalScenario"]
        records.append(
            {
                "scenarioId": scenario_id,
                "title": final_scenario["title"],
                "scenarioType": final_scenario["scenarioType"],
                "description": final_scenario["core"],
                "relevantThemeIds": theme_ids,
                "relevantTensionIds": tension_ids,
                "relevantFutureTrendFamilyIds": [
                    fid for fid in family_ids if fid.startswith("FTP-")
                ],
                "relevantKeyConceptFamilyIds": [
                    fid for fid in family_ids if fid.startswith("KCF-")
                ],
                "triggerConditions": trigger_conditions,
                "branchPoints": branch_points,
                "plausiblePathways": [row["pathwayStep"] for row in pathways],
                "indicators": [row["indicator"] for row in indicators],
                "counterSignposts": counter_signposts,
                "mitigatingConditions": mitigating_conditions,
                "tensionPoleDynamics": governed["tensionPoleDynamics"],
                "relationshipsToOtherScenarios": governed["relationships"],
                "strategicImplications": list(source.get("strategicImplications", ())),
                "responseOptions": [
                    action["policyOrPracticeAction"] for action in actions
                ],
                "researchQuestions": list(source.get("researchQuestions", ())),
                "uncertaintyStatement": (
                    "This scenario is a plausibility exercise, not a prediction. Its "
                    "pathway, indicators, and implications are contingent on interacting "
                    "technical, institutional, social, and adversarial conditions."
                ),
                "corpusSupportProfile": profile,
                "adjudicationStatus": "analyst-reviewed-draft",
                "adjudicationConfidence": governed["confidence"],
                "reviewRequired": bool(governed["reviewRequired"]),
                "reviewFlags": governed["reviewFlags"],
                "historicalScenarioId": source_id,
                "evidenceSelection": {
                    "pathwayOrdinals": sorted(included_pathways),
                    "indicatorOrdinals": sorted(included_indicators),
                    "actionOrdinals": sorted(included_actions),
                    "selectionRationales": {
                        "pathways": selections["pathways"]["selectionRationale"],
                        "indicators": selections["indicators"]["selectionRationale"],
                        "actions": selections["actions"]["selectionRationale"],
                        "triggers": governed["triggerRationale"],
                        "branches": governed["branchRationale"],
                        "counterSignposts": governed["counterSignpostRationale"],
                        "mitigation": governed["mitigationRationale"],
                    },
                },
                "limitations": profile["limitations"],
            }
        )
    return records, scenario_items


def _jaccard(left: Iterable[str], right: Iterable[str]) -> float:
    left_set = set(left)
    right_set = set(right)
    union = left_set | right_set
    return len(left_set & right_set) / len(union) if union else 0.0


def build_redundancy_audit(
    families: Sequence[Mapping[str, Any]],
    family_items: Mapping[str, set[str]],
    themes: Sequence[Mapping[str, Any]],
    theme_items: Mapping[str, set[str]],
    tensions: Sequence[Mapping[str, Any]],
    tension_items: Mapping[str, set[str]],
    narratives: Sequence[Mapping[str, Any]],
    narrative_items: Mapping[str, set[str]],
    findings: Sequence[Mapping[str, Any]],
    finding_items: Mapping[str, set[str]],
    scenarios: Sequence[Mapping[str, Any]],
    scenario_items: Mapping[str, set[str]],
) -> dict[str, Any]:
    specifications = (
        (
            "family",
            families,
            "familyId",
            lambda row: row["memberClusterIds"],
            family_items,
        ),
        (
            "theme",
            themes,
            "themeId",
            lambda row: row["primaryFamilyIds"] + row["secondaryFamilyIds"],
            theme_items,
        ),
        (
            "tension",
            tensions,
            "tensionId",
            lambda row: row["supportingFamilyIds"],
            tension_items,
        ),
        (
            "narrative",
            narratives,
            "narrativeId",
            lambda row: row["integratesThemeIds"] + row["integratesTensionIds"],
            narrative_items,
        ),
        (
            "finding",
            findings,
            "findingId",
            lambda row: row["supportingFamilyIds"],
            finding_items,
        ),
        (
            "scenario",
            scenarios,
            "scenarioId",
            lambda row: row["relevantThemeIds"] + row["relevantTensionIds"],
            scenario_items,
        ),
    )
    comparisons = []
    flagged = []
    for entity_type, rows, id_field, structure, item_map in specifications:
        ordered = sorted(rows, key=lambda row: _natural(row[id_field]))
        for index, left in enumerate(ordered):
            for right in ordered[index + 1 :]:
                left_id = str(left[id_field])
                right_id = str(right[id_field])
                structural = _jaccard(structure(left), structure(right))
                evidence = _jaccard(
                    item_map.get(left_id, set()), item_map.get(right_id, set())
                )
                potential = (
                    evidence >= 0.75
                    if entity_type == "family"
                    else structural >= 0.75 and evidence >= 0.75
                )
                complementary_finding_pair = (
                    potential
                    and entity_type == "finding"
                    and str(left.get("categoryId")) == str(right.get("categoryId"))
                    and {
                        str(left.get("findingType")),
                        str(right.get("findingType")),
                    }
                    == {"integrative-category-finding", "open-question"}
                )
                record = {
                    "entityType": entity_type,
                    "leftId": left_id,
                    "rightId": right_id,
                    "structuralJaccard": _round(structural),
                    "evidenceJaccard": _round(evidence),
                    "potentialRedundancy": potential,
                    "reviewDisposition": (
                        "retain-distinct-complementary-finding-types"
                        if complementary_finding_pair
                        else ("requires-human-review" if potential else "not-flagged")
                    ),
                    "unresolvedPotentialRedundancy": bool(
                        potential and not complementary_finding_pair
                    ),
                    "reviewRationale": (
                        "The category synthesis and open question intentionally share "
                        "the same evidence envelope but perform different analytical "
                        "functions: one states the integrated finding and the other "
                        "marks the remaining uncertainty."
                        if complementary_finding_pair
                        else None
                    ),
                }
                comparisons.append(record)
                if record["potentialRedundancy"]:
                    flagged.append(record)
    return {
        "method": (
            "Pairwise within-level audit using transparent Jaccard overlap of governed "
            "structural relationships and supporting corrected items. No pair is merged "
            "on this diagnostic alone."
        ),
        "comparisonCount": len(comparisons),
        "potentialRedundancyCount": len(flagged),
        "resolvedDistinctCount": sum(
            not row["unresolvedPotentialRedundancy"] for row in flagged
        ),
        "unresolvedPotentialRedundancyCount": sum(
            row["unresolvedPotentialRedundancy"] for row in flagged
        ),
        "flaggedPairs": flagged,
        "comparisons": comparisons,
    }


def _support_profile_catalog(
    clusters: Sequence[Mapping[str, Any]],
    families: Sequence[Mapping[str, Any]],
    themes: Sequence[Mapping[str, Any]],
    tensions: Sequence[Mapping[str, Any]],
    narratives: Sequence[Mapping[str, Any]],
    findings: Sequence[Mapping[str, Any]],
    scenarios: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    specifications = (
        ("cluster", clusters, "clusterId"),
        ("family", families, "familyId"),
        ("theme", themes, "themeId"),
        ("tension", tensions, "tensionId"),
        ("narrative", narratives, "narrativeId"),
        ("finding", findings, "findingId"),
        ("scenario", scenarios, "scenarioId"),
    )
    records = []
    for entity_type, rows, id_field in specifications:
        for row in rows:
            records.append(
                {
                    "entityType": entity_type,
                    "entityId": row[id_field],
                    "corpusSupportProfile": row["corpusSupportProfile"],
                }
            )
    return {
        "schemaVersion": SCHEMA_VERSION,
        "interpretation": SUPPORT_INTERPRETATION,
        "compositeScoreProhibited": True,
        "records": sorted(
            records, key=lambda row: (row["entityType"], _natural(row["entityId"]))
        ),
    }


def build_review_queue(
    clusters: Sequence[Mapping[str, Any]],
    family_adjudication: Mapping[str, Any],
    theme_adjudication: Mapping[str, Any],
    tensions: Sequence[Mapping[str, Any]],
    tension_allocation: Mapping[str, Any],
    narratives: Sequence[Mapping[str, Any]],
    scenarios: Sequence[Mapping[str, Any]],
    redundancy_audit: Mapping[str, Any],
) -> list[dict[str, Any]]:
    queue = []
    for cluster in clusters:
        if cluster["sensitivityStatus"] in {"lost-all-support", "review-required"}:
            queue.append(
                {
                    "reviewId": f"RQ-CLUSTER-{cluster['clusterId']}",
                    "entityType": "cluster",
                    "entityId": cluster["clusterId"],
                    "priority": "high"
                    if cluster["sensitivityStatus"] == "lost-all-support"
                    else "medium",
                    "reason": "; ".join(cluster["investigationTriggers"]),
                    "status": "requires-human-confirmation",
                }
            )
    for decision in family_adjudication["mappingDecisions"]:
        if (
            decision["mappingConfidence"] == "medium"
            or decision["adjudicatedConfidence"] == "moderate"
        ):
            queue.append(
                {
                    "reviewId": f"RQ-FAMILY-{decision['clusterId']}",
                    "entityType": "cluster-family-mapping",
                    "entityId": decision["clusterId"],
                    "priority": "medium",
                    "reason": (
                        "The mapping received deep item-level evidence review and was "
                        f"confirmed with {decision['adjudicatedConfidence']} final "
                        "confidence; human approval remains required."
                    ),
                    "status": "analyst-resolved-awaiting-human-approval",
                }
            )
    for decision in theme_adjudication["decisions"]:
        if decision["confidence"] == "medium":
            queue.append(
                {
                    "reviewId": f"RQ-THEME-{decision['themeId']}",
                    "entityType": "theme",
                    "entityId": decision["themeId"],
                    "priority": "medium",
                    "reason": (
                        "The theme remains distinct after evidence and boundary review, "
                        "but its specialized or revised scope requires human approval."
                    ),
                    "status": "analyst-resolved-awaiting-human-approval",
                }
            )
    for narrative in narratives:
        if narrative["adjudicationConfidence"] == "medium":
            queue.append(
                {
                    "reviewId": f"RQ-NARRATIVE-{narrative['narrativeId']}",
                    "entityType": "narrative",
                    "entityId": narrative["narrativeId"],
                    "priority": "medium",
                    "reason": (
                        "The integrative storyline and historical consolidation are "
                        "analyst-resolved but require human approval."
                    ),
                    "status": "analyst-resolved-awaiting-human-approval",
                }
            )
    for scenario in scenarios:
        if scenario["reviewRequired"]:
            queue.append(
                {
                    "reviewId": f"RQ-SCENARIO-{scenario['scenarioId']}",
                    "entityType": "scenario",
                    "entityId": scenario["scenarioId"],
                    "priority": "medium",
                    "reason": "; ".join(scenario["reviewFlags"]),
                    "status": "requires-human-confirmation",
                }
            )
    allocation_counts = tension_allocation["counts"]
    if (
        allocation_counts["counterpartMediumConfidenceRecords"]
        or allocation_counts["counterpartReviewRecords"]
    ):
        queue.append(
            {
                "reviewId": "RQ-TENSION-ALIAS-COUNTERPARTS",
                "entityType": "tension-evidence-lineage",
                "entityId": "governed-alias-counterpart-adjudication",
                "priority": "medium",
                "reason": (
                    "All alias counterparts were transcript- and lineage-reviewed; "
                    f"{allocation_counts['counterpartMediumConfidenceRecords']} medium "
                    "and "
                    f"{allocation_counts['counterpartReviewRecords']} compound-review "
                    "decision(s) remain for human approval."
                ),
                "status": "analyst-resolved-awaiting-human-approval",
            }
        )
    if allocation_counts["splitReviewRecords"]:
        queue.append(
            {
                "reviewId": "RQ-TENSION-SPLIT-ALLOCATIONS",
                "entityType": "tension-evidence-allocation",
                "entityId": "governed-split-tension-adjudication",
                "priority": "medium",
                "reason": (
                    "Every split-tension pole occurrence was explicitly reviewed; "
                    f"{allocation_counts['splitReviewRecords']} bridge/review decision(s) "
                    "remain for human approval."
                ),
                "status": "analyst-resolved-awaiting-human-approval",
            }
        )
    if allocation_counts["crossTensionCollisionReviewRequiredRecords"]:
        queue.append(
            {
                "reviewId": "RQ-TENSION-CROSS-COLLISIONS",
                "entityType": "tension-evidence-allocation",
                "entityId": "governed-cross-tension-collision-adjudication",
                "priority": "medium",
                "reason": (
                    "Every cross-tension item collision received a semantic, "
                    "single-use allocation; "
                    f"{allocation_counts['crossTensionCollisionReviewRequiredRecords']} "
                    "decision(s) remain explicit human-review checkpoints."
                ),
                "status": "analyst-resolved-awaiting-human-approval",
            }
        )
    if allocation_counts["sameTensionDuplicateReviewRequiredRecords"]:
        queue.append(
            {
                "reviewId": "RQ-TENSION-SAME-TENSION-POLE-CONFLICTS",
                "entityType": "tension-evidence-allocation",
                "entityId": "same-tension-pole-conflicts",
                "priority": "medium",
                "reason": (
                    f"{allocation_counts['sameTensionDuplicateReviewRequiredRecords']} "
                    "governed repeated-lineage decision(s) require human confirmation."
                ),
                "status": "requires-human-confirmation",
            }
        )
    for tension in tensions:
        balance = tension["evidenceBalanceAcrossPoles"]
        imbalance = (
            not balance["bothPolesDirectlySupported"]
            or min(balance["poleAShare"], balance["poleBShare"]) < 0.2
        )
        if tension["reviewRequired"] or imbalance:
            reasons = list(tension["reviewFlags"])
            if imbalance:
                reasons.append("missing-or-materially-imbalanced-direct-pole-evidence")
            queue.append(
                {
                    "reviewId": f"RQ-TENSION-{tension['tensionId']}",
                    "entityType": "tension",
                    "entityId": tension["tensionId"],
                    "priority": "high" if imbalance else "medium",
                    "reason": "; ".join(reasons),
                    "status": (
                        "requires-human-confirmation"
                        if imbalance
                        else "analyst-resolved-awaiting-human-approval"
                    ),
                }
            )
    for record in redundancy_audit["flaggedPairs"]:
        if not record["unresolvedPotentialRedundancy"]:
            continue
        queue.append(
            {
                "reviewId": f"RQ-REDUNDANCY-{record['leftId']}-{record['rightId']}",
                "entityType": record["entityType"],
                "entityId": f"{record['leftId']}|{record['rightId']}",
                "priority": "medium",
                "reason": "High structural and evidence overlap in the redundancy diagnostic.",
                "status": "requires-human-confirmation",
            }
        )
    return sorted(queue, key=lambda row: _natural(row["reviewId"]))


def _quality_gates(
    selection: Mapping[str, Any],
    clusters: Sequence[Mapping[str, Any]],
    families: Sequence[Mapping[str, Any]],
    family_adjudication: Mapping[str, Any],
    themes: Sequence[Mapping[str, Any]],
    theme_adjudication: Mapping[str, Any],
    tensions: Sequence[Mapping[str, Any]],
    tension_allocation: Mapping[str, Any],
    tension_item_ids: Mapping[str, set[str]],
    narratives: Sequence[Mapping[str, Any]],
    findings: Sequence[Mapping[str, Any]],
    historical_finding_lineage: Sequence[Mapping[str, Any]],
    scenarios: Sequence[Mapping[str, Any]],
    relationship_semantics: Sequence[Mapping[str, Any]],
    redundancy_audit: Mapping[str, Any],
) -> list[dict[str, Any]]:
    counts = selection["counts"]
    selected_sources = set(selection["selectedCanonicalContentUnitIds"])
    alias_contributors = [
        row
        for row in selection["historicalIdentitySelection"]
        if row["analyticalStatus"] == "confirmed-alias-excluded"
        and row["contributesAnalyticalWeight"]
    ]
    excluded_identity_ids = {
        str(row["sourceIdentityId"])
        for row in selection["historicalIdentitySelection"]
        if not row["contributesAnalyticalWeight"]
    }
    selected_item_records = [
        row
        for row in selection["historicalItemSelection"]
        if row["contributesAnalyticalWeight"]
    ]
    transcript_audit = selection["transcriptManifestAudit"]
    reconciliation_audit = selection["reconciliationAudit"]
    included_allocations = [
        row for row in tension_allocation["records"] if row.get("included")
    ]
    included_item_ids = [str(row["itemId"]) for row in included_allocations]
    allocation_counts = tension_allocation["counts"]
    counterpart_bridge_ids = {
        str(row["historicalItemId"])
        for row in tension_allocation["records"]
        if row.get("counterpartBothPoleBridge")
    }
    td024_allocations = [
        row
        for row in tension_allocation["records"]
        if row.get("historicalTensionId") == "TD-024"
    ]
    required_roles = {row[0] for row in RELATIONSHIP_SEMANTICS}
    actual_roles = {str(row["semanticRole"]) for row in relationship_semantics}
    cluster_ids = {str(row["clusterId"]) for row in clusters}
    family_ids = {str(row["familyId"]) for row in families}
    theme_ids = {str(row["themeId"]) for row in themes}
    tension_ids = {str(row["tensionId"]) for row in tensions}
    scenario_ids = {str(row["scenarioId"]) for row in scenarios}
    endpoints_resolve = (
        all(set(row["memberClusterIds"]) <= cluster_ids for row in families)
        and all(
            set(row["primaryFamilyIds"] + row["secondaryFamilyIds"]) <= family_ids
            and set(row["primaryClusterIds"] + row["secondaryClusterIds"])
            <= cluster_ids
            for row in themes
        )
        and all(
            set(row["supportingFamilyIds"]) <= family_ids
            and set(row["supportingClusterIds"]) <= cluster_ids
            for row in tensions
        )
        and all(
            set(row["integratesThemeIds"]) <= theme_ids
            and set(row["integratesTensionIds"]) <= tension_ids
            for row in narratives
        )
        and all(
            set(row["relevantThemeIds"]) <= theme_ids
            and set(row["relevantTensionIds"]) <= tension_ids
            and {
                str(relation["targetScenarioId"])
                for relation in row["relationshipsToOtherScenarios"]
            }
            <= scenario_ids
            for row in scenarios
        )
    )
    gates = (
        (
            "canonical-content-unit-count",
            len(selected_sources) == 241,
            len(selected_sources),
        ),
        (
            "public-release-count",
            counts["publicReleaseCount"] == 242,
            counts["publicReleaseCount"],
        ),
        (
            "historical-source-identity-count",
            counts["historicalSourceIdentityCount"] == 269,
            counts["historicalSourceIdentityCount"],
        ),
        (
            "historical-higher-order-provenance-retained",
            (
                counts["historicalClusterCount"] == 127
                and counts["historicalMetaClusterCount"] == 36
                and counts["historicalThemeCount"] == 11
                and counts["historicalTensionCount"] == 30
                and counts["historicalNarrativeCount"] == 7
                and counts["historicalCategoryFindingCount"] == 42
                and counts["historicalScenarioCount"] == 6
            ),
            {
                "clusters": counts["historicalClusterCount"],
                "metaClusters": counts["historicalMetaClusterCount"],
                "themes": counts["historicalThemeCount"],
                "tensions": counts["historicalTensionCount"],
                "narratives": counts["historicalNarrativeCount"],
                "categoryFindings": counts["historicalCategoryFindingCount"],
                "scenarios": counts["historicalScenarioCount"],
            },
        ),
        (
            "exact-corrected-item-count",
            counts["canonicalItemCount"] == 12933,
            counts["canonicalItemCount"],
        ),
        (
            "no-alias-analytical-weight",
            not alias_contributors
            and not (selected_sources & excluded_identity_ids)
            and all(
                str(row["sourceIdentityId"]) not in excluded_identity_ids
                for row in selected_item_records
            ),
            len(alias_contributors)
            + len(selected_sources & excluded_identity_ids)
            + sum(
                str(row["sourceIdentityId"]) in excluded_identity_ids
                for row in selected_item_records
            ),
        ),
        (
            "per-item-selection-lineage-complete",
            len(selection["historicalItemSelection"]) == counts["historicalItemCount"]
            and len(selected_item_records) == counts["canonicalItemCount"]
            and len(selection["canonicalContentUnitSelection"])
            == counts["canonicalAnalyticalContentUnitCount"],
            {
                "historicalItems": len(selection["historicalItemSelection"]),
                "selectedItems": len(selected_item_records),
                "contentUnits": len(selection["canonicalContentUnitSelection"]),
            },
        ),
        (
            "reconciliation-report-cross-validated",
            reconciliation_audit.get("crossValidationStatus") == "pass"
            and reconciliation_audit.get("releaseEligible") is True,
            reconciliation_audit.get("crossValidationStatus"),
        ),
        (
            "transcript-manifest-governance-complete",
            transcript_audit
            == {
                "status": "pass",
                "coverageComplete": True,
                "sourceIdentityCount": 269,
                "canonicalReleaseCount": 242,
                "selectedTranscriptCount": 242,
                "strictUniqueContentUnits": 241,
                "excludedAliasIdentityCount": 27,
                "excludedOutsideGovernedCorpusCount": 2,
                "filenameDriftResolutions": 1,
                "contentReuseGroupCount": 1,
            },
            transcript_audit,
        ),
        (
            "shared-content-weighted-once",
            counts["excludedSharedContentIdentityCount"] == 1
            and counts["excludedSharedContentItemCount"] == 45,
            counts["excludedSharedContentItemCount"],
        ),
        (
            "all-127-clusters-accounted-for",
            len(clusters) == 127 and len({row["clusterId"] for row in clusters}) == 127,
            len(clusters),
        ),
        (
            "no-cluster-lost-all-support",
            all(
                row["canonicalPrimaryItemCount"] + row["canonicalSecondaryItemCount"]
                > 0
                for row in clusters
            ),
            sum(
                row["canonicalPrimaryItemCount"] + row["canonicalSecondaryItemCount"]
                == 0
                for row in clusters
            ),
        ),
        (
            "all-families-have-members",
            bool(families) and all(row["memberClusterIds"] for row in families),
            len(families),
        ),
        (
            "no-orphan-canonical-cluster",
            len(
                [
                    cluster_id
                    for row in families
                    for cluster_id in row["memberClusterIds"]
                ]
            )
            == len(
                {
                    cluster_id
                    for row in families
                    for cluster_id in row["memberClusterIds"]
                }
            )
            == 127,
            sum(len(row["memberClusterIds"]) for row in families),
        ),
        (
            "all-medium-family-mappings-evidence-reviewed",
            family_adjudication["decisionSummary"]["mediumConfidenceEvidenceReviews"]
            == 33
            and all(
                row["reviewedEvidence"]
                for row in family_adjudication["mappingDecisions"]
                if row["mappingConfidence"] == "medium"
            ),
            family_adjudication["decisionSummary"]["mediumConfidenceEvidenceReviews"],
        ),
        (
            "all-family-mappings-evidence-profile-reviewed",
            len(family_adjudication["mappingDecisions"]) == 127
            and all(
                row["reviewedEvidence"]
                for row in family_adjudication["mappingDecisions"]
            ),
            len(family_adjudication["mappingDecisions"]),
        ),
        (
            "independent-family-adjudication-complete",
            family_adjudication["decisionSummary"]["confirmedWithoutFamilyChange"]
            == 127
            and family_adjudication["decisionSummary"]["reassignedAfterReview"] == 0
            and family_adjudication["decisionSummary"]["deepItemLevelEvidenceReviews"]
            == 34
            and family_adjudication["decisionSummary"]["finalHighConfidenceMappings"]
            == 119
            and family_adjudication["decisionSummary"][
                "finalModerateConfidenceMappings"
            ]
            == 8
            and family_adjudication["governedValidation"].get("passed") is True
            and any(
                row.get("metaClusterId") == "CRB-M05"
                and row.get("decision") == "do-not-revive"
                for row in family_adjudication["historicalMetaClusterDecisions"]
            ),
            {
                "confirmed": family_adjudication["decisionSummary"][
                    "confirmedWithoutFamilyChange"
                ],
                "reassigned": family_adjudication["decisionSummary"][
                    "reassignedAfterReview"
                ],
                "deepReviews": family_adjudication["decisionSummary"][
                    "deepItemLevelEvidenceReviews"
                ],
                "finalHigh": family_adjudication["decisionSummary"][
                    "finalHighConfidenceMappings"
                ],
                "finalModerate": family_adjudication["decisionSummary"][
                    "finalModerateConfidenceMappings"
                ],
            },
        ),
        (
            "all-tension-source-candidates-and-pole-occurrences-accounted",
            allocation_counts["sourceCandidateRows"] == 43
            and allocation_counts["sourceCandidateIdsRepresented"] == 43
            and allocation_counts["allocationOccurrences"] == 452,
            {
                "candidateRows": allocation_counts["sourceCandidateRows"],
                "candidateIdsRepresented": allocation_counts[
                    "sourceCandidateIdsRepresented"
                ],
                "poleOccurrences": allocation_counts["allocationOccurrences"],
            },
        ),
        (
            "all-alias-tension-counterparts-governed",
            allocation_counts["counterpartAdjudicationRecords"] == 74
            and allocation_counts["uniqueGovernedCounterpartMappingsUsed"] == 74
            and allocation_counts["distinctGovernedCounterpartTargets"] == 72
            and allocation_counts["governedCanonicalCounterpartSubstitutionOccurrences"]
            == 76,
            {
                "records": allocation_counts["counterpartAdjudicationRecords"],
                "used": allocation_counts["uniqueGovernedCounterpartMappingsUsed"],
                "targets": allocation_counts["distinctGovernedCounterpartTargets"],
                "occurrences": allocation_counts[
                    "governedCanonicalCounterpartSubstitutionOccurrences"
                ],
            },
        ),
        (
            "alias-counterpart-confidence-and-drift-audited",
            allocation_counts["counterpartHighConfidenceRecords"] == 62
            and allocation_counts["counterpartMediumConfidenceRecords"] == 11
            and allocation_counts["counterpartReviewRecords"] == 1
            and allocation_counts["counterpartCategoryDriftRecords"] == 18
            and allocation_counts["counterpartCompoundOverlapRecords"] == 1,
            {
                "high": allocation_counts["counterpartHighConfidenceRecords"],
                "medium": allocation_counts["counterpartMediumConfidenceRecords"],
                "review": allocation_counts["counterpartReviewRecords"],
                "categoryDrift": allocation_counts["counterpartCategoryDriftRecords"],
                "compoundOverlap": allocation_counts[
                    "counterpartCompoundOverlapRecords"
                ],
                "approvedClusterLineageBridges": allocation_counts[
                    "counterpartApprovedClusterLineageBridgeRecords"
                ],
            },
        ),
        (
            "contextual-counterpart-cluster-lineage-explicitly-approved",
            allocation_counts["counterpartApprovedClusterLineageBridgeRecords"] == 4,
            allocation_counts["counterpartApprovedClusterLineageBridgeRecords"],
        ),
        (
            "both-pole-alias-bridges-counted-once",
            allocation_counts["counterpartBothPoleBridgeRecords"] == 2
            and allocation_counts["counterpartBothPoleBridgeOccurrences"] == 4
            and all(
                sum(
                    float(row.get("analyticalSupportWeight", 0.0))
                    for row in tension_allocation["records"]
                    if str(row.get("historicalItemId")) == historical_item_id
                )
                <= 1
                for historical_item_id in counterpart_bridge_ids
            ),
            {
                "bridgeRecords": allocation_counts["counterpartBothPoleBridgeRecords"],
                "sourcePoleOccurrences": allocation_counts[
                    "counterpartBothPoleBridgeOccurrences"
                ],
            },
        ),
        (
            "excluded-tension-lineage-fully-resolved",
            allocation_counts["unresolvedExcludedSourceIdentityOccurrences"] == 0,
            allocation_counts["unresolvedExcludedSourceIdentityOccurrences"],
        ),
        (
            "split-tension-adjudication-complete",
            allocation_counts["splitAdjudicationRecords"] == 98
            and allocation_counts["splitAdjudicationRecordsUsed"] == 98
            and allocation_counts["splitSourcePoleOccurrences"] == 98
            and allocation_counts["splitOccurrencesIncludedByAdjudication"] == 96
            and allocation_counts[
                "splitDuplicateBridgeOccurrencesExcludedByAdjudication"
            ]
            == 2
            and allocation_counts["splitHighConfidenceRecords"] == 87
            and allocation_counts["splitMediumConfidenceRecords"] == 9
            and allocation_counts["splitReviewRecords"] == 2
            and allocation_counts["splitOccurrencesDepartingFromProposedLegacyTargets"]
            == 5
            and len(tension_allocation["splitDecisionSummary"]) == 4,
            {
                "records": allocation_counts["splitAdjudicationRecords"],
                "used": allocation_counts["splitAdjudicationRecordsUsed"],
                "included": allocation_counts["splitOccurrencesIncludedByAdjudication"],
                "excludedBridgeOccurrences": allocation_counts[
                    "splitDuplicateBridgeOccurrencesExcludedByAdjudication"
                ],
                "high": allocation_counts["splitHighConfidenceRecords"],
                "medium": allocation_counts["splitMediumConfidenceRecords"],
                "review": allocation_counts["splitReviewRecords"],
                "designDepartures": allocation_counts[
                    "splitOccurrencesDepartingFromProposedLegacyTargets"
                ],
            },
        ),
        (
            "no-lexical-counterpart-or-split-routing",
            not tension_allocation["allocationRules"][
                "lexicalSimilarityUsedForCounterpartIdentity"
            ]
            and not tension_allocation["allocationRules"][
                "lexicalOrKeywordRoutingUsedForSplitTensions"
            ],
            False,
        ),
        (
            "cross-tension-collision-adjudication-complete",
            allocation_counts["crossTensionCollisionAdjudicationRecords"] == 12
            and allocation_counts["crossTensionCollisionEligibleOccurrences"] == 24
            and allocation_counts["crossTensionCollisionExcludedOccurrences"] == 12
            and allocation_counts["crossTensionCollisionHighConfidenceRecords"] == 9
            and allocation_counts["crossTensionCollisionMediumConfidenceRecords"] == 2
            and allocation_counts["crossTensionCollisionReviewConfidenceRecords"] == 1
            and allocation_counts["crossTensionCollisionReviewRequiredRecords"] == 3
            and allocation_counts[
                "crossTensionCollisionDecisionsChangingMechanicalWinner"
            ]
            == 7
            and len(tension_allocation["crossTensionCollisionDecisionSummary"]) == 12,
            {
                "records": allocation_counts[
                    "crossTensionCollisionAdjudicationRecords"
                ],
                "occurrences": allocation_counts[
                    "crossTensionCollisionEligibleOccurrences"
                ],
                "excludedOccurrences": allocation_counts[
                    "crossTensionCollisionExcludedOccurrences"
                ],
                "high": allocation_counts["crossTensionCollisionHighConfidenceRecords"],
                "medium": allocation_counts[
                    "crossTensionCollisionMediumConfidenceRecords"
                ],
                "review": allocation_counts[
                    "crossTensionCollisionReviewConfidenceRecords"
                ],
                "reviewRequired": allocation_counts[
                    "crossTensionCollisionReviewRequiredRecords"
                ],
                "changedMechanicalWinner": allocation_counts[
                    "crossTensionCollisionDecisionsChangingMechanicalWinner"
                ],
            },
        ),
        (
            "same-tension-repeated-lineage-adjudication-complete",
            allocation_counts["sameTensionDuplicateAdjudicationRecords"] == 8
            and allocation_counts["sameTensionDuplicateEligibleOccurrences"] == 16
            and allocation_counts["sameTensionDuplicateHighConfidenceRecords"] == 8
            and allocation_counts["sameTensionDuplicateReviewRequiredRecords"] == 0
            and allocation_counts["sameTensionSharedProvenanceRecords"] == 4
            and allocation_counts["sameTensionDualPoleBridgeRecords"] == 4
            and allocation_counts["sameTensionPoleConflictItemCount"] == 4
            and len(tension_allocation["sameTensionDuplicateDecisionSummary"]) == 8,
            {
                "records": allocation_counts["sameTensionDuplicateAdjudicationRecords"],
                "occurrences": allocation_counts[
                    "sameTensionDuplicateEligibleOccurrences"
                ],
                "samePoleSharedProvenance": allocation_counts[
                    "sameTensionSharedProvenanceRecords"
                ],
                "dualPoleBridges": allocation_counts[
                    "sameTensionDualPoleBridgeRecords"
                ],
                "reviewRequired": allocation_counts[
                    "sameTensionDuplicateReviewRequiredRecords"
                ],
            },
        ),
        (
            "tension-evidence-not-double-counted",
            all(
                isinstance(row.get("analyticalSupportWeight"), (int, float))
                for row in tension_allocation["records"]
            )
            and len(set(included_item_ids))
            == sum(len(value) for value in tension_item_ids.values())
            and math.isclose(
                sum(
                    float(row.get("analyticalSupportWeight", 0.0))
                    for row in included_allocations
                ),
                float(len(set(included_item_ids))),
                abs_tol=1e-9,
            )
            and all(
                len(
                    {
                        str(row["canonicalTensionId"])
                        for row in included_allocations
                        if str(row["itemId"]) == item_id
                    }
                )
                == 1
                for item_id in set(included_item_ids)
            )
            and all(
                sum(
                    float(row.get("analyticalSupportWeight", 0.0))
                    for row in tension_allocation["records"]
                    if str(row["itemId"]) == item_id
                )
                <= 1.0
                for item_id in {
                    str(row["itemId"]) for row in tension_allocation["records"]
                }
            ),
            {
                "totalAllocationRows": len(tension_allocation["records"]),
                "positiveWeightAllocationRows": len(included_item_ids),
                "uniqueItems": len(set(included_item_ids)),
                "totalAnalyticalWeight": _round(
                    sum(
                        float(row.get("analyticalSupportWeight", 0.0))
                        for row in included_allocations
                    )
                ),
            },
        ),
        (
            "tension-evidence-uses-selected-content-only",
            all(
                str(row.get("canonicalContentUnitId")) in selected_sources
                for row in included_allocations
            ),
            sum(
                str(row.get("canonicalContentUnitId")) not in selected_sources
                for row in included_allocations
            ),
        ),
        (
            "tension-evidence-cluster-family-lineage-complete",
            all(
                str(row.get("primaryClusterId")) in cluster_ids
                and str(row.get("primaryFamilyId")) in family_ids
                for row in included_allocations
            ),
            sum(
                str(row.get("primaryClusterId")) not in cluster_ids
                or str(row.get("primaryFamilyId")) not in family_ids
                for row in included_allocations
            ),
        ),
        (
            "td024-orientation-normalized",
            tension_allocation["allocationRules"]["td024OrientationNormalized"]
            and len(td024_allocations) == 12
            and all(
                row.get("orientationTreatment") == "reversed-to-canonical-orientation"
                and (
                    (
                        row.get("sourcePoleOccurrence") == "A"
                        and row.get("normalizedPole") == "B"
                    )
                    or (
                        row.get("sourcePoleOccurrence") == "B"
                        and row.get("normalizedPole") == "A"
                    )
                )
                for row in td024_allocations
            ),
            len(td024_allocations),
        ),
        (
            "key-concepts-in-cross-level-audit",
            theme_adjudication["keyConceptsIncludedInEveryCrossLevelAudit"],
            True,
        ),
        (
            "future-trends-in-cross-level-audit",
            theme_adjudication["futureTrendsIncludedInEveryCrossLevelAudit"],
            True,
        ),
        (
            "theme-category-breadth-uses-direct-support-only",
            all(
                row["categoryBreadth"]
                == len(
                    {
                        str(family_id).split("-", 1)[0]
                        for family_id in (
                            row["primaryFamilyIds"] + row["secondaryFamilyIds"]
                        )
                        if not str(family_id).startswith(("KCF-", "FTP-"))
                    }
                )
                for row in themes
            ),
            {row["themeId"]: row["categoryBreadth"] for row in themes},
        ),
        (
            "theme-narrative-governed-adjudication-complete",
            len(theme_adjudication["decisions"]) == 11
            and Counter(
                str(row["confidence"]) for row in theme_adjudication["decisions"]
            )
            == Counter({"high": 7, "medium": 4})
            and len(narratives) == 5
            and Counter(str(row["adjudicationConfidence"]) for row in narratives)
            == Counter({"high": 3, "medium": 2})
            and {
                str(value)
                for row in narratives
                for value in row["historicalNarrativeIds"]
            }
            == {f"N{number:02d}" for number in range(1, 8)},
            {
                "themes": len(theme_adjudication["decisions"]),
                "themeConfidence": dict(
                    Counter(
                        str(row["confidence"])
                        for row in theme_adjudication["decisions"]
                    )
                ),
                "narratives": len(narratives),
                "narrativeConfidence": dict(
                    Counter(str(row["adjudicationConfidence"]) for row in narratives)
                ),
            },
        ),
        (
            "historical-cross-level-lineage-complete",
            len(theme_adjudication["historicalThemeLineage"]) == 11
            and len(theme_adjudication["historicalMetaClusterLineage"]) == 36
            and len(theme_adjudication["historicalNarrativeLineage"]) == 7
            and len(theme_adjudication["historicalScenarioLineage"]) == 6,
            {
                "themes": len(theme_adjudication["historicalThemeLineage"]),
                "metaClusters": len(theme_adjudication["historicalMetaClusterLineage"]),
                "narratives": len(theme_adjudication["historicalNarrativeLineage"]),
                "scenarios": len(theme_adjudication["historicalScenarioLineage"]),
            },
        ),
        (
            "historical-generation-artifact-caveats-removed",
            not any(
                phrase in json.dumps(narratives).lower()
                for phrase in ("truncated prompt", "missing source context")
            ),
            True,
        ),
        (
            "all-themes-traceable",
            bool(themes)
            and all(
                row["corpusSupportProfile"]["itemSupportCount"] > 0 for row in themes
            ),
            len(themes),
        ),
        (
            "all-tensions-traceable",
            bool(tensions)
            and all(
                row["corpusSupportProfile"]["itemSupportCount"] > 0 for row in tensions
            ),
            len(tensions),
        ),
        (
            "all-tensions-have-both-poles",
            bool(tensions)
            and all(
                row["evidenceBalanceAcrossPoles"]["bothPolesDirectlySupported"]
                for row in tensions
            ),
            len(tensions),
        ),
        (
            "canonical-tension-construct-adjudication-complete",
            len(tensions) == 20
            and Counter(str(row["adjudicationDecision"]) for row in tensions)
            == Counter({"retain": 20})
            and Counter(str(row["adjudicationConfidence"]) for row in tensions)
            == Counter({"high": 10, "medium": 10})
            and sum(bool(row["reviewRequired"]) for row in tensions) == 10
            and all(
                row["neighborDistinctions"]
                and row["conditionsFavoringA"]
                and row["conditionsFavoringB"]
                and row["adjudicationRationale"]
                for row in tensions
            )
            and tension_allocation["canonicalTensionConstructAdjudication"][
                "decisionSummary"
            ]["unresolvedStructuralDecisionCount"]
            == 0,
            {
                "retained": sum(
                    row["adjudicationDecision"] == "retain" for row in tensions
                ),
                "confidence": dict(
                    Counter(row["adjudicationConfidence"] for row in tensions)
                ),
                "reviewRequired": sum(bool(row["reviewRequired"]) for row in tensions),
                "unresolvedStructuralDecisions": tension_allocation[
                    "canonicalTensionConstructAdjudication"
                ]["decisionSummary"]["unresolvedStructuralDecisionCount"],
            },
        ),
        (
            "all-narratives-traceable",
            bool(narratives)
            and all(
                row["corpusSupportProfile"]["itemSupportCount"] > 0
                and row["unresolvedIssue"]
                for row in narratives
            ),
            len(narratives),
        ),
        (
            "all-findings-traceable",
            bool(findings)
            and all(
                row["corpusSupportProfile"]["itemSupportCount"] > 0 for row in findings
            ),
            len(findings),
        ),
        (
            "historical-category-finding-lineage-complete",
            len(historical_finding_lineage) == 42
            and len(
                {str(row["historicalFindingId"]) for row in historical_finding_lineage}
            )
            == 42
            and all(
                row["canonicalFamilyIds"]
                and row["canonicalFindingIds"]
                and row["contributesAdditionalAnalyticalWeight"] is False
                for row in historical_finding_lineage
            ),
            len(historical_finding_lineage),
        ),
        (
            "all-scenarios-traceable",
            bool(scenarios)
            and all(
                row["corpusSupportProfile"]["itemSupportCount"] > 0 for row in scenarios
            ),
            len(scenarios),
        ),
        (
            "scenario-portfolio-governed-and-not-forced",
            len(scenarios) == 6
            and Counter(str(row["adjudicationConfidence"]) for row in scenarios)
            == Counter({"high": 5, "medium": 1})
            and {str(row["scenarioId"]) for row in scenarios if row["reviewRequired"]}
            == {"SC-04"}
            and {str(row["historicalScenarioId"]) for row in scenarios}
            == {f"S{number:02d}" for number in range(1, 7)}
            and {
                str(row["scenarioId"]): (
                    len(row["plausiblePathways"]),
                    len(row["indicators"]),
                    len(row["responseOptions"]),
                )
                for row in scenarios
            }
            == {
                "SC-01": (7, 6, 5),
                "SC-02": (6, 6, 4),
                "SC-03": (6, 6, 4),
                "SC-04": (7, 6, 5),
                "SC-05": (7, 6, 5),
                "SC-06": (6, 5, 5),
            }
            and all(
                row["corpusSupportProfile"]["historicalToCorrectedSensitivity"][
                    "status"
                ]
                == "unassessable"
                and all(
                    dynamic.get("direction")
                    and dynamic.get("rationale")
                    and dynamic.get("dynamic")
                    != "Scenario conditions may shift the practical balance between the poles; direction remains context-dependent."
                    for dynamic in row["tensionPoleDynamics"]
                )
                and all(
                    relation.get("causalClaim") is False and relation.get("rationale")
                    for relation in row["relationshipsToOtherScenarios"]
                )
                for row in scenarios
            ),
            {
                str(row["scenarioId"]): {
                    "pathways": len(row["plausiblePathways"]),
                    "indicators": len(row["indicators"]),
                    "actions": len(row["responseOptions"]),
                    "confidence": row["adjudicationConfidence"],
                    "reviewRequired": row["reviewRequired"],
                }
                for row in scenarios
            },
        ),
        (
            "relationship-vocabulary-complete",
            required_roles <= actual_roles,
            len(actual_roles),
        ),
        (
            "canonical-relationship-endpoints-resolve",
            endpoints_resolve,
            endpoints_resolve,
        ),
        (
            "redundancy-audit-complete",
            redundancy_audit["comparisonCount"] == 3511
            and redundancy_audit["potentialRedundancyCount"] == 7
            and redundancy_audit["resolvedDistinctCount"] == 7
            and redundancy_audit["unresolvedPotentialRedundancyCount"] == 0,
            {
                "comparisons": redundancy_audit["comparisonCount"],
                "flagged": redundancy_audit["potentialRedundancyCount"],
                "resolvedDistinct": redundancy_audit["resolvedDistinctCount"],
                "unresolved": redundancy_audit["unresolvedPotentialRedundancyCount"],
            },
        ),
    )
    return [
        {"gate": name, "status": "pass" if passed else "fail", "observed": observed}
        for name, passed, observed in gates
    ]


def _public_surface_hashes(repo_root: Path) -> dict[str, str]:
    roots = (
        repo_root / "data" / "cognitive-security",
        repo_root / "cognitive-security",
    )
    return {
        path.relative_to(repo_root).as_posix(): _sha256(path)
        for root in roots
        for path in sorted(root.rglob("*"), key=lambda row: row.as_posix())
        if path.is_file()
    }


def _require_git_ignored_output(repo_root: Path, output_dir: Path) -> None:
    """Fail closed unless Git confirms the private output path is ignored."""

    try:
        relative = output_dir.resolve().relative_to(repo_root.resolve()).as_posix()
    except ValueError as exc:
        raise ValueError("Canonical output is outside the repository.") from exc
    check = subprocess.run(
        ["git", "check-ignore", "--quiet", "--", relative],
        cwd=repo_root,
        check=False,
        capture_output=True,
        text=True,
    )
    if check.returncode != 0:
        raise ValueError(
            "Canonical output must be ignored by Git before private artifacts "
            "can be written."
        )


def _render_report(
    selection_report: Mapping[str, Any],
    clusters: Sequence[Mapping[str, Any]],
    families: Sequence[Mapping[str, Any]],
    themes: Sequence[Mapping[str, Any]],
    tensions: Sequence[Mapping[str, Any]],
    narratives: Sequence[Mapping[str, Any]],
    findings: Sequence[Mapping[str, Any]],
    scenarios: Sequence[Mapping[str, Any]],
    redundancy_audit: Mapping[str, Any],
    review_queue: Sequence[Mapping[str, Any]],
    quality_gates: Sequence[Mapping[str, Any]],
) -> str:
    counts = selection_report["counts"]
    sensitive = sorted(
        [row for row in clusters if row["sensitivityStatus"] != "stable"],
        key=lambda row: (
            row["historicalToCorrectedChange"]["weightedPercentChange"] or 0,
            _natural(row["clusterId"]),
        ),
    )
    finding_counts = Counter(row["findingType"] for row in findings)
    lines = [
        "# Cognitive Security canonical re-synthesis — analytical checkpoint",
        "",
        "This ignored report describes a draft analytical reconstruction. It does not change the public Explorer.",
        "",
        "## Corpus selection",
        "",
        f"- Historical source identities retained as provenance: {counts['historicalSourceIdentityCount']}",
        f"- Public releases retained: {counts['publicReleaseCount']}",
        f"- Canonical analytical content units: {counts['canonicalAnalyticalContentUnitCount']}",
        f"- Corrected items: {counts['canonicalItemCount']} ({counts['canonicalFocalItemCount']} focal; {counts['canonicalContextualItemCount']} contextual)",
        f"- Excluded identities/items: {counts['excludedIdentityCount']} / {counts['excludedItemCount']}",
        "- Episode 83: both releases retained; the re-release inherits from the original recording and adds no analytical weight.",
        "",
        "## Canonical draft architecture",
        "",
        f"- Clusters recomputed: {len(clusters)}",
        f"- Families: {len(families)}",
        f"- Equal-level themes: {len(themes)}",
        f"- Tensions: {len(tensions)}",
        f"- Narratives: {len(narratives)}",
        f"- Category findings: {len(findings)} ({dict(sorted(finding_counts.items()))})",
        f"- Scenarios: {len(scenarios)}",
        "",
        "## Cluster sensitivity",
        "",
    ]
    if sensitive:
        for row in sensitive[:25]:
            change = row["historicalToCorrectedChange"]["weightedPercentChange"]
            lines.append(
                f"- `{row['clusterId']}`: {row['sensitivityStatus']}; weighted change {change}%"
            )
    else:
        lines.append("- No cluster crossed a review threshold.")
    lines.extend(
        [
            "",
            "## Redundancy and review",
            "",
            f"- Pairwise comparisons completed: {redundancy_audit['comparisonCount']}",
            f"- Potential redundancy pairs: {redundancy_audit['potentialRedundancyCount']}",
            f"- Resolved as intentionally distinct: {redundancy_audit['resolvedDistinctCount']}",
            f"- Unresolved redundancy pairs: {redundancy_audit['unresolvedPotentialRedundancyCount']}",
            f"- Review-queue records: {len(review_queue)}",
            "",
            "## Quality gates",
            "",
        ]
    )
    for gate in quality_gates:
        lines.append(f"- {gate['status'].upper()}: {gate['gate']} ({gate['observed']})")
    lines.extend(["", "> " + SUPPORT_INTERPRETATION])
    return "\n".join(lines)


def build_canonical_resynthesis(
    *,
    repo_root: Path,
    normalized_dir: Path,
    reconciliation_dir: Path,
    transcript_summary_dir: Path,
    source_workbook_dir: Path,
    design_dir: Path,
    output_dir: Path,
) -> dict[str, Any]:
    """Build and write all ignored analytical checkpoint artifacts."""

    private_output_root = (
        repo_root / "analysis" / "cognitive-security" / "canonical-resynthesis"
    ).resolve()
    resolved_output_dir = output_dir.resolve()
    if not (
        resolved_output_dir == private_output_root
        or private_output_root in resolved_output_dir.parents
    ):
        raise ValueError(
            "Canonical re-synthesis output must remain beneath the repository's "
            "ignored analysis/cognitive-security/canonical-resynthesis directory."
        )
    private_input_root = (private_output_root / "inputs").resolve()
    if (
        resolved_output_dir == private_input_root
        or private_input_root in resolved_output_dir.parents
    ):
        raise ValueError("Canonical output cannot overwrite the governed input tree.")
    _require_git_ignored_output(repo_root, resolved_output_dir)

    tension_workbook = source_workbook_dir / "tensions_debates_rebuilt.xlsx"
    public_before = _public_surface_hashes(repo_root)
    input_before = _input_manifest(
        normalized_dir,
        reconciliation_dir,
        design_dir,
        tension_workbook,
        transcript_summary_dir,
    )
    dataset = load_private_dataset(normalized_dir)
    design = load_design(design_dir)
    reconciliation_report = _read_json(
        reconciliation_dir / "corpus_reconciliation_report.json"
    )
    counterpart_adjudication = _read_json(design_dir / COUNTERPART_ADJUDICATION_FILE)
    split_allocation_adjudication = _read_json(
        design_dir / SPLIT_ALLOCATION_ADJUDICATION_FILE
    )
    collision_adjudication = _read_json(design_dir / COLLISION_ADJUDICATION_FILE)
    same_tension_adjudication = _read_json(design_dir / SAME_TENSION_ADJUDICATION_FILE)
    tension_adjudication = _read_json(design_dir / TENSION_ADJUDICATION_FILE)
    cross_level_adjudication = _read_json(
        design_dir / THEME_NARRATIVE_ADJUDICATION_FILE
    )
    scenario_adjudication = _read_json(design_dir / SCENARIO_ADJUDICATION_FILE)
    family_mapping_adjudication = _read_json(design_dir / FAMILY_ADJUDICATION_FILE)
    transcript_report = _read_json(transcript_summary_dir / "manifest_report.json")

    selection, selection_report, selected_item_ids, release_by_content = (
        build_corpus_selection(dataset, transcript_report, reconciliation_report)
    )
    selection_report["inputHashes"] = input_before
    clusters, historical_by_cluster, retained_by_cluster = build_cluster_support(
        dataset, selected_item_ids, release_by_content
    )
    families, family_adjudication, cluster_to_family, family_items = build_families(
        dataset,
        design,
        historical_by_cluster,
        retained_by_cluster,
        release_by_content,
        family_mapping_adjudication,
    )
    themes, theme_adjudication, theme_items = build_themes(
        dataset,
        design,
        families,
        cluster_to_family,
        historical_by_cluster,
        retained_by_cluster,
        release_by_content,
        cross_level_adjudication,
    )
    tensions, tension_allocation, tension_items = build_tensions(
        dataset,
        design,
        tension_workbook,
        selected_item_ids,
        cluster_to_family,
        release_by_content,
        counterpart_adjudication,
        split_allocation_adjudication,
        collision_adjudication,
        same_tension_adjudication,
        tension_adjudication,
    )
    narratives, narrative_items = build_narratives(
        dataset,
        design,
        themes,
        theme_items,
        tensions,
        tension_items,
        historical_by_cluster,
        release_by_content,
        cross_level_adjudication,
    )
    findings, finding_items, historical_finding_lineage = build_category_findings(
        dataset,
        families,
        family_items,
        historical_by_cluster,
        release_by_content,
    )
    scenarios, scenario_items = build_scenarios(
        dataset,
        design,
        themes,
        theme_items,
        tension_items,
        cluster_to_family,
        retained_by_cluster,
        release_by_content,
        scenario_adjudication,
    )
    relationship_semantics = [
        {"semanticRole": role, "meaning": meaning, "causalClaim": False}
        for role, meaning in RELATIONSHIP_SEMANTICS
    ]
    redundancy_audit = build_redundancy_audit(
        families,
        family_items,
        themes,
        theme_items,
        tensions,
        tension_items,
        narratives,
        narrative_items,
        findings,
        finding_items,
        scenarios,
        scenario_items,
    )
    support_profiles = _support_profile_catalog(
        clusters, families, themes, tensions, narratives, findings, scenarios
    )
    review_queue = build_review_queue(
        clusters,
        family_adjudication,
        theme_adjudication,
        tensions,
        tension_allocation,
        narratives,
        scenarios,
        redundancy_audit,
    )
    quality_gates = _quality_gates(
        selection,
        clusters,
        families,
        family_adjudication,
        themes,
        theme_adjudication,
        tensions,
        tension_allocation,
        tension_items,
        narratives,
        findings,
        historical_finding_lineage,
        scenarios,
        relationship_semantics,
        redundancy_audit,
    )
    failed = [row for row in quality_gates if row["status"] != "pass"]
    if failed:
        names = ", ".join(row["gate"] for row in failed)
        raise ValueError(f"Canonical re-synthesis quality gates failed: {names}")

    outputs: dict[str, Any] = {
        "canonical_corpus_selection.json": selection,
        "corpus_selection_report.json": selection_report,
        "cluster_support_recomputed.json": {
            "schemaVersion": SCHEMA_VERSION,
            "methodVersion": METHOD_VERSION,
            "weightingFormula": "2 * canonicalPrimaryItemCount + canonicalSecondaryItemCount",
            "records": clusters,
        },
        "family_adjudication.json": family_adjudication,
        "canonical_families_draft.json": {
            "schemaVersion": SCHEMA_VERSION,
            "status": "analyst-reviewed-draft",
            "records": families,
        },
        "theme_adjudication.json": theme_adjudication,
        "canonical_themes_draft.json": {
            "schemaVersion": SCHEMA_VERSION,
            "status": "analyst-reviewed-draft",
            "publicLevels": ["theme"],
            "records": themes,
        },
        "tension_evidence_allocation.json": tension_allocation,
        "canonical_tensions_draft.json": {
            "schemaVersion": SCHEMA_VERSION,
            "status": "analyst-reviewed-draft",
            "constructAdjudication": {
                "methodVersion": tension_adjudication["methodVersion"],
                "decisionSummary": tension_adjudication["decisionSummary"],
            },
            "records": tensions,
        },
        "canonical_narratives_draft.json": {
            "schemaVersion": SCHEMA_VERSION,
            "status": "analyst-reviewed-draft",
            "adjudicationDecisions": cross_level_adjudication["narrativeDecisions"],
            "historicalNarrativeLineage": cross_level_adjudication[
                "historicalNarrativeLineage"
            ],
            "records": narratives,
        },
        "canonical_category_findings_draft.json": {
            "schemaVersion": SCHEMA_VERSION,
            "status": "analyst-reviewed-draft",
            "historicalFindingLineage": historical_finding_lineage,
            "records": findings,
        },
        "canonical_scenarios_draft.json": {
            "schemaVersion": SCHEMA_VERSION,
            "status": "analyst-reviewed-draft",
            "portfolioAdjudication": scenario_adjudication["portfolioReview"],
            "records": scenarios,
        },
        "relationship_semantics.json": {
            "schemaVersion": SCHEMA_VERSION,
            "records": relationship_semantics,
        },
        "support_profiles.json": support_profiles,
        "canonicalization_review_queue.json": {
            "schemaVersion": SCHEMA_VERSION,
            "humanApprovalRequired": True,
            "records": review_queue,
            "redundancyAudit": redundancy_audit,
        },
    }
    report = _render_report(
        selection_report,
        clusters,
        families,
        themes,
        tensions,
        narratives,
        findings,
        scenarios,
        redundancy_audit,
        review_queue,
        quality_gates,
    )
    input_prewrite = _input_manifest(
        normalized_dir,
        reconciliation_dir,
        design_dir,
        tension_workbook,
        transcript_summary_dir,
    )
    public_prewrite = _public_surface_hashes(repo_root)
    if input_before != input_prewrite:
        raise RuntimeError("A read-only governed input changed before output writing.")
    if public_before != public_prewrite:
        raise RuntimeError("The public Explorer surface changed before output writing.")

    for filename, payload in outputs.items():
        _write_json(resolved_output_dir / filename, payload)
    _write_text(resolved_output_dir / "canonical_resynthesis_report.md", report)

    input_after = _input_manifest(
        normalized_dir,
        reconciliation_dir,
        design_dir,
        tension_workbook,
        transcript_summary_dir,
    )
    public_after = _public_surface_hashes(repo_root)
    if input_before != input_after:
        raise RuntimeError("A read-only governed input changed during the build.")
    if public_before != public_after:
        raise RuntimeError(
            "The public Explorer surface changed during the private build."
        )
    return {
        "counts": selection_report["counts"],
        "qualityGates": quality_gates,
        "reviewQueueCount": len(review_queue),
        "redundancyAudit": {
            "comparisonCount": redundancy_audit["comparisonCount"],
            "potentialRedundancyCount": redundancy_audit["potentialRedundancyCount"],
            "resolvedDistinctCount": redundancy_audit["resolvedDistinctCount"],
            "unresolvedPotentialRedundancyCount": redundancy_audit[
                "unresolvedPotentialRedundancyCount"
            ],
            "flaggedPairs": redundancy_audit["flaggedPairs"],
        },
        "outputHashes": {
            filename: hashlib.sha256(_json_bytes(payload)).hexdigest()
            for filename, payload in outputs.items()
        }
        | {
            "canonical_resynthesis_report.md": hashlib.sha256(
                (report.rstrip() + "\n").encode("utf-8")
            ).hexdigest()
        },
        "publicExplorerUnchanged": True,
        "readOnlyInputsUnchanged": True,
    }
