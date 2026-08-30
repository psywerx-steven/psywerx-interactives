"""Lossless normalized-cell extraction from the governed XLSX package."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .sources import (
    ARTIFACT_BY_ID,
    ARTIFACT_SPECS,
    SHEETS_BY_ARTIFACT,
    SourceValidationError,
    artifact_manifest,
    find_header_row,
    normalized_headers,
    resolve_source_paths,
)
from .utils import normalize_cell


def _row_is_empty(values) -> bool:
    return not any(value is not None for value in values)


def _read_table(worksheet, spec, header_row: int, headers: tuple[str, ...]) -> list[dict]:
    records: list[dict] = []
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        header_row + 1,
    ):
        normalized = tuple(normalize_cell(value) for value in values[: len(headers)])
        if _row_is_empty(normalized):
            continue
        record = {
            header: normalized[index] if index < len(normalized) else None
            for index, header in enumerate(headers)
            if header
        }
        record["_source"] = {
            "artifactId": spec.artifact_id,
            "fileName": next(
                artifact.file_name
                for artifact in ARTIFACT_SPECS
                if artifact.artifact_id == spec.artifact_id
            ),
            "sheet": spec.sheet_name,
            "rowNumber": row_number,
        }
        records.append(record)
    if not records and not spec.allow_empty:
        raise SourceValidationError(
            f"{ARTIFACT_BY_ID[spec.artifact_id].file_name} / "
            f"{spec.sheet_name}: canonical table contains no data rows."
        )
    return records


def _inventory_header(worksheet) -> tuple[int | None, tuple[str, ...]]:
    candidates: list[tuple[int, tuple[str, ...], int]] = []
    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 25), values_only=True),
        1,
    ):
        headers = normalized_headers(values)
        while headers and not headers[-1]:
            headers = headers[:-1]
        candidates.append((row_number, headers, sum(bool(header) for header in headers)))
    if not candidates or max(candidate[2] for candidate in candidates) == 0:
        return None, ()
    row_number, headers, _ = max(candidates, key=lambda candidate: (candidate[2], -candidate[0]))
    return row_number, headers


def extract_sources(source_dir: Path | str) -> dict[str, Any]:
    """Verify and extract all canonical source tables.

    Returned rows retain exact source headers and portable ``_source``
    provenance.  No workbook path or wall-clock value enters the result.
    """

    paths = resolve_source_paths(source_dir)
    artifacts = artifact_manifest(paths)
    tables: dict[str, list[dict]] = {}
    inventory: list[dict] = []
    errors: list[str] = []

    for artifact in ARTIFACT_SPECS:
        path = paths[artifact.artifact_id]
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as error:  # openpyxl emits several format-specific exceptions
            errors.append(f"{artifact.file_name}: could not open workbook ({type(error).__name__}).")
            continue
        try:
            required_specs = SHEETS_BY_ARTIFACT[artifact.artifact_id]
            required_names = {spec.sheet_name for spec in required_specs}
            missing = sorted(required_names - set(workbook.sheetnames))
            if missing:
                errors.append(
                    f"{artifact.file_name}: missing required worksheets: {missing}. "
                    f"Discovered: {workbook.sheetnames}."
                )
                continue

            canonical_by_sheet = {spec.sheet_name: spec.table_key for spec in required_specs}
            for worksheet in workbook.worksheets:
                header_row, headers = _inventory_header(worksheet)
                inventory.append(
                    {
                        "artifactId": artifact.artifact_id,
                        "fileName": artifact.file_name,
                        "sheet": worksheet.title,
                        "rowCount": worksheet.max_row,
                        "columnCount": worksheet.max_column,
                        "headerRow": header_row,
                        "headers": list(headers),
                        "canonicalTable": canonical_by_sheet.get(worksheet.title),
                    }
                )

            for spec in required_specs:
                worksheet = workbook[spec.sheet_name]
                try:
                    header_row, headers = find_header_row(worksheet, spec)
                    tables[spec.table_key] = _read_table(
                        worksheet, spec, header_row, headers
                    )
                except SourceValidationError as error:
                    errors.extend(error.messages)
        finally:
            workbook.close()

    if errors:
        raise SourceValidationError(errors)
    expected_tables = {spec.table_key for specs in SHEETS_BY_ARTIFACT.values() for spec in specs}
    missing_tables = sorted(expected_tables - set(tables))
    if missing_tables:
        raise SourceValidationError(f"Canonical extraction omitted tables: {missing_tables}.")
    return {
        "artifacts": artifacts,
        "tables": tables,
        "sheetInventory": inventory,
    }
