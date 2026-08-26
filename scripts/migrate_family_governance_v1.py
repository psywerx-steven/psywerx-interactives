"""Apply canonical Family Schema v1.0 governance to all workbook Codebooks."""

from __future__ import annotations

import hashlib
import os
import sys
import tempfile
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from build_families import (
    CODEBOOK_DRIVER_COUNT_ROW,
    CODEBOOK_FAMILY_ID_ROW,
    CODEBOOK_REPRESENTATIVE_ROW,
    FAMILY_HEADERS,
    clean,
)


GOVERNED_FIELDS = {
    "Family ID",
    "Representative Drivers",
    "Driver Count",
}


@dataclass
class Baseline:
    sheet_hashes: dict[str, str]
    preserved_codebook_rows: tuple[tuple[str, ...], ...]
    family_ids: tuple[str, ...]


def hash_sheet_values(sheet: Any) -> str:
    digest = hashlib.sha256(sheet.title.encode("utf-8"))
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            payload = (cell.coordinate, cell.data_type, repr(cell.value))
            digest.update(repr(payload).encode("utf-8"))
    digest.update(
        repr(sorted(str(item) for item in sheet.merged_cells.ranges)).encode("utf-8")
    )
    return digest.hexdigest()


def codebook_rows(sheet: Any) -> tuple[tuple[str, ...], ...]:
    return tuple(
        tuple(clean(cell.value) for cell in row)
        for row in sheet.iter_rows()
        if any(clean(cell.value) for cell in row)
    )


def preserved_codebook_rows(sheet: Any) -> tuple[tuple[str, ...], ...]:
    return tuple(
        row
        for row in codebook_rows(sheet)
        if not (row[0] == "Families" and row[1] in GOVERNED_FIELDS)
    )


def collect_baseline(workbook: Any, path: Path) -> Baseline:
    if "Families" not in workbook.sheetnames or "Codebook" not in workbook.sheetnames:
        raise ValueError(f"{path.name}: missing Families or Codebook worksheet")
    families = workbook["Families"]
    headers = tuple(clean(cell.value) for cell in families[4] if clean(cell.value))
    if headers != FAMILY_HEADERS:
        raise ValueError(f"{path.name}: Families headers do not match Schema v1.0")
    id_column = headers.index("Family ID") + 1
    ids = tuple(
        str(row[id_column - 1].value)
        for row in families.iter_rows(min_row=5)
        if row[id_column - 1].value is not None
        and str(row[id_column - 1].value).strip()
    )
    if len(ids) != len(set(ids)):
        raise ValueError(f"{path.name}: duplicate Family IDs")
    return Baseline(
        sheet_hashes={
            name: hash_sheet_values(workbook[name])
            for name in workbook.sheetnames
            if name != "Codebook"
        },
        preserved_codebook_rows=preserved_codebook_rows(workbook["Codebook"]),
        family_ids=ids,
    )


def copy_row_style(sheet: Any, source_row: int, target_row: int) -> None:
    sheet.row_dimensions[target_row].height = sheet.row_dimensions[source_row].height
    for column in range(1, 7):
        source = sheet.cell(row=source_row, column=column)
        target = sheet.cell(row=target_row, column=column)
        if source.has_style:
            target._style = copy(source._style)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def write_row(sheet: Any, row_number: int, values: tuple[str, ...]) -> None:
    for column, value in enumerate(values, start=1):
        sheet.cell(row=row_number, column=column).value = value


def migrate_codebook(sheet: Any) -> None:
    fields = {
        clean(cell.value): cell.row for cell in sheet["B"] if clean(cell.value)
    }
    for required in ("Family ID", "Representative Drivers"):
        if required not in fields:
            raise ValueError(f"Codebook is missing {required!r}")
    if "Driver Count" not in fields:
        count_row = fields["Representative Drivers"] + 1
        sheet.insert_rows(count_row, 1)
        copy_row_style(sheet, count_row - 1, count_row)
    else:
        count_row = fields["Driver Count"]

    fields = {
        clean(cell.value): cell.row for cell in sheet["B"] if clean(cell.value)
    }
    write_row(sheet, fields["Family ID"], CODEBOOK_FAMILY_ID_ROW)
    write_row(
        sheet,
        fields["Representative Drivers"],
        CODEBOOK_REPRESENTATIVE_ROW,
    )
    write_row(sheet, count_row, CODEBOOK_DRIVER_COUNT_ROW)


def validate_staged(path: Path, baseline: Baseline) -> tuple[tuple[str, ...], ...]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        for name, expected_hash in baseline.sheet_hashes.items():
            if hash_sheet_values(workbook[name]) != expected_hash:
                raise ValueError(f"{path.name}: unrelated worksheet {name!r} changed")
        codebook = workbook["Codebook"]
        if preserved_codebook_rows(codebook) != baseline.preserved_codebook_rows:
            raise ValueError(f"{path.name}: unrelated Codebook content changed")
        rows = codebook_rows(codebook)
        governed = {
            row[1]: row for row in rows if row[0] == "Families" and len(row) >= 6
        }
        expected = {
            "Family ID": CODEBOOK_FAMILY_ID_ROW,
            "Representative Drivers": CODEBOOK_REPRESENTATIVE_ROW,
            "Driver Count": CODEBOOK_DRIVER_COUNT_ROW,
        }
        for field_name, expected_row in expected.items():
            if governed.get(field_name) != expected_row:
                raise ValueError(
                    f"{path.name}: Family governance row {field_name!r} is invalid"
                )
        return rows
    finally:
        workbook.close()


def cleanup_staging(directory: Path) -> None:
    resolved = directory.resolve()
    source_root = directory.parent.resolve()
    if resolved.parent != source_root or not resolved.name.startswith(
        ".family-governance-v1-"
    ):
        raise RuntimeError(f"Refusing to clean unexpected staging path: {resolved}")
    for child in resolved.iterdir():
        if child.is_file():
            child.unlink()
        else:
            raise RuntimeError(f"Unexpected directory in staging area: {child}")
    resolved.rmdir()


def main() -> int:
    root = Path(__file__).resolve().parent.parent
    source_dir = root / "source-data"
    paths = sorted(
        path
        for path in source_dir.glob("*.xlsx")
        if not path.name.startswith("~$")
    )
    if len(paths) != 8:
        print(f"ERROR: expected 8 XLSX workbooks; found {len(paths)}", file=sys.stderr)
        return 1

    staging = Path(
        tempfile.mkdtemp(prefix=".family-governance-v1-", dir=source_dir)
    )
    baselines: dict[str, Baseline] = {}
    try:
        global_ids: set[str] = set()
        for path in paths:
            workbook = load_workbook(path, read_only=False, data_only=False)
            try:
                baseline = collect_baseline(workbook, path)
                overlap = global_ids.intersection(baseline.family_ids)
                if overlap:
                    raise ValueError(
                        f"{path.name}: globally duplicate Family IDs: {sorted(overlap)}"
                    )
                global_ids.update(baseline.family_ids)
                baselines[path.name] = baseline
                migrate_codebook(workbook["Codebook"])
                workbook.save(staging / path.name)
            finally:
                workbook.close()

        if len(global_ids) != 105:
            raise ValueError(f"Expected 105 Family IDs; found {len(global_ids)}")
        signatures = {
            validate_staged(staging / path.name, baselines[path.name])
            for path in paths
        }
        if len(signatures) != 1:
            raise ValueError("Staged Codebook worksheets are not identical")
        for path in paths:
            os.replace(staging / path.name, path)

        print("Family governance Codebook migration complete")
        print(f"  Workbooks: {len(paths)}")
        print(f"  Globally unique Family IDs: {len(global_ids)}")
        print("  Identical Codebooks: yes")
        print("  Unrelated worksheet changes: 0")
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        print("No source workbook was replaced.", file=sys.stderr)
        return 1
    finally:
        if staging.exists():
            cleanup_staging(staging)


if __name__ == "__main__":
    raise SystemExit(main())
