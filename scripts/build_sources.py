"""Build the public PSYWERX Source Registry from Evidence Library worksheets.

The eight canonical workbooks are authoritative for Evidence records. This
builder preserves each permanent Evidence ID and citation, derives consumer
links from the public Driver and Relationship datasets, validates the complete
registry in memory, and atomically replaces data/sources.json only after
validation succeeds. Private analytical notes remain in the workbooks and are
not copied into the public registry.

Link resolution is deliberately conservative: explicit DOI, then explicit URL,
then a supported explicit scholarly identifier. A record without any of those
receives a clearly labelled scholarly-search fallback; no DOI or source URL is
guessed.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import quote, urlencode, urlparse

from openpyxl import load_workbook


SCHEMA_VERSION = "1.0"
SHEET = "Evidence Library"
HEADER_ROW = 4
FIRST_DATA_ROW = 5
HEADERS = (
    "Evidence ID",
    "Linked Driver IDs",
    "Family / Domain",
    "Citation / Title",
    "Year",
    "Evidence Type",
    "Population",
    "Finding",
    "Direction / Effect",
    "Limitations",
    "Evidence Strength",
    "Source URL",
    "Verification Note",
)
OUTPUT_KEYS = (
    "id",
    "citationText",
    "year",
    "evidenceType",
    "evidenceStrength",
    "sourceUrl",
    "resolutionType",
    "resolvedIdentifier",
    "href",
    "linkLabel",
    "driverIds",
    "relationshipIds",
    "source",
)
LAYERS = (
    "Biological",
    "Psychological",
    "Social",
    "Cultural",
    "Physical / Environmental",
    "Institutional / Structural",
    "Informational",
    "Technological",
)
FILENAME_LAYER_TOKENS = (
    ("layer_1_biological", "Biological"),
    ("layer_2_psychological", "Psychological"),
    ("layer_3_social", "Social"),
    ("layer_4_cultural", "Cultural"),
    ("layer_5_physical_environmental", "Physical / Environmental"),
    ("layer_6_institutional_structural", "Institutional / Structural"),
    ("layer_7_informational", "Informational"),
    ("layer_8_technological", "Technological"),
)
SOURCE_ID = re.compile(r"^SRC(?:-)?\d{3}$")
DOI = re.compile(
    r"(?i)(?:https?://(?:dx\.)?doi\.org/|\bdoi\s*:\s*)"
    r"(10\.\d{4,9}/[-._;()/:A-Z0-9]+)"
)
URL = re.compile(r"https?://[^\s<>\"']+", re.I)
PMID = re.compile(r"(?i)\bPMID\s*:?\s*(\d{5,9})\b")
PMCID = re.compile(r"(?i)\bPMCID\s*:?\s*(PMC\d+)\b")
ARXIV = re.compile(r"(?i)\barXiv\s*:?\s*(\d{4}\.\d{4,5}(?:v\d+)?)\b")
LOCAL_PATH = re.compile(
    r"(?:(?<![A-Za-z])\b[A-Za-z]:[\\/]|\\\\|file://|"
    r"source-data[\\/]|analysis[\\/])",
    re.I,
)
RESOLUTION_TYPES = {"DOI", "URL", "IDENTIFIER", "SEARCH", "UNRESOLVED"}


@dataclass
class Summary:
    workbooks: int = 0
    evidence_sheets: int = 0
    layers: set[str] = field(default_factory=set)
    source_records: int = 0
    linked_driver_references: int = 0
    driver_references: int = 0
    relationship_references: int = 0
    duplicate_urls: int = 0
    linked_driver_set_differences: int = 0
    resolution_counts: Counter[str] = field(default_factory=Counter)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def text(value: Any) -> str:
    """Trim cell boundaries without rewriting source-authored internal text."""

    return "" if value is None else str(value).strip()


def split_semicolon(value: Any) -> list[str]:
    return [item for part in str(value or "").split(";") if (item := part.strip())]


def workbook_layer(name: str) -> str | None:
    folded = name.casefold()
    matches = {layer for token, layer in FILENAME_LAYER_TOKENS if token in folded}
    return next(iter(matches)) if len(matches) == 1 else None


def source_sort_key(source_id: str) -> tuple[int, str]:
    match = re.search(r"\d+", source_id)
    return (int(match.group()) if match else sys.maxsize, source_id)


def parse_year(value: Any, location: str, summary: Summary) -> int | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, bool):
        summary.errors.append(f"{location}: Year must be a four-digit year or empty.")
        return None
    if isinstance(value, int):
        year = value
    elif isinstance(value, float) and value.is_integer():
        year = int(value)
    elif re.fullmatch(r"\d{4}", text(value)):
        year = int(text(value))
    else:
        summary.errors.append(f"{location}: Year {value!r} is not a four-digit year.")
        return None
    if not 1000 <= year <= 9999:
        summary.errors.append(f"{location}: Year {year!r} is not a four-digit year.")
        return None
    return year


def valid_http_url(value: str) -> bool:
    parsed = urlparse(value)
    return (
        parsed.scheme.casefold() in {"http", "https"}
        and bool(parsed.netloc)
        and not any(character.isspace() for character in value)
    )


def trim_identifier_punctuation(value: str) -> str:
    result = value.rstrip(".,;:")
    while result.endswith(")") and result.count("(") < result.count(")"):
        result = result[:-1]
    return result


def explicit_dois(values: Iterable[str]) -> list[str]:
    found: dict[str, str] = {}
    for value in values:
        for match in DOI.finditer(value):
            candidate = trim_identifier_punctuation(match.group(1))
            found.setdefault(candidate.casefold(), candidate)
    return list(found.values())


def explicit_urls(values: Iterable[str]) -> list[str]:
    found: dict[str, str] = {}
    for value in values:
        for match in URL.finditer(value):
            candidate = match.group().rstrip(".,;:")
            while candidate.endswith(")") and candidate.count("(") < candidate.count(")"):
                candidate = candidate[:-1]
            found.setdefault(candidate, candidate)
    return list(found.values())


def supported_identifiers(values: Iterable[str]) -> list[tuple[str, str, str]]:
    """Return only explicitly labelled identifiers with standard resolvers."""

    found: dict[str, tuple[str, str, str]] = {}
    for value in values:
        for match in PMID.finditer(value):
            identifier = match.group(1)
            found.setdefault(
                f"PMID:{identifier}",
                (f"PMID:{identifier}", f"https://pubmed.ncbi.nlm.nih.gov/{identifier}/", "Open PMID"),
            )
        for match in PMCID.finditer(value):
            identifier = match.group(1).upper()
            found.setdefault(
                f"PMCID:{identifier}",
                (f"PMCID:{identifier}", f"https://pmc.ncbi.nlm.nih.gov/articles/{identifier}/", "Open PMCID"),
            )
        for match in ARXIV.finditer(value):
            identifier = match.group(1)
            found.setdefault(
                f"arXiv:{identifier}",
                (f"arXiv:{identifier}", f"https://arxiv.org/abs/{identifier}", "Open arXiv record"),
            )
    return list(found.values())


def resolve_link(
    source_id: str,
    citation: str,
    source_url: str,
    summary: Summary,
) -> tuple[str, str | None, str | None, str | None]:
    """Resolve a public link without guessing a publication identifier."""

    searchable_fields = (source_url, citation)
    dois = explicit_dois(searchable_fields)
    if len(dois) > 1:
        summary.errors.append(
            f"{source_id}: multiple explicit DOI values conflict: {dois}."
        )
        return "UNRESOLVED", None, None, None
    if dois:
        doi = dois[0]
        href = None
        for candidate in explicit_urls((source_url,)):
            parsed = urlparse(candidate)
            if parsed.netloc.casefold() in {"doi.org", "dx.doi.org"}:
                href = candidate
                break
        if href is None:
            href = f"https://doi.org/{quote(doi, safe='/:;()')}"
        return "DOI", doi, href, "Open DOI"

    urls = explicit_urls(searchable_fields)
    if source_url:
        # The designated Source URL cell is authoritative when it has no DOI.
        href = source_url
    elif len(urls) == 1:
        href = urls[0]
    elif len(urls) > 1:
        summary.errors.append(
            f"{source_id}: multiple explicit URLs conflict and Source URL is empty."
        )
        return "UNRESOLVED", None, None, None
    else:
        href = None
    if href is not None:
        if not valid_http_url(href):
            summary.errors.append(f"{source_id}: invalid explicit Source URL {href!r}.")
            return "UNRESOLVED", None, None, None
        return "URL", None, href, "Open source"

    identifiers = supported_identifiers((citation,))
    if len(identifiers) > 1:
        summary.errors.append(
            f"{source_id}: multiple supported scholarly identifiers conflict."
        )
        return "UNRESOLVED", None, None, None
    if identifiers:
        identifier, href, label = identifiers[0]
        return "IDENTIFIER", identifier, href, label

    if citation:
        href = "https://scholar.google.com/scholar?" + urlencode({"q": citation})
        summary.warnings.append(
            f"{source_id}: no explicit DOI, URL, or supported scholarly identifier; "
            "using a marked scholarly-search fallback."
        )
        return "SEARCH", None, href, "Search Google Scholar"

    summary.warnings.append(
        f"{source_id}: no citation text or explicit resolvable source information."
    )
    return "UNRESOLVED", None, None, None


def load_codebook_strengths(path: Path, summary: Summary) -> set[str]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        summary.errors.append(f"Could not read Codebook {path}: {exc}")
        return set()
    entries = payload.get("entries") if isinstance(payload, dict) else None
    if not isinstance(entries, list):
        summary.errors.append(f"{path}: expected an entries array.")
        return set()
    matches = [row for row in entries if row.get("id") == "CB-EVI-EVIDENCE-STRENGTH"]
    if len(matches) != 1 or not isinstance(matches[0].get("allowedValues"), list):
        summary.errors.append(
            f"{path}: CB-EVI-EVIDENCE-STRENGTH is missing or malformed."
        )
        return set()
    return set(matches[0]["allowedValues"])


def load_drivers(
    path: Path, summary: Summary
) -> tuple[dict[str, dict[str, Any]], dict[str, list[str]], dict[str, int]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        summary.errors.append(f"Could not read Drivers {path}: {exc}")
        return {}, {}, {}
    if not isinstance(payload, list):
        summary.errors.append(f"{path}: expected a top-level Driver array.")
        return {}, {}, {}
    drivers: dict[str, dict[str, Any]] = {}
    reverse: dict[str, list[str]] = defaultdict(list)
    order: dict[str, int] = {}
    for index, record in enumerate(payload):
        if not isinstance(record, dict) or not isinstance(record.get("id"), str):
            summary.errors.append(f"{path}: Driver record {index} lacks a valid ID.")
            continue
        driver_id = record["id"]
        if driver_id in drivers:
            summary.errors.append(f"{path}: duplicate Driver ID {driver_id!r}.")
            continue
        drivers[driver_id] = record
        order[driver_id] = index
        references = record.get("keySources")
        if not isinstance(references, list):
            summary.errors.append(f"{driver_id}: keySources must be an array.")
            continue
        if len(references) != len(set(references)):
            summary.errors.append(f"{driver_id}: keySources contains duplicate IDs.")
        for source_id in references:
            if not isinstance(source_id, str) or not source_id:
                summary.errors.append(f"{driver_id}: keySources contains an invalid ID.")
                continue
            reverse[source_id].append(driver_id)
            summary.driver_references += 1
    return drivers, reverse, order


def load_relationships(
    path: Path, summary: Summary
) -> tuple[dict[str, dict[str, Any]], dict[str, list[str]]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        summary.errors.append(f"Could not read Relationships {path}: {exc}")
        return {}, {}
    records = payload.get("relationships") if isinstance(payload, dict) else None
    if not isinstance(records, list):
        summary.errors.append(f"{path}: expected a relationships array.")
        return {}, {}
    relationships: dict[str, dict[str, Any]] = {}
    reverse: dict[str, list[str]] = defaultdict(list)
    for index, record in enumerate(records):
        if not isinstance(record, dict) or not isinstance(record.get("id"), str):
            summary.errors.append(
                f"{path}: Relationship record {index} lacks a valid ID."
            )
            continue
        relationship_id = record["id"]
        if relationship_id in relationships:
            summary.errors.append(
                f"{path}: duplicate Relationship ID {relationship_id!r}."
            )
            continue
        relationships[relationship_id] = record
        references = record.get("supportingEvidenceIds")
        if not isinstance(references, list):
            summary.errors.append(
                f"{relationship_id}: supportingEvidenceIds must be an array."
            )
            continue
        if len(references) != len(set(references)):
            summary.errors.append(
                f"{relationship_id}: supportingEvidenceIds contains duplicate IDs."
            )
        for source_id in references:
            if not isinstance(source_id, str) or not source_id:
                summary.errors.append(
                    f"{relationship_id}: supportingEvidenceIds contains an invalid ID."
                )
                continue
            reverse[source_id].append(relationship_id)
            summary.relationship_references += 1
    return relationships, reverse


def read_sources(
    source_dir: Path,
    strengths: set[str],
    drivers: dict[str, dict[str, Any]],
    driver_references: dict[str, list[str]],
    driver_order: dict[str, int],
    relationship_references: dict[str, list[str]],
    summary: Summary,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    paths = sorted(
        path
        for path in source_dir.glob("*.xlsx")
        if not path.name.startswith("~$")
    )
    summary.workbooks = len(paths)
    if len(paths) != 8:
        summary.errors.append(f"Expected 8 XLSX workbooks; found {len(paths)}.")

    detected_layers: dict[str, Path] = {}
    for path in paths:
        layer = workbook_layer(path.name)
        if layer is None:
            summary.errors.append(
                f"{path.name}: filename does not identify exactly one canonical layer."
            )
        elif layer in detected_layers:
            summary.errors.append(
                f"{path.name}: duplicates the {layer} workbook {detected_layers[layer].name}."
            )
        else:
            detected_layers[layer] = path
            summary.layers.add(layer)

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            summary.errors.append(f"{path.name}: could not open workbook: {exc}")
            continue
        try:
            if SHEET not in workbook.sheetnames:
                summary.errors.append(f"{path.name}: missing {SHEET} worksheet.")
                continue
            sheet = workbook[SHEET]
            summary.evidence_sheets += 1
            actual_headers = tuple(
                text(cell.value) for cell in sheet[HEADER_ROW][: len(HEADERS)]
            )
            if actual_headers != HEADERS or sheet.max_column != len(HEADERS):
                summary.errors.append(
                    f"{path.name} / {SHEET}: expected exact headers {list(HEADERS)}; "
                    f"found {list(actual_headers)} with {sheet.max_column} columns."
                )
                continue

            for row_number, row in enumerate(
                sheet.iter_rows(
                    min_row=FIRST_DATA_ROW,
                    max_col=len(HEADERS),
                    values_only=True,
                ),
                start=FIRST_DATA_ROW,
            ):
                if not any(text(value) for value in row):
                    continue
                location = f"{path.name} / {SHEET} / row {row_number}"
                source_id = text(row[0])
                linked_driver_ids = split_semicolon(row[1])
                family_domain = text(row[2])
                citation = text(row[3])
                year = parse_year(row[4], location, summary)
                evidence_type = text(row[5])
                population = text(row[6])
                finding = text(row[7])
                direction_effect = text(row[8])
                limitations = text(row[9])
                evidence_strength = text(row[10])
                source_url = text(row[11])
                verification_note = text(row[12])

                required = {
                    "Evidence ID": source_id,
                    "Family / Domain": family_domain,
                    "Citation / Title": citation,
                    "Evidence Type": evidence_type,
                    "Population": population,
                    "Finding": finding,
                    "Direction / Effect": direction_effect,
                    "Limitations": limitations,
                    "Evidence Strength": evidence_strength,
                    "Verification Note": verification_note,
                }
                for header, value in required.items():
                    if not value:
                        summary.errors.append(
                            f"{location}: required field {header!r} is empty."
                        )
                if source_id and not SOURCE_ID.fullmatch(source_id):
                    summary.errors.append(
                        f"{location}: Evidence ID {source_id!r} is not a canonical ID."
                    )
                if strengths and evidence_strength not in strengths:
                    summary.errors.append(
                        f"{location}: Evidence Strength {evidence_strength!r} is not "
                        "allowed by CB-EVI-EVIDENCE-STRENGTH."
                    )
                if len(linked_driver_ids) != len(set(linked_driver_ids)):
                    summary.errors.append(
                        f"{location}: Linked Driver IDs contains duplicates."
                    )
                for driver_id in linked_driver_ids:
                    if driver_id not in drivers:
                        summary.errors.append(
                            f"{location}: Linked Driver ID {driver_id!r} does not exist."
                        )
                summary.linked_driver_references += len(linked_driver_ids)

                resolution_type, identifier, href, link_label = resolve_link(
                    source_id,
                    citation,
                    source_url,
                    summary,
                )
                derived_driver_ids = sorted(
                    driver_references.get(source_id, []),
                    key=lambda driver_id: driver_order.get(driver_id, sys.maxsize),
                )
                derived_relationship_ids = sorted(
                    relationship_references.get(source_id, [])
                )
                if set(linked_driver_ids) != set(derived_driver_ids):
                    summary.linked_driver_set_differences += 1

                record = {
                    "id": source_id or None,
                    "citationText": citation or None,
                    "year": year,
                    "evidenceType": evidence_type or None,
                    "evidenceStrength": evidence_strength or None,
                    "sourceUrl": source_url or None,
                    "resolutionType": resolution_type,
                    "resolvedIdentifier": identifier,
                    "href": href,
                    "linkLabel": link_label,
                    "driverIds": derived_driver_ids,
                    "relationshipIds": derived_relationship_ids,
                    "source": {
                        "workbook": path.name,
                        "worksheet": SHEET,
                        "row": row_number,
                    },
                }
                records.append(record)
        except Exception as exc:
            summary.errors.append(f"{path.name}: error while reading workbook: {exc}")
        finally:
            workbook.close()

    missing_layers = sorted(set(LAYERS) - summary.layers)
    if missing_layers:
        summary.errors.append(
            "Missing canonical layer workbook(s): " + ", ".join(missing_layers) + "."
        )
    return records


def validate_records(
    records: list[dict[str, Any]],
    driver_references: dict[str, list[str]],
    relationship_references: dict[str, list[str]],
    summary: Summary,
) -> None:
    ids = Counter(record["id"] for record in records)
    for source_id, count in sorted(ids.items(), key=lambda pair: str(pair[0])):
        if count > 1:
            summary.errors.append(
                f"Duplicate permanent Evidence ID {source_id!r} ({count} records)."
            )
    known = {source_id for source_id in ids if isinstance(source_id, str)}
    # Avoid hundreds of derivative "unknown ID" errors when an input workbook
    # could not be opened. Cross-dataset completeness is meaningful only after
    # all eight Evidence Library tables were read successfully.
    if summary.evidence_sheets == 8:
        for consumer, references in (
            ("Driver keySources", driver_references),
            ("Relationship supportingEvidenceIds", relationship_references),
        ):
            unknown = sorted(set(references) - known, key=source_sort_key)
            for source_id in unknown:
                summary.errors.append(
                    f"{consumer} references unknown Evidence ID {source_id!r}."
                )

    urls = Counter(
        record["sourceUrl"] for record in records if record.get("sourceUrl")
    )
    summary.duplicate_urls = sum(count > 1 for count in urls.values())
    if summary.duplicate_urls:
        summary.warnings.append(
            f"{summary.duplicate_urls} explicit Source URLs are reused by more than "
            "one Evidence record; records were preserved and not merged."
        )
    if summary.linked_driver_set_differences:
        summary.warnings.append(
            f"{summary.linked_driver_set_differences} Evidence Library Linked Driver "
            "ID sets differ from reverse Driver keySources references; public "
            "driverIds were derived from Driver keySources without rewriting the "
            "workbooks."
        )

    for record in records:
        source_id = record.get("id") or "unknown Evidence record"
        if tuple(record) != OUTPUT_KEYS:
            summary.errors.append(f"{source_id}: non-canonical output key structure.")
        resolution_type = record.get("resolutionType")
        if resolution_type not in RESOLUTION_TYPES:
            summary.errors.append(
                f"{source_id}: invalid resolutionType {resolution_type!r}."
            )
        href = record.get("href")
        if href is not None and not valid_http_url(href):
            summary.errors.append(f"{source_id}: generated href is not a valid URL.")
        if resolution_type == "DOI" and not record.get("resolvedIdentifier"):
            summary.errors.append(f"{source_id}: DOI resolution lacks its explicit DOI.")
        if resolution_type in {"DOI", "URL", "IDENTIFIER", "SEARCH"} and not href:
            summary.errors.append(
                f"{source_id}: {resolution_type} resolution lacks an href."
            )
        if resolution_type == "UNRESOLVED" and href is not None:
            summary.errors.append(f"{source_id}: unresolved record must have null href.")
        path_checked = {
            key: value
            for key, value in record.items()
            if key not in {"sourceUrl", "href", "source"}
        }
        if LOCAL_PATH.search(json.dumps(path_checked, ensure_ascii=False)):
            summary.errors.append(f"{source_id}: local/private path detected in output.")
        provenance = record.get("source") or {}
        workbook = provenance.get("workbook")
        if (
            not isinstance(workbook, str)
            or Path(workbook).name != workbook
            or provenance.get("worksheet") != SHEET
            or not isinstance(provenance.get("row"), int)
        ):
            summary.errors.append(f"{source_id}: invalid public provenance object.")
        summary.resolution_counts[resolution_type] += 1
    summary.source_records = len(records)


def write_atomically(payload: dict[str, Any], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    content = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            newline="\n",
            dir=output.parent,
            prefix=f".{output.name}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            handle.write(content)
            temporary = Path(handle.name)
        temporary.replace(output)
    finally:
        if temporary is not None and temporary.exists():
            temporary.unlink()


def resolve_argument(path: Path, root: Path) -> Path:
    return path if path.is_absolute() else root / path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--source-dir",
        type=Path,
        default=Path("source-data"),
        help="XLSX directory relative to the repository root (default: source-data).",
    )
    parser.add_argument(
        "--drivers",
        type=Path,
        default=Path("data/drivers.json"),
        help="Driver JSON relative to the repository root.",
    )
    parser.add_argument(
        "--relationships",
        type=Path,
        default=Path("data/relationships.json"),
        help="Relationship JSON relative to the repository root.",
    )
    parser.add_argument(
        "--codebook",
        type=Path,
        default=Path("data/codebook.json"),
        help="Codebook JSON relative to the repository root.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/sources.json"),
        help="Output JSON relative to the repository root.",
    )
    return parser.parse_args()


def print_summary(summary: Summary, output: Path, wrote: bool) -> None:
    print("Source Registry import statistics")
    print(f"  Schema: {SCHEMA_VERSION}")
    print(f"  Workbooks: {summary.workbooks}")
    print(f"  Evidence Library worksheets: {summary.evidence_sheets}")
    print(f"  Layers: {len(summary.layers)}")
    print(f"  Source records: {summary.source_records}")
    print(f"  Evidence Library linked-Driver references: {summary.linked_driver_references}")
    print(f"  Driver keySources references: {summary.driver_references}")
    print(f"  Relationship supporting-evidence references: {summary.relationship_references}")
    for resolution_type in ("DOI", "URL", "IDENTIFIER", "SEARCH", "UNRESOLVED"):
        print(
            f"  Resolution / {resolution_type}: "
            f"{summary.resolution_counts[resolution_type]}"
        )
    print(f"  Duplicate explicit URLs (not merged): {summary.duplicate_urls}")
    print(f"  Warnings: {len(summary.warnings)}")
    print(f"  Errors: {len(summary.errors)}")
    for warning in summary.warnings:
        print(f"WARNING: {warning}")
    for error in summary.errors:
        print(f"ERROR: {error}", file=sys.stderr)
    if wrote:
        print(f"Wrote validated data to {output}.")
    else:
        print(f"Did not modify {output}.", file=sys.stderr)


def main() -> int:
    root = Path(__file__).resolve().parent.parent
    args = parse_args()
    source_dir = resolve_argument(args.source_dir, root)
    drivers_path = resolve_argument(args.drivers, root)
    relationships_path = resolve_argument(args.relationships, root)
    codebook_path = resolve_argument(args.codebook, root)
    output = resolve_argument(args.output, root)
    summary = Summary()

    strengths = load_codebook_strengths(codebook_path, summary)
    drivers, driver_references, driver_order = load_drivers(drivers_path, summary)
    _relationships, relationship_references = load_relationships(
        relationships_path, summary
    )
    records = read_sources(
        source_dir,
        strengths,
        drivers,
        driver_references,
        driver_order,
        relationship_references,
        summary,
    )
    validate_records(records, driver_references, relationship_references, summary)
    records.sort(key=lambda record: source_sort_key(record["id"] or ""))
    payload = {"schemaVersion": SCHEMA_VERSION, "sources": records}

    wrote = False
    if not summary.errors:
        try:
            write_atomically(payload, output)
            wrote = True
        except OSError as exc:
            summary.errors.append(f"Could not replace {output}: {exc}")
    print_summary(summary, output, wrote)
    return 0 if wrote else 1


if __name__ == "__main__":
    raise SystemExit(main())
