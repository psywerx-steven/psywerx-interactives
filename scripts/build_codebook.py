"""Build data/codebook.json from identical canonical PSYWERX Codebooks."""

from __future__ import annotations

import json
import re
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCHEMA_VERSION = "1.0"
WORKBOOK_COUNT = 8
WORKSHEET_POPULATED_ROWS = 59
STRUCTURAL_ROWS = 3
TERM_RECORDS = 56
HEADER_ROW = 4
FIRST_TERM_ROW = 5
CODEBOOK_SHEET = "Codebook"
HEADERS = (
    "Codebook Term ID",
    "Sheet",
    "Field",
    "Definition",
    "Allowed Values / Format",
    "Required",
    "Rule / Guidance",
)
OUTPUT_KEYS = (
    "id",
    "sheet",
    "field",
    "definition",
    "allowedValuesOrFormat",
    "allowedValues",
    "required",
    "guidance",
    "source",
)
SHEET_ORDER = (
    "Layer Summary",
    "Drivers",
    "Families",
    "Evidence Library",
    "Relationships",
    "Operationalizations",
    "Cautions & Exclusions",
)
ID_PATTERN = re.compile(r"^CB-(?:LAY|DRV|FAM|EVI|REL|OPR|CAU)-[A-Z0-9]+(?:-[A-Z0-9]+)*$")
LOCAL_PATH_PATTERN = re.compile(r"(?:[A-Za-z]:[\\/]|\\\\)")

# These rows contain true controlled vocabularies rather than narrative formats.
CONTROLLED_TERM_IDS = {
    "CB-DRV-LAYER",
    "CB-DRV-DATA-TYPE",
    "CB-DRV-MODIFIABILITY",
    "CB-DRV-VOLATILITY",
    "CB-DRV-TIME-SCALE",
    "CB-DRV-ONSET-LAG",
    "CB-DRV-OBSERVABILITY",
    "CB-DRV-EVIDENCE-STRENGTH",
    "CB-EVI-EVIDENCE-STRENGTH",
    "CB-REL-EXPECTED-DIRECTION",
    "CB-REL-EVIDENCE-STRENGTH",
    "CB-REL-CAUSAL-ROLE",
    "CB-REL-POLARITY",
    "CB-REL-DIRECTNESS",
    "CB-REL-LEVEL",
    "CB-REL-LAG-PROFILE",
    "CB-REL-EXPOSURE-PATTERN",
    "CB-REL-CONFIDENCE",
    "CB-REL-GOVERNANCE-CLASS",
    "CB-OPR-METHOD-TYPE",
    "CB-CAU-DISPOSITION",
}
TEMPORAL_FORMAT_DIRECTIVE = "semicolon-separated canonical combinations"
DRIVER_VOCABULARIES = (
    ("layer", "CB-DRV-LAYER"),
    ("dataType", "CB-DRV-DATA-TYPE"),
    ("modifiability", "CB-DRV-MODIFIABILITY"),
    ("volatility", "CB-DRV-VOLATILITY"),
    ("timeScaleOfChange", "CB-DRV-TIME-SCALE"),
    ("onsetCausalLag", "CB-DRV-ONSET-LAG"),
    ("observability", "CB-DRV-OBSERVABILITY"),
    ("evidenceStrength", "CB-DRV-EVIDENCE-STRENGTH"),
)
TEMPORAL_FIELDS = {"timeScaleOfChange", "onsetCausalLag"}
EXCLUSIVE_TEMPORAL_VALUES = {
    "Mixed / Context-dependent",
    "Stable / Not applicable",
}


@dataclass
class VocabularyAudit:
    allowed: list[str]
    used: list[str]
    allowed_unused: list[str]
    used_undefined: list[str]


@dataclass
class Summary:
    workbooks: int = 0
    codebook_sheets: int = 0
    worksheet_populated_rows: int = 0
    structural_rows: int = 0
    term_records: int = 0
    duplicate_ids: int = 0
    missing_ids: int = 0
    driver_records: int = 0
    codebooks_identical: bool = False
    vocabularies: dict[str, VocabularyAudit] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def text(value: Any) -> str:
    """Trim surrounding whitespace while preserving source wording internally."""
    return "" if value is None else str(value).strip()


def row_values(row: tuple[Any, ...], width: int = len(HEADERS)) -> tuple[str, ...]:
    return tuple(text(row[index]) if index < len(row) else "" for index in range(width))


def populated_signature(sheet: Any) -> tuple[tuple[str, ...], ...]:
    return tuple(
        row_values(tuple(cell.value for cell in row))
        for row in sheet.iter_rows()
        if any(text(cell.value) for cell in row)
    )


def parse_allowed_values(identifier: str, source_text: str) -> list[str]:
    if identifier not in CONTROLLED_TERM_IDS:
        return []
    values = [text(part) for part in source_text.split(";") if text(part)]
    if identifier in {"CB-DRV-TIME-SCALE", "CB-DRV-ONSET-LAG"}:
        values = [
            value for value in values if value != TEMPORAL_FORMAT_DIRECTIVE
        ]
    return values


def build_entry(
    values: tuple[str, ...], row_number: int, path: Path, summary: Summary
) -> dict[str, Any]:
    location = f"{path.name} / {CODEBOOK_SHEET} / row {row_number}"
    identifier, sheet_name, field_name, definition, allowed, required, guidance = values
    if not identifier:
        summary.missing_ids += 1
        summary.errors.append(f"{location}: Codebook Term ID is missing.")
    elif not ID_PATTERN.fullmatch(identifier):
        summary.errors.append(
            f"{location}: Codebook Term ID {identifier!r} is not URL-safe/canonical."
        )
    required_values = {
        "Sheet": sheet_name,
        "Field": field_name,
        "Definition": definition,
        "Allowed Values / Format": allowed,
        "Required": required,
        "Rule / Guidance": guidance,
    }
    for label, value in required_values.items():
        if not value:
            summary.errors.append(f"{location}: required source field {label!r} is empty.")
    if sheet_name not in SHEET_ORDER:
        summary.errors.append(
            f"{location}: source Sheet classification {sheet_name!r} is not canonical."
        )
    if required not in {"Yes", "No"}:
        summary.errors.append(
            f"{location}: Required must be exactly 'Yes' or 'No'; found {required!r}."
        )
    allowed_values = parse_allowed_values(identifier, allowed)
    if identifier in CONTROLLED_TERM_IDS:
        if not allowed_values:
            summary.errors.append(
                f"{location}: controlled vocabulary has no parsed allowed values."
            )
        if len(allowed_values) != len(set(allowed_values)):
            summary.errors.append(
                f"{location}: controlled vocabulary repeats an allowed value."
            )
    return {
        "id": identifier or None,
        "sheet": sheet_name or None,
        "field": field_name or None,
        "definition": definition or None,
        "allowedValuesOrFormat": allowed or None,
        "allowedValues": allowed_values,
        "required": required == "Yes" if required in {"Yes", "No"} else None,
        "guidance": guidance or None,
        "source": {"sheet": CODEBOOK_SHEET, "row": row_number},
    }


def read_codebooks(
    source_dir: Path, summary: Summary
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    paths = sorted(
        path
        for path in source_dir.glob("*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    )
    summary.workbooks = len(paths)
    if len(paths) != WORKBOOK_COUNT:
        summary.errors.append(
            f"Expected {WORKBOOK_COUNT} XLSX workbooks; found {len(paths)}."
        )
        return [], {}

    signatures: dict[str, tuple[tuple[str, ...], ...]] = {}
    canonical_entries: list[dict[str, Any]] = []
    canonical_metadata: dict[str, Any] = {}
    for index, path in enumerate(paths):
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
        except Exception as exc:
            summary.errors.append(f"{path.name}: could not open workbook: {exc}")
            continue
        try:
            if CODEBOOK_SHEET not in workbook.sheetnames:
                summary.errors.append(f"{path.name}: missing Codebook worksheet.")
                continue
            sheet = workbook[CODEBOOK_SHEET]
            summary.codebook_sheets += 1
            signature = populated_signature(sheet)
            signatures[path.name] = signature
            header = row_values(
                tuple(cell.value for cell in sheet[HEADER_ROW]), len(HEADERS)
            )
            if header != HEADERS:
                summary.errors.append(
                    f"{path.name} / Codebook: expected exact headers {HEADERS}; "
                    f"found {header}."
                )
            if any(text(cell.value) for cell in sheet[3]):
                summary.errors.append(
                    f"{path.name} / Codebook: structural row 3 must remain blank."
                )

            semantic_rows: list[tuple[int, tuple[str, ...]]] = []
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=FIRST_TERM_ROW, values_only=True),
                start=FIRST_TERM_ROW,
            ):
                values = row_values(row)
                if any(values):
                    semantic_rows.append((row_number, values))
            worksheet_populated = len(signature)
            term_count = len(semantic_rows)
            structural_count = worksheet_populated - term_count
            counts = (worksheet_populated, structural_count, term_count)
            expected_counts = (
                WORKSHEET_POPULATED_ROWS,
                STRUCTURAL_ROWS,
                TERM_RECORDS,
            )
            if counts != expected_counts:
                summary.errors.append(
                    f"{path.name} / Codebook: expected worksheetPopulatedRows="
                    f"{WORKSHEET_POPULATED_ROWS}, structuralRows={STRUCTURAL_ROWS}, "
                    f"termRecords={TERM_RECORDS}; found {counts[0]}, {counts[1]}, "
                    f"{counts[2]}."
                )
            if index == 0:
                summary.worksheet_populated_rows = worksheet_populated
                summary.structural_rows = structural_count
                summary.term_records = term_count
                canonical_metadata = {
                    "title": text(sheet["A1"].value),
                    "governanceInstruction": text(sheet["A2"].value),
                    "worksheetPopulatedRows": worksheet_populated,
                    "structuralRows": structural_count,
                    "termRecords": term_count,
                }
                canonical_entries = [
                    build_entry(values, row_number, path, summary)
                    for row_number, values in semantic_rows
                ]
        except Exception as exc:
            summary.errors.append(f"{path.name}: error reading Codebook: {exc}")
        finally:
            workbook.close()

    summary.codebooks_identical = (
        len(signatures) == WORKBOOK_COUNT and len(set(signatures.values())) == 1
    )
    if not summary.codebooks_identical:
        summary.errors.append("Codebook worksheet content differs across workbooks.")
    return canonical_entries, canonical_metadata


def validate_entries(entries: list[dict[str, Any]], summary: Summary) -> None:
    if len(entries) != TERM_RECORDS:
        summary.errors.append(
            f"Expected {TERM_RECORDS} semantic Codebook entries; found {len(entries)}."
        )
    identifiers: dict[str, int] = {}
    for entry in entries:
        if tuple(entry) != OUTPUT_KEYS:
            summary.errors.append(
                f"Codebook entry {entry.get('id')!r} has non-canonical keys."
            )
        identifier = entry.get("id")
        if isinstance(identifier, str):
            identifiers[identifier] = identifiers.get(identifier, 0) + 1
    duplicate_ids = sorted(
        identifier for identifier, count in identifiers.items() if count > 1
    )
    summary.duplicate_ids = len(duplicate_ids)
    for identifier in duplicate_ids:
        summary.errors.append(f"Duplicate Codebook Term ID {identifier!r}.")
    missing_controlled = sorted(CONTROLLED_TERM_IDS - set(identifiers))
    if missing_controlled:
        summary.errors.append(
            "Missing controlled-vocabulary Codebook term(s): "
            + ", ".join(missing_controlled)
        )


def load_drivers(path: Path, summary: Summary) -> list[dict[str, Any]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        summary.errors.append(f"Could not load {path}: {exc}")
        return []
    if not isinstance(payload, list):
        summary.errors.append(f"{path}: expected a top-level JSON array.")
        return []
    drivers: list[dict[str, Any]] = []
    for index, value in enumerate(payload):
        if not isinstance(value, dict):
            summary.errors.append(f"{path}: Driver record {index} is not an object.")
            continue
        drivers.append(value)
    summary.driver_records = len(drivers)
    return drivers


def validate_temporal_values(
    driver_id: Any,
    field_name: str,
    values: Any,
    allowed: list[str],
    summary: Summary,
) -> list[str]:
    if not isinstance(values, list) or not values:
        summary.errors.append(
            f"Driver {driver_id!r}: {field_name} must be a non-empty array."
        )
        return []
    if any(not isinstance(value, str) or not value for value in values):
        summary.errors.append(
            f"Driver {driver_id!r}: {field_name} contains a non-string/empty value."
        )
        return [value for value in values if isinstance(value, str) and value]
    if len(values) != len(set(values)):
        summary.errors.append(
            f"Driver {driver_id!r}: {field_name} contains duplicate values."
        )
    if EXCLUSIVE_TEMPORAL_VALUES.intersection(values) and len(values) != 1:
        summary.errors.append(
            f"Driver {driver_id!r}: {field_name} combines an exclusive value."
        )
    known = [value for value in values if value in allowed]
    if known != sorted(known, key=allowed.index):
        summary.errors.append(
            f"Driver {driver_id!r}: {field_name} values are out of canonical order."
        )
    return values


def validate_driver_vocabularies(
    entries: list[dict[str, Any]],
    drivers: list[dict[str, Any]],
    summary: Summary,
) -> None:
    entries_by_id = {entry.get("id"): entry for entry in entries}
    for field_name, term_id in DRIVER_VOCABULARIES:
        entry = entries_by_id.get(term_id)
        if entry is None:
            continue
        allowed = entry["allowedValues"]
        used_values: set[str] = set()
        for index, driver in enumerate(drivers):
            driver_id = driver.get("id", f"record {index}")
            value = driver.get(field_name)
            if field_name in TEMPORAL_FIELDS:
                values = validate_temporal_values(
                    driver_id, field_name, value, allowed, summary
                )
                used_values.update(
                    item for item in values if isinstance(item, str) and item
                )
            elif not isinstance(value, str) or not value:
                summary.errors.append(
                    f"Driver {driver_id!r}: {field_name} must be a non-empty string."
                )
            else:
                used_values.add(value)
        used = [value for value in allowed if value in used_values]
        undefined = sorted(used_values - set(allowed))
        unused = [value for value in allowed if value not in used_values]
        summary.vocabularies[field_name] = VocabularyAudit(
            allowed=allowed,
            used=used + undefined,
            allowed_unused=unused,
            used_undefined=undefined,
        )
        if undefined:
            summary.errors.append(
                f"{field_name}: used-but-undefined controlled value(s): "
                + ", ".join(repr(value) for value in undefined)
            )


def contains_local_path(value: Any) -> bool:
    if isinstance(value, str):
        return bool(LOCAL_PATH_PATTERN.search(value))
    if isinstance(value, list):
        return any(contains_local_path(item) for item in value)
    if isinstance(value, dict):
        return any(contains_local_path(item) for item in value.values())
    return False


def entry_sort_key(entry: dict[str, Any]) -> tuple[int, int, str]:
    sheet_name = entry.get("sheet")
    sheet_index = (
        SHEET_ORDER.index(sheet_name)
        if sheet_name in SHEET_ORDER
        else len(SHEET_ORDER)
    )
    row = entry.get("source", {}).get("row", 0)
    return sheet_index, row, str(entry.get("id") or "")


def write_atomically(payload: dict[str, Any], output: Path) -> None:
    content = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    temporary: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            newline="\n",
            dir=output.parent,
            prefix=f".{output.name}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            handle.write(content)
            temporary = Path(handle.name)
        temporary.replace(output)
    finally:
        if temporary is not None and temporary.exists():
            temporary.unlink()


def print_summary(summary: Summary, output: Path, wrote: bool) -> None:
    print("Codebook import statistics")
    print(f"  Workbooks: {summary.workbooks}")
    print(f"  Codebook worksheets: {summary.codebook_sheets}")
    print(
        "  Codebooks identical: "
        + ("yes" if summary.codebooks_identical else "no")
    )
    print(f"  worksheetPopulatedRows: {summary.worksheet_populated_rows}")
    print(f"  structuralRows: {summary.structural_rows}")
    print(f"  termRecords: {summary.term_records}")
    print(f"  Permanent Codebook Term IDs: {summary.term_records - summary.missing_ids}")
    print(f"  Duplicate IDs: {summary.duplicate_ids}")
    print(f"  Missing IDs: {summary.missing_ids}")
    print(f"  Driver records validated: {summary.driver_records}")
    print("Controlled vocabulary validation")
    for field_name, audit in summary.vocabularies.items():
        print(f"  {field_name}")
        print("    Allowed: " + "; ".join(audit.allowed))
        print("    Used: " + ("; ".join(audit.used) or "(none)"))
        print(
            "    Allowed but unused: "
            + ("; ".join(audit.allowed_unused) or "(none)")
        )
        print(
            "    Used but undefined: "
            + ("; ".join(audit.used_undefined) or "(none)")
        )
    print(f"  Warnings: {len(summary.warnings)}")
    print(f"  Errors: {len(summary.errors)}")
    for warning in summary.warnings:
        print(f"WARNING: {warning}")
    for error in summary.errors:
        print(f"ERROR: {error}", file=sys.stderr)
    if wrote:
        print(f"Wrote validated data to {output}.")
    else:
        print(f"Did not modify {output}.", file=sys.stderr)


def main() -> int:
    root = Path(__file__).resolve().parent.parent
    output = root / "data" / "codebook.json"
    summary = Summary()
    entries, metadata = read_codebooks(root / "source-data", summary)
    validate_entries(entries, summary)
    drivers = load_drivers(root / "data" / "drivers.json", summary)
    validate_driver_vocabularies(entries, drivers, summary)
    entries.sort(key=entry_sort_key)
    payload = {
        "schemaVersion": SCHEMA_VERSION,
        "metadata": metadata,
        "entries": entries,
    }
    if contains_local_path(payload):
        summary.errors.append("Generated Codebook payload contains a local path.")

    wrote = False
    if not summary.errors:
        try:
            write_atomically(payload, output)
            wrote = True
        except OSError as exc:
            summary.errors.append(f"Could not replace {output}: {exc}")
    print_summary(summary, output, wrote)
    return 0 if wrote else 1


if __name__ == "__main__":
    raise SystemExit(main())
