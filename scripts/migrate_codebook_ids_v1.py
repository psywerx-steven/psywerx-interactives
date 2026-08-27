"""Add permanent Codebook Term IDs to all canonical PSYWERX workbooks.

The migration treats rows 1, 2, and 4 as structural worksheet rows. Only the
48 semantic records in rows 5-52 receive IDs. All workbooks are staged and
validated before any source workbook is replaced.
"""

from __future__ import annotations

import hashlib
import os
import re
import shutil
import sys
import tempfile
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


WORKBOOK_COUNT = 8
WORKSHEET_POPULATED_ROWS = 51
STRUCTURAL_ROWS = 3
TERM_RECORDS = 48
HEADER_ROW = 4
FIRST_TERM_ROW = 5
CODEBOOK_SHEET = "Codebook"
ID_HEADER = "Codebook Term ID"
SOURCE_HEADERS = (
    "Sheet",
    "Field",
    "Definition",
    "Allowed Values / Format",
    "Required",
    "Rule / Guidance",
)
CANONICAL_HEADERS = (ID_HEADER, *SOURCE_HEADERS)
SOURCE_TABLE_REF = "A4:F50"
CANONICAL_TABLE_REF = "A4:G50"
ID_PATTERN = re.compile(r"^CB-(?:LAY|DRV|FAM|EVI|REL|OPR|CAU)-[A-Z0-9]+(?:-[A-Z0-9]+)*$")

# Permanent semantic identities. Never derive these IDs from row positions.
TERM_ID_ASSIGNMENTS = (
    ("Layer Summary", "Layer definition", "CB-LAY-DEFINITION"),
    ("Layer Summary", "Scope", "CB-LAY-SCOPE"),
    ("Layer Summary", "Boundary with other layers", "CB-LAY-BOUNDARY"),
    (
        "Layer Summary",
        "Ontology Standard Version",
        "CB-LAY-ONTOLOGY-STANDARD-VERSION",
    ),
    ("Layer Summary", "Layer Taxonomy Version", "CB-LAY-TAXONOMY-VERSION"),
    (
        "Layer Summary",
        "Creation Date / Last Review Date",
        "CB-LAY-REVIEW-DATES",
    ),
    ("Layer Summary", "Known gaps", "CB-LAY-KNOWN-GAPS"),
    ("Drivers", "ID", "CB-DRV-ID"),
    ("Drivers", "Name", "CB-DRV-NAME"),
    ("Drivers", "Other Names / Aliases", "CB-DRV-ALIASES"),
    ("Drivers", "Layer", "CB-DRV-LAYER"),
    ("Drivers", "Family", "CB-DRV-FAMILY"),
    ("Drivers", "Definition", "CB-DRV-DEFINITION"),
    ("Drivers", "Data Type", "CB-DRV-DATA-TYPE"),
    ("Drivers", "Representation / Scale", "CB-DRV-REPRESENTATION-SCALE"),
    (
        "Drivers",
        "Polarity / Direction of Interpretation",
        "CB-DRV-POLARITY-DIRECTION",
    ),
    ("Drivers", "Mechanism", "CB-DRV-MECHANISM"),
    (
        "Drivers",
        "Likely Upstream Influences",
        "CB-DRV-UPSTREAM-INFLUENCES",
    ),
    (
        "Drivers",
        "Likely Downstream Influences",
        "CB-DRV-DOWNSTREAM-INFLUENCES",
    ),
    (
        "Drivers",
        "Moderators / Boundary Conditions",
        "CB-DRV-MODERATORS",
    ),
    (
        "Drivers",
        "Typical Interaction Candidates",
        "CB-DRV-INTERACTION-CANDIDATES",
    ),
    (
        "Drivers",
        "Modifiability / Malleability",
        "CB-DRV-MODIFIABILITY",
    ),
    ("Drivers", "Volatility", "CB-DRV-VOLATILITY"),
    ("Drivers", "Time Scale of Change", "CB-DRV-TIME-SCALE"),
    (
        "Drivers",
        "Time Scale Qualifier",
        "CB-DRV-TIME-SCALE-QUALIFIER",
    ),
    ("Drivers", "Onset / Causal Lag", "CB-DRV-ONSET-LAG"),
    (
        "Drivers",
        "Persistence / Recovery",
        "CB-DRV-PERSISTENCE-RECOVERY",
    ),
    ("Drivers", "Indicators", "CB-DRV-INDICATORS"),
    (
        "Drivers",
        "Measurement / Assessment Methods",
        "CB-DRV-MEASUREMENT-METHODS",
    ),
    ("Drivers", "Observability", "CB-DRV-OBSERVABILITY"),
    (
        "Drivers",
        "Measurement Caveats",
        "CB-DRV-MEASUREMENT-CAVEATS",
    ),
    (
        "Drivers",
        "Evidence Strength",
        "CB-DRV-EVIDENCE-STRENGTH",
    ),
    ("Drivers", "Evidence Notes", "CB-DRV-EVIDENCE-NOTES"),
    (
        "Drivers",
        "Common Misinterpretations",
        "CB-DRV-MISINTERPRETATIONS",
    ),
    ("Drivers", "Key Sources", "CB-DRV-KEY-SOURCES"),
    ("Families", "Family ID", "CB-FAM-ID"),
    ("Families", "Family Name", "CB-FAM-NAME"),
    ("Families", "Layer", "CB-FAM-LAYER"),
    ("Families", "Definition", "CB-FAM-DEFINITION"),
    ("Families", "Inclusion Rule", "CB-FAM-INCLUSION-RULE"),
    (
        "Families",
        "Exclusion / Boundary Rule",
        "CB-FAM-EXCLUSION-RULE",
    ),
    (
        "Families",
        "Representative Drivers",
        "CB-FAM-REPRESENTATIVE-DRIVERS",
    ),
    ("Families", "Driver Count", "CB-FAM-DRIVER-COUNT"),
    (
        "Evidence Library",
        "Evidence Strength",
        "CB-EVI-EVIDENCE-STRENGTH",
    ),
    (
        "Relationships",
        "Expected Direction",
        "CB-REL-EXPECTED-DIRECTION",
    ),
    (
        "Relationships",
        "Evidence Strength",
        "CB-REL-EVIDENCE-STRENGTH",
    ),
    ("Operationalizations", "Method Type", "CB-OPR-METHOD-TYPE"),
    ("Cautions & Exclusions", "Disposition", "CB-CAU-DISPOSITION"),
)


@dataclass(frozen=True)
class Baseline:
    sheet_hashes: dict[str, str]
    full_codebook_signature: tuple[tuple[str, ...], ...]
    substantive_signature: tuple[tuple[str, ...], ...]
    worksheet_populated_rows: int
    structural_rows: int
    term_records: int


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def hash_sheet_values(sheet: Any) -> str:
    digest = hashlib.sha256(sheet.title.encode("utf-8"))
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            payload = (cell.coordinate, cell.data_type, repr(cell.value))
            digest.update(repr(payload).encode("utf-8"))
    digest.update(
        repr(sorted(str(item) for item in sheet.merged_cells.ranges)).encode("utf-8")
    )
    return digest.hexdigest()


def populated_rows(sheet: Any) -> tuple[tuple[str, ...], ...]:
    return tuple(
        tuple(text(cell.value) for cell in row)
        for row in sheet.iter_rows()
        if any(text(cell.value) for cell in row)
    )


def header_values(sheet: Any) -> tuple[str, ...]:
    return tuple(text(cell.value) for cell in sheet[HEADER_ROW] if text(cell.value))


def term_rows(sheet: Any) -> tuple[tuple[str, ...], ...]:
    headers = header_values(sheet)
    if headers == SOURCE_HEADERS:
        start_column = 1
    elif headers == CANONICAL_HEADERS:
        start_column = 2
    else:
        raise ValueError(f"Unexpected Codebook headers: {headers}")
    rows: list[tuple[str, ...]] = []
    for row in sheet.iter_rows(min_row=FIRST_TERM_ROW):
        values = tuple(
            text(row[column - 1].value)
            for column in range(start_column, start_column + len(SOURCE_HEADERS))
        )
        if any(values):
            rows.append(values)
    return tuple(rows)


def diagnostic_counts(sheet: Any) -> tuple[int, int, int]:
    populated = len(populated_rows(sheet))
    terms = len(term_rows(sheet))
    structural = populated - terms
    return populated, structural, terms


def validate_assignment_map() -> None:
    if len(TERM_ID_ASSIGNMENTS) != TERM_RECORDS:
        raise ValueError(
            f"Expected {TERM_RECORDS} ID assignments; found {len(TERM_ID_ASSIGNMENTS)}"
        )
    keys = [(sheet, field) for sheet, field, _ in TERM_ID_ASSIGNMENTS]
    ids = [identifier for _, _, identifier in TERM_ID_ASSIGNMENTS]
    if len(keys) != len(set(keys)):
        raise ValueError("The permanent ID map contains duplicate Sheet/Field keys")
    if len(ids) != len(set(ids)):
        raise ValueError("The permanent ID map contains duplicate IDs")
    invalid = [identifier for identifier in ids if not ID_PATTERN.fullmatch(identifier)]
    if invalid:
        raise ValueError(f"The permanent ID map contains invalid IDs: {invalid}")


def collect_baseline(workbook: Any, path: Path) -> Baseline:
    if CODEBOOK_SHEET not in workbook.sheetnames:
        raise ValueError(f"{path.name}: missing Codebook worksheet")
    sheet = workbook[CODEBOOK_SHEET]
    full_signature = populated_rows(sheet)
    substantive = term_rows(sheet)
    counts = diagnostic_counts(sheet)
    expected_keys = tuple((sheet_name, field) for sheet_name, field, _ in TERM_ID_ASSIGNMENTS)
    actual_keys = tuple((row[0], row[1]) for row in substantive)
    if actual_keys != expected_keys:
        raise ValueError(
            f"{path.name}: semantic Codebook rows do not match the permanent ID map"
        )
    if counts != (WORKSHEET_POPULATED_ROWS, STRUCTURAL_ROWS, TERM_RECORDS):
        raise ValueError(
            f"{path.name}: expected worksheetPopulatedRows={WORKSHEET_POPULATED_ROWS}, "
            f"structuralRows={STRUCTURAL_ROWS}, termRecords={TERM_RECORDS}; found "
            f"{counts[0]}, {counts[1]}, {counts[2]}"
        )
    return Baseline(
        sheet_hashes={
            name: hash_sheet_values(workbook[name])
            for name in workbook.sheetnames
            if name != CODEBOOK_SHEET
        },
        full_codebook_signature=full_signature,
        substantive_signature=substantive,
        worksheet_populated_rows=counts[0],
        structural_rows=counts[1],
        term_records=counts[2],
    )


def copy_cell_style(source: Any, target: Any) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def migrate_sheet(sheet: Any) -> None:
    headers = header_values(sheet)
    if headers == SOURCE_HEADERS:
        title = sheet["A1"].value
        instruction = sheet["A2"].value
        old_widths = {
            column: sheet.column_dimensions[get_column_letter(column)].width
            for column in range(1, len(SOURCE_HEADERS) + 1)
        }
        for merged in ("A1:F1", "A2:F2"):
            if merged not in {str(item) for item in sheet.merged_cells.ranges}:
                raise ValueError(f"Codebook is missing expected merged range {merged}")
            sheet.unmerge_cells(merged)
        sheet.insert_cols(1, 1)
        sheet["A1"] = title
        sheet["B1"] = None
        sheet["A2"] = instruction
        sheet["B2"] = None
        sheet.merge_cells("A1:G1")
        sheet.merge_cells("A2:G2")
        for old_column, width in old_widths.items():
            sheet.column_dimensions[get_column_letter(old_column + 1)].width = width
        sheet.column_dimensions["A"].width = 32.0
        copy_cell_style(sheet["B4"], sheet["A4"])
        sheet["A4"] = ID_HEADER
        for row_number in range(FIRST_TERM_ROW, sheet.max_row + 1):
            copy_cell_style(sheet.cell(row_number, 2), sheet.cell(row_number, 1))
        for table in sheet.tables.values():
            if table.ref == SOURCE_TABLE_REF:
                table.ref = CANONICAL_TABLE_REF
            else:
                raise ValueError(f"Unexpected Codebook table range {table.ref}")
    elif headers != CANONICAL_HEADERS:
        raise ValueError(f"Codebook has unexpected headers: {headers}")

    rows = term_rows(sheet)
    row_numbers = [
        row_number
        for row_number in range(FIRST_TERM_ROW, sheet.max_row + 1)
        if any(text(cell.value) for cell in sheet[row_number][1:])
    ]
    if len(rows) != TERM_RECORDS or len(row_numbers) != TERM_RECORDS:
        raise ValueError("Codebook does not contain exactly 48 semantic term rows")
    for row_number, row, (sheet_name, field, identifier) in zip(
        row_numbers, rows, TERM_ID_ASSIGNMENTS
    ):
        if (row[0], row[1]) != (sheet_name, field):
            raise ValueError(
                f"Codebook row {row_number} does not match ID assignment {identifier}"
            )
        existing = text(sheet.cell(row_number, 1).value)
        if existing and existing != identifier:
            raise ValueError(
                f"Codebook row {row_number} has conflicting ID {existing!r}"
            )
        sheet.cell(row_number, 1).value = identifier


def validate_staged(path: Path, baseline: Baseline) -> tuple[tuple[str, ...], ...]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        for name, expected_hash in baseline.sheet_hashes.items():
            if hash_sheet_values(workbook[name]) != expected_hash:
                raise ValueError(f"{path.name}: unrelated worksheet {name!r} changed")
        sheet = workbook[CODEBOOK_SHEET]
        if header_values(sheet) != CANONICAL_HEADERS:
            raise ValueError(f"{path.name}: canonical Codebook header was not created")
        if term_rows(sheet) != baseline.substantive_signature:
            raise ValueError(f"{path.name}: substantive Codebook content changed")
        counts = diagnostic_counts(sheet)
        if counts != (WORKSHEET_POPULATED_ROWS, STRUCTURAL_ROWS, TERM_RECORDS):
            raise ValueError(f"{path.name}: staged row diagnostics are invalid: {counts}")
        expected_ids = tuple(identifier for _, _, identifier in TERM_ID_ASSIGNMENTS)
        actual_ids = tuple(
            text(row[0].value)
            for row in sheet.iter_rows(min_row=FIRST_TERM_ROW)
            if any(text(cell.value) for cell in row[1:])
        )
        if actual_ids != expected_ids:
            raise ValueError(f"{path.name}: staged Codebook IDs do not match the map")
        if len(actual_ids) != len(set(actual_ids)):
            raise ValueError(f"{path.name}: staged Codebook IDs are not unique")
        if sorted(str(item) for item in sheet.merged_cells.ranges) != [
            "A1:G1",
            "A2:G2",
        ]:
            raise ValueError(f"{path.name}: staged Codebook merged ranges are invalid")
        table_refs = {table.ref for table in sheet.tables.values()}
        if table_refs != {CANONICAL_TABLE_REF}:
            raise ValueError(f"{path.name}: staged Codebook table range is invalid")
        return populated_rows(sheet)
    finally:
        workbook.close()


def cleanup_staging(directory: Path) -> None:
    resolved = directory.resolve()
    source_root = directory.parent.resolve()
    if resolved.parent != source_root or not resolved.name.startswith(
        ".codebook-ids-v1-"
    ):
        raise RuntimeError(f"Refusing to clean unexpected staging path: {resolved}")
    for child in resolved.iterdir():
        if not child.is_file():
            raise RuntimeError(f"Unexpected directory in staging area: {child}")
        child.unlink()
    resolved.rmdir()


def main() -> int:
    root = Path(__file__).resolve().parent.parent
    source_dir = root / "source-data"
    paths = sorted(
        path
        for path in source_dir.glob("*.xlsx")
        if not path.name.startswith("~$")
    )
    if len(paths) != WORKBOOK_COUNT:
        print(
            f"ERROR: expected {WORKBOOK_COUNT} XLSX workbooks; found {len(paths)}",
            file=sys.stderr,
        )
        return 1

    staging: Path | None = None
    try:
        validate_assignment_map()
        baselines: dict[str, Baseline] = {}
        for path in paths:
            workbook = load_workbook(path, read_only=False, data_only=False)
            try:
                baselines[path.name] = collect_baseline(workbook, path)
            finally:
                workbook.close()

        signatures = {
            baseline.full_codebook_signature for baseline in baselines.values()
        }
        if len(signatures) != 1:
            raise ValueError("Codebook worksheets are not identical before migration")

        staging = Path(
            tempfile.mkdtemp(prefix=".codebook-ids-v1-", dir=source_dir)
        )
        for path in paths:
            workbook = load_workbook(path, read_only=False, data_only=False)
            try:
                migrate_sheet(workbook[CODEBOOK_SHEET])
                workbook.save(staging / path.name)
            finally:
                workbook.close()

        staged_signatures = {
            validate_staged(staging / path.name, baselines[path.name])
            for path in paths
        }
        if len(staged_signatures) != 1:
            raise ValueError("Staged Codebook worksheets are not identical")

        backups: dict[Path, Path] = {}
        for path in paths:
            backup = staging / f"backup-{path.name}"
            shutil.copy2(path, backup)
            backups[path] = backup
        try:
            for path in paths:
                os.replace(staging / path.name, path)
        except Exception:
            for path, backup in backups.items():
                if backup.exists():
                    os.replace(backup, path)
            raise

        print("Codebook Term ID migration complete")
        print(f"  Workbooks: {len(paths)}")
        print(f"  Codebook worksheets identical: yes")
        print(f"  worksheetPopulatedRows: {WORKSHEET_POPULATED_ROWS}")
        print(f"  structuralRows: {STRUCTURAL_ROWS}")
        print(f"  termRecords: {TERM_RECORDS}")
        print(f"  Permanent Codebook Term IDs: {len(TERM_ID_ASSIGNMENTS)}")
        print("  Duplicate IDs: 0")
        print("  Missing IDs: 0")
        print("  Unrelated worksheet changes: 0")
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        print("No source workbook was replaced.", file=sys.stderr)
        return 1
    finally:
        if staging is not None and staging.exists():
            cleanup_staging(staging)


if __name__ == "__main__":
    raise SystemExit(main())
